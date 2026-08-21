import io
import json
import os
import secrets
import sqlite3
import urllib.error
import urllib.request
from datetime import date
from functools import wraps

from flask import (
    Blueprint, Response, current_app, jsonify, redirect, render_template,
    request, send_file, session, url_for,
)
from openpyxl import Workbook, load_workbook

from .database import get_db

try:
    from pywebpush import WebPushException, webpush
except Exception:  # pragma: no cover - 배포 설정 전에도 관리화면은 동작한다.
    WebPushException = Exception
    webpush = None


parent_notification_bp = Blueprint('parent_notifications', __name__)

NOTICE_KINDS = {
    '출석': ('출석 알림', '{student} 학생이 수업에 출석했습니다.'),
    '지각': ('지각 알림', '{student} 학생이 수업에 지각했습니다.'),
    '결석': ('결석 알림', '{student} 학생이 수업에 결석했습니다.'),
    '하원': ('하원 알림', '{student} 학생이 안전하게 하원했습니다.'),
    '휴강': ('휴강 안내', '수업 휴강 안내입니다.'),
    '보강': ('보강 안내', '수업 보강 일정을 확인해주세요.'),
    '준비물': ('준비물 안내', '다음 수업 준비물을 확인해주세요.'),
    '시간변경': ('수업시간 변경', '수업시간이 변경되었습니다.'),
    '강사공지': ('강사 공지', '강사 공지가 도착했습니다.'),
    '수강료': ('수강료 안내', '수강료 안내를 확인해주세요.'),
    '만족도조사': ('만족도조사', '방과후학교 만족도조사에 참여해주세요.'),
    '일반': ('새담 방과후학교', '새로운 안내가 도착했습니다.'),
}
ATTENDANCE_KINDS = {'출석', '지각', '결석', '하원'}


def ensure_parent_notification_schema(conn=None):
    owns_connection = conn is None
    if owns_connection:
        conn = get_db()
    conn.executescript('''
        CREATE TABLE IF NOT EXISTS parent_guardians (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            phone TEXT NOT NULL,
            email TEXT,
            memo TEXT,
            is_active INTEGER NOT NULL DEFAULT 1,
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            updated_at DATETIME DEFAULT CURRENT_TIMESTAMP
        );
        CREATE INDEX IF NOT EXISTS idx_parent_guardians_phone
            ON parent_guardians(phone);

        CREATE TABLE IF NOT EXISTS parent_students (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            school_id INTEGER,
            school_name TEXT,
            grade TEXT,
            classroom TEXT,
            memo TEXT,
            is_active INTEGER NOT NULL DEFAULT 1,
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            updated_at DATETIME DEFAULT CURRENT_TIMESTAMP
        );

        CREATE TABLE IF NOT EXISTS parent_guardian_students (
            guardian_id INTEGER NOT NULL,
            student_id INTEGER NOT NULL,
            relationship TEXT DEFAULT '보호자',
            PRIMARY KEY (guardian_id, student_id),
            FOREIGN KEY (guardian_id) REFERENCES parent_guardians(id) ON DELETE CASCADE,
            FOREIGN KEY (student_id) REFERENCES parent_students(id) ON DELETE CASCADE
        );

        CREATE TABLE IF NOT EXISTS parent_classes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            school_id INTEGER,
            school_name TEXT NOT NULL,
            department TEXT,
            class_name TEXT NOT NULL,
            instructor_emp_no TEXT,
            instructor_name TEXT,
            access_token TEXT NOT NULL UNIQUE,
            is_active INTEGER NOT NULL DEFAULT 1,
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            updated_at DATETIME DEFAULT CURRENT_TIMESTAMP
        );
        CREATE INDEX IF NOT EXISTS idx_parent_classes_school
            ON parent_classes(school_name, department, class_name);

        CREATE TABLE IF NOT EXISTS parent_class_students (
            class_id INTEGER NOT NULL,
            student_id INTEGER NOT NULL,
            PRIMARY KEY (class_id, student_id),
            FOREIGN KEY (class_id) REFERENCES parent_classes(id) ON DELETE CASCADE,
            FOREIGN KEY (student_id) REFERENCES parent_students(id) ON DELETE CASCADE
        );

        CREATE TABLE IF NOT EXISTS parent_invites (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            guardian_id INTEGER NOT NULL,
            token TEXT NOT NULL UNIQUE,
            is_active INTEGER NOT NULL DEFAULT 1,
            sms_sent_at DATETIME,
            sms_send_count INTEGER NOT NULL DEFAULT 0,
            registered_at DATETIME,
            last_opened_at DATETIME,
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (guardian_id) REFERENCES parent_guardians(id) ON DELETE CASCADE
        );
        CREATE INDEX IF NOT EXISTS idx_parent_invites_guardian
            ON parent_invites(guardian_id, is_active);

        CREATE TABLE IF NOT EXISTS parent_push_subscriptions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            guardian_id INTEGER NOT NULL,
            endpoint TEXT NOT NULL UNIQUE,
            p256dh TEXT NOT NULL,
            auth TEXT NOT NULL,
            user_agent TEXT,
            is_active INTEGER NOT NULL DEFAULT 1,
            failure_count INTEGER NOT NULL DEFAULT 0,
            last_success_at DATETIME,
            last_error TEXT,
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            updated_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (guardian_id) REFERENCES parent_guardians(id) ON DELETE CASCADE
        );
        CREATE INDEX IF NOT EXISTS idx_parent_push_guardian
            ON parent_push_subscriptions(guardian_id, is_active);

        CREATE TABLE IF NOT EXISTS parent_notifications (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            kind TEXT NOT NULL,
            title TEXT NOT NULL,
            body TEXT NOT NULL,
            class_id INTEGER,
            student_id INTEGER,
            target_type TEXT,
            created_by TEXT,
            total_count INTEGER NOT NULL DEFAULT 0,
            sent_count INTEGER NOT NULL DEFAULT 0,
            failed_count INTEGER NOT NULL DEFAULT 0,
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (class_id) REFERENCES parent_classes(id) ON DELETE SET NULL,
            FOREIGN KEY (student_id) REFERENCES parent_students(id) ON DELETE SET NULL
        );

        CREATE TABLE IF NOT EXISTS parent_notification_recipients (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            notification_id INTEGER NOT NULL,
            guardian_id INTEGER NOT NULL,
            student_id INTEGER,
            status TEXT NOT NULL,
            subscription_count INTEGER NOT NULL DEFAULT 0,
            sent_count INTEGER NOT NULL DEFAULT 0,
            error_message TEXT,
            sent_at DATETIME,
            FOREIGN KEY (notification_id) REFERENCES parent_notifications(id) ON DELETE CASCADE,
            FOREIGN KEY (guardian_id) REFERENCES parent_guardians(id) ON DELETE CASCADE,
            FOREIGN KEY (student_id) REFERENCES parent_students(id) ON DELETE SET NULL
        );

        CREATE TABLE IF NOT EXISTS parent_attendance (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            class_id INTEGER NOT NULL,
            student_id INTEGER NOT NULL,
            status TEXT NOT NULL,
            event_date TEXT NOT NULL,
            note TEXT,
            recorded_by TEXT,
            notification_id INTEGER,
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (class_id) REFERENCES parent_classes(id) ON DELETE CASCADE,
            FOREIGN KEY (student_id) REFERENCES parent_students(id) ON DELETE CASCADE,
            FOREIGN KEY (notification_id) REFERENCES parent_notifications(id) ON DELETE SET NULL
        );
    ''')
    conn.commit()
    if owns_connection:
        conn.close()


def _level():
    try:
        return int(session.get('user_level', 99))
    except (TypeError, ValueError):
        return 99


def _is_manager():
    return session.get('user_name') == 'admin' or (
        bool(session.get('emp_no')) and 1 <= _level() <= 7
    )


def manager_required(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        if not _is_manager():
            if request.path.startswith('/parent-notifications/api/'):
                return jsonify(ok=False, message='학부모알림 관리 권한이 없습니다.'), 403
            return '학부모알림 관리 권한이 없습니다.', 403
        return view(*args, **kwargs)
    return wrapped


def _text(value, limit=500):
    return str(value or '').strip()[:limit]


def _phone(value):
    return ''.join(ch for ch in str(value or '') if ch.isdigit())[:20]


def _token():
    return secrets.token_urlsafe(32)


def _base_url():
    return str(os.getenv('PUBLIC_BASE_URL') or request.url_root).rstrip('/')


def _push_config():
    return {
        'public_key': _text(os.getenv('VAPID_PUBLIC_KEY'), 500),
        'private_key': _text(os.getenv('VAPID_PRIVATE_KEY'), 2000),
        'subject': _text(os.getenv('VAPID_SUBJECT'), 500),
    }


def _push_ready():
    cfg = _push_config()
    return bool(webpush and cfg['public_key'] and cfg['private_key'] and cfg['subject'])


def _active_invite(conn, guardian_id):
    row = conn.execute('''
        SELECT * FROM parent_invites
        WHERE guardian_id=? AND is_active=1
        ORDER BY id DESC LIMIT 1
    ''', (guardian_id,)).fetchone()
    if row:
        return row
    token = _token()
    cursor = conn.execute(
        'INSERT INTO parent_invites(guardian_id, token) VALUES (?, ?)',
        (guardian_id, token),
    )
    conn.commit()
    return conn.execute('SELECT * FROM parent_invites WHERE id=?', (cursor.lastrowid,)).fetchone()


def _invite_url(conn, guardian_id):
    invite = _active_invite(conn, guardian_id)
    return f"{_base_url()}/parent/register/{invite['token']}"


def _invite_message(url):
    return (
        '[새담 방과후학교]\n'
        '앞으로 출결 알림이나 방과후 안내를 무료 알림으로 받아보실 수 있습니다.\n'
        '아래에서 알림을 등록해주세요.\n'
        f'{url}'
    )


def _send_sms_webhook(phone, message):
    endpoint = _text(os.getenv('PARENT_SMS_WEBHOOK_URL'), 2000)
    if not endpoint:
        return False, 'SMS 연동값이 없어 문자 문구만 생성했습니다.'
    payload = json.dumps({
        'to': phone,
        'from': _phone(os.getenv('PARENT_SMS_SENDER')),
        'text': message,
    }, ensure_ascii=False).encode('utf-8')
    headers = {'Content-Type': 'application/json; charset=utf-8'}
    token = _text(os.getenv('PARENT_SMS_WEBHOOK_TOKEN'), 2000)
    if token:
        headers['Authorization'] = f'Bearer {token}'
    req = urllib.request.Request(endpoint, data=payload, headers=headers, method='POST')
    try:
        with urllib.request.urlopen(req, timeout=10) as response:
            if 200 <= response.status < 300:
                return True, '최초 푸시등록 안내문자를 발송했습니다.'
            return False, f'SMS 연동 서버 응답 오류({response.status})'
    except urllib.error.HTTPError as exc:
        return False, f'SMS 발송 실패({exc.code})'
    except Exception as exc:
        return False, f'SMS 발송 실패: {exc}'


def _send_subscription(conn, row, payload):
    if not _push_ready():
        return False, 'VAPID 설정이 필요합니다.'
    cfg = _push_config()
    try:
        webpush(
            subscription_info={
                'endpoint': row['endpoint'],
                'keys': {'p256dh': row['p256dh'], 'auth': row['auth']},
            },
            data=json.dumps(payload, ensure_ascii=False),
            vapid_private_key=cfg['private_key'],
            vapid_claims={'sub': cfg['subject']},
            ttl=86400,
            timeout=7,
        )
        conn.execute('''
            UPDATE parent_push_subscriptions
            SET last_success_at=CURRENT_TIMESTAMP, failure_count=0,
                last_error=NULL, updated_at=CURRENT_TIMESTAMP, is_active=1
            WHERE id=?
        ''', (row['id'],))
        return True, ''
    except WebPushException as exc:
        status = getattr(getattr(exc, 'response', None), 'status_code', None)
        if status in (404, 410):
            conn.execute(
                'UPDATE parent_push_subscriptions SET is_active=0, last_error=?, updated_at=CURRENT_TIMESTAMP WHERE id=?',
                (f'만료된 구독({status})', row['id']),
            )
        else:
            conn.execute('''
                UPDATE parent_push_subscriptions
                SET failure_count=failure_count+1, last_error=?, updated_at=CURRENT_TIMESTAMP
                WHERE id=?
            ''', (_text(exc, 1000), row['id']))
        return False, f'Web Push 오류({status or "unknown"})'
    except Exception as exc:
        conn.execute('''
            UPDATE parent_push_subscriptions
            SET failure_count=failure_count+1, last_error=?, updated_at=CURRENT_TIMESTAMP
            WHERE id=?
        ''', (_text(exc, 1000), row['id']))
        return False, _text(exc, 500)


def _notification_url(conn, guardian_id):
    invite = _active_invite(conn, guardian_id)
    return f"/parent/register/{invite['token']}"


def _send_notification(conn, guardian_ids, kind, title, body, *, class_id=None,
                       student_id=None, target_type='선택', created_by=None):
    guardian_ids = sorted({int(value) for value in guardian_ids if value})
    kind = kind if kind in NOTICE_KINDS else '일반'
    title = _text(title, 120) or NOTICE_KINDS[kind][0]
    body = _text(body, 1000) or NOTICE_KINDS[kind][1].format(student='학생')
    cursor = conn.execute('''
        INSERT INTO parent_notifications(
            kind, title, body, class_id, student_id, target_type, created_by, total_count
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
    ''', (kind, title, body, class_id, student_id, target_type,
          created_by or session.get('user_name') or '시스템', len(guardian_ids)))
    notification_id = cursor.lastrowid
    successful_parents = 0
    failed_parents = 0
    for guardian_id in guardian_ids:
        subscriptions = conn.execute('''
            SELECT * FROM parent_push_subscriptions
            WHERE guardian_id=? AND is_active=1 ORDER BY id
        ''', (guardian_id,)).fetchall()
        sent = 0
        errors = []
        payload = {
            'title': title,
            'body': body,
            'kind': kind,
            'tag': f'saedam-parent-{notification_id}-{guardian_id}',
            'url': _notification_url(conn, guardian_id),
        }
        for subscription in subscriptions:
            ok, error = _send_subscription(conn, subscription, payload)
            if ok:
                sent += 1
            elif error:
                errors.append(error)
        if sent:
            status = '발송'
            successful_parents += 1
        elif not subscriptions:
            status = '푸시미등록'
            failed_parents += 1
        else:
            status = '실패'
            failed_parents += 1
        conn.execute('''
            INSERT INTO parent_notification_recipients(
                notification_id, guardian_id, student_id, status,
                subscription_count, sent_count, error_message, sent_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, CASE WHEN ? > 0 THEN CURRENT_TIMESTAMP END)
        ''', (notification_id, guardian_id, student_id, status,
              len(subscriptions), sent, '; '.join(errors)[:1000], sent))
    conn.execute('''
        UPDATE parent_notifications
        SET sent_count=?, failed_count=? WHERE id=?
    ''', (successful_parents, failed_parents, notification_id))
    conn.commit()
    return notification_id, successful_parents, failed_parents


def _guardian_ids_for_target(conn, target_type, target_id=None):
    if target_type == 'all':
        return [row['id'] for row in conn.execute(
            'SELECT id FROM parent_guardians WHERE is_active=1 ORDER BY id'
        ).fetchall()]
    if target_type == 'parent' and target_id:
        row = conn.execute(
            'SELECT id FROM parent_guardians WHERE id=? AND is_active=1',
            (target_id,),
        ).fetchone()
        return [row['id']] if row else []
    if target_type == 'class' and target_id:
        return [row['guardian_id'] for row in conn.execute('''
            SELECT DISTINCT gs.guardian_id
            FROM parent_class_students cs
            JOIN parent_guardian_students gs ON gs.student_id=cs.student_id
            JOIN parent_guardians g ON g.id=gs.guardian_id AND g.is_active=1
            WHERE cs.class_id=?
        ''', (target_id,)).fetchall()]
    return []


def _children_by_guardian(conn):
    result = {}
    rows = conn.execute('''
        SELECT gs.guardian_id, s.*,
               GROUP_CONCAT(DISTINCT c.class_name) AS class_names,
               GROUP_CONCAT(DISTINCT c.id) AS class_ids
        FROM parent_guardian_students gs
        JOIN parent_students s ON s.id=gs.student_id
        LEFT JOIN parent_class_students cs ON cs.student_id=s.id
        LEFT JOIN parent_classes c ON c.id=cs.class_id AND c.is_active=1
        WHERE s.is_active=1
        GROUP BY gs.guardian_id, s.id
        ORDER BY s.school_name, s.name
    ''').fetchall()
    for row in rows:
        result.setdefault(row['guardian_id'], []).append(dict(row))
    return result


def _bootstrap_payload(conn):
    children = _children_by_guardian(conn)
    guardians = []
    for row in conn.execute('''
        SELECT g.*,
               COUNT(DISTINCT CASE WHEN ps.is_active=1 THEN ps.id END) AS subscription_count,
               MAX(i.sms_sent_at) AS sms_sent_at,
               MAX(i.registered_at) AS registered_at,
               MAX(i.token) AS invite_token
        FROM parent_guardians g
        LEFT JOIN parent_push_subscriptions ps ON ps.guardian_id=g.id
        LEFT JOIN parent_invites i ON i.guardian_id=g.id AND i.is_active=1
        WHERE g.is_active=1
        GROUP BY g.id ORDER BY g.name, g.id
    ''').fetchall():
        item = dict(row)
        item['children'] = children.get(row['id'], [])
        item['invite_url'] = (
            f"{_base_url()}/parent/register/{row['invite_token']}"
            if row['invite_token'] else ''
        )
        guardians.append(item)
    classes = []
    for row in conn.execute('''
        SELECT c.*, COUNT(DISTINCT cs.student_id) AS student_count
        FROM parent_classes c
        LEFT JOIN parent_class_students cs ON cs.class_id=c.id
        WHERE c.is_active=1 GROUP BY c.id
        ORDER BY c.school_name, c.department, c.class_name
    ''').fetchall():
        item = dict(row)
        item['instructor_url'] = f"{_base_url()}/parent-notifications/instructor/{row['access_token']}"
        classes.append(item)
    histories = [dict(row) for row in conn.execute('''
        SELECT n.*, c.class_name, c.school_name
        FROM parent_notifications n
        LEFT JOIN parent_classes c ON c.id=n.class_id
        ORDER BY n.id DESC LIMIT 100
    ''').fetchall()]
    total = conn.execute(
        'SELECT COUNT(*) FROM parent_guardians WHERE is_active=1'
    ).fetchone()[0]
    registered = conn.execute('''
        SELECT COUNT(DISTINCT guardian_id) FROM parent_push_subscriptions WHERE is_active=1
    ''').fetchone()[0]
    sms_sent = conn.execute(
        'SELECT COUNT(DISTINCT guardian_id) FROM parent_invites WHERE sms_sent_at IS NOT NULL'
    ).fetchone()[0]
    schools = [dict(row) for row in conn.execute('''
        SELECT id, school_name, year FROM schools
        WHERE COALESCE(is_active, 1)=1 ORDER BY year DESC, school_name
    ''').fetchall()]
    return {
        'ok': True,
        'stats': {
            'parents': total,
            'registered': registered,
            'unregistered': max(0, total - registered),
            'sms_sent': sms_sent,
            'classes': len(classes),
        },
        'guardians': guardians,
        'classes': classes,
        'histories': histories,
        'schools': schools,
        'notice_kinds': list(NOTICE_KINDS),
        'push_configured': _push_ready(),
        'sms_configured': bool(_text(os.getenv('PARENT_SMS_WEBHOOK_URL'))),
    }


@parent_notification_bp.route('/parent-notifications')
@manager_required
def manager_page():
    return render_template('parent_notifications/admin.html')


@parent_notification_bp.route('/parent-notifications/api/bootstrap')
@manager_required
def bootstrap_api():
    conn = get_db()
    ensure_parent_notification_schema(conn)
    payload = _bootstrap_payload(conn)
    conn.close()
    return jsonify(payload)


@parent_notification_bp.route('/parent-notifications/api/parents', methods=['POST'])
@manager_required
def save_parent():
    data = request.get_json(silent=True) or {}
    parent_id = data.get('id')
    name = _text(data.get('name'), 100)
    phone = _phone(data.get('phone'))
    if not name or len(phone) < 9:
        return jsonify(ok=False, message='학부모명과 올바른 휴대폰 번호를 입력해주세요.'), 400
    conn = get_db()
    ensure_parent_notification_schema(conn)
    if parent_id:
        row = conn.execute('SELECT id FROM parent_guardians WHERE id=?', (parent_id,)).fetchone()
        if not row:
            conn.close()
            return jsonify(ok=False, message='학부모 정보를 찾을 수 없습니다.'), 404
        conn.execute('''
            UPDATE parent_guardians SET name=?, phone=?, email=?, memo=?,
                updated_at=CURRENT_TIMESTAMP WHERE id=?
        ''', (name, phone, _text(data.get('email'), 200), _text(data.get('memo'), 1000), parent_id))
    else:
        cursor = conn.execute('''
            INSERT INTO parent_guardians(name, phone, email, memo) VALUES (?, ?, ?, ?)
        ''', (name, phone, _text(data.get('email'), 200), _text(data.get('memo'), 1000)))
        parent_id = cursor.lastrowid
    if 'children' in data:
        conn.execute('DELETE FROM parent_guardian_students WHERE guardian_id=?', (parent_id,))
        for child in data.get('children') or []:
            child_name = _text(child.get('name'), 100)
            if not child_name:
                continue
            student_id = child.get('id')
            school_id = child.get('school_id') or None
            school_name = _text(child.get('school_name'), 200)
            if school_id and not school_name:
                school = conn.execute('SELECT school_name FROM schools WHERE id=?', (school_id,)).fetchone()
                school_name = school['school_name'] if school else ''
            if student_id and conn.execute('SELECT 1 FROM parent_students WHERE id=?', (student_id,)).fetchone():
                conn.execute('''
                    UPDATE parent_students SET name=?, school_id=?, school_name=?, grade=?,
                        classroom=?, memo=?, updated_at=CURRENT_TIMESTAMP WHERE id=?
                ''', (child_name, school_id, school_name, _text(child.get('grade'), 30),
                      _text(child.get('classroom'), 30), _text(child.get('memo'), 500), student_id))
            else:
                cursor = conn.execute('''
                    INSERT INTO parent_students(name, school_id, school_name, grade, classroom, memo)
                    VALUES (?, ?, ?, ?, ?, ?)
                ''', (child_name, school_id, school_name, _text(child.get('grade'), 30),
                      _text(child.get('classroom'), 30), _text(child.get('memo'), 500)))
                student_id = cursor.lastrowid
            conn.execute('''
                INSERT OR IGNORE INTO parent_guardian_students(guardian_id, student_id)
                VALUES (?, ?)
            ''', (parent_id, student_id))
            if 'class_ids' in child or 'class_id' in child:
                class_ids = child.get('class_ids') or []
                if child.get('class_id'):
                    class_ids = [*class_ids, child.get('class_id')]
                conn.execute('DELETE FROM parent_class_students WHERE student_id=?', (student_id,))
                for class_id in {str(value) for value in class_ids if value}:
                    conn.execute('''
                        INSERT OR IGNORE INTO parent_class_students(class_id, student_id)
                        SELECT id, ? FROM parent_classes WHERE id=? AND is_active=1
                    ''', (student_id, class_id))
    _active_invite(conn, parent_id)
    conn.commit()
    conn.close()
    return jsonify(ok=True, message='학부모 정보를 저장했습니다.', id=parent_id)


@parent_notification_bp.route('/parent-notifications/api/parents/<int:parent_id>', methods=['DELETE'])
@manager_required
def delete_parent(parent_id):
    conn = get_db()
    ensure_parent_notification_schema(conn)
    conn.execute('DELETE FROM parent_guardians WHERE id=?', (parent_id,))
    conn.commit()
    conn.close()
    return jsonify(ok=True, message='학부모 정보를 삭제했습니다.')


@parent_notification_bp.route('/parent-notifications/api/parents/<int:parent_id>/invite', methods=['POST'])
@manager_required
def parent_invite(parent_id):
    conn = get_db()
    ensure_parent_notification_schema(conn)
    parent = conn.execute('SELECT * FROM parent_guardians WHERE id=? AND is_active=1', (parent_id,)).fetchone()
    if not parent:
        conn.close()
        return jsonify(ok=False, message='학부모 정보를 찾을 수 없습니다.'), 404
    url = _invite_url(conn, parent_id)
    message = _invite_message(url)
    conn.close()
    return jsonify(ok=True, url=url, sms_message=message)


@parent_notification_bp.route('/parent-notifications/api/parents/<int:parent_id>/sms', methods=['POST'])
@manager_required
def send_parent_invite_sms(parent_id):
    data = request.get_json(silent=True) or {}
    conn = get_db()
    ensure_parent_notification_schema(conn)
    parent = conn.execute('SELECT * FROM parent_guardians WHERE id=? AND is_active=1', (parent_id,)).fetchone()
    if not parent:
        conn.close()
        return jsonify(ok=False, message='학부모 정보를 찾을 수 없습니다.'), 404
    invite = _active_invite(conn, parent_id)
    url = f"{_base_url()}/parent/register/{invite['token']}"
    message = _invite_message(url)
    if invite['sms_sent_at'] and not data.get('force'):
        conn.close()
        return jsonify(ok=False, already_sent=True, message='최초 안내문자는 이미 발송했습니다.',
                       url=url, sms_message=message), 409
    ok, result_message = _send_sms_webhook(parent['phone'], message)
    if ok:
        conn.execute('''
            UPDATE parent_invites SET sms_sent_at=CURRENT_TIMESTAMP,
                sms_send_count=sms_send_count+1 WHERE id=?
        ''', (invite['id'],))
        conn.commit()
    conn.close()
    status = 200 if ok else 503
    return jsonify(ok=ok, configured=bool(_text(os.getenv('PARENT_SMS_WEBHOOK_URL'))),
                   message=result_message, url=url, sms_message=message), status


@parent_notification_bp.route('/parent-notifications/api/classes', methods=['POST'])
@manager_required
def save_class():
    data = request.get_json(silent=True) or {}
    class_id = data.get('id')
    school_id = data.get('school_id') or None
    school_name = _text(data.get('school_name'), 200)
    class_name = _text(data.get('class_name'), 200)
    conn = get_db()
    ensure_parent_notification_schema(conn)
    if school_id and not school_name:
        school = conn.execute('SELECT school_name FROM schools WHERE id=?', (school_id,)).fetchone()
        school_name = school['school_name'] if school else ''
    if not school_name or not class_name:
        conn.close()
        return jsonify(ok=False, message='학교명과 강좌명을 입력해주세요.'), 400
    values = (school_id, school_name, _text(data.get('department'), 100), class_name,
              _text(data.get('instructor_emp_no'), 100), _text(data.get('instructor_name'), 100))
    if class_id:
        conn.execute('''
            UPDATE parent_classes SET school_id=?, school_name=?, department=?, class_name=?,
                instructor_emp_no=?, instructor_name=?, updated_at=CURRENT_TIMESTAMP WHERE id=?
        ''', (*values, class_id))
    else:
        cursor = conn.execute('''
            INSERT INTO parent_classes(
                school_id, school_name, department, class_name,
                instructor_emp_no, instructor_name, access_token
            ) VALUES (?, ?, ?, ?, ?, ?, ?)
        ''', (*values, _token()))
        class_id = cursor.lastrowid
    conn.commit()
    row = conn.execute('SELECT access_token FROM parent_classes WHERE id=?', (class_id,)).fetchone()
    conn.close()
    return jsonify(ok=True, message='강좌 전용페이지를 저장했습니다.', id=class_id,
                   url=f"{_base_url()}/parent-notifications/instructor/{row['access_token']}")


@parent_notification_bp.route('/parent-notifications/api/classes/<int:class_id>/sms', methods=['POST'])
@manager_required
def send_instructor_link_sms(class_id):
    conn = get_db()
    ensure_parent_notification_schema(conn)
    class_row = conn.execute('''
        SELECT c.*, u.phone AS instructor_phone, u.name AS member_name
        FROM parent_classes c
        LEFT JOIN users u ON u.emp_no=c.instructor_emp_no
        WHERE c.id=? AND c.is_active=1
    ''', (class_id,)).fetchone()
    if not class_row:
        conn.close()
        return jsonify(ok=False, message='강좌 정보를 찾을 수 없습니다.'), 404
    phone = _phone(class_row['instructor_phone'])
    url = f"{_base_url()}/parent-notifications/instructor/{class_row['access_token']}"
    message = (
        '[새담 방과후학교]\n'
        f'{class_row["school_name"]} {class_row["class_name"]} 강사 전용 출결 페이지입니다.\n'
        '인트라넷 로그인 후 이용해주세요.\n'
        f'{url}'
    )
    if not class_row['instructor_emp_no']:
        conn.close()
        return jsonify(ok=False, message='강사 사번을 먼저 지정해주세요.',
                       url=url, sms_message=message), 400
    if len(phone) < 9:
        conn.close()
        return jsonify(ok=False, message='지정 강사의 회원정보에 휴대폰 번호가 없습니다.',
                       url=url, sms_message=message), 400
    ok, result_message = _send_sms_webhook(phone, message)
    conn.close()
    return jsonify(ok=ok,
                   configured=bool(_text(os.getenv('PARENT_SMS_WEBHOOK_URL'))),
                   message=result_message, url=url,
                   sms_message=message), 200 if ok else 503


@parent_notification_bp.route('/parent-notifications/api/classes/<int:class_id>', methods=['DELETE'])
@manager_required
def delete_class(class_id):
    conn = get_db()
    ensure_parent_notification_schema(conn)
    conn.execute('UPDATE parent_classes SET is_active=0, updated_at=CURRENT_TIMESTAMP WHERE id=?', (class_id,))
    conn.commit()
    conn.close()
    return jsonify(ok=True, message='강좌 전용페이지를 사용 중지했습니다.')


@parent_notification_bp.route('/parent-notifications/api/send', methods=['POST'])
@manager_required
def manager_send_notification():
    data = request.get_json(silent=True) or {}
    target_type = _text(data.get('target_type'), 20)
    target_id = data.get('target_id')
    conn = get_db()
    ensure_parent_notification_schema(conn)
    guardian_ids = _guardian_ids_for_target(conn, target_type, target_id)
    if not guardian_ids:
        conn.close()
        return jsonify(ok=False, message='발송할 학부모가 없습니다.'), 400
    notification_id, sent, failed = _send_notification(
        conn, guardian_ids, _text(data.get('kind'), 30),
        _text(data.get('title'), 120), _text(data.get('body'), 1000),
        class_id=target_id if target_type == 'class' else None,
        target_type=target_type, created_by=session.get('user_name'),
    )
    conn.close()
    return jsonify(ok=True, message=f'푸시 발송 완료: 성공 {sent}명, 미등록·실패 {failed}명',
                   id=notification_id, sent=sent, failed=failed)


@parent_notification_bp.route('/parent-notifications/api/history/<int:notification_id>')
@manager_required
def history_detail(notification_id):
    conn = get_db()
    ensure_parent_notification_schema(conn)
    notification = conn.execute(
        'SELECT * FROM parent_notifications WHERE id=?', (notification_id,)
    ).fetchone()
    recipients = [dict(row) for row in conn.execute('''
        SELECT r.*, g.name AS guardian_name, g.phone, s.name AS student_name
        FROM parent_notification_recipients r
        JOIN parent_guardians g ON g.id=r.guardian_id
        LEFT JOIN parent_students s ON s.id=r.student_id
        WHERE r.notification_id=? ORDER BY g.name
    ''', (notification_id,)).fetchall()]
    conn.close()
    if not notification:
        return jsonify(ok=False, message='발송 기록을 찾을 수 없습니다.'), 404
    return jsonify(ok=True, notification=dict(notification), recipients=recipients)


EXCEL_ALIASES = {
    'parent_name': ('학부모명', '보호자명', '학부모', '보호자'),
    'phone': ('휴대폰', '휴대폰번호', '전화번호', '학부모연락처', '보호자연락처'),
    'student_name': ('학생명', '학생이름', '학생'),
    'school_name': ('학교명', '학교'),
    'grade': ('학년',),
    'classroom': ('반', '학급'),
    'class_name': ('강좌명', '수업명', '프로그램명', '방과후강좌'),
    'department': ('부서', '분야'),
    'instructor_emp_no': ('강사사번', '강사ID', '강사아이디'),
    'instructor_name': ('강사명', '강사이름'),
}


def _header_map(values):
    normalized = {str(value or '').strip().replace(' ', ''): index for index, value in enumerate(values)}
    result = {}
    for key, aliases in EXCEL_ALIASES.items():
        for alias in aliases:
            index = normalized.get(alias.replace(' ', ''))
            if index is not None:
                result[key] = index
                break
    return result


def _cell(values, mapping, key):
    index = mapping.get(key)
    return values[index] if index is not None and index < len(values) else ''


@parent_notification_bp.route('/parent-notifications/api/upload', methods=['POST'])
@manager_required
def upload_parents():
    uploaded = request.files.get('file')
    if not uploaded or not uploaded.filename.lower().endswith('.xlsx'):
        return jsonify(ok=False, message='xlsx 엑셀 파일을 업로드해주세요.'), 400
    try:
        workbook = load_workbook(uploaded.stream, read_only=True, data_only=True)
    except Exception:
        return jsonify(ok=False, message='엑셀 파일을 읽을 수 없습니다.'), 400
    conn = get_db()
    ensure_parent_notification_schema(conn)
    counts = {'parents': 0, 'students': 0, 'links': 0, 'classes': 0, 'rows': 0}
    errors = []
    for sheet in workbook.worksheets:
        rows = sheet.iter_rows(values_only=True)
        headers = next(rows, None)
        if not headers:
            continue
        mapping = _header_map(headers)
        if not {'parent_name', 'phone', 'student_name'}.issubset(mapping):
            errors.append(f'{sheet.title}: 학부모명·휴대폰·학생명 열이 필요합니다.')
            continue
        for excel_row, values in enumerate(rows, 2):
            parent_name = _text(_cell(values, mapping, 'parent_name'), 100)
            phone = _phone(_cell(values, mapping, 'phone'))
            student_name = _text(_cell(values, mapping, 'student_name'), 100)
            if not parent_name and not phone and not student_name:
                continue
            if not parent_name or len(phone) < 9 or not student_name:
                errors.append(f'{sheet.title} {excel_row}행: 학부모명·휴대폰·학생명을 확인해주세요.')
                continue
            counts['rows'] += 1
            guardian = conn.execute('''
                SELECT id FROM parent_guardians WHERE phone=? AND name=? AND is_active=1 LIMIT 1
            ''', (phone, parent_name)).fetchone()
            if guardian:
                guardian_id = guardian['id']
            else:
                guardian_id = conn.execute(
                    'INSERT INTO parent_guardians(name, phone) VALUES (?, ?)',
                    (parent_name, phone),
                ).lastrowid
                counts['parents'] += 1
            school_name = _text(_cell(values, mapping, 'school_name'), 200)
            grade = _text(_cell(values, mapping, 'grade'), 30)
            classroom = _text(_cell(values, mapping, 'classroom'), 30)
            school = conn.execute('SELECT id FROM schools WHERE school_name=? ORDER BY year DESC LIMIT 1',
                                  (school_name,)).fetchone() if school_name else None
            student = conn.execute('''
                SELECT id FROM parent_students
                WHERE name=? AND COALESCE(school_name, '')=? AND COALESCE(grade, '')=?
                  AND COALESCE(classroom, '')=? AND is_active=1 LIMIT 1
            ''', (student_name, school_name, grade, classroom)).fetchone()
            if student:
                student_id = student['id']
            else:
                student_id = conn.execute('''
                    INSERT INTO parent_students(name, school_id, school_name, grade, classroom)
                    VALUES (?, ?, ?, ?, ?)
                ''', (student_name, school['id'] if school else None, school_name, grade, classroom)).lastrowid
                counts['students'] += 1
            before = conn.total_changes
            conn.execute('INSERT OR IGNORE INTO parent_guardian_students(guardian_id, student_id) VALUES (?, ?)',
                         (guardian_id, student_id))
            counts['links'] += int(conn.total_changes > before)
            class_name = _text(_cell(values, mapping, 'class_name'), 200)
            if class_name:
                department = _text(_cell(values, mapping, 'department'), 100)
                class_row = conn.execute('''
                    SELECT id FROM parent_classes WHERE school_name=? AND COALESCE(department, '')=?
                      AND class_name=? AND is_active=1 LIMIT 1
                ''', (school_name, department, class_name)).fetchone()
                if class_row:
                    class_id = class_row['id']
                else:
                    class_id = conn.execute('''
                        INSERT INTO parent_classes(
                            school_id, school_name, department, class_name,
                            instructor_emp_no, instructor_name, access_token
                        ) VALUES (?, ?, ?, ?, ?, ?, ?)
                    ''', (school['id'] if school else None, school_name or '미지정 학교', department,
                          class_name, _text(_cell(values, mapping, 'instructor_emp_no'), 100),
                          _text(_cell(values, mapping, 'instructor_name'), 100), _token())).lastrowid
                    counts['classes'] += 1
                conn.execute('INSERT OR IGNORE INTO parent_class_students(class_id, student_id) VALUES (?, ?)',
                             (class_id, student_id))
            _active_invite(conn, guardian_id)
    conn.commit()
    conn.close()
    return jsonify(ok=True, message=f"엑셀 {counts['rows']}행을 처리했습니다.", counts=counts,
                   errors=errors[:30])


@parent_notification_bp.route('/parent-notifications/template.xlsx')
@manager_required
def excel_template():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = '학부모명단'
    sheet.append(['학부모명', '휴대폰', '학생명', '학교명', '학년', '반',
                  '강좌명', '부서', '강사사번', '강사명'])
    sheet.append(['홍길동', '01012345678', '김민준', '새담초등학교', '3', '1',
                  '로봇과학 A반', '과학', 'T001', '이강사'])
    sheet.append(['홍길동', '01012345678', '김서연', '새담초등학교', '1', '2',
                  '창의미술', '예술', 'T002', '박강사'])
    stream = io.BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return send_file(stream, as_attachment=True, download_name='학부모알림_업로드양식.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


def _invite_from_token(conn, token):
    return conn.execute('''
        SELECT i.*, g.name AS guardian_name, g.phone, g.email
        FROM parent_invites i
        JOIN parent_guardians g ON g.id=i.guardian_id AND g.is_active=1
        WHERE i.token=? AND i.is_active=1
    ''', (token,)).fetchone()


@parent_notification_bp.route('/parent/register/<string:token>')
def parent_register(token):
    conn = get_db()
    ensure_parent_notification_schema(conn)
    invite = _invite_from_token(conn, token)
    if not invite:
        conn.close()
        return render_template('parent_notifications/register.html', invalid=True), 404
    conn.execute('UPDATE parent_invites SET last_opened_at=CURRENT_TIMESTAMP WHERE id=?', (invite['id'],))
    conn.commit()
    children = [dict(row) for row in conn.execute('''
        SELECT s.*, GROUP_CONCAT(DISTINCT c.class_name) AS class_names
        FROM parent_guardian_students gs
        JOIN parent_students s ON s.id=gs.student_id
        LEFT JOIN parent_class_students cs ON cs.student_id=s.id
        LEFT JOIN parent_classes c ON c.id=cs.class_id AND c.is_active=1
        WHERE gs.guardian_id=? GROUP BY s.id ORDER BY s.name
    ''', (invite['guardian_id'],)).fetchall()]
    notices = [dict(row) for row in conn.execute('''
        SELECT n.kind, n.title, n.body, n.created_at, r.status
        FROM parent_notification_recipients r
        JOIN parent_notifications n ON n.id=r.notification_id
        WHERE r.guardian_id=? ORDER BY r.id DESC LIMIT 30
    ''', (invite['guardian_id'],)).fetchall()]
    registered = conn.execute('''
        SELECT COUNT(*) FROM parent_push_subscriptions
        WHERE guardian_id=? AND is_active=1
    ''', (invite['guardian_id'],)).fetchone()[0] > 0
    conn.close()
    return render_template('parent_notifications/register.html', invalid=False, invite=dict(invite),
                           children=children, notices=notices, registered=registered,
                           token=token, push_configured=_push_ready())


@parent_notification_bp.route('/parent/api/push/public-key/<string:token>')
def parent_push_public_key(token):
    conn = get_db()
    ensure_parent_notification_schema(conn)
    valid = _invite_from_token(conn, token) is not None
    conn.close()
    if not valid:
        return jsonify(ok=False, message='유효하지 않은 등록 링크입니다.'), 404
    cfg = _push_config()
    if not _push_ready():
        return jsonify(ok=False, configured=False, message='서버의 Web Push 설정이 필요합니다.'), 503
    response = jsonify(ok=True, configured=True, public_key=cfg['public_key'])
    response.headers['Cache-Control'] = 'no-store'
    return response


@parent_notification_bp.route('/parent/api/push/subscribe/<string:token>', methods=['POST'])
def parent_push_subscribe(token):
    conn = get_db()
    ensure_parent_notification_schema(conn)
    invite = _invite_from_token(conn, token)
    if not invite:
        conn.close()
        return jsonify(ok=False, message='유효하지 않은 등록 링크입니다.'), 404
    data = request.get_json(silent=True) or {}
    keys = data.get('keys') or {}
    endpoint = _text(data.get('endpoint'), 3000)
    p256dh = _text(keys.get('p256dh'), 1000)
    auth = _text(keys.get('auth'), 1000)
    if not endpoint.startswith('https://') or not p256dh or not auth:
        conn.close()
        return jsonify(ok=False, message='푸시 구독 정보가 올바르지 않습니다.'), 400
    conn.execute('''
        INSERT INTO parent_push_subscriptions(
            guardian_id, endpoint, p256dh, auth, user_agent
        ) VALUES (?, ?, ?, ?, ?)
        ON CONFLICT(endpoint) DO UPDATE SET
            guardian_id=excluded.guardian_id, p256dh=excluded.p256dh,
            auth=excluded.auth, user_agent=excluded.user_agent,
            is_active=1, failure_count=0, last_error=NULL,
            updated_at=CURRENT_TIMESTAMP
    ''', (invite['guardian_id'], endpoint, p256dh, auth,
          _text(request.headers.get('User-Agent'), 500)))
    conn.execute('''
        UPDATE parent_invites SET registered_at=COALESCE(registered_at, CURRENT_TIMESTAMP),
            last_opened_at=CURRENT_TIMESTAMP WHERE id=?
    ''', (invite['id'],))
    conn.commit()
    subscription = conn.execute(
        'SELECT * FROM parent_push_subscriptions WHERE endpoint=?', (endpoint,)
    ).fetchone()
    _send_subscription(conn, subscription, {
        'title': '새담 방과후학교',
        'body': '무료 출결·학부모 알림 등록이 완료되었습니다.',
        'tag': f"saedam-parent-welcome-{invite['guardian_id']}",
        'url': f'/parent/register/{token}',
    })
    conn.commit()
    conn.close()
    return jsonify(ok=True, message='출결·학부모 알림 등록이 완료되었습니다.')


@parent_notification_bp.route('/parent/push-sw.js')
def parent_push_worker():
    response = current_app.send_static_file('js/parent_push_sw.js')
    response.headers['Content-Type'] = 'application/javascript; charset=utf-8'
    response.headers['Service-Worker-Allowed'] = '/parent/'
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    return response


def _load_instructor_class(conn, token):
    return conn.execute('''
        SELECT * FROM parent_classes WHERE access_token=? AND is_active=1
    ''', (token,)).fetchone()


def _instructor_authorized(conn, class_row, bind=False):
    emp_no = _text(session.get('emp_no'), 100)
    if not emp_no:
        return False
    if _is_manager():
        return True
    assigned = _text(class_row['instructor_emp_no'], 100)
    if assigned:
        return assigned == emp_no
    if bind:
        conn.execute('''
            UPDATE parent_classes SET instructor_emp_no=?, instructor_name=?,
                updated_at=CURRENT_TIMESTAMP WHERE id=? AND COALESCE(instructor_emp_no, '')=''
        ''', (emp_no, _text(session.get('user_name'), 100), class_row['id']))
        conn.commit()
        return True
    return False


@parent_notification_bp.route('/parent-notifications/instructor/<string:token>')
def instructor_page(token):
    conn = get_db()
    ensure_parent_notification_schema(conn)
    class_row = _load_instructor_class(conn, token)
    if not class_row:
        conn.close()
        return render_template('parent_notifications/instructor.html', invalid=True), 404
    logged_in = bool(session.get('emp_no'))
    authorized = _instructor_authorized(conn, class_row, bind=logged_in)
    if logged_in and not authorized:
        conn.close()
        return render_template('parent_notifications/instructor.html', invalid=False,
                               forbidden=True, class_info=dict(class_row)), 403
    students = []
    if authorized:
        students = [dict(row) for row in conn.execute('''
            SELECT s.*,
                   (SELECT a.status FROM parent_attendance a
                    WHERE a.class_id=? AND a.student_id=s.id AND a.event_date=?
                    ORDER BY a.id DESC LIMIT 1) AS today_status,
                   COUNT(DISTINCT gs.guardian_id) AS guardian_count,
                   COUNT(DISTINCT CASE WHEN ps.is_active=1 THEN ps.guardian_id END) AS registered_guardian_count
            FROM parent_class_students cs
            JOIN parent_students s ON s.id=cs.student_id AND s.is_active=1
            LEFT JOIN parent_guardian_students gs ON gs.student_id=s.id
            LEFT JOIN parent_push_subscriptions ps ON ps.guardian_id=gs.guardian_id
            WHERE cs.class_id=? GROUP BY s.id ORDER BY s.name
        ''', (class_row['id'], date.today().isoformat(), class_row['id'])).fetchall()]
    conn.close()
    return render_template('parent_notifications/instructor.html', invalid=False,
                           logged_in=logged_in, forbidden=False, authorized=authorized,
                           class_info=dict(class_row), students=students, token=token,
                           notice_kinds=list(NOTICE_KINDS))


@parent_notification_bp.route('/parent-notifications/instructor/<string:token>/attendance', methods=['POST'])
def instructor_attendance(token):
    data = request.get_json(silent=True) or {}
    status = _text(data.get('status'), 30)
    student_id = data.get('student_id')
    if status not in ATTENDANCE_KINDS:
        return jsonify(ok=False, message='출석·지각·결석·하원 중에서 선택해주세요.'), 400
    conn = get_db()
    ensure_parent_notification_schema(conn)
    class_row = _load_instructor_class(conn, token)
    if not class_row or not _instructor_authorized(conn, class_row):
        conn.close()
        return jsonify(ok=False, message='강좌 접근 권한이 없습니다.'), 403
    student = conn.execute('''
        SELECT s.* FROM parent_class_students cs
        JOIN parent_students s ON s.id=cs.student_id
        WHERE cs.class_id=? AND s.id=? AND s.is_active=1
    ''', (class_row['id'], student_id)).fetchone()
    if not student:
        conn.close()
        return jsonify(ok=False, message='강좌 학생을 찾을 수 없습니다.'), 404
    guardian_ids = [row['guardian_id'] for row in conn.execute('''
        SELECT guardian_id FROM parent_guardian_students WHERE student_id=?
    ''', (student_id,)).fetchall()]
    title = f'[{status}] {student["name"]} 학생'
    body = (
        f'{class_row["school_name"]} {class_row["class_name"]} · '
        f'{NOTICE_KINDS[status][1].format(student=student["name"])}'
    )
    notification_id, sent, failed = _send_notification(
        conn, guardian_ids, status, title, body, class_id=class_row['id'],
        student_id=student_id, target_type='학생', created_by=session.get('user_name'),
    )
    conn.execute('''
        INSERT INTO parent_attendance(
            class_id, student_id, status, event_date, note, recorded_by, notification_id
        ) VALUES (?, ?, ?, ?, ?, ?, ?)
    ''', (class_row['id'], student_id, status, date.today().isoformat(),
          _text(data.get('note'), 500), session.get('user_name'), notification_id))
    conn.commit()
    conn.close()
    return jsonify(ok=True, message=f'{student["name"]} 학생 {status} 처리 · 푸시 성공 {sent}명, 미등록·실패 {failed}명',
                   status=status, sent=sent, failed=failed)


@parent_notification_bp.route('/parent-notifications/instructor/<string:token>/send', methods=['POST'])
def instructor_send(token):
    data = request.get_json(silent=True) or {}
    conn = get_db()
    ensure_parent_notification_schema(conn)
    class_row = _load_instructor_class(conn, token)
    if not class_row or not _instructor_authorized(conn, class_row):
        conn.close()
        return jsonify(ok=False, message='강좌 접근 권한이 없습니다.'), 403
    guardian_ids = _guardian_ids_for_target(conn, 'class', class_row['id'])
    if not guardian_ids:
        conn.close()
        return jsonify(ok=False, message='강좌에 등록된 학부모가 없습니다.'), 400
    kind = _text(data.get('kind'), 30)
    notification_id, sent, failed = _send_notification(
        conn, guardian_ids, kind, _text(data.get('title'), 120),
        _text(data.get('body'), 1000), class_id=class_row['id'],
        target_type='강좌', created_by=session.get('user_name'),
    )
    conn.close()
    return jsonify(ok=True, id=notification_id,
                   message=f'강좌 알림 발송 완료: 성공 {sent}명, 미등록·실패 {failed}명')
