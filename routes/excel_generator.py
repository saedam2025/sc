from flask import Blueprint, request, send_file, render_template, abort, current_app, jsonify
import io, os, re, zipfile, uuid
import pandas as pd
from datetime import datetime
from decimal import Decimal, InvalidOperation
from numbers import Number
from openpyxl import load_workbook

# 블루프린트 이름 지정
excel_bp = Blueprint("excel_generator", __name__)
DEPOSIT_CHUNK_SIZE = int(os.environ.get("CHUNK_SIZE", "200"))  # 시트 누적 인원 기준 (기본 200)
DEPOSIT_TARGET_COLUMNS = ["은행", "계좌번호", "예금주", "입금액"]

def deposit_clean_account(acc):
    if pd.isna(acc):
        return ""
    s = str(acc).strip()
    if re.fullmatch(r"\d+(\.\d+)?[eE][+-]?\d+", s):
        try:
            s = format(Decimal(s), "f")
            if "." in s:
                s = s.rstrip("0").rstrip(".")
        except Exception:
            pass
    if s.endswith(".0"):
        s = s[:-2]
    s = s.replace(" ", "")
    # s = s.replace("-", "")  # 필요 시 하이픈 제거
    return s

def deposit_is_missing(value):
    if value is None:
        return True
    try:
        return bool(pd.isna(value))
    except (TypeError, ValueError):
        return False

def deposit_clean_text(value):
    return "" if deposit_is_missing(value) else str(value).strip()

def deposit_amount_decimal(value):
    if deposit_is_missing(value):
        return None
    text = str(value).strip().replace(",", "").replace("원", "").replace(" ", "")
    if not text:
        return None
    try:
        return Decimal(text)
    except (InvalidOperation, ValueError):
        return None

def deposit_clean_amount(value):
    amount = deposit_amount_decimal(value)
    if amount is None:
        return None
    return int(amount)

def deposit_display_amount(value):
    amount = deposit_amount_decimal(value)
    if amount is None:
        return deposit_clean_text(value)
    return f"{int(amount):,}"

def deposit_format_numeric_account(value, number_format):
    """숫자형 계좌번호의 000-0000 형식을 실제 문자열로 복원한다."""
    raw = deposit_clean_account(value)
    if not raw or not re.fullmatch(r"\d+", raw):
        return raw, False

    pattern = str(number_format or "").split(";")[0]
    pattern = re.sub(r"\[[^\]]*\]", "", pattern)
    pattern = pattern.replace("\\", "").replace('"', "")
    zero_count = pattern.count("0")
    if zero_count < 4 or not re.fullmatch(r"[0\-\s().]+", pattern) or len(raw) > zero_count:
        return raw, False

    padded = raw.zfill(zero_count)
    digits = iter(padded)
    restored = "".join(next(digits) if char == "0" else char for char in pattern)
    return deposit_clean_account(restored), padded != raw

def deposit_account_from_cell(value, number_format):
    cleaned = deposit_clean_account(value)
    if not isinstance(value, Number) or isinstance(value, bool):
        return cleaned, "", False

    restored, leading_zero_restored = deposit_format_numeric_account(value, number_format)
    digits = re.sub(r"\D", "", cleaned)
    if leading_zero_restored:
        return restored, "엑셀 숫자 서식에서 선행 0을 자동 복원했습니다.", False

    warnings = ["원본 통장번호가 숫자형이어서 맨 앞 0 존재 여부를 자동 확정할 수 없습니다."]
    if len(digits) >= 15:
        warnings.append("엑셀 숫자 15자리 제한으로 정밀도 손실 가능성이 있습니다.")
    return restored, " ".join(warnings), True

def deposit_upload_path():
    upload_root = os.path.join(os.getcwd(), "uploads_deposit")
    os.makedirs(upload_root, exist_ok=True)
    return os.path.join(upload_root, f"_upload_{uuid.uuid4().hex}.xlsx")

def deposit_extract_generated_rows(xlsx_bytes, expected_count):
    workbook = load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    try:
        ws = workbook.active
        headers = {
            deposit_clean_text(cell.value): cell.column
            for cell in ws[1]
            if deposit_clean_text(cell.value)
        }
        rows = []
        for excel_row in range(2, expected_count + 2):
            rows.append({
                column: ws.cell(excel_row, headers[column]).value if column in headers else None
                for column in DEPOSIT_TARGET_COLUMNS
            })
        return rows
    finally:
        workbook.close()

def deposit_build_verification(sheet_dfs):
    merged = pd.concat([df for _, df in sheet_dfs], ignore_index=True)
    generated_bytes = deposit_build_excel_bytes(merged, file_label="preview")
    generated_rows = deposit_extract_generated_rows(generated_bytes, len(merged))
    verification_rows = []

    for index, (_, source_row) in enumerate(merged.iterrows()):
        generated = generated_rows[index] if index < len(generated_rows) else {}
        before = {
            "school": deposit_clean_text(source_row.get("_원본학교")),
            "person": deposit_clean_text(source_row.get("_원본예금주")),
            "amount": deposit_display_amount(source_row.get("_원본입금액")),
            "account": deposit_clean_account(source_row.get("_원본계좌번호")),
        }
        after = {
            "school": deposit_clean_text(source_row.get("_시트")),
            "person": deposit_clean_text(generated.get("예금주")),
            "amount": deposit_display_amount(generated.get("입금액")),
            "account": deposit_clean_account(generated.get("계좌번호")),
        }

        fields = {
            "school": bool(before["school"]) and before["school"] == after["school"],
            "person": bool(before["person"]) and before["person"] == after["person"],
            "amount": deposit_clean_amount(source_row.get("_원본입금액")) is not None
                      and deposit_clean_amount(source_row.get("_원본입금액")) == deposit_clean_amount(generated.get("입금액")),
            "account": bool(before["account"]) and before["account"] == after["account"],
        }

        issues = []
        if not fields["school"]:
            issues.append("학교명 불일치")
        if not fields["person"]:
            issues.append("사람(예금주) 불일치")
        if not fields["amount"]:
            issues.append("입금액 불일치 또는 숫자 변환 실패")
        if not fields["account"]:
            if before["account"].startswith("0") and not after["account"].startswith("0"):
                issues.append("통장번호 선행 0 손실")
            else:
                issues.append("통장번호 불일치 또는 누락")

        account_note = deposit_clean_text(source_row.get("_계좌검증메모"))
        account_warning = bool(source_row.get("_계좌검증주의"))
        account_restored = bool(source_row.get("_계좌자동복원"))
        status = "error" if issues else ("warning" if account_warning else ("restored" if account_restored else "ok"))
        notes = list(issues)
        if account_note:
            notes.append(account_note)

        verification_rows.append({
            "no": index + 1,
            "source_row": int(source_row.get("_원본행", index + 4)),
            "status": status,
            "auto_restored": account_restored,
            "before": before,
            "after": after,
            "fields": fields,
            "message": " ".join(notes) if notes else "4개 항목 일치",
        })

    summary = {
        "total": len(verification_rows),
        "ok": sum(row["status"] == "ok" for row in verification_rows),
        "restored": sum(row["status"] == "restored" for row in verification_rows),
        "warning": sum(row["status"] == "warning" for row in verification_rows),
        "error": sum(row["status"] == "error" for row in verification_rows),
        "before_total": deposit_display_amount(sum((deposit_clean_amount(row.get("_원본입금액")) or 0) for _, row in merged.iterrows())),
        "after_total": deposit_display_amount(sum((deposit_clean_amount(row.get("입금액")) or 0) for row in generated_rows)),
    }
    sheet_summaries = []
    offset = 0
    for sheet_name, sheet_df in sheet_dfs:
        row_count = len(sheet_df)
        sheet_checks = verification_rows[offset:offset + row_count]
        sheet_generated = generated_rows[offset:offset + row_count]
        sheet_summaries.append({
            "sheet": sheet_name,
            "rows": row_count,
            "before_total": deposit_display_amount(sum((deposit_clean_amount(row.get("_원본입금액")) or 0) for _, row in sheet_df.iterrows())),
            "after_total": deposit_display_amount(sum((deposit_clean_amount(row.get("입금액")) or 0) for row in sheet_generated)),
            "ok": sum(row["status"] == "ok" for row in sheet_checks),
            "restored": sum(row["status"] == "restored" for row in sheet_checks),
            "warning": sum(row["status"] == "warning" for row in sheet_checks),
            "error": sum(row["status"] == "error" for row in sheet_checks),
        })
        offset += row_count

    return {
        "summary": summary,
        "sheets": sheet_summaries,
        "rows": verification_rows,
        "can_download": summary["error"] == 0,
    }

@excel_bp.route("/excel-generator", methods=["GET"])
def deposit_index():
    # 분리된 HTML 템플릿 렌더링
    return render_template('excel_generator.html', chunk_size=DEPOSIT_CHUNK_SIZE)

@excel_bp.route("/excel-generator/preview", methods=["POST"])
def deposit_preview():
    up = request.files.get("file")
    if not up or not up.filename:
        return jsonify({"ok": False, "message": "파일이 없습니다."}), 400
    if not up.filename.lower().endswith(".xlsx"):
        return jsonify({"ok": False, "message": "xlsx 형식만 지원합니다."}), 400

    input_path = deposit_upload_path()
    up.save(input_path)
    try:
        sheets = deposit_read_sheets_as_list(input_path)
        if not sheets:
            return jsonify({"ok": False, "message": "추출할 데이터가 없습니다. (열 이름/헤더 3행 확인)"}), 400
        verification = deposit_build_verification(sheets)
        return jsonify({"ok": True, **verification})
    except Exception as exc:
        current_app.logger.exception("입금용 엑셀 미리보기 검증 실패")
        return jsonify({"ok": False, "message": f"파일 검증 중 오류가 발생했습니다: {exc}"}), 400
    finally:
        try:
            os.remove(input_path)
        except OSError:
            pass

@excel_bp.route("/excel-generator/process", methods=["POST"])
def deposit_process():
    up = request.files.get("file")
    if not up:
        abort(400, "파일이 없습니다.")
    if not up.filename.lower().endswith(".xlsx"):
        abort(400, "xlsx 형식만 지원합니다.")

    input_path = deposit_upload_path()
    up.save(input_path)

    try:
        sheets = deposit_read_sheets_as_list(input_path)
        if not sheets:
            abort(400, "추출할 데이터가 없습니다. (열 이름/헤더 3행 확인)")

        verification = deposit_build_verification(sheets)
        if verification["summary"]["error"]:
            abort(400, f"변환 전·후 검증에서 {verification['summary']['error']}건의 불일치가 발견되어 다운로드를 중단했습니다.")
        if verification["summary"]["warning"] and request.form.get("warning_confirmed") != "1":
            abort(400, "통장번호 숫자형 주의 내용을 확인한 후 다운로드해주세요.")

        parts = deposit_split_by_sheet_boundary(sheets, DEPOSIT_CHUNK_SIZE)
        today = datetime.today().strftime("%Y-%m-%d")

        # 파트 1개면 단일 파일 반환
        if len(parts) == 1:
            merged = pd.concat([df for _, df in parts[0]], ignore_index=True)
            xlsx_bytes = deposit_build_excel_bytes(merged, file_label=today)
            return send_file(
                io.BytesIO(xlsx_bytes),
                as_attachment=True,
                download_name=f"입금내역_{today}.xlsx",
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        # 여러 파트면 ZIP으로 압축하여 반환
        buff = io.BytesIO()
        with zipfile.ZipFile(buff, "w", compression=zipfile.ZIP_DEFLATED) as zf:
            for i, part in enumerate(parts, start=1):
                merged = pd.concat([df for _, df in part], ignore_index=True)
                part_no = f"{i:02d}"
                fname = f"입금내역_{today}_part{part_no}.xlsx"
                xlsx_bytes = deposit_build_excel_bytes(merged, file_label=f"{today} part {part_no}")
                zf.writestr(fname, xlsx_bytes)

        buff.seek(0)
        return send_file(buff, as_attachment=True, download_name=f"입금내역_{today}_split.zip", mimetype="application/zip")

    finally:
        try:
            os.remove(input_path)
        except Exception:
            pass

def deposit_read_sheets_as_list(input_path: str):
    """[(sheet_name, df_with__시트), ...]"""
    out = []
    workbook = load_workbook(input_path, read_only=True, data_only=True)
    try:
        with pd.ExcelFile(input_path) as xlsx:
            for sheet_name in xlsx.sheet_names:
                df = pd.read_excel(xlsx, sheet_name=sheet_name, header=2, dtype={"계좌번호": str})
                df.columns = [deposit_clean_text(col) for col in df.columns]
                if not all(col in df.columns for col in DEPOSIT_TARGET_COLUMNS):
                    continue

                ws = workbook[sheet_name]
                header_columns = {
                    deposit_clean_text(cell.value): cell.column
                    for cell in ws[3]
                    if deposit_clean_text(cell.value)
                }
                records = []
                for source_index, row in df.iterrows():
                    excel_row = int(source_index) + 4
                    person = deposit_clean_text(row.get("예금주"))
                    if not person:
                        continue

                    def original_cell_value(column_name):
                        column_index = header_columns.get(column_name)
                        return ws.cell(excel_row, column_index).value if column_index else row.get(column_name)

                    account_column = header_columns.get("계좌번호")
                    account_cell = ws.cell(excel_row, account_column) if account_column else None
                    account_value = account_cell.value if account_cell else row.get("계좌번호")
                    account_format = account_cell.number_format if account_cell else "General"
                    account, account_note, account_warning = deposit_account_from_cell(account_value, account_format)
                    account_restored = account_note.startswith("엑셀 숫자 서식에서 선행 0을 자동 복원")

                    records.append({
                        "은행": deposit_clean_text(row.get("은행")),
                        "계좌번호": account,
                        "예금주": person,
                        "입금액": deposit_clean_amount(row.get("입금액")),
                        "_시트": sheet_name,
                        "_원본학교": sheet_name,
                        "_원본예금주": original_cell_value("예금주"),
                        "_원본입금액": original_cell_value("입금액"),
                        "_원본계좌번호": account,
                        "_원본행": excel_row,
                        "_계좌검증메모": account_note,
                        "_계좌검증주의": account_warning,
                        "_계좌자동복원": account_restored,
                    })

                if records:
                    out.append((sheet_name, pd.DataFrame(records)))
    finally:
        workbook.close()
    return out

def deposit_split_by_sheet_boundary(sheet_dfs, chunk_size: int):
    """시트 경계 기준 누적 인원 chunk_size 초과 시 파트 분할 (시트는 절대 쪼개지지 않음)."""
    parts, current, count = [], [], 0
    for sheet_name, df in sheet_dfs:
        rows = len(df)
        if rows == 0:
            current.append((sheet_name, df))
            continue
        if count > 0 and (count + rows) > chunk_size:
            parts.append(current)
            current, count = [], 0
        current.append((sheet_name, df))
        count += rows
        if rows > chunk_size:
            parts.append(current)
            current, count = [], 0
    if current:
        parts.append(current)
    return parts

def deposit_build_excel_bytes(df_with_sheet: pd.DataFrame, file_label: str) -> bytes:
    """검증용 메타 열은 제외하고 4개 입금 열과 F열 입금자 문구만 출력한다."""
    visible_cols = [c for c in DEPOSIT_TARGET_COLUMNS if c in df_with_sheet.columns]
    result_df = df_with_sheet[visible_cols].copy()

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
        result_df.to_excel(writer, index=False)
        workbook = writer.book
        ws = writer.sheets["Sheet1"]

        for idx, col in enumerate(result_df.columns):
            if col == "계좌번호":
                ws.set_column(idx, idx, 25, workbook.add_format({"num_format": "@"}))
            elif col == "입금액":
                ws.set_column(idx, idx, 15, workbook.add_format({"num_format": "#,##0"}))
            else:
                ws.set_column(idx, idx, 20)

        # F열 "새담청소년교육" (헤더 제외)
        tf = workbook.add_format({"num_format": "@"})
        for r in range(1, len(result_df) + 1):
            ws.write(r, 5, "새담청소년교육", tf)
        ws.set_column(5, 5, 20, tf)

        # 계좌번호 문자열 강제
        if "계좌번호" in result_df.columns:
            acc_idx = list(result_df.columns).index("계좌번호")
            for r, v in enumerate(result_df["계좌번호"].fillna(""), start=1):
                ws.write_string(r, acc_idx, str(v))

    output.seek(0)
    return output.getvalue()
