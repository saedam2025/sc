"""기존 전자계약과 완전히 분리된 인증전자계약 시스템."""

from __future__ import annotations

import base64
import hashlib
import hmac
import io
import json
import mimetypes
import os
import re
import secrets
import shutil
import threading
import zipfile
import tempfile
from datetime import datetime, timedelta, timezone
from functools import wraps
from html import escape
from pathlib import Path
from urllib.request import Request, urlopen

import pdfkit
import pandas as pd
import yagmail
from cryptography.fernet import Fernet, InvalidToken
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from PIL import Image
from flask import (
    Blueprint,
    jsonify,
    redirect,
    render_template,
    request,
    send_file,
    session,
    url_for,
)
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.utils import secure_filename

from .database import get_db
from .security import load_credential_secret, menu_permission_required
from .storage import (
    APP_ROOT,
    COMPANY_STAMP_ROOT,
    DATA_ROOT,
    TERMS_ROOT,
    VERIFIED_CONTRACT_ROOT,
    VERIFIED_CONTRACTS_ROOT,
    VERIFIED_LOGO_ROOT,
    VERIFIED_PDF_FONT_ROOT,
    VERIFIED_SIGNATURE_ROOT,
    VERIFIED_STAMP_ROOT,
    VERIFIED_TERMS_ROOT,
)
from .verified_contract_repository import (
    add_verified_contract_event,
    insert_verified_contract,
    update_verified_contract,
)
from .secure_files import (
    delete_file,
    encrypted_response,
    encrypted_storage_name,
    encrypt_bytes,
    encrypt_stream,
    encrypt_upload,
    original_filename,
    read_decrypted,
    temporary_decrypted_path,
)


verified_contract_bp = Blueprint("verified_contract", __name__)
_settings_file_lock = threading.RLock()

KST = timezone(timedelta(hours=9))
UTC = timezone.utc
DEFAULT_CATEGORIES = [
    "방과후강사",
    "맞춤형강사",
    "코디사업자",
    "코디근로자",
    "안전코디",
    "직원근로자",
    "직원사업자",
    "원어민근로자",
    "원어민사업자",
]
MAX_COMPANY_PROFILES = 20
MAX_MAIL_ACCOUNTS = 10
MONEY_FIELDS = ("수수료", "보조금", "경력수당", "직책수당", "기타")
CONTRACT_FIELDS = (
    "계약구분",
    "수탁학교명",
    "부서명",
    "성명",
    "수수료",
    "보조금",
    "경력수당",
    "직책수당",
    "기타",
    "근무시간",
    "계약기간",
    "비고1",
    "비고2",
    "비고3",
    "비고4",
    "email",
    "연락처",
    "거주지",
    "주민번호",
    "은행",
    "계좌번호",
    "회사명",
    "대표직함",
    "대표자명",
    "계약서제목",
)
AGREEMENTS = (
    {
        "key": "contract_content",
        "title": "계약 내용 및 전자서명 동의",
        "text": (
            "본인은 계약서 전체 내용을 확인하였으며 전자문서 및 전자서명 방식으로 "
            "계약을 체결하는 데 동의합니다. 본인이 입력한 성명, 직접 작성한 서명과 "
            "계약완료 버튼 클릭을 본인의 계약 체결 의사표시로 사용하는 데 동의합니다."
        ),
    },
    {
        "key": "identity",
        "title": "본인 확인",
        "text": "본인은 계약서에 표시 및 개인정보를 입력한 계약 당사자 본인이며 타인에게 인증번호와 계약 링크를 제공하지 않았습니다.",
    },
    {
        "key": "privacy",
        "title": "개인정보 수집·이용 안내 확인",
        "text": (
            "계약 체결·이행·보관 및 대금 지급을 위해 성명, 이메일, 연락처, 주소, 주민등록번호, "
            "은행명, 계좌번호, 접속기록, 서명정보를 "
            "수집·이용하며 관계 법령과 내부 보존기준에 따른 기간 동안 보관하는 안내를 확인했습니다."
        ),
    },
)

DEFAULTS_ROOT = APP_ROOT / "verified_contract_defaults"
VERIFIED_CATEGORIES_FILE = VERIFIED_CONTRACT_ROOT / "categories.json"
VERIFIED_TITLES_FILE = VERIFIED_CONTRACT_ROOT / "contract_titles.json"
VERIFIED_COMPANY_FILE = VERIFIED_CONTRACT_ROOT / "company_settings.json"
VERIFIED_MAIL_FILE = VERIFIED_CONTRACT_ROOT / "mail_settings.json"


def _now() -> datetime:
    return datetime.now(UTC)


def _iso(value: datetime | None = None) -> str:
    return (value or _now()).isoformat(timespec="seconds")


def _parse_iso(value: object) -> datetime | None:
    try:
        parsed = datetime.fromisoformat(str(value))
        return parsed if parsed.tzinfo else parsed.replace(tzinfo=UTC)
    except (TypeError, ValueError):
        return None


def _format_kst(value: object) -> str:
    parsed = _parse_iso(value)
    if not parsed:
        text = str(value or "")
        try:
            parsed = datetime.fromisoformat(text.replace(" ", "T")).replace(tzinfo=UTC)
        except ValueError:
            return text
    return parsed.astimezone(KST).strftime("%Y-%m-%d %H:%M")


def _sensitive_cipher() -> Fernet:
    """서버에 자동 보관되는 자격증명 키로 계약 민감정보를 암호화한다."""
    secret = load_credential_secret()
    digest = hashlib.sha256(
        f"verified-contract-sensitive-data:{secret}".encode("utf-8")
    ).digest()
    return Fernet(base64.urlsafe_b64encode(digest))


def _encrypt_sensitive(value: object) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    return _sensitive_cipher().encrypt(text.encode("utf-8")).decode("ascii")


def _decrypt_sensitive(value: object) -> str:
    token = str(value or "").strip()
    if not token:
        return ""
    try:
        return _sensitive_cipher().decrypt(token.encode("ascii")).decode("utf-8")
    except (InvalidToken, UnicodeDecodeError, ValueError, TypeError):
        return "[암호화 자료 확인 필요]"


def _normalize_rrn(value: object) -> str:
    digits = re.sub(r"\D", "", str(value or ""))
    if len(digits) != 13 or digits[6] not in "12345678":
        raise ValueError("주민등록번호를 13자리로 정확히 입력해 주세요.")
    try:
        datetime.strptime(digits[:6], "%y%m%d")
    except ValueError as exc:
        raise ValueError("주민등록번호 앞자리의 생년월일을 확인해 주세요.") from exc
    return f"{digits[:6]}-{digits[6:]}"


def _normalize_account(value: object) -> str:
    account = re.sub(r"\s+", "", str(value or "").strip())
    if not re.fullmatch(r"[0-9-]{5,80}", account) or not re.search(r"\d", account):
        raise ValueError("계좌번호는 숫자와 하이픈(-)으로 정확히 입력해 주세요.")
    return account


def _client_ip() -> str:
    forwarded = request.headers.get("X-Forwarded-For", "")
    return (forwarded.split(",")[0].strip() if forwarded else request.remote_addr or "")[:100]


def _user_agent() -> str:
    return str(request.headers.get("User-Agent", ""))[:500]


def _json_file(path: Path, default):
    backup = path.with_suffix(path.suffix + ".bak")
    with _settings_file_lock:
        for candidate in (path, backup):
            try:
                with candidate.open("r", encoding="utf-8") as handle:
                    return json.load(handle)
            except (OSError, ValueError, TypeError):
                continue
        return default


def _save_json(path: Path, value) -> None:
    """설정 파일을 중간 손상 없이 교체하고 직전 정상본도 보존한다."""
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_suffix(path.suffix + ".tmp")
    backup = path.with_suffix(path.suffix + ".bak")
    backup_temporary = path.with_suffix(path.suffix + ".bak.tmp")
    with _settings_file_lock:
        with temporary.open("w", encoding="utf-8") as handle:
            json.dump(value, handle, ensure_ascii=False, indent=2)
            handle.flush()
            os.fsync(handle.fileno())
        if path.is_file():
            shutil.copy2(path, backup_temporary)
            os.replace(backup_temporary, backup)
        os.replace(temporary, path)


def _copy_if_missing(source: Path, target: Path) -> bool:
    if target.exists() or not source.is_file():
        return False
    target.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(source, target)
    return True


def _bootstrap_verified_storage() -> None:
    """기존 값을 최초 한 번만 복제하고 이후에는 인증계약 전용 파일만 사용한다."""
    for directory in (
        VERIFIED_CONTRACT_ROOT,
        VERIFIED_CONTRACTS_ROOT,
        VERIFIED_TERMS_ROOT,
        VERIFIED_STAMP_ROOT,
        VERIFIED_LOGO_ROOT,
        VERIFIED_SIGNATURE_ROOT,
        VERIFIED_PDF_FONT_ROOT,
    ):
        directory.mkdir(parents=True, exist_ok=True)

    default_terms = DEFAULTS_ROOT / "terms"
    for category in DEFAULT_CATEGORIES:
        for suffix in ("", "2"):
            filename = f"{category}{suffix}.txt"
            target = VERIFIED_TERMS_ROOT / filename
            if target.exists():
                continue
            # Render에 기존 관리자가 수정한 양식이 있으면 그 상태를 최초 복제한다.
            source = TERMS_ROOT / filename
            if not source.is_file():
                source = default_terms / filename
            _copy_if_missing(source, target)

    if not VERIFIED_CATEGORIES_FILE.exists():
        legacy = DATA_ROOT / "categories.json"
        categories = _json_file(legacy, DEFAULT_CATEGORIES)
        _save_json(
            VERIFIED_CATEGORIES_FILE,
            categories if isinstance(categories, list) and categories else DEFAULT_CATEGORIES,
        )

    if not VERIFIED_TITLES_FILE.exists():
        legacy = DATA_ROOT / "contract_titles.json"
        titles = _json_file(legacy, {})
        _save_json(VERIFIED_TITLES_FILE, titles if isinstance(titles, dict) else {})

    if not VERIFIED_COMPANY_FILE.exists():
        legacy = DATA_ROOT / "company_settings.json"
        settings = _json_file(legacy, {})
        if not settings:
            settings = {
                "active_profile_id": "verified-default",
                "profiles": [
                    {
                        "id": "verified-default",
                        "label": "기본 회사",
                        "company_name": "(사)새담청소년교육문화원",
                        "representative_title": "이사장",
                        "representative_name": "",
                        "stamp_filename": "verified_default_stamp.png",
                    }
                ],
            }
        _save_json(VERIFIED_COMPANY_FILE, settings)

        for profile in settings.get("profiles", []):
            filename = os.path.basename(str(profile.get("stamp_filename", "")))
            if not filename:
                continue
            source = COMPANY_STAMP_ROOT / filename
            if source.is_file():
                _copy_if_missing(source, VERIFIED_STAMP_ROOT / filename)

    default_stamp = VERIFIED_STAMP_ROOT / "verified_default_stamp.png"
    _copy_if_missing(DEFAULTS_ROOT / "stamp.png", default_stamp)

    for font_name in ("NanumGothic-Regular.ttf", "NanumGothic-Bold.ttf"):
        font_target = VERIFIED_PDF_FONT_ROOT / font_name
        if font_target.exists():
            continue
        font_source = DATA_ROOT / "pdf_fonts" / font_name
        if not font_source.is_file():
            font_source = DEFAULTS_ROOT / "pdf_fonts" / font_name
        _copy_if_missing(font_source, font_target)

    if not VERIFIED_MAIL_FILE.exists():
        legacy_file = APP_ROOT / "mail_settings.json"
        settings = _json_file(legacy_file, {})
        username = str(settings.get("MAIL_USERNAME") or os.environ.get("MAIL_USERNAME") or "").strip()
        password = str(settings.get("MAIL_PASSWORD") or os.environ.get("MAIL_PASSWORD") or "").strip()
        account_id = "verified-mail-default"
        _save_json(
            VERIFIED_MAIL_FILE,
            {
                "active_account_id": account_id if username else "",
                "accounts": (
                    [
                        {
                            "id": account_id,
                            "label": "기본 발송계정",
                            "email": username,
                            "encrypted_password": _encrypt_sensitive(password),
                        }
                    ]
                    if username
                    else []
                ),
            },
        )


_bootstrap_verified_storage()


@verified_contract_bp.after_request
def _verified_contract_security_headers(response):
    if request.path.startswith("/verified-contract/sign/"):
        response.headers["Cache-Control"] = "no-store, max-age=0"
        response.headers["Pragma"] = "no-cache"
        response.headers["Referrer-Policy"] = "no-referrer"
        response.headers["X-Frame-Options"] = "DENY"
        response.headers["X-Content-Type-Options"] = "nosniff"
    return response


def _categories() -> list[str]:
    value = _json_file(VERIFIED_CATEGORIES_FILE, DEFAULT_CATEGORIES)
    if not isinstance(value, list):
        return DEFAULT_CATEGORIES.copy()
    result = []
    for item in value:
        name = str(item).strip()
        if (
            name
            and len(name) <= 40
            and not any(character in name for character in '\\/:*?"<>|')
        ):
            result.append(name)
    return result or DEFAULT_CATEGORIES.copy()


def _titles() -> dict[str, str]:
    value = _json_file(VERIFIED_TITLES_FILE, {})
    return value if isinstance(value, dict) else {}


def _normalize_companies(value) -> dict:
    default_profile = {
        "id": "verified-default",
        "label": "기본 회사",
        "company_name": "(사)새담청소년교육문화원",
        "representative_title": "이사장",
        "representative_name": "",
        "business_number": "",
        "address": "",
        "phone": "",
        "logo_filename": "",
        "logo_original_name": "",
        "stamp_filename": "verified_default_stamp.png",
        "stamp_original_name": "기본 도장.png",
    }
    profiles = []
    if isinstance(value, dict):
        for index, item in enumerate(value.get("profiles", [])):
            if not isinstance(item, dict):
                continue
            profile = default_profile.copy()
            profile.update(item)
            profile["id"] = str(profile.get("id") or f"verified-{index + 1}")
            profiles.append(profile)
    if not profiles:
        profiles = [default_profile]
    active_id = str((value or {}).get("active_profile_id", "")) if isinstance(value, dict) else ""
    if not any(profile["id"] == active_id for profile in profiles):
        active_id = profiles[0]["id"]
    return {
        "active_profile_id": active_id,
        "profiles": profiles[:MAX_COMPANY_PROFILES],
    }


def _company_settings() -> dict:
    stored = _json_file(VERIFIED_COMPANY_FILE, {})
    normalized = _normalize_companies(stored)
    if stored != normalized:
        _save_json(VERIFIED_COMPANY_FILE, normalized)
    return normalized


def _company_profile(profile_id: str | None = None) -> dict:
    settings = _company_settings()
    target_id = profile_id or settings["active_profile_id"]
    return next(
        (profile for profile in settings["profiles"] if profile["id"] == target_id),
        settings["profiles"][0],
    )


def _company_snapshot(profile_id: str | None = None) -> dict:
    profile = _company_profile(profile_id)
    return {
        "profile_id": profile["id"],
        "label": str(profile.get("label", "")),
        "company_name": str(profile.get("company_name", "")).strip(),
        "representative_title": str(profile.get("representative_title", "")).strip(),
        "representative_name": str(profile.get("representative_name", "")).strip(),
        "business_number": str(profile.get("business_number", "")).strip(),
        "address": str(profile.get("address", "")).strip(),
        "phone": str(profile.get("phone", "")).strip(),
        "logo_filename": os.path.basename(str(profile.get("logo_filename", ""))),
        "logo_original_name": str(profile.get("logo_original_name", "")).strip(),
        "stamp_filename": os.path.basename(str(profile.get("stamp_filename", ""))),
        "stamp_original_name": str(profile.get("stamp_original_name", "")).strip(),
    }


def _default_title(contract_type: str, company_name: str) -> str:
    mapping = {
        "방과후강사": f"{company_name} 위탁교육계약서",
        "맞춤형강사": f"{company_name} 위탁교육계약서",
        "코디근로자": f"{company_name} 센터장 계약서",
        "코디사업자": f"{company_name} 센터장 계약서",
        "원어민근로자": "방과후 영어 원어민 강사 위탁 계약서",
        "원어민사업자": "방과후 영어 원어민 강사 위탁 계약서",
        "안전코디": f"{company_name} 위수탁계약서",
        "직원근로자": f"{company_name} 근로계약서",
        "직원사업자": f"{company_name} 위탁업무계약서",
    }
    return mapping.get(contract_type, f"{company_name} 계약서 ({contract_type})")


def _contract_title(contract_type: str, company_name: str) -> str:
    saved = str(_titles().get(contract_type, "")).strip()
    if not saved:
        return _default_title(contract_type, company_name)
    for marker in (
        "{{ data.회사명 }}",
        "{{data.회사명}}",
        "{{ company.name }}",
        "{{company.name}}",
    ):
        saved = saved.replace(marker, company_name)
    return saved


def _read_terms(contract_type: str) -> tuple[str, str]:
    def read(suffix: str) -> str:
        path = VERIFIED_TERMS_ROOT / f"{contract_type}{suffix}.txt"
        try:
            return path.read_text(encoding="utf-8")
        except OSError:
            return ""

    return read(""), read("2")


def _format_value(value: object) -> str:
    raw = str(value or "").strip()
    if not raw:
        return ""
    try:
        number = float(raw.replace(",", ""))
        if 0 < number < 1:
            return f"{int(number * 100)}%"
        if number >= 100:
            return f"{int(number):,}"
    except ValueError:
        pass
    return raw


def _empty_value(value: object) -> bool:
    cleaned = str(value or "").strip().lower().replace(",", "").replace("원", "")
    if cleaned in {"", "none", "nan", "null", "-", "0", "0.0", "0.00"}:
        return True
    try:
        return float(cleaned) == 0
    except ValueError:
        return False


def _render_terms(raw_html: str, values: dict[str, str]) -> str:
    result = _sanitize_contract_html(raw_html)
    row_pattern = re.compile(r"<tr\b[^>]*>.*?</tr>", flags=re.IGNORECASE | re.DOTALL)
    for field in MONEY_FIELDS:
        if _empty_value(values.get(field)):
            marker = re.compile(
                r"\{\{\s*(?:data|style)\." + re.escape(field) + r"\s*\}\}",
                flags=re.IGNORECASE,
            )
            result = row_pattern.sub(
                lambda match: "" if marker.search(match.group(0)) else match.group(0),
                result,
            )
    result = re.sub(
        r"\sdata-show-if\s*=\s*([\"']).*?\1",
        "",
        result,
        flags=re.IGNORECASE | re.DOTALL,
    )
    for field in CONTRACT_FIELDS:
        value = _format_value(values.get(field)) if field in MONEY_FIELDS else str(values.get(field, ""))
        result = result.replace(f"{{{{ data.{field} }}}}", value)
        result = result.replace(
            f"{{{{ style.{field} }}}}",
            "display:none" if _empty_value(values.get(field)) else "display:table-row",
        )
    return result


def _sanitize_contract_html(raw_html: object) -> str:
    """양식 편집에 불필요한 실행 코드와 외부 프레임을 제거한다."""
    result = str(raw_html or "")
    result = re.sub(
        r"<\s*(script|iframe|object|embed|form|meta)\b[^>]*>.*?<\s*/\s*\1\s*>",
        "",
        result,
        flags=re.IGNORECASE | re.DOTALL,
    )
    result = re.sub(
        r"<\s*(script|iframe|object|embed|form|meta)\b[^>]*/?\s*>",
        "",
        result,
        flags=re.IGNORECASE | re.DOTALL,
    )
    result = re.sub(
        r"\s+on[a-z]+\s*=\s*([\"']).*?\1",
        "",
        result,
        flags=re.IGNORECASE | re.DOTALL,
    )
    result = re.sub(
        r"\s+(href|src)\s*=\s*([\"'])\s*javascript:.*?\2",
        "",
        result,
        flags=re.IGNORECASE | re.DOTALL,
    )
    return result


def _mail_account_store() -> dict:
    """기존 단일 메일 설정도 자동으로 다중 계정 형식으로 전환한다."""
    value = _json_file(VERIFIED_MAIL_FILE, {})
    accounts = []
    active_id = ""
    needs_save = False
    if isinstance(value, dict) and isinstance(value.get("accounts"), list):
        active_id = str(value.get("active_account_id", "")).strip()
        for index, item in enumerate(value["accounts"]):
            if not isinstance(item, dict):
                needs_save = True
                continue
            email = str(item.get("email", "")).strip().lower()
            encrypted_password = str(item.get("encrypted_password", "")).strip()
            if not email:
                needs_save = True
                continue
            if not encrypted_password and item.get("password"):
                encrypted_password = _encrypt_sensitive(item["password"])
                needs_save = True
            accounts.append(
                {
                    "id": str(item.get("id") or f"verified-mail-{index + 1}"),
                    "label": str(item.get("label") or email).strip()[:80],
                    "email": email[:254],
                    "encrypted_password": encrypted_password,
                }
            )
    elif isinstance(value, dict):
        username = str(value.get("MAIL_USERNAME", "")).strip().lower()
        password = str(value.get("MAIL_PASSWORD", "")).strip()
        if username:
            account_id = "verified-mail-default"
            accounts = [
                {
                    "id": account_id,
                    "label": "기본 발송계정",
                    "email": username,
                    "encrypted_password": _encrypt_sensitive(password),
                }
            ]
            active_id = account_id
        needs_save = True
    if len(accounts) > MAX_MAIL_ACCOUNTS:
        accounts = accounts[:MAX_MAIL_ACCOUNTS]
        needs_save = True
    if not any(item["id"] == active_id for item in accounts):
        active_id = accounts[0]["id"] if accounts else ""
        needs_save = True
    store = {
        "active_account_id": active_id,
        "accounts": accounts,
    }
    if needs_save:
        _save_json(VERIFIED_MAIL_FILE, store)
    return store


def _mail_accounts_for_view(store: dict | None = None) -> list[dict[str, str]]:
    return [
        {
            "id": item["id"],
            "label": item["label"],
            "email": item["email"],
            "has_password": bool(item.get("encrypted_password")),
        }
        for item in (store or _mail_account_store())["accounts"]
    ]


def _mail_settings(account_id: str | None = None) -> dict[str, str]:
    store = _mail_account_store()
    target_id = str(account_id or store["active_account_id"]).strip()
    account = next(
        (item for item in store["accounts"] if item["id"] == target_id),
        None,
    )
    if not account:
        return {"MAIL_USERNAME": "", "MAIL_PASSWORD": ""}
    encrypted_password = str(account.get("encrypted_password", "")).strip()
    password = ""
    if encrypted_password:
        try:
            password = _sensitive_cipher().decrypt(
                encrypted_password.encode("ascii")
            ).decode("utf-8")
        except (InvalidToken, UnicodeDecodeError, ValueError, TypeError) as exc:
            raise RuntimeError(
                "발송계정 비밀번호를 복호화할 수 없습니다. 비밀번호를 다시 저장해 주세요."
            ) from exc
    return {
        "MAIL_USERNAME": account["email"],
        "MAIL_PASSWORD": password,
    }


def _send_mail(to, subject: str, contents, attachments=None) -> None:
    settings = _mail_settings()
    if not settings["MAIL_USERNAME"] or not settings["MAIL_PASSWORD"]:
        raise RuntimeError("인증전자계약 전용 메일 계정이 설정되지 않았습니다.")
    smtp = yagmail.SMTP(settings["MAIL_USERNAME"], settings["MAIL_PASSWORD"])
    smtp.send(to=to, subject=subject, contents=contents, attachments=attachments)


def _csrf_token() -> str:
    token = session.get("verified_contract_csrf")
    if not token:
        token = secrets.token_urlsafe(32)
        session["verified_contract_csrf"] = token
    return token


def _csrf_required(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        supplied = request.headers.get("X-CSRF-Token") or request.form.get("csrf_token", "")
        expected = session.get("verified_contract_csrf", "")
        if not expected or not supplied or not hmac.compare_digest(str(expected), str(supplied)):
            return jsonify({"status": "error", "message": "보안 확인값이 만료되었습니다. 화면을 새로고침해 주세요."}), 403
        return view(*args, **kwargs)

    return wrapped


def _token_hash(token: str) -> str:
    return hashlib.sha256(token.encode("utf-8")).hexdigest()


def _request_scheme() -> str:
    forwarded = request.headers.get("X-Forwarded-Proto", "").split(",")[0].strip()
    return forwarded or request.scheme


def _external_url(token: str) -> str:
    return url_for(
        "verified_contract.public_contract",
        token=token,
        _external=True,
        _scheme=_request_scheme(),
    )


def _load_by_token(conn, token: str):
    return conn.execute(
        "SELECT * FROM verified_contracts WHERE invitation_token_hash=?",
        (_token_hash(token),),
    ).fetchone()


def _contract_available(row) -> tuple[bool, str]:
    if not row:
        return False, "유효하지 않은 계약 링크입니다."
    if row["status"] == "revoked":
        return False, "관리자가 취소한 계약 링크입니다."
    if row["status"] == "expired":
        return False, "유효기간이 만료된 계약 링크입니다."
    expiry = _parse_iso(row["invitation_expires_at"])
    if row["status"] == "pending" and expiry and expiry < _now():
        return False, "유효기간이 만료된 계약 링크입니다."
    return True, ""


def _public_verified(row, token: str) -> bool:
    access = session.get("verified_contract_access", {})
    return (
        isinstance(access, dict)
        and access.get(str(row["id"])) == _token_hash(token)[:24]
        and bool(row["verified_at"])
    )


def _mask_email(email: str) -> str:
    local, separator, domain = str(email).partition("@")
    if not separator:
        return "-"
    visible = local[:2] if len(local) > 2 else local[:1]
    return f"{visible}{'*' * max(2, len(local) - len(visible))}@{domain}"


def _row_for_view(row) -> dict:
    item = dict(row)
    try:
        contract_data = json.loads(item.get("contract_data_json") or "{}")
    except (TypeError, ValueError):
        contract_data = {}
    for field in (
        "수수료",
        "보조금",
        "경력수당",
        "직책수당",
        "기타",
        "근무시간",
        "계약기간",
        "비고1",
        "비고2",
        "비고3",
        "비고4",
    ):
        item[field] = contract_data.get(field, "")
    item["주민번호"] = _decrypt_sensitive(item.pop("signer_rrn_encrypted", ""))
    item["은행"] = _decrypt_sensitive(item.pop("signer_bank_encrypted", ""))
    item["계좌번호"] = _decrypt_sensitive(item.pop("signer_account_encrypted", ""))
    date_source = item.get("signed_at") or item.get("created_at") or ""
    item["연도"] = str(date_source)[:4]
    item["created_display"] = _format_kst(item.get("created_at"))
    item["expires_display"] = _format_kst(item.get("invitation_expires_at"))
    item["signed_display"] = _format_kst(item.get("signed_at"))
    status_map = {
        "draft": "등록대기",
        "pending": "서명대기",
        "completed": "계약완료",
        "revoked": "취소",
        "expired": "기간만료",
        "voided": "폐기",
        "superseded": "변경계약",
    }
    item["status_label"] = status_map.get(item.get("status"), item.get("status"))
    return item


def _record_event(conn, row_or_id, event_type: str, details=None) -> None:
    contract_id = int(row_or_id["id"] if hasattr(row_or_id, "keys") else row_or_id)
    add_verified_contract_event(
        conn,
        contract_id,
        event_type,
        _iso(),
        ip_address=_client_ip(),
        user_agent=_user_agent(),
        details=details,
    )


def _invitation_mail(row, invitation_url: str) -> None:
    _send_mail(
        row["signer_email"],
        f"[전자계약 요청] {row['title_snapshot']}",
        _invitation_html(row, invitation_url),
    )


def _invitation_html(row, invitation_url: str) -> str:
    return f"""
    <div style="font-family:Arial,'Malgun Gothic',sans-serif;line-height:1.7;color:#1f2937">
      <h2 style="color:#123b6d">새담 인증전자계약 요청</h2>
      <p><b>{escape(row['signer_name'])}</b>님, {escape(row['title_snapshot'])} 확인과 서명을 요청드립니다.</p>
      <p>아래 버튼을 누른 후 이메일 인증번호를 확인하면 계약서를 작성할 수 있습니다.</p>
      <p style="margin:28px 0">
        <a href="{escape(invitation_url)}" style="background:#123b6d;color:#fff;padding:13px 24px;border-radius:7px;text-decoration:none;font-weight:bold">계약서 확인하기</a>
      </p>
      <p style="font-size:13px;color:#64748b">유효기간: {escape(_format_kst(row['invitation_expires_at']))}<br>
      본인이 요청한 계약이 아니라면 링크를 열지 말고 새담 담당자에게 알려주세요.</p>
    </div>
    """


def _void_notice_html(row, reason_label: str) -> str:
    return f"""
    <div style="font-family:Arial,'Malgun Gothic',sans-serif;line-height:1.7;color:#1f2937">
      <h2 style="color:#a82d2d">계약 {escape(reason_label)} 안내</h2>
      <p><b>{escape(row['signer_name'])}</b>님, 아래 계약 건이 <b>{escape(reason_label)}</b> 처리되었습니다.</p>
      <p style="background:#f8f9fa;border:1px solid #dee2e6;border-radius:8px;padding:14px">{escape(row['title_snapshot'])}<br>
      계약구분: {escape(row['contract_type'])} · 수탁학교: {escape(row['school_name'] or '-')} · 부서: {escape(row['department'] or '-')}</p>
      <p><b>본 안내 이후로 위 계약서는 더 이상 법적 효력이 없습니다.</b><br>
      동일한 내용으로 다시 계약이 필요한 경우 새담 담당자로부터 별도의 계약 요청 메일을 받게 됩니다.</p>
      <p style="font-size:13px;color:#64748b">문의사항은 새담 계약 담당자에게 연락해 주세요.</p>
    </div>
    """


def _excel_text(value: object) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except (TypeError, ValueError):
        pass
    return str(value).strip()


def _normalized_excel_records(frame: pd.DataFrame) -> list[dict[str, str]]:
    aliases = {
        "이름": "성명",
        "계약자명": "성명",
        "이메일": "email",
        "계약자이메일": "email",
        "학교명": "수탁학교명",
        "부서": "부서명",
        "담당업무": "부서명",
        "강의과목": "부서명",
        "과목": "부서명",
    }
    frame = frame.dropna(how="all").copy()
    frame.columns = [str(column).replace("\ufeff", "").strip() for column in frame.columns]
    for source, target in aliases.items():
        if target not in frame.columns and source in frame.columns:
            frame.rename(columns={source: target}, inplace=True)
    return [
        {str(key): _excel_text(value) for key, value in record.items()}
        for record in frame.to_dict("records")
    ]


def _resolve_company_snapshot(value: object) -> dict:
    requested = str(value or "").strip()
    settings = _company_settings()
    if not requested:
        return _company_snapshot(settings["active_profile_id"])
    for profile in settings["profiles"]:
        if requested in {
            str(profile.get("id", "")),
            str(profile.get("label", "")),
            str(profile.get("company_name", "")),
        }:
            return _company_snapshot(str(profile["id"]))
    raise ValueError(f"회사정보 '{requested}'를 찾을 수 없습니다.")


@verified_contract_bp.route("/admin")
@menu_permission_required("verified_contract_admin")
def admin_page():
    page = max(1, request.args.get("page", 1, type=int))
    status_filter = str(request.args.get("status", "")).strip()
    query = str(request.args.get("q", "")).strip()
    year_filter = str(request.args.get("year", "")).strip()
    category_filter = str(request.args.get("category", "")).strip()
    school_filter = str(request.args.get("school", "")).strip()
    department_filter = str(request.args.get("dept", "")).strip()
    name_filter = str(request.args.get("name", "")).strip()
    params: list[object] = []
    where = []
    if category_filter == "미작성":
        where.append("status IN ('draft','pending','expired')")
    elif category_filter:
        where.append("contract_type=?")
        params.append(category_filter)
    if status_filter:
        where.append("status=?")
        params.append(status_filter)
    if year_filter:
        where.append("substr(COALESCE(signed_at, created_at), 1, 4)=?")
        params.append(year_filter)
    if school_filter:
        where.append("school_name=?")
        params.append(school_filter)
    if department_filter:
        where.append("department=?")
        params.append(department_filter)
    if name_filter:
        where.append("signer_name LIKE ?")
        params.append(f"%{name_filter}%")
    if query:
        where.append(
            "(signer_name LIKE ? OR signer_email LIKE ? OR school_name LIKE ? OR department LIKE ?)"
        )
        params.extend([f"%{query}%"] * 4)
    where_sql = f"WHERE {' AND '.join(where)}" if where else ""
    sort_columns = {
        "id": "id",
        "year": "COALESCE(signed_at, created_at)",
        "category": "contract_type",
        "school": "school_name",
        "dept": "department",
        "name": "signer_name",
        "email": "signer_email",
        "status": "status",
        "created": "created_at",
        "signed": "signed_at",
    }
    sort_key = request.args.get("sort", "")
    if sort_key not in sort_columns:
        sort_key = "id"
    sort_dir = "asc" if request.args.get("dir") == "asc" else "desc"
    sort_column = sort_columns[sort_key]
    sort_dir_sql = "ASC" if sort_dir == "asc" else "DESC"
    conn = get_db()
    try:
        now_text = _iso()
        expired_rows = conn.execute(
            """
            SELECT id FROM verified_contracts
            WHERE status='pending' AND invitation_expires_at < ?
            """,
            (now_text,),
        ).fetchall()
        for expired_row in expired_rows:
            update_verified_contract(conn, expired_row["id"], {"status": "expired"})
            _record_event(conn, expired_row["id"], "EXPIRED")
        if expired_rows:
            conn.commit()
        counts = {
            row["status"]: row["count"]
            for row in conn.execute(
                "SELECT status, COUNT(*) AS count FROM verified_contracts GROUP BY status"
            ).fetchall()
        }
        filter_values = conn.execute(
            """
            SELECT DISTINCT substr(COALESCE(signed_at, created_at),1,4) AS year,
                            school_name, department
            FROM verified_contracts
            """
        ).fetchall()
        total = int(
            conn.execute(
                f"SELECT COUNT(*) FROM verified_contracts {where_sql}",
                params,
            ).fetchone()[0]
        )
        # 강사 20~30명을 한 화면에서 전체 선택해 일괄 발송할 수 있도록 넉넉히 표시한다.
        per_page = 50
        rows = conn.execute(
            f"""
            SELECT * FROM verified_contracts
            {where_sql}
            ORDER BY {sort_column} {sort_dir_sql}, id DESC LIMIT ? OFFSET ?
            """,
            [*params, per_page, (page - 1) * per_page],
        ).fetchall()
        dup_counts = {
            (
                dup_row["signer_name"],
                dup_row["contract_type"],
                dup_row["school_name"],
                dup_row["department"],
            ): dup_row["cnt"]
            for dup_row in conn.execute(
                """
                SELECT signer_name, contract_type, school_name, department, COUNT(*) AS cnt
                FROM verified_contracts
                WHERE status NOT IN ('voided','superseded','revoked')
                GROUP BY signer_name, contract_type, school_name, department
                HAVING COUNT(*) > 1
                """
            ).fetchall()
        }
    finally:
        conn.close()
    total_pages = max(1, (total + per_page - 1) // per_page)
    all_total = sum(int(value) for value in counts.values())
    completed_count = int(counts.get("completed", 0))
    pending_count = int(counts.get("draft", 0)) + int(counts.get("pending", 0)) + int(counts.get("expired", 0))
    completion_rate = round(completed_count / all_total * 100, 1) if all_total else 0
    years = sorted({row["year"] for row in filter_values if row["year"]}, reverse=True)
    schools = sorted({row["school_name"] for row in filter_values if row["school_name"]})
    departments = sorted({row["department"] for row in filter_values if row["department"]})
    items = []
    for row in rows:
        item = _row_for_view(row)
        key = (item["signer_name"], item["contract_type"], item["school_name"], item["department"])
        item["is_duplicate"] = (
            dup_counts.get(key, 0) > 1
            and item["status"] not in ("voided", "superseded", "revoked")
        )
        items.append(item)
    companies = _company_settings()
    mail_store = _mail_account_store()
    return render_template(
        "verified_contract/admin.html",
        items=items,
        counts=counts,
        total_count=all_total,
        completed_count=completed_count,
        pending_count=pending_count,
        completion_rate=completion_rate,
        total=total,
        page=page,
        total_pages=total_pages,
        status_filter=status_filter,
        query=query,
        categories=_categories(),
        categories_list=_categories(),
        years=years,
        schools=schools,
        depts=departments,
        companies=companies,
        mail_accounts=_mail_accounts_for_view(mail_store),
        active_mail_account_id=mail_store["active_account_id"],
        csrf_token=_csrf_token(),
        sort_key=sort_key,
        sort_dir=sort_dir,
        current_page=page,
        start_page=max(1, ((page - 1) // 20) * 20 + 1),
        end_page=min(total_pages, ((page - 1) // 20) * 20 + 20),
    )


@verified_contract_bp.route("/admin/settings")
@menu_permission_required("verified_contract_admin")
def settings_page():
    """계약 목록과 분리된 인증계약 양식·발송 리소스 관리 화면."""
    companies = _company_settings()
    mail_store = _mail_account_store()
    return render_template(
        "verified_contract/settings.html",
        categories_list=_categories(),
        companies=companies,
        mail_accounts=_mail_accounts_for_view(mail_store),
        active_mail_account_id=mail_store["active_account_id"],
        csrf_token=_csrf_token(),
    )


@verified_contract_bp.route(
    "/admin/settings/company/<string:profile_id>/<string:asset_kind>"
)
@menu_permission_required("verified_contract_admin")
def company_asset(profile_id: str, asset_kind: str):
    """회사관리 카드에서 암호화된 로고·도장을 안전하게 미리보기한다."""
    if asset_kind not in {"logo", "stamp"}:
        return "지원하지 않는 회사 이미지입니다.", 404
    profile = next(
        (
            item
            for item in _company_settings()["profiles"]
            if str(item.get("id", "")) == str(profile_id)
        ),
        None,
    )
    if not profile:
        return "회사정보를 찾을 수 없습니다.", 404
    filename = os.path.basename(str(profile.get(f"{asset_kind}_filename", "")))
    if not filename:
        return "등록된 이미지가 없습니다.", 404
    root = VERIFIED_LOGO_ROOT if asset_kind == "logo" else VERIFIED_STAMP_ROOT
    path = root / filename
    if not path.is_file():
        return "회사 이미지 파일을 찾을 수 없습니다.", 404
    display_name = str(
        profile.get(f"{asset_kind}_original_name") or filename
    ).strip()
    return encrypted_response(
        path,
        display_name,
        as_attachment=False,
        mimetype=mimetypes.guess_type(display_name)[0] or "image/png",
    )


@verified_contract_bp.route("/admin/create", methods=["POST"])
@menu_permission_required("verified_contract_admin")
@_csrf_required
def create_contract():
    data = request.get_json(silent=True) or {}
    contract_type = str(data.get("contract_type", "")).strip()
    signer_name = str(data.get("signer_name", "")).strip()
    signer_email = str(data.get("signer_email", "")).strip().lower()
    if contract_type not in _categories():
        return jsonify({"status": "error", "message": "올바른 계약구분을 선택해 주세요."}), 400
    if not signer_name or not re.fullmatch(r"[^@\s]{1,80}", signer_name):
        return jsonify({"status": "error", "message": "계약자 성명을 확인해 주세요."}), 400
    if not re.fullmatch(r"[^@\s]+@[^@\s]+\.[^@\s]+", signer_email):
        return jsonify({"status": "error", "message": "계약자 이메일을 확인해 주세요."}), 400

    profile = _company_snapshot(str(data.get("company_profile_id", "")).strip() or None)
    terms1, terms2 = _read_terms(contract_type)
    if not terms1.strip():
        return jsonify({"status": "error", "message": "선택한 계약구분의 1페이지 양식이 비어 있습니다."}), 400

    contract_data = {
        "계약구분": contract_type,
        "수탁학교명": str(data.get("school_name", "")).strip(),
        "부서명": str(data.get("department", "")).strip(),
        "성명": signer_name,
        "email": signer_email,
    }
    for key in (
        "수수료",
        "보조금",
        "경력수당",
        "직책수당",
        "기타",
        "근무시간",
        "계약기간",
        "비고1",
        "비고2",
        "비고3",
        "비고4",
    ):
        contract_data[key] = str(data.get(key, "")).strip()

    conn = get_db()
    try:
        duplicate_rows = conn.execute(
            """
            SELECT id FROM verified_contracts
            WHERE signer_name=? AND contract_type=? AND school_name=? AND department=?
              AND status NOT IN ('voided','superseded','revoked')
            """,
            (
                signer_name,
                contract_type,
                contract_data["수탁학교명"],
                contract_data["부서명"],
            ),
        ).fetchall()
    finally:
        conn.close()
    duplicate_ids = [int(r["id"]) for r in duplicate_rows]

    token = secrets.token_urlsafe(32)
    expires_days = min(30, max(1, int(data.get("expires_days") or 7)))
    expires_at = _iso(_now() + timedelta(days=expires_days))
    values = {
        "contract_type": contract_type,
        "school_name": contract_data["수탁학교명"],
        "department": contract_data["부서명"],
        "signer_name": signer_name,
        "signer_email": signer_email,
        "contract_data_json": json.dumps(contract_data, ensure_ascii=False),
        "status": "pending",
        "title_snapshot": _contract_title(contract_type, profile["company_name"]),
        "terms1_snapshot": terms1,
        "terms2_snapshot": terms2,
        "company_snapshot_json": json.dumps(profile, ensure_ascii=False),
        "agreement_snapshot_json": json.dumps(AGREEMENTS, ensure_ascii=False),
        "invitation_token_hash": _token_hash(token),
        "invitation_expires_at": expires_at,
        "created_by": str(session.get("user_name") or session.get("emp_no") or "admin"),
    }
    conn = get_db()
    try:
        contract_id = insert_verified_contract(conn, values)
        _record_event(conn, contract_id, "CREATED", {"expires_days": expires_days})
        conn.commit()
        row = conn.execute(
            "SELECT * FROM verified_contracts WHERE id=?", (contract_id,)
        ).fetchone()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()

    invitation_url = _external_url(token)
    mail_status = "sent"
    mail_error = ""
    try:
        _invitation_mail(row, invitation_url)
    except Exception as exc:
        mail_status = "failed"
        mail_error = str(exc)[:500]

    conn = get_db()
    try:
        update_verified_contract(
            conn,
            contract_id,
            {
                "invitation_sent_at": _iso() if mail_status == "sent" else None,
                "invite_mail_status": mail_status,
                "invite_mail_error": mail_error,
            },
        )
        _record_event(
            conn,
            contract_id,
            "INVITATION_SENT" if mail_status == "sent" else "INVITATION_FAILED",
            {"recipient": signer_email, "error": mail_error},
        )
        conn.commit()
    finally:
        conn.close()

    message = "계약을 등록하고 인증 링크를 이메일로 발송했습니다."
    if mail_status == "failed":
        message = f"계약은 등록했지만 메일 발송에 실패했습니다: {mail_error}"
    if duplicate_ids:
        message += (
            f"\n\n⚠ 동일 인물·계약구분·학교·부서 조합의 기존 계약이 {len(duplicate_ids)}건 있습니다."
            " 목록에서 기존 계약을 확인하고 [폐기] 또는 [변경계약] 처리해 주세요."
        )
    return jsonify(
        {
            "status": "success" if mail_status == "sent" else "warning",
            "message": message,
            "contract_id": contract_id,
            "invitation_url": invitation_url,
            "duplicate_ids": duplicate_ids,
            "duplicate_name": signer_name if duplicate_ids else "",
        }
    )


@verified_contract_bp.route("/admin/excel-template")
@menu_permission_required("verified_contract_admin")
def download_excel_template():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "계약자 일괄등록"
    headers = [
        "계약구분",
        "성명",
        "email",
        "수탁학교명",
        "부서명",
        "계약기간",
        "수수료",
        "보조금",
        "경력수당",
        "직책수당",
        "기타",
        "근무시간",
        "비고1",
        "비고2",
        "비고3",
        "비고4",
        "회사정보",
    ]
    sample = [
        _categories()[0],
        "홍길동",
        "hong@example.com",
        "새담초등학교",
        "수학",
        "2026.03.01 ~ 2027.02.28",
        "50000",
        "",
        "",
        "",
        "",
        "주 2회",
        "",
        "",
        "",
        "",
        _company_profile().get("label", "기본 회사"),
    ]
    sheet.append(headers)
    sheet.append(sample)
    header_fill = PatternFill("solid", fgColor="173F6A")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
    widths = {
        "A": 18,
        "B": 14,
        "C": 28,
        "D": 24,
        "E": 22,
        "F": 27,
        "G": 14,
        "H": 14,
        "I": 14,
        "J": 14,
        "K": 14,
        "L": 18,
        "M": 18,
        "N": 18,
        "O": 18,
        "P": 18,
        "Q": 20,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    sheet.freeze_panes = "A2"

    guide = workbook.create_sheet("작성안내")
    guide_rows = [
        ("필수 열", "계약구분, 성명, email"),
        ("학교·업무", "수탁학교명, 부서명(담당업무·강의과목·직책)"),
        ("회사정보", "비우면 현재 기본 회사정보 사용. 등록된 구분명·회사명도 입력 가능"),
        ("민감정보 입력", "주민번호·은행·계좌번호는 계약자가 계약 작성 화면에서 직접 입력합니다."),
        ("민감정보 보관", "입력값은 서버에서 암호화하여 저장되며 엑셀에는 작성하지 않습니다."),
        ("대량발송", "업로드 후 목록에서 대상을 선택하고 '선택 계약링크 발송'을 누릅니다."),
        ("계약구분", ", ".join(_categories())),
    ]
    for row in guide_rows:
        guide.append(row)
    guide.column_dimensions["A"].width = 18
    guide.column_dimensions["B"].width = 100
    for cell in guide["A"]:
        cell.font = Font(bold=True, color="173F6A")

    memory = io.BytesIO()
    workbook.save(memory)
    memory.seek(0)
    return send_file(
        memory,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="인증전자계약_일괄등록_양식.xlsx",
    )


@verified_contract_bp.route("/admin/upload-excel", methods=["POST"])
@menu_permission_required("verified_contract_admin")
@_csrf_required
def upload_excel():
    uploaded = request.files.get("excel_file")
    if not uploaded or not uploaded.filename:
        return jsonify({"status": "error", "message": "업로드할 엑셀 파일을 선택해 주세요."}), 400
    if Path(uploaded.filename).suffix.lower() != ".xlsx":
        return jsonify({"status": "error", "message": "엑셀 파일은 .xlsx 형식만 사용할 수 있습니다."}), 400
    if request.content_length and request.content_length > 10 * 1024 * 1024:
        return jsonify({"status": "error", "message": "엑셀 파일은 10MB 이하만 사용할 수 있습니다."}), 413
    try:
        records = _normalized_excel_records(pd.read_excel(uploaded, dtype=str))
    except Exception as exc:
        return jsonify({"status": "error", "message": f"엑셀 파일을 읽을 수 없습니다: {str(exc)[:200]}"}), 400
    if not records:
        return jsonify({"status": "error", "message": "등록할 계약자 행이 없습니다."}), 400
    if len(records) > 500:
        return jsonify({"status": "error", "message": "한 번에 최대 500명까지 등록할 수 있습니다."}), 400

    categories = set(_categories())
    existing_keys = set()
    conn = get_db()
    try:
        for existing in conn.execute(
            """
            SELECT contract_type, signer_name, signer_email, school_name, department
            FROM verified_contracts WHERE status IN ('draft','pending')
            """
        ).fetchall():
            existing_keys.add(
                (
                    existing["contract_type"],
                    existing["signer_name"],
                    existing["signer_email"].lower(),
                    existing["school_name"],
                    existing["department"],
                )
            )

        inserted = 0
        skipped = 0
        duplicated = 0
        errors = []
        seen_keys = set()
        terms_cache: dict[str, tuple[str, str]] = {}
        created_by = str(session.get("user_name") or session.get("emp_no") or "admin")
        for excel_row, record in enumerate(records, start=2):
            contract_type = _excel_text(record.get("계약구분"))
            signer_name = _excel_text(record.get("성명"))
            signer_email = _excel_text(record.get("email")).lower()
            school_name = _excel_text(record.get("수탁학교명"))
            department = _excel_text(record.get("부서명"))
            row_errors = []
            if contract_type not in categories:
                row_errors.append(f"등록되지 않은 계약구분: {contract_type or '빈 값'}")
            if not signer_name or len(signer_name) > 80:
                row_errors.append("성명 확인 필요")
            if not re.fullmatch(r"[^@\s]+@[^@\s]+\.[^@\s]+", signer_email):
                row_errors.append("이메일 형식 확인 필요")
            try:
                profile = _resolve_company_snapshot(record.get("회사정보"))
            except ValueError as exc:
                profile = None
                row_errors.append(str(exc))
            key = (contract_type, signer_name, signer_email, school_name, department)
            # 같은 파일 안에서의 완전 중복 행만 막고, 이미 등록된 서명대기/등록대기 건은
            # 재계약(폐기·변경계약) 시나리오일 수 있으므로 차단하지 않고 등록 후 안내한다.
            if key in seen_keys:
                row_errors.append("업로드한 엑셀 파일 안에서 같은 행이 중복됨")
            if row_errors:
                skipped += 1
                errors.append({"row": excel_row, "name": signer_name or "-", "reason": ", ".join(row_errors)})
                continue
            if key in existing_keys:
                duplicated += 1
            terms1, terms2 = terms_cache.setdefault(contract_type, _read_terms(contract_type))
            if not terms1.strip():
                skipped += 1
                errors.append({"row": excel_row, "name": signer_name, "reason": "계약양식 내용 1이 비어 있음"})
                continue

            contract_data = {
                "계약구분": contract_type,
                "수탁학교명": school_name,
                "부서명": department,
                "성명": signer_name,
                "email": signer_email,
            }
            for field in (
                "수수료",
                "보조금",
                "경력수당",
                "직책수당",
                "기타",
                "근무시간",
                "계약기간",
                "비고1",
                "비고2",
                "비고3",
                "비고4",
            ):
                contract_data[field] = _excel_text(record.get(field))
            placeholder_token = secrets.token_urlsafe(32)
            contract_id = insert_verified_contract(
                conn,
                {
                    "contract_type": contract_type,
                    "school_name": school_name,
                    "department": department,
                    "signer_name": signer_name,
                    "signer_email": signer_email,
                    "contract_data_json": json.dumps(contract_data, ensure_ascii=False),
                    "status": "draft",
                    "title_snapshot": _contract_title(contract_type, profile["company_name"]),
                    "terms1_snapshot": terms1,
                    "terms2_snapshot": terms2,
                    "company_snapshot_json": json.dumps(profile, ensure_ascii=False),
                    "agreement_snapshot_json": json.dumps(AGREEMENTS, ensure_ascii=False),
                    "invitation_token_hash": _token_hash(placeholder_token),
                    "invitation_expires_at": _iso(),
                    "invite_mail_status": "not_sent",
                    "created_by": created_by,
                },
            )
            _record_event(
                conn,
                contract_id,
                "BULK_IMPORTED",
                {"excel_row": excel_row, "source_filename": secure_filename(uploaded.filename)},
            )
            seen_keys.add(key)
            inserted += 1
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()

    if inserted == 0:
        return jsonify(
            {
                "status": "error",
                "message": "등록된 계약자가 없습니다. 오류 내용을 확인해 주세요.",
                "inserted": 0,
                "skipped": skipped,
                "errors": errors[:50],
            }
        ), 400
    message = f"{inserted}명 등록 완료" + (f", {skipped}명 제외" if skipped else "")
    if duplicated:
        message += (
            f"\n\n⚠ {duplicated}명은 기존 서명대기/등록대기 계약과 동일한 조합(성명·계약구분·학교·부서)입니다."
            " 목록에서 기존 계약을 확인하고 [폐기] 또는 [변경계약] 처리해 주세요."
        )
    return jsonify(
        {
            "status": "warning" if (skipped or duplicated) else "success",
            "message": message,
            "inserted": inserted,
            "skipped": skipped,
            "duplicated": duplicated,
            "errors": errors[:50],
        }
    )


@verified_contract_bp.route("/admin/bulk-send", methods=["POST"])
@menu_permission_required("verified_contract_admin")
@_csrf_required
def bulk_send_invitations():
    data = request.get_json(silent=True) or {}
    try:
        ids = sorted({int(value) for value in data.get("ids", [])})
    except (TypeError, ValueError):
        return jsonify({"status": "error", "message": "선택한 계약번호를 확인해 주세요."}), 400
    if not ids:
        return jsonify({"status": "error", "message": "발송할 계약자를 선택해 주세요."}), 400
    if len(ids) > 100:
        return jsonify({"status": "error", "message": "한 번에 최대 100명까지 발송할 수 있습니다."}), 400
    expires_days = min(30, max(1, int(data.get("expires_days") or 7)))
    placeholders = ",".join("?" for _ in ids)
    conn = get_db()
    queued = []
    try:
        rows = conn.execute(
            f"SELECT * FROM verified_contracts WHERE id IN ({placeholders}) ORDER BY id",
            ids,
        ).fetchall()
        for row in rows:
            if row["status"] not in {"draft", "pending", "expired"}:
                continue
            token = secrets.token_urlsafe(32)
            expires_at = _iso(_now() + timedelta(days=expires_days))
            update_verified_contract(
                conn,
                row["id"],
                {
                    "status": "pending",
                    "invitation_token_hash": _token_hash(token),
                    "invitation_expires_at": expires_at,
                    "opened_at": None,
                    "otp_hash": None,
                    "otp_expires_at": None,
                    "otp_attempts": 0,
                    "otp_sent_at": None,
                    "verified_at": None,
                    "invite_mail_status": "sending",
                    "invite_mail_error": "",
                },
            )
            queued_row = dict(row)
            queued_row["invitation_expires_at"] = expires_at
            queued.append((queued_row, token, _external_url(token)))
            _record_event(
                conn,
                row,
                "BULK_SEND_QUEUED",
                {"expires_days": expires_days},
            )
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()
    if not queued:
        return jsonify({"status": "error", "message": "발송 가능한 등록대기·서명대기 계약이 없습니다."}), 400

    sent_count = 0
    failed = []
    settings = _mail_settings()
    smtp = None
    setup_error = ""
    try:
        if not settings["MAIL_USERNAME"] or not settings["MAIL_PASSWORD"]:
            raise RuntimeError("인증전자계약 전용 메일 계정이 설정되지 않았습니다.")
        smtp = yagmail.SMTP(settings["MAIL_USERNAME"], settings["MAIL_PASSWORD"])
    except Exception as exc:
        setup_error = str(exc)[:500]

    conn = get_db()
    try:
        for row, token, invitation_url in queued:
            error = setup_error
            if smtp is not None:
                try:
                    smtp.send(
                        to=row["signer_email"],
                        subject=f"[전자계약 요청] {row['title_snapshot']}",
                        contents=_invitation_html(row, invitation_url),
                    )
                    error = ""
                except Exception as exc:
                    error = str(exc)[:500]
            if error:
                failed.append({"id": row["id"], "name": row["signer_name"], "reason": error})
                update_verified_contract(
                    conn,
                    row["id"],
                    {"invite_mail_status": "failed", "invite_mail_error": error},
                )
                _record_event(
                    conn,
                    row,
                    "INVITATION_FAILED",
                    {"recipient": row["signer_email"], "error": error},
                )
            else:
                sent_count += 1
                update_verified_contract(
                    conn,
                    row["id"],
                    {
                        "invitation_sent_at": _iso(),
                        "invite_mail_status": "sent",
                        "invite_mail_error": "",
                    },
                )
                _record_event(
                    conn,
                    row,
                    "INVITATION_SENT",
                    {"recipient": row["signer_email"], "bulk": True},
                )
        conn.commit()
    finally:
        conn.close()
    return jsonify(
        {
            "status": "warning" if failed else "success",
            "message": f"{sent_count}명 발송 완료" + (f", {len(failed)}명 실패" if failed else ""),
            "sent": sent_count,
            "failed": failed,
        }
    )


@verified_contract_bp.route("/admin/bulk-revoke", methods=["POST"])
@menu_permission_required("verified_contract_admin")
@_csrf_required
def bulk_revoke():
    try:
        ids = sorted({int(value) for value in (request.get_json(silent=True) or {}).get("ids", [])})
    except (TypeError, ValueError):
        return jsonify({"status": "error", "message": "선택 항목을 확인해 주세요."}), 400
    if not ids:
        return jsonify({"status": "error", "message": "취소할 계약을 선택해 주세요."}), 400
    placeholders = ",".join("?" for _ in ids)
    conn = get_db()
    try:
        rows = conn.execute(
            f"SELECT id, status FROM verified_contracts WHERE id IN ({placeholders})",
            ids,
        ).fetchall()
        changed = 0
        for row in rows:
            if row["status"] == "completed":
                continue
            update_verified_contract(conn, row["id"], {"status": "revoked"})
            _record_event(conn, row, "REVOKED", {"bulk": True})
            changed += 1
        conn.commit()
    finally:
        conn.close()
    return jsonify({"status": "success", "message": f"{changed}건의 계약 링크를 취소했습니다."})


@verified_contract_bp.route("/admin/bulk-delete", methods=["POST"])
@menu_permission_required("verified_contract_admin")
@_csrf_required
def bulk_delete():
    try:
        ids = sorted({int(value) for value in (request.get_json(silent=True) or {}).get("ids", [])})
    except (TypeError, ValueError):
        return jsonify({"status": "error", "message": "선택 항목을 확인해 주세요."}), 400
    if not ids:
        return jsonify({"status": "error", "message": "삭제할 계약을 선택해 주세요."}), 400
    placeholders = ",".join("?" for _ in ids)
    conn = get_db()
    try:
        rows = conn.execute(
            f"SELECT id, pdf_filename, signature_filename FROM verified_contracts WHERE id IN ({placeholders})",
            ids,
        ).fetchall()
        for row in rows:
            delete_file(VERIFIED_CONTRACTS_ROOT / os.path.basename(row["pdf_filename"] or ""))
            delete_file(VERIFIED_SIGNATURE_ROOT / os.path.basename(row["signature_filename"] or ""))
        conn.execute(f"DELETE FROM verified_contract_events WHERE contract_id IN ({placeholders})", ids)
        conn.execute(f"DELETE FROM verified_contracts WHERE id IN ({placeholders})", ids)
        conn.commit()
        deleted = len(rows)
    finally:
        conn.close()
    return jsonify({"status": "success", "message": f"{deleted}건의 계약 등록정보와 계약서 파일을 완전히 삭제했습니다."})


@verified_contract_bp.route("/admin/bulk-void", methods=["POST"])
@menu_permission_required("verified_contract_admin")
@_csrf_required
def bulk_void():
    data = request.get_json(silent=True) or {}
    action = str(data.get("action", "")).strip()
    if action not in ("discard", "amend"):
        return jsonify({"status": "error", "message": "처리 방식을 확인해 주세요."}), 400
    notify = bool(data.get("notify"))
    try:
        ids = sorted({int(value) for value in data.get("ids", [])})
    except (TypeError, ValueError):
        return jsonify({"status": "error", "message": "선택 항목을 확인해 주세요."}), 400
    if not ids:
        return jsonify({"status": "error", "message": "처리할 계약을 선택해 주세요."}), 400
    new_status = "voided" if action == "discard" else "superseded"
    reason_label = "폐기" if action == "discard" else "변경계약"
    placeholders = ",".join("?" for _ in ids)
    conn = get_db()
    try:
        rows = conn.execute(
            f"SELECT * FROM verified_contracts WHERE id IN ({placeholders})",
            ids,
        ).fetchall()
        processed = 0
        mail_sent = 0
        mail_failed = []
        for row in rows:
            # 상태만 변경하며, 계약 정보나 계약서·서명 파일은 그대로 보존한다.
            if row["status"] in ("voided", "superseded"):
                continue
            update_verified_contract(conn, row["id"], {"status": new_status})
            _record_event(
                conn,
                row,
                "VOIDED" if action == "discard" else "SUPERSEDED",
                {"reason": reason_label, "notify": notify},
            )
            processed += 1
            if not notify:
                continue
            try:
                pdf_filename = str(row["pdf_filename"] or "")
                if pdf_filename:
                    pdf_path = VERIFIED_CONTRACTS_ROOT / os.path.basename(pdf_filename)
                    with temporary_decrypted_path(pdf_path, pdf_path.name) as mail_pdf_path:
                        _send_mail(
                            row["signer_email"],
                            f"[전자계약 {reason_label} 안내] {row['title_snapshot']}",
                            _void_notice_html(row, reason_label),
                            attachments=mail_pdf_path,
                        )
                else:
                    _send_mail(
                        row["signer_email"],
                        f"[전자계약 {reason_label} 안내] {row['title_snapshot']}",
                        _void_notice_html(row, reason_label),
                    )
                mail_sent += 1
            except Exception as exc:
                mail_failed.append({"name": row["signer_name"], "reason": str(exc)[:200]})
        conn.commit()
    finally:
        conn.close()
    message = f"{processed}건을 {reason_label} 처리했습니다. (계약 정보·파일은 삭제되지 않았습니다)"
    if notify:
        message += f" 안내메일 {mail_sent}건 발송."
    if mail_failed:
        message += f" (안내메일 발송 실패 {len(mail_failed)}건)"
    return jsonify({"status": "success", "message": message, "failed": mail_failed})


@verified_contract_bp.route("/admin/download-selected")
@menu_permission_required("verified_contract_admin")
def download_selected():
    try:
        ids = sorted({int(value) for value in request.args.get("ids", "").split(",") if value})
    except ValueError:
        return "선택한 계약번호를 확인해 주세요.", 400
    if not ids or len(ids) > 500:
        return "완료 계약서를 1~500건 선택해 주세요.", 400
    placeholders = ",".join("?" for _ in ids)
    conn = get_db()
    try:
        rows = conn.execute(
            f"""
            SELECT id, signer_name, pdf_filename
            FROM verified_contracts
            WHERE id IN ({placeholders}) AND status='completed'
            ORDER BY id
            """,
            ids,
        ).fetchall()
    finally:
        conn.close()
    memory = io.BytesIO()
    count = 0
    with zipfile.ZipFile(memory, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        for row in rows:
            path = VERIFIED_CONTRACTS_ROOT / os.path.basename(row["pdf_filename"] or "")
            if not path.is_file():
                continue
            safe_signer = re.sub(r'[\\/:*?"<>|]+', "_", str(row["signer_name"]))[:60]
            archive.writestr(
                f"{row['id']}_{safe_signer}_{path.name}", read_decrypted(path)
            )
            count += 1
    if count == 0:
        return "선택한 항목에 완료된 계약서 파일이 없습니다.", 404
    memory.seek(0)
    return send_file(
        memory,
        mimetype="application/zip",
        as_attachment=True,
        download_name=f"인증전자계약_{datetime.now(KST).strftime('%Y%m%d_%H%M%S')}.zip",
    )


@verified_contract_bp.route("/admin/<int:contract_id>/resend", methods=["POST"])
@menu_permission_required("verified_contract_admin")
@_csrf_required
def resend_invitation(contract_id: int):
    token = secrets.token_urlsafe(32)
    expires_at = _iso(_now() + timedelta(days=7))
    conn = get_db()
    try:
        row = conn.execute(
            "SELECT * FROM verified_contracts WHERE id=?", (contract_id,)
        ).fetchone()
        if not row:
            return jsonify({"status": "error", "message": "계약을 찾을 수 없습니다."}), 404
        if row["status"] == "completed":
            return jsonify({"status": "error", "message": "완료된 계약은 다시 발송할 수 없습니다."}), 400
        update_verified_contract(
            conn,
            contract_id,
            {
                "status": "pending",
                "invitation_token_hash": _token_hash(token),
                "invitation_expires_at": expires_at,
                "opened_at": None,
                "otp_hash": None,
                "otp_expires_at": None,
                "otp_attempts": 0,
                "otp_sent_at": None,
                "verified_at": None,
                "invite_mail_status": "waiting",
                "invite_mail_error": "",
            },
        )
        conn.commit()
        row = conn.execute(
            "SELECT * FROM verified_contracts WHERE id=?", (contract_id,)
        ).fetchone()
    finally:
        conn.close()

    invitation_url = _external_url(token)
    try:
        _invitation_mail(row, invitation_url)
        status, error = "sent", ""
    except Exception as exc:
        status, error = "failed", str(exc)[:500]
    conn = get_db()
    try:
        update_verified_contract(
            conn,
            contract_id,
            {
                "invitation_sent_at": _iso() if status == "sent" else None,
                "invite_mail_status": status,
                "invite_mail_error": error,
            },
        )
        _record_event(
            conn,
            contract_id,
            "INVITATION_RESENT" if status == "sent" else "INVITATION_FAILED",
            {"recipient": row["signer_email"], "error": error},
        )
        conn.commit()
    finally:
        conn.close()
    return jsonify(
        {
            "status": "success" if status == "sent" else "warning",
            "message": "새 인증 링크를 발송했습니다." if status == "sent" else f"메일 발송 실패: {error}",
            "invitation_url": invitation_url,
        }
    )


@verified_contract_bp.route("/admin/<int:contract_id>/revoke", methods=["POST"])
@menu_permission_required("verified_contract_admin")
@_csrf_required
def revoke_contract(contract_id: int):
    conn = get_db()
    try:
        row = conn.execute(
            "SELECT status FROM verified_contracts WHERE id=?", (contract_id,)
        ).fetchone()
        if not row:
            return jsonify({"status": "error", "message": "계약을 찾을 수 없습니다."}), 404
        if row["status"] == "completed":
            return jsonify({"status": "error", "message": "완료된 계약은 취소할 수 없습니다."}), 400
        update_verified_contract(conn, contract_id, {"status": "revoked"})
        _record_event(conn, contract_id, "REVOKED")
        conn.commit()
    finally:
        conn.close()
    return jsonify({"status": "success", "message": "계약 링크를 취소했습니다."})


@verified_contract_bp.route("/admin/<int:contract_id>/evidence")
@menu_permission_required("verified_contract_admin")
def evidence(contract_id: int):
    conn = get_db()
    try:
        row = conn.execute(
            "SELECT * FROM verified_contracts WHERE id=?", (contract_id,)
        ).fetchone()
        events = conn.execute(
            """
            SELECT event_type, event_at, ip_address, user_agent, details_json
            FROM verified_contract_events WHERE contract_id=? ORDER BY id
            """,
            (contract_id,),
        ).fetchall()
    finally:
        conn.close()
    if not row:
        return jsonify({"status": "error", "message": "계약을 찾을 수 없습니다."}), 404
    return jsonify(
        {
            "status": "success",
            "contract": _row_for_view(row),
            "events": [
                {
                    **dict(event),
                    "event_display": _format_kst(event["event_at"]),
                    "details": json.loads(event["details_json"] or "{}"),
                }
                for event in events
            ],
            "agreements": json.loads(row["agreement_snapshot_json"] or "[]"),
        }
    )


@verified_contract_bp.route("/admin/<int:contract_id>/download")
@menu_permission_required("verified_contract_admin")
def admin_download(contract_id: int):
    conn = get_db()
    try:
        row = conn.execute(
            "SELECT pdf_filename FROM verified_contracts WHERE id=?", (contract_id,)
        ).fetchone()
    finally:
        conn.close()
    if not row or not row["pdf_filename"]:
        return "완료된 계약서가 없습니다.", 404
    path = VERIFIED_CONTRACTS_ROOT / os.path.basename(row["pdf_filename"])
    if not path.is_file():
        return "계약서 파일을 찾을 수 없습니다.", 404
    return encrypted_response(path, path.name, as_attachment=True, mimetype='application/pdf')


@verified_contract_bp.route("/admin/terms")
@menu_permission_required("verified_contract_admin")
def get_terms():
    contract_type = str(request.args.get("type", "")).strip()
    if contract_type not in _categories():
        return jsonify({"status": "error", "message": "계약구분을 확인해 주세요."}), 400
    content1, content2 = _read_terms(contract_type)
    profile = _company_snapshot()
    title = _titles().get(contract_type) or _default_title(contract_type, profile["company_name"])
    return jsonify(
        {
            "status": "success",
            "content1": content1,
            "content2": content2,
            "title": title,
        }
    )


@verified_contract_bp.route("/admin/terms", methods=["POST"])
@menu_permission_required("verified_contract_admin")
@_csrf_required
def save_terms():
    data = request.get_json(silent=True) or {}
    contract_type = str(data.get("type", "")).strip()
    if contract_type not in _categories():
        return jsonify({"status": "error", "message": "계약구분을 확인해 주세요."}), 400
    (VERIFIED_TERMS_ROOT / f"{contract_type}.txt").write_text(
        _sanitize_contract_html(data.get("content1", "")), encoding="utf-8"
    )
    (VERIFIED_TERMS_ROOT / f"{contract_type}2.txt").write_text(
        _sanitize_contract_html(data.get("content2", "")), encoding="utf-8"
    )
    titles = _titles()
    title = str(data.get("title", "")).strip()
    if title:
        titles[contract_type] = title
    else:
        titles.pop(contract_type, None)
    _save_json(VERIFIED_TITLES_FILE, titles)
    return jsonify({"status": "success", "message": "인증전자계약 전용 양식을 저장했습니다."})


@verified_contract_bp.route("/admin/categories", methods=["GET", "POST"])
@menu_permission_required("verified_contract_admin")
def add_category():
    if request.method == "GET":
        return jsonify({"status": "success", "categories": _categories()})
    supplied = request.headers.get("X-CSRF-Token") or ""
    expected = session.get("verified_contract_csrf", "")
    if not expected or not supplied or not hmac.compare_digest(str(expected), str(supplied)):
        return jsonify({"status": "error", "message": "보안 확인값이 만료되었습니다."}), 403
    data = request.get_json(silent=True) or {}
    categories = _categories()
    if isinstance(data.get("categories"), list):
        requested = []
        for item in data["categories"]:
            name = str(item).strip()
            if (
                not name
                or len(name) > 40
                or any(character in name for character in '\\/:*?"<>|')
            ):
                return jsonify({"status": "error", "message": f"사용할 수 없는 계약구분: {name}"}), 400
            if name not in requested:
                requested.append(name)
        if not requested:
            return jsonify({"status": "error", "message": "계약구분은 한 개 이상 필요합니다."}), 400
        removed = set(categories) - set(requested)
        if removed:
            conn = get_db()
            try:
                used = {
                    row["contract_type"]
                    for row in conn.execute(
                        "SELECT DISTINCT contract_type FROM verified_contracts"
                    ).fetchall()
                }
            finally:
                conn.close()
            blocked = sorted(removed & used)
            if blocked:
                return jsonify(
                    {
                        "status": "error",
                        "message": "계약 기록이 있는 구분은 삭제할 수 없습니다: " + ", ".join(blocked),
                    }
                ), 400
        categories = requested
        for name in categories:
            (VERIFIED_TERMS_ROOT / f"{name}.txt").touch(exist_ok=True)
            (VERIFIED_TERMS_ROOT / f"{name}2.txt").touch(exist_ok=True)
        _save_json(VERIFIED_CATEGORIES_FILE, categories)
        return jsonify({"status": "success", "message": "계약구분을 저장했습니다.", "categories": categories})

    name = str(data.get("name", "")).strip()
    if not name or len(name) > 40 or any(character in name for character in '\\/:*?"<>|'):
        return jsonify({"status": "error", "message": "사용할 수 없는 계약구분 이름입니다."}), 400
    if name not in categories:
        categories.append(name)
        _save_json(VERIFIED_CATEGORIES_FILE, categories)
        (VERIFIED_TERMS_ROOT / f"{name}.txt").touch(exist_ok=True)
        (VERIFIED_TERMS_ROOT / f"{name}2.txt").touch(exist_ok=True)
    return jsonify({"status": "success", "message": "새 계약구분을 추가했습니다.", "categories": categories})


@verified_contract_bp.route("/admin/settings/mail", methods=["POST"])
@menu_permission_required("verified_contract_admin")
@_csrf_required
def save_mail_settings():
    data = request.get_json(silent=True) or {}
    action = str(data.get("action", "save")).strip()
    account_id = str(data.get("account_id", "")).strip()
    store = _mail_account_store()
    accounts = store["accounts"]

    if action == "select":
        if not any(item["id"] == account_id for item in accounts):
            return jsonify({"status": "error", "message": "발송계정을 찾을 수 없습니다."}), 404
        store["active_account_id"] = account_id
        _save_json(VERIFIED_MAIL_FILE, store)
        return jsonify(
            {
                "status": "success",
                "message": "선택한 계정을 계약 발송계정으로 적용했습니다.",
                "active_account_id": account_id,
            }
        )

    if action == "delete":
        account = next((item for item in accounts if item["id"] == account_id), None)
        if not account:
            return jsonify({"status": "error", "message": "삭제할 발송계정을 찾을 수 없습니다."}), 404
        accounts = [item for item in accounts if item["id"] != account_id]
        if store["active_account_id"] == account_id:
            store["active_account_id"] = accounts[0]["id"] if accounts else ""
        store["accounts"] = accounts
        _save_json(VERIFIED_MAIL_FILE, store)
        return jsonify(
            {
                "status": "success",
                "message": "발송계정을 삭제했습니다.",
                "active_account_id": store["active_account_id"],
                "accounts": _mail_accounts_for_view(store),
            }
        )

    label = str(data.get("label", "")).strip()
    email = str(data.get("email", data.get("username", ""))).strip().lower()
    password = re.sub(r"\s+", "", str(data.get("password", "")))
    if not label or len(label) > 80:
        return jsonify({"status": "error", "message": "계정 이름을 입력해 주세요."}), 400
    if not re.fullmatch(r"[^@\s]+@[^@\s]+\.[^@\s]+", email) or len(email) > 254:
        return jsonify({"status": "error", "message": "발송 메일주소를 확인해 주세요."}), 400
    duplicate = next(
        (
            item for item in accounts
            if item["email"].lower() == email and item["id"] != account_id
        ),
        None,
    )
    if duplicate:
        return jsonify({"status": "error", "message": "같은 발송 메일주소가 이미 등록되어 있습니다."}), 409

    if action == "add":
        if len(accounts) >= MAX_MAIL_ACCOUNTS:
            return jsonify(
                {
                    "status": "error",
                    "message": f"발송계정은 최대 {MAX_MAIL_ACCOUNTS}개까지 등록할 수 있습니다.",
                }
            ), 400
        if not password:
            return jsonify({"status": "error", "message": "새 계정의 앱 비밀번호를 입력해 주세요."}), 400
        account_id = f"verified-mail-{secrets.token_hex(8)}"
        account = {"id": account_id}
        accounts.append(account)
    else:
        account = next((item for item in accounts if item["id"] == account_id), None)
        if not account:
            return jsonify({"status": "error", "message": "수정할 발송계정을 선택해 주세요."}), 404
        if not password and not account.get("encrypted_password"):
            return jsonify({"status": "error", "message": "앱 비밀번호를 입력해 주세요."}), 400

    account.update(
        {
            "label": label,
            "email": email,
            "encrypted_password": (
                _encrypt_sensitive(password)
                if password
                else account.get("encrypted_password", "")
            ),
        }
    )
    store["active_account_id"] = account_id
    store["accounts"] = accounts
    _save_json(VERIFIED_MAIL_FILE, store)
    return jsonify(
        {
            "status": "success",
            "message": (
                "새 발송계정을 저장하고 적용했습니다."
                if action == "add"
                else "발송계정 정보를 수정하고 적용했습니다."
            ),
            "active_account_id": account_id,
            "accounts": _mail_accounts_for_view(store),
        }
    )


@verified_contract_bp.route("/admin/settings/company", methods=["POST"])
@menu_permission_required("verified_contract_admin")
@_csrf_required
def save_company_settings():
    settings = _company_settings()
    profile_id = str(request.form.get("profile_id", "")).strip()
    action = str(request.form.get("action", "save")).strip()
    profiles = settings["profiles"]
    if action == "select":
        if not any(item["id"] == profile_id for item in profiles):
            return jsonify({"status": "error", "message": "적용할 발송회사를 찾을 수 없습니다."}), 404
        settings["active_profile_id"] = profile_id
        _save_json(VERIFIED_COMPANY_FILE, settings)
        return jsonify({"status": "success", "message": "선택한 회사를 기본 발송회사로 적용했습니다."})
    if action == "delete":
        if len(profiles) <= 1:
            return jsonify({"status": "error", "message": "발송회사는 최소 1개 이상 필요합니다."}), 400
        profile = next((item for item in profiles if item["id"] == profile_id), None)
        if not profile:
            return jsonify({"status": "error", "message": "삭제할 발송회사를 찾을 수 없습니다."}), 404
        profiles = [item for item in profiles if item["id"] != profile_id]
        if settings["active_profile_id"] == profile_id:
            settings["active_profile_id"] = profiles[0]["id"]
        settings["profiles"] = profiles
        _save_json(VERIFIED_COMPANY_FILE, settings)
        for key, root in (
            ("logo_filename", VERIFIED_LOGO_ROOT),
            ("stamp_filename", VERIFIED_STAMP_ROOT),
        ):
            if profile.get(key):
                delete_file(root / os.path.basename(profile[key]))
        return jsonify({"status": "success", "message": "발송회사를 삭제했습니다."})
    if action == "add":
        if len(profiles) >= MAX_COMPANY_PROFILES:
            return jsonify(
                {
                    "status": "error",
                    "message": f"회사정보는 최대 {MAX_COMPANY_PROFILES}개까지 등록할 수 있습니다.",
                }
            ), 400
        profile_id = f"verified-{secrets.token_hex(8)}"
        profile = {"id": profile_id, "logo_filename": "", "stamp_filename": ""}
        profiles.append(profile)
    else:
        profile = next((item for item in profiles if item["id"] == profile_id), None)
        if not profile:
            return jsonify({"status": "error", "message": "회사정보를 찾을 수 없습니다."}), 404
    profile.update(
        {
            "label": str(request.form.get("label", "")).strip() or "회사정보",
            "company_name": str(request.form.get("company_name", "")).strip(),
            "representative_title": str(request.form.get("representative_title", "")).strip(),
            "representative_name": str(request.form.get("representative_name", "")).strip(),
        }
    )
    for optional_field in ("business_number", "address", "phone"):
        if optional_field in request.form:
            profile[optional_field] = str(request.form.get(optional_field, "")).strip()
    if not profile["company_name"]:
        return jsonify({"status": "error", "message": "회사명을 입력해 주세요."}), 400
    if not profile["representative_name"]:
        return jsonify({"status": "error", "message": "대표자 이름을 입력해 주세요."}), 400

    new_asset_paths = []
    replaced_assets = []
    try:
        for asset_kind, upload_key, root, fallback, label in (
            ("logo", "logo_file", VERIFIED_LOGO_ROOT, "logo.png", "회사 로고"),
            ("stamp", "stamp_file", VERIFIED_STAMP_ROOT, "stamp.png", "회사 도장"),
        ):
            upload = request.files.get(upload_key)
            if not upload or not upload.filename:
                continue
            display_name = original_filename(upload.filename, fallback)
            extension = Path(display_name).suffix.lower()
            if extension not in {".png", ".jpg", ".jpeg", ".gif", ".webp"}:
                raise ValueError(f"{label}는 PNG, JPG, GIF 또는 WEBP 파일만 사용할 수 있습니다.")
            filename = encrypted_storage_name(display_name)
            new_path = root / filename
            old_filename = str(profile.get(f"{asset_kind}_filename") or "")
            encrypt_upload(upload, new_path)
            new_asset_paths.append(new_path)
            if old_filename and old_filename != filename:
                replaced_assets.append((root, old_filename))
            profile[f"{asset_kind}_filename"] = filename
            profile[f"{asset_kind}_original_name"] = display_name
    except ValueError as exc:
        for path in new_asset_paths:
            delete_file(path)
        return jsonify({"status": "error", "message": str(exc)}), 400
    except Exception:
        for path in new_asset_paths:
            delete_file(path)
        raise
    settings["active_profile_id"] = profile_id
    settings["profiles"] = profiles
    try:
        _save_json(VERIFIED_COMPANY_FILE, settings)
    except Exception:
        for path in new_asset_paths:
            delete_file(path)
        raise
    for root, old_filename in replaced_assets:
        delete_file(root / os.path.basename(old_filename))
    return jsonify({"status": "success", "message": "인증전자계약 전용 회사정보를 저장했습니다."})


@verified_contract_bp.route("/sign/<string:token>")
def public_contract(token: str):
    conn = get_db()
    try:
        row = _load_by_token(conn, token)
        available, message = _contract_available(row)
        if row and not available and row["status"] == "pending":
            update_verified_contract(conn, row["id"], {"status": "expired"})
            _record_event(conn, row, "EXPIRED")
            conn.commit()
        if row and available and not row["opened_at"]:
            update_verified_contract(conn, row["id"], {"opened_at": _iso()})
            _record_event(conn, row, "LINK_OPENED")
            conn.commit()
            row = _load_by_token(conn, token)
    finally:
        conn.close()
    if not available:
        return render_template("verified_contract/public.html", state="error", message=message), 410
    if row["status"] == "completed":
        return render_template(
            "verified_contract/public.html",
            state="completed",
            data=_row_for_view(row),
            token=token,
            can_download=_public_verified(row, token),
        )
    if not _public_verified(row, token):
        return render_template(
            "verified_contract/public.html",
            state="verify",
            data=_row_for_view(row),
            masked_email=_mask_email(row["signer_email"]),
            token=token,
            csrf_token=_csrf_token(),
        )
    contract_data = json.loads(row["contract_data_json"] or "{}")
    company = json.loads(row["company_snapshot_json"] or "{}")
    values = _public_values(row, contract_data, company)
    return render_template(
        "verified_contract/public.html",
        state="sign",
        data=_row_for_view(row),
        token=token,
        csrf_token=_csrf_token(),
        agreements=json.loads(row["agreement_snapshot_json"] or "[]"),
        content1=_render_terms(row["terms1_snapshot"], values),
        content2=_render_terms(row["terms2_snapshot"], values),
        contract_values=contract_data,
        company=company,
    )


@verified_contract_bp.route("/sign/<string:token>/send-code", methods=["POST"])
@_csrf_required
def send_otp(token: str):
    conn = get_db()
    try:
        row = _load_by_token(conn, token)
        available, message = _contract_available(row)
        if not available or row["status"] != "pending":
            return jsonify({"status": "error", "message": message or "인증할 수 없는 계약입니다."}), 410
        sent_at = _parse_iso(row["otp_sent_at"])
        if sent_at and (_now() - sent_at).total_seconds() < 60:
            return jsonify({"status": "error", "message": "인증번호는 1분 후 다시 요청할 수 있습니다."}), 429
        code = f"{secrets.randbelow(900000) + 100000:06d}"
        update_verified_contract(
            conn,
            row["id"],
            {
                "otp_hash": generate_password_hash(code),
                "otp_expires_at": _iso(_now() + timedelta(minutes=5)),
                "otp_attempts": 0,
                "otp_sent_at": _iso(),
            },
        )
        conn.commit()
    finally:
        conn.close()
    try:
        _send_mail(
            row["signer_email"],
            "[새담 인증전자계약] 이메일 인증번호",
            f"""
            <div style="font-family:Arial,'Malgun Gothic',sans-serif;line-height:1.7">
              <h2>이메일 인증번호</h2>
              <p>{escape(row['signer_name'])}님의 인증번호는 다음과 같습니다.</p>
              <div style="font-size:30px;font-weight:bold;letter-spacing:8px;color:#123b6d">{code}</div>
              <p>5분 안에 계약 화면에 입력해 주세요. 타인에게 알려주지 마세요.</p>
            </div>
            """,
        )
    except Exception as exc:
        return jsonify({"status": "error", "message": f"인증번호 메일 발송 실패: {str(exc)[:200]}"}), 500
    conn = get_db()
    try:
        _record_event(conn, row, "OTP_SENT", {"recipient": _mask_email(row["signer_email"])})
        conn.commit()
    finally:
        conn.close()
    return jsonify({"status": "success", "message": f"{_mask_email(row['signer_email'])}로 인증번호를 보냈습니다."})


@verified_contract_bp.route("/sign/<string:token>/verify-code", methods=["POST"])
@_csrf_required
def verify_otp(token: str):
    code = str((request.get_json(silent=True) or {}).get("code", "")).strip()
    conn = get_db()
    try:
        row = _load_by_token(conn, token)
        available, message = _contract_available(row)
        if not available or row["status"] != "pending":
            return jsonify({"status": "error", "message": message or "인증할 수 없는 계약입니다."}), 410
        expiry = _parse_iso(row["otp_expires_at"])
        if not row["otp_hash"] or not expiry or expiry < _now():
            return jsonify({"status": "error", "message": "인증번호가 만료되었습니다. 새로 받아주세요."}), 400
        attempts = int(row["otp_attempts"] or 0)
        if attempts >= 5:
            return jsonify({"status": "error", "message": "입력 횟수를 초과했습니다. 인증번호를 새로 받아주세요."}), 429
        if not re.fullmatch(r"\d{6}", code) or not check_password_hash(row["otp_hash"], code):
            update_verified_contract(conn, row["id"], {"otp_attempts": attempts + 1})
            _record_event(conn, row, "OTP_FAILED", {"attempt": attempts + 1})
            conn.commit()
            return jsonify({"status": "error", "message": "인증번호가 일치하지 않습니다."}), 400
        verified_at = _iso()
        update_verified_contract(
            conn,
            row["id"],
            {"verified_at": verified_at, "otp_hash": None, "otp_expires_at": None},
        )
        _record_event(conn, row, "OTP_VERIFIED")
        conn.commit()
    finally:
        conn.close()
    access = session.get("verified_contract_access", {})
    if not isinstance(access, dict):
        access = {}
    access[str(row["id"])] = _token_hash(token)[:24]
    session["verified_contract_access"] = access
    return jsonify({"status": "success", "message": "본인 이메일 인증이 완료되었습니다."})


def _public_values(row, contract_data: dict, company: dict) -> dict[str, str]:
    values = {
        key: escape(str(value or ""))
        for key, value in contract_data.items()
    }
    values.update(
        {
            "성명": escape(row["signer_name"]),
            "email": escape(row["signer_email"]),
            "회사명": escape(company.get("company_name", "")),
            "대표직함": escape(company.get("representative_title", "")),
            "대표자명": escape(company.get("representative_name", "")),
            "계약서제목": escape(row["title_snapshot"]),
        }
    )
    return values


def _decode_signature(data_url: str) -> bytes:
    match = re.fullmatch(r"data:image/png;base64,([A-Za-z0-9+/=\r\n]+)", str(data_url or ""))
    if not match:
        raise ValueError("서명 형식이 올바르지 않습니다.")
    raw = base64.b64decode(match.group(1), validate=True)
    if not 200 <= len(raw) <= 1_500_000:
        raise ValueError("서명 데이터 크기가 올바르지 않습니다.")
    try:
        image = Image.open(io.BytesIO(raw)).convert("RGBA")
        if image.width > 1600 or image.height > 800:
            raise ValueError("서명 이미지가 너무 큽니다.")
        alpha = image.getchannel("A")
        if alpha.getbbox() is None:
            raise ValueError("서명란이 비어 있습니다.")
        colors = image.getcolors(maxcolors=image.width * image.height)
        if colors is not None and len(colors) < 2:
            raise ValueError("서명란이 비어 있습니다.")
    except (OSError, ValueError) as exc:
        if isinstance(exc, ValueError):
            raise
        raise ValueError("서명 이미지를 확인할 수 없습니다.") from exc
    return raw


_PDF_FONT_CSS = None


def _pdf_font_css() -> str:
    global _PDF_FONT_CSS
    if _PDF_FONT_CSS is not None:
        return _PDF_FONT_CSS
    files = {
        "regular": (
            VERIFIED_PDF_FONT_ROOT / "NanumGothic-Regular.ttf",
            "https://cdn.jsdelivr.net/gh/google/fonts@main/ofl/nanumgothic/NanumGothic-Regular.ttf",
        ),
        "bold": (
            VERIFIED_PDF_FONT_ROOT / "NanumGothic-Bold.ttf",
            "https://cdn.jsdelivr.net/gh/google/fonts@main/ofl/nanumgothic/NanumGothic-Bold.ttf",
        ),
    }
    encoded = {}
    for weight, (path, url) in files.items():
        if not path.is_file():
            try:
                req = Request(url, headers={"User-Agent": "Mozilla/5.0"})
                with urlopen(req, timeout=30) as response:
                    content = response.read()
                if len(content) > 100_000:
                    path.write_bytes(content)
            except Exception:
                pass
        if path.is_file():
            encoded[weight] = base64.b64encode(path.read_bytes()).decode("ascii")
    rules = []
    if encoded.get("regular"):
        rules.append(
            "@font-face{font-family:'VerifiedNanum';font-weight:400;"
            f"src:url(data:font/ttf;base64,{encoded['regular']}) format('truetype');}}"
        )
    if encoded.get("bold"):
        rules.append(
            "@font-face{font-family:'VerifiedNanum';font-weight:700;"
            f"src:url(data:font/ttf;base64,{encoded['bold']}) format('truetype');}}"
        )
    rules.append("html,body,body *{font-family:'VerifiedNanum','Malgun Gothic',sans-serif!important;}")
    _PDF_FONT_CSS = "\n".join(rules)
    return _PDF_FONT_CSS


def _pdf_configuration():
    candidates = [
        os.environ.get("VERIFIED_WKHTMLTOPDF_PATH", ""),
        "/usr/bin/wkhtmltopdf",
        "/usr/local/bin/wkhtmltopdf",
        "/opt/render/project/src/.apt/usr/bin/wkhtmltopdf",
        r"C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe",
        r"C:\Program Files (x86)\wkhtmltopdf\bin\wkhtmltopdf.exe",
        shutil.which("wkhtmltopdf") or "",
    ]
    path = next((candidate for candidate in candidates if candidate and os.path.exists(candidate)), None)
    return pdfkit.configuration(wkhtmltopdf=path) if path else None


def _stamp_data_uri(company: dict) -> str:
    filename = os.path.basename(str(company.get("stamp_filename", "")))
    path = VERIFIED_STAMP_ROOT / filename if filename else VERIFIED_STAMP_ROOT / "verified_default_stamp.png"
    if not path.is_file():
        return ""
    mime, _ = mimetypes.guess_type(path.name)
    return f"data:{mime or 'image/png'};base64,{base64.b64encode(read_decrypted(path)).decode('ascii')}"


def _build_pdf(row, contract_data: dict, company: dict, signature_uri: str, signed_at: datetime):
    values = _public_values(row, contract_data, company)
    values["연락처"] = escape(str(contract_data.get("연락처", "")))
    values["거주지"] = escape(str(contract_data.get("거주지", "")))
    content1 = _render_terms(row["terms1_snapshot"], values)
    content2 = _render_terms(row["terms2_snapshot"], values)
    agreements = json.loads(row["agreement_snapshot_json"] or "[]")
    agreement_html = "".join(
        f"<li><b>{escape(item['title'])}</b>: {escape(item['text'])}</li>"
        for item in agreements
    )
    stamp = _stamp_data_uri(company)
    company_text = " ".join(
        escape(str(company.get(key, "")))
        for key in ("company_name", "representative_title", "representative_name")
        if company.get(key)
    )
    html = f"""
    <!doctype html><html><head><meta charset="utf-8"><style>
    {_pdf_font_css()}
    body{{color:#111;font-size:15px;line-height:1.72;word-break:keep-all}}
    h1{{font-size:24px;text-align:center;text-decoration:underline;margin:10px 0 28px}}
    .info{{width:100%;border-collapse:collapse;margin-bottom:24px;table-layout:fixed}}
    .info th,.info td{{border-bottom:1px solid #ccc;padding:8px;text-align:left}}
    .info th{{width:110px}}
    .terms table{{width:100%;border-collapse:collapse;margin:12px 0}}
    .terms th,.terms td{{border:1px solid #333;padding:7px}}
    .terms p{{margin:0 0 8px}}
    .sign{{margin-top:35px;min-height:170px;page-break-inside:avoid}}
    .party{{width:48%;display:inline-block;vertical-align:top;position:relative}}
    .evidence{{border:1px solid #9fb3c8;background:#f5f8fb;padding:14px;margin-top:25px;font-size:12px}}
    .evidence li{{margin:5px 0}}
    </style></head><body>
      <div style="text-align:center;margin-bottom:14px"><img src="https://www.saedam.org/img/logo01.gif" style="max-width:112px"></div>
      <h1>{escape(row['title_snapshot'])}</h1>
      <table class="info">
        <tr><th>학교/부서</th><td>{values.get('수탁학교명','')} / {values.get('부서명','')}</td><th>계약자</th><td>{escape(row['signer_name'])}</td></tr>
        <tr><th>이메일</th><td>{escape(row['signer_email'])}</td><th>연락처</th><td>{values.get('연락처','')}</td></tr>
        <tr><th>주민번호</th><td>{values.get('주민번호','')}</td><th>은행</th><td>{values.get('은행','')}</td></tr>
        <tr><th>계좌번호</th><td colspan="3">{values.get('계좌번호','')}</td></tr>
        <tr><th>주소</th><td colspan="3">{values.get('거주지','')}</td></tr>
      </table>
      <div class="terms">{content1}</div>
      <div class="sign">
        <p style="text-align:center">{signed_at.astimezone(KST).strftime('%Y년 %m월 %d일')}</p>
        <div class="party"><b>[계약기관]</b><p>{company_text}</p>
          {f'<img src="{stamp}" style="width:85px;position:absolute;right:45px;top:18px">' if stamp else ''}
        </div>
        <div class="party"><b>[계약자]</b><p>성명: {escape(row['signer_name'])}<br>
          서명: <img src="{signature_uri}" style="width:150px;max-height:70px;border-bottom:1px solid #222;vertical-align:middle"></p>
        </div>
      </div>
      {f'<div style="page-break-before:always"></div><div class="terms">{content2}</div>' if content2.strip() else ''}
      <div class="evidence"><b>전자계약 확인기록</b><ul>{agreement_html}</ul>
        <p>이메일 인증 완료: {escape(_format_kst(row['verified_at']))}<br>
        전자서명 완료: {signed_at.astimezone(KST).strftime('%Y-%m-%d %H:%M:%S KST')}<br>
        계약서 버전: {int(row['version'])}</p>
      </div>
    </body></html>
    """
    configuration = _pdf_configuration()
    if not configuration:
        raise RuntimeError("인증전자계약 전용 PDF 엔진(wkhtmltopdf)을 찾을 수 없습니다.")
    filename = f"verified_contract_{row['id']}_v{row['version']}_{signed_at.astimezone(KST).strftime('%Y%m%d_%H%M%S')}.pdf"
    path = VERIFIED_CONTRACTS_ROOT / filename
    descriptor, temporary_path = tempfile.mkstemp(prefix='saedam-verified-', suffix='.pdf')
    os.close(descriptor)
    try:
        pdfkit.from_string(
            html,
            temporary_path,
            configuration=configuration,
            options={
                "encoding": "UTF-8",
                "enable-local-file-access": None,
                "print-media-type": None,
                "page-size": "A4",
                "margin-top": "18mm",
                "margin-right": "17mm",
                "margin-bottom": "18mm",
                "margin-left": "17mm",
            },
        )
        with open(temporary_path, 'rb') as source:
            pdf_hash = hashlib.sha256(source.read()).hexdigest()
            source.seek(0)
            encrypt_stream(source, path)
        return path, pdf_hash, temporary_path
    except Exception:
        delete_file(temporary_path)
        delete_file(path)
        raise


@verified_contract_bp.route("/sign/<string:token>/complete", methods=["POST"])
@_csrf_required
def complete_contract(token: str):
    data = request.get_json(silent=True) or {}
    conn = get_db()
    try:
        row = _load_by_token(conn, token)
    finally:
        conn.close()
    available, message = _contract_available(row)
    if not available or row["status"] != "pending":
        return jsonify({"status": "error", "message": message or "완료할 수 없는 계약입니다."}), 410
    if not _public_verified(row, token):
        return jsonify({"status": "error", "message": "이메일 인증을 먼저 완료해 주세요."}), 403

    agreement_keys = {item["key"] for item in json.loads(row["agreement_snapshot_json"] or "[]")}
    accepted = {
        str(key)
        for key, value in (data.get("agreements") or {}).items()
        if value is True
    }
    if agreement_keys != accepted:
        return jsonify({"status": "error", "message": "모든 확인사항에 동의해 주세요."}), 400
    confirmed_name = re.sub(r"\s+", "", str(data.get("confirmed_name", "")))
    expected_name = re.sub(r"\s+", "", str(row["signer_name"]))
    if not confirmed_name or not hmac.compare_digest(
        confirmed_name.encode("utf-8"),
        expected_name.encode("utf-8"),
    ):
        return jsonify({"status": "error", "message": "직접 입력한 성명이 계약자 성명과 일치하지 않습니다."}), 400
    phone = str(data.get("phone", "")).strip()
    address = str(data.get("address", "")).strip()
    if not phone or len(phone) > 50 or not address or len(address) > 300:
        return jsonify({"status": "error", "message": "연락처와 주소를 정확히 입력해 주세요."}), 400
    try:
        resident_number = _normalize_rrn(data.get("resident_number"))
        bank_name = str(data.get("bank_name", "")).strip()
        if not bank_name or len(bank_name) > 50:
            raise ValueError("은행명을 정확히 입력해 주세요.")
        account_number = _normalize_account(data.get("account_number"))
    except ValueError as exc:
        return jsonify({"status": "error", "message": str(exc)}), 400
    try:
        signature_bytes = _decode_signature(str(data.get("signature", "")))
    except ValueError as exc:
        return jsonify({"status": "error", "message": str(exc)}), 400

    signature_filename = f"signature_{row['id']}_{secrets.token_hex(8)}.png"
    signature_path = VERIFIED_SIGNATURE_ROOT / signature_filename
    encrypt_bytes(signature_bytes, signature_path)
    signature_uri = f"data:image/png;base64,{base64.b64encode(signature_bytes).decode('ascii')}"
    contract_data = json.loads(row["contract_data_json"] or "{}")
    contract_data["연락처"] = phone
    contract_data["거주지"] = address
    pdf_contract_data = {
        **contract_data,
        "주민번호": resident_number,
        "은행": bank_name,
        "계좌번호": account_number,
    }
    company = json.loads(row["company_snapshot_json"] or "{}")
    signed_at = _now()
    try:
        pdf_path, pdf_hash, temporary_pdf_path = _build_pdf(
            row, pdf_contract_data, company, signature_uri, signed_at
        )
    except Exception:
        delete_file(signature_path)
        raise

    agreement_evidence = [
        {**item, "accepted": True, "accepted_at": _iso(signed_at)}
        for item in json.loads(row["agreement_snapshot_json"] or "[]")
    ]
    conn = get_db()
    try:
        current = conn.execute(
            "SELECT status FROM verified_contracts WHERE id=?", (row["id"],)
        ).fetchone()
        if not current or current["status"] != "pending":
            delete_file(pdf_path)
            delete_file(signature_path)
            delete_file(temporary_pdf_path)
            return jsonify({"status": "error", "message": "이미 처리된 계약입니다."}), 409
        update_verified_contract(
            conn,
            row["id"],
            {
                "status": "completed",
                "signer_phone": phone,
                "signer_address": address,
                "signer_rrn_encrypted": _encrypt_sensitive(resident_number),
                "signer_bank_encrypted": _encrypt_sensitive(bank_name),
                "signer_account_encrypted": _encrypt_sensitive(account_number),
                "contract_data_json": json.dumps(contract_data, ensure_ascii=False),
                "agreement_snapshot_json": json.dumps(agreement_evidence, ensure_ascii=False),
                "confirmed_name": str(data.get("confirmed_name", "")).strip(),
                "signature_filename": signature_filename,
                "signed_at": _iso(signed_at),
                "ip_address": _client_ip(),
                "user_agent": _user_agent(),
                "pdf_filename": pdf_path.name,
                "pdf_sha256": pdf_hash,
                "completion_mail_status": "waiting",
            },
        )
        _record_event(
            conn,
            row,
            "COMPLETED",
            {
                "pdf_sha256": pdf_hash,
                "version": int(row["version"]),
                "agreement_keys": sorted(agreement_keys),
                "confirmed_name": str(data.get("confirmed_name", "")).strip(),
            },
        )
        conn.commit()
    except Exception:
        conn.rollback()
        delete_file(pdf_path)
        delete_file(signature_path)
        delete_file(temporary_pdf_path)
        raise
    finally:
        conn.close()

    sender = _mail_settings()["MAIL_USERNAME"]
    recipients = [row["signer_email"]]
    if sender and sender.lower() != row["signer_email"].lower():
        recipients.append(sender)
    try:
        with temporary_decrypted_path(pdf_path, pdf_path.name) as mail_pdf_path:
            _send_mail(
                recipients,
                f"[계약완료] {row['title_snapshot']}",
                (
                    f"{row['signer_name']}님의 인증전자계약이 완료되었습니다.<br>"
                    "첨부된 최종 계약서를 확인해 주세요.<br><br>"
                    "계약서 위변조 확인용 고유번호(SHA-256): "
                    f"<span style='font-family:monospace'>{pdf_hash}</span><br>"
                    "<span style='font-size:12px;color:#64748b'>"
                    "첨부 계약서가 이후 변경되지 않았는지 확인할 때 사용하는 번호이며, "
                    "별도로 입력하실 필요는 없습니다.</span>"
                ),
                attachments=mail_pdf_path,
            )
        mail_status, mail_error = "sent", ""
    except Exception as exc:
        mail_status, mail_error = "failed", str(exc)[:500]
    conn = get_db()
    try:
        update_verified_contract(
            conn,
            row["id"],
            {
                "completion_mail_status": mail_status,
                "completion_mail_error": mail_error,
            },
        )
        _record_event(
            conn,
            row,
            "COMPLETION_MAIL_SENT" if mail_status == "sent" else "COMPLETION_MAIL_FAILED",
            {"error": mail_error},
        )
        conn.commit()
    finally:
        conn.close()
        delete_file(temporary_pdf_path)
    return jsonify(
        {
            "status": "success",
            "message": (
                "계약이 완료되었고 최종본을 이메일로 발송했습니다."
                if mail_status == "sent"
                else "계약은 완료되었지만 최종본 이메일 발송에 실패했습니다. 화면에서 내려받을 수 있습니다."
            ),
        }
    )


@verified_contract_bp.route("/sign/<string:token>/download")
def public_download(token: str):
    conn = get_db()
    try:
        row = _load_by_token(conn, token)
    finally:
        conn.close()
    if not row or row["status"] != "completed" or not _public_verified(row, token):
        return "이메일 인증을 완료한 계약자만 내려받을 수 있습니다.", 403
    path = VERIFIED_CONTRACTS_ROOT / os.path.basename(row["pdf_filename"] or "")
    if not path.is_file():
        return "계약서 파일을 찾을 수 없습니다.", 404
    return encrypted_response(path, path.name, as_attachment=True, mimetype='application/pdf')
