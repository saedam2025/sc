import sqlite3
import os
import secrets
from datetime import date, datetime

from .storage import (
    AI_MAIL_UPLOADS as _AI_MAIL_UPLOADS,
    APP_ROOT,
    DATA_ROOT,
    GALLERY_ROOT as _GALLERY_ROOT,
    GALLERY_THUMBS as _GALLERY_THUMBS,
    GALLERY_UPLOADS as _GALLERY_UPLOADS,
    MAIN_DB_FILE,
    PROFILE_ROOT as _PROFILE_ROOT,
    SCHOOL_UPLOADS as _SCHOOL_UPLOADS,
    bootstrap_legacy_files,
)


BASE_DIR = str(DATA_ROOT)
DB_FILE = str(MAIN_DB_FILE)
GALLERY_ROOT = str(_GALLERY_ROOT)
GALLERY_UPLOADS = str(_GALLERY_UPLOADS)
GALLERY_THUMBS = str(_GALLERY_THUMBS)
PROFILE_ROOT = str(_PROFILE_ROOT)
SCHOOL_UPLOADS = str(_SCHOOL_UPLOADS)
AI_MAIL_UPLOADS = str(_AI_MAIL_UPLOADS)


def _bootstrap_main_database():
    """빈 영구 디스크에는 프로젝트 DB를 자동으로 최초 복제한다."""
    bundled_db = APP_ROOT / 'saedam.db'
    target_db = MAIN_DB_FILE
    if target_db.exists() or not bundled_db.is_file():
        return
    target_db.parent.mkdir(parents=True, exist_ok=True)
    source_conn = sqlite3.connect(str(bundled_db))
    target_conn = sqlite3.connect(str(target_db))
    try:
        source_conn.backup(target_conn)
    finally:
        target_conn.close()
        source_conn.close()


bootstrap_legacy_files()
_bootstrap_main_database()

def get_db():
    conn = sqlite3.connect(DB_FILE)
    conn.row_factory = sqlite3.Row
    return conn


def ensure_certificate_schema(conn):
    """증명발급 신청을 saedam.db에 저장하기 위한 표준 스키마."""
    conn.execute('''
        CREATE TABLE IF NOT EXISTS certificate_requests (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            applied_date TEXT,
            applicant_type TEXT,
            certificate_type TEXT NOT NULL,
            applicant_name TEXT NOT NULL,
            resident_number TEXT,
            home_address TEXT,
            work_start_date TEXT,
            work_end_date TEXT,
            workplace TEXT,
            subject_or_duty TEXT,
            purpose TEXT,
            position TEXT,
            email TEXT,
            status TEXT NOT NULL DEFAULT '대기',
            issued_date TEXT,
            issue_number TEXT,
            termination_reason TEXT,
            filename TEXT,
            legacy_row_number INTEGER UNIQUE,
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            updated_at DATETIME DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    request_columns = {
        row['name'] if hasattr(row, 'keys') else row[1]
        for row in conn.execute('PRAGMA table_info(certificate_requests)').fetchall()
    }
    request_additions = {
        'workgroup_id': 'INTEGER',
        'company_id': 'INTEGER',
        'workgroup_name': "TEXT NOT NULL DEFAULT ''",
        'company_name': "TEXT NOT NULL DEFAULT ''",
    }
    for column_name, column_ddl in request_additions.items():
        if column_name not in request_columns:
            conn.execute(
                f'ALTER TABLE certificate_requests ADD COLUMN {column_name} {column_ddl}'
            )

    conn.execute('''
        CREATE TABLE IF NOT EXISTS certificate_companies (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            company_name TEXT NOT NULL,
            representative_name TEXT NOT NULL,
            business_number TEXT NOT NULL DEFAULT '',
            address TEXT NOT NULL DEFAULT '',
            phone TEXT NOT NULL DEFAULT '',
            seal_filename TEXT NOT NULL DEFAULT '',
            seal_path TEXT NOT NULL DEFAULT '',
            logo_filename TEXT NOT NULL DEFAULT '',
            logo_path TEXT NOT NULL DEFAULT '',
            is_active INTEGER NOT NULL DEFAULT 1,
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            updated_at DATETIME DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    company_columns = {
        row['name'] if hasattr(row, 'keys') else row[1]
        for row in conn.execute('PRAGMA table_info(certificate_companies)').fetchall()
    }
    for column_name in ('logo_filename', 'logo_path'):
        if column_name not in company_columns:
            conn.execute(
                f"ALTER TABLE certificate_companies ADD COLUMN {column_name} TEXT NOT NULL DEFAULT ''"
            )
    conn.execute('''
        CREATE TABLE IF NOT EXISTS certificate_workgroups (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            company_id INTEGER NOT NULL,
            sender_id INTEGER,
            access_token TEXT NOT NULL UNIQUE,
            allow_instructor INTEGER NOT NULL DEFAULT 1,
            allow_employee INTEGER NOT NULL DEFAULT 1,
            is_active INTEGER NOT NULL DEFAULT 1,
            created_by TEXT NOT NULL DEFAULT '',
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            updated_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (company_id) REFERENCES certificate_companies(id),
            FOREIGN KEY (sender_id) REFERENCES ai_mail_senders(id)
        )
    ''')
    conn.execute(
        'CREATE INDEX IF NOT EXISTS idx_certificate_requests_status '
        'ON certificate_requests(status)'
    )
    conn.execute(
        'CREATE INDEX IF NOT EXISTS idx_certificate_requests_type '
        'ON certificate_requests(certificate_type)'
    )
    conn.execute(
        'CREATE INDEX IF NOT EXISTS idx_certificate_requests_name '
        'ON certificate_requests(applicant_name)'
    )
    conn.execute(
        'CREATE INDEX IF NOT EXISTS idx_certificate_requests_workgroup '
        'ON certificate_requests(workgroup_id)'
    )
    conn.execute(
        'CREATE INDEX IF NOT EXISTS idx_certificate_workgroups_active '
        'ON certificate_workgroups(is_active, company_id)'
    )


def _legacy_certificate_value(value):
    if value is None:
        return ''
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    return str(value).strip()


def migrate_legacy_certificates(conn):
    """기존 certificates.xlsx를 최초 1회만 SQLite로 안전하게 이관한다."""
    conn.execute('''
        CREATE TABLE IF NOT EXISTS admin_settings (
            key TEXT PRIMARY KEY,
            value TEXT,
            updated_at DATETIME DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    ensure_certificate_schema(conn)

    migration_key = 'certificate_sqlite_migration_v1'
    migrated = conn.execute(
        'SELECT value FROM admin_settings WHERE key=?',
        (migration_key,),
    ).fetchone()
    if migrated:
        return int(conn.execute(
            'SELECT COUNT(*) FROM certificate_requests'
        ).fetchone()[0])

    existing_count = int(conn.execute(
        'SELECT COUNT(*) FROM certificate_requests'
    ).fetchone()[0])
    if existing_count:
        conn.execute('''
            INSERT INTO admin_settings (key, value, updated_at)
            VALUES (?, ?, CURRENT_TIMESTAMP)
            ON CONFLICT(key) DO UPDATE SET
                value=excluded.value,
                updated_at=CURRENT_TIMESTAMP
        ''', (migration_key, f'existing:{existing_count}'))
        return existing_count

    candidates = [
        os.path.join(BASE_DIR, 'certificates.xlsx'),
        str(APP_ROOT / 'certificates.xlsx'),
    ]
    legacy_path = next(
        (path for path in candidates if os.path.exists(path)),
        None,
    )
    if not legacy_path:
        conn.execute('''
            INSERT INTO admin_settings (key, value, updated_at)
            VALUES (?, 'no_legacy_file', CURRENT_TIMESTAMP)
            ON CONFLICT(key) DO UPDATE SET
                value=excluded.value,
                updated_at=CURRENT_TIMESTAMP
        ''', (migration_key,))
        return 0

    from openpyxl import load_workbook

    workbook = load_workbook(legacy_path, read_only=True, data_only=True)
    worksheet = workbook.active
    rows = worksheet.iter_rows(values_only=True)
    headers = [_legacy_certificate_value(value) for value in next(rows, ())]
    inserted_count = 0

    try:
        for row_number, values in enumerate(rows, start=2):
            record = {
                header: _legacy_certificate_value(value)
                for header, value in zip(headers, values)
                if header
            }
            if not any(record.values()):
                continue
            certificate_type = record.get('증명서종류', '')
            applicant_name = record.get('성명', '')
            if not certificate_type or not applicant_name:
                continue
            applicant_type = record.get('신청구분', '')
            if not applicant_type:
                applicant_type = '강사' if '강사' in certificate_type else '임직원'

            conn.execute('''
                INSERT OR IGNORE INTO certificate_requests (
                    applied_date, applicant_type, certificate_type,
                    applicant_name, resident_number, home_address,
                    work_start_date, work_end_date, workplace,
                    subject_or_duty, purpose, position, email, status,
                    issued_date, issue_number, termination_reason, filename,
                    legacy_row_number
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                record.get('신청일', ''),
                applicant_type,
                certificate_type,
                applicant_name,
                record.get('주민번호', ''),
                record.get('자택주소', ''),
                record.get('근무시작일', ''),
                record.get('근무종료일', ''),
                record.get('근무장소', ''),
                record.get('강의과목', ''),
                record.get('용도', ''),
                record.get('직책', ''),
                record.get('이메일주소', ''),
                record.get('상태', '') or '대기',
                record.get('발급일', ''),
                record.get('발급번호', ''),
                record.get('종료사유', ''),
                record.get('파일명', ''),
                row_number,
            ))
            inserted_count += int(conn.execute(
                'SELECT changes()'
            ).fetchone()[0])
    finally:
        workbook.close()

    conn.execute('''
        INSERT INTO admin_settings (key, value, updated_at)
        VALUES (?, ?, CURRENT_TIMESTAMP)
        ON CONFLICT(key) DO UPDATE SET
            value=excluded.value,
            updated_at=CURRENT_TIMESTAMP
    ''', (migration_key, f'imported:{inserted_count}'))
    return inserted_count


def init_db():
    os.makedirs(GALLERY_UPLOADS, exist_ok=True)
    os.makedirs(GALLERY_THUMBS, exist_ok=True)
    os.makedirs(PROFILE_ROOT, exist_ok=True)
    os.makedirs(SCHOOL_UPLOADS, exist_ok=True)
    os.makedirs(AI_MAIL_UPLOADS, exist_ok=True)
    
    conn = get_db()
    c = conn.cursor()
    ensure_certificate_schema(conn)
    
    c.execute('''CREATE TABLE IF NOT EXISTS tasks (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        year TEXT, date TEXT, owner TEXT,
        cat_meeting_title TEXT, cat_meeting_time TEXT,
        cat_interview_title TEXT, cat_interview_time TEXT,
        cat_miting_title TEXT, cat_miting_time TEXT,
        cat_out_title TEXT, cat_out_time TEXT,
        cat_etc_title TEXT, cat_etc_time TEXT,
        note TEXT
    )''')
    
    c.execute('''CREATE TABLE IF NOT EXISTS attendance (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        owner TEXT, type TEXT, start_date TEXT, end_date TEXT, status TEXT,
        approval_id INTEGER
    )''')

    c.execute('''CREATE TABLE IF NOT EXISTS daily_attendance (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        emp_no TEXT NOT NULL,
        date TEXT NOT NULL,
        clock_in_time TEXT NOT NULL,
        clock_out_time TEXT,
        status TEXT NOT NULL,
        reason TEXT,
        position TEXT
    )''')
    
    c.execute('''CREATE TABLE IF NOT EXISTS board (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        title TEXT, content TEXT, author TEXT, 
        created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
        filename TEXT, filepath TEXT
    )''')
    
    c.execute('''CREATE TABLE IF NOT EXISTS messages (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        room_id TEXT,
        sender TEXT, receiver TEXT, content TEXT, 
        sent_at DATETIME DEFAULT CURRENT_TIMESTAMP,
        is_read INTEGER DEFAULT 0,
        filename TEXT, filepath TEXT,
        message_uid TEXT,
        reply_to_uid TEXT,
        edited_at DATETIME,
        deleted_for_all INTEGER DEFAULT 0,
        deleted_at DATETIME
    )''')

    c.execute('''CREATE TABLE IF NOT EXISTS chat_rooms (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT,                 
        created_by TEXT,           
        created_at DATETIME DEFAULT CURRENT_TIMESTAMP
    )''')

    c.execute('''CREATE TABLE IF NOT EXISTS chat_members (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        room_id INTEGER NOT NULL,
        emp_no TEXT NOT NULL,      
        joined_at DATETIME DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (room_id) REFERENCES chat_rooms(id) ON DELETE CASCADE
    )''')

    c.execute('''CREATE TABLE IF NOT EXISTS message_reads (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        message_id INTEGER NOT NULL,
        emp_no TEXT NOT NULL,
        is_read INTEGER DEFAULT 0,
        read_at DATETIME,
        FOREIGN KEY (message_id) REFERENCES messages(id) ON DELETE CASCADE
    )''')

    c.execute('''CREATE TABLE IF NOT EXISTS users (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        emp_no TEXT, name TEXT, password TEXT, position TEXT, level INTEGER,
        rrn TEXT, email TEXT, phone TEXT,
        address TEXT, bank_account TEXT, department TEXT, profile_path TEXT,
        profile_icon TEXT DEFAULT '👤',
        join_date TEXT, retire_date TEXT, status TEXT DEFAULT '대기',
        applied_at DATETIME, approved_at DATETIME,
        rejection_reason TEXT DEFAULT '', rejected_at DATETIME
    )''')

    # 인사관리에서 직급명과 권한 레벨을 화면으로 관리한다.
    c.execute('''CREATE TABLE IF NOT EXISTS hr_positions (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL UNIQUE,
        level INTEGER NOT NULL CHECK(level BETWEEN 0 AND 99),
        sort_order INTEGER NOT NULL DEFAULT 0,
        created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
        updated_at DATETIME DEFAULT CURRENT_TIMESTAMP
    )''')
    default_positions = (
        ('최고관리자', 0), ('대표이사', 1), ('이사', 2), ('실장', 3),
        ('팀장', 4), ('사원', 5), ('계약직', 6), ('센터장(팀장)', 7),
        ('센터장', 8), ('전담코디', 9), ('보조코디', 10), ('안전코디', 11),
        ('방과후강사', 12), ('맞춤형강사', 13), ('임시회원', 14),
    )
    c.executemany('''
        INSERT OR IGNORE INTO hr_positions (name, level, sort_order)
        VALUES (?, ?, ?)
    ''', ((name, level, order) for order, (name, level) in enumerate(default_positions, 1)))

    c.execute('''CREATE TABLE IF NOT EXISTS contact_center_teams (
        emp_no TEXT PRIMARY KEY,
        team_no INTEGER NOT NULL,
        updated_at DATETIME DEFAULT CURRENT_TIMESTAMP
    )''')

    c.execute('''CREATE TABLE IF NOT EXISTS saved_contact_directories (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        owner_emp_no TEXT NOT NULL,
        name TEXT NOT NULL,
        settings_json TEXT NOT NULL DEFAULT '{}',
        created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
        updated_at DATETIME DEFAULT CURRENT_TIMESTAMP,
        UNIQUE(owner_emp_no, name)
    )''')
    c.execute('''CREATE INDEX IF NOT EXISTS idx_saved_contact_directories_owner
                 ON saved_contact_directories(owner_emp_no, updated_at)''')

    c.execute('''CREATE TABLE IF NOT EXISTS login_activity (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        emp_no TEXT,
        user_name TEXT,
        action TEXT NOT NULL,
        ip_address TEXT,
        user_agent TEXT,
        created_at DATETIME DEFAULT CURRENT_TIMESTAMP
    )''')

    c.execute('''CREATE TABLE IF NOT EXISTS usage_logs (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        emp_no TEXT,
        user_name TEXT,
        menu_name TEXT,
        endpoint TEXT,
        path TEXT,
        method TEXT,
        ip_address TEXT,
        session_id TEXT,
        created_at DATETIME DEFAULT CURRENT_TIMESTAMP
    )''')

    c.execute('''CREATE TABLE IF NOT EXISTS usage_user_totals (
        emp_no TEXT PRIMARY KEY,
        user_name TEXT,
        access_count INTEGER NOT NULL DEFAULT 0,
        login_count INTEGER NOT NULL DEFAULT 0,
        logout_count INTEGER NOT NULL DEFAULT 0,
        first_used DATETIME,
        last_used DATETIME,
        last_login DATETIME,
        last_logout DATETIME,
        updated_at DATETIME DEFAULT CURRENT_TIMESTAMP
    )''')

    c.execute('''CREATE TABLE IF NOT EXISTS usage_user_menu_totals (
        emp_no TEXT NOT NULL,
        user_name TEXT,
        menu_name TEXT NOT NULL,
        access_count INTEGER NOT NULL DEFAULT 0,
        first_used DATETIME,
        last_used DATETIME,
        PRIMARY KEY (emp_no, menu_name)
    )''')

    c.execute('''CREATE TABLE IF NOT EXISTS usage_page_sessions (
        session_id TEXT PRIMARY KEY,
        emp_no TEXT NOT NULL,
        path TEXT NOT NULL,
        last_seen DATETIME DEFAULT CURRENT_TIMESTAMP
    )''')

    c.execute('''CREATE TABLE IF NOT EXISTS admin_settings (
        key TEXT PRIMARY KEY,
        value TEXT,
        updated_at DATETIME DEFAULT CURRENT_TIMESTAMP
    )''')

    # 상단 주메뉴와 서브메뉴의 회원 레벨별 접근 기준.
    c.execute('''CREATE TABLE IF NOT EXISTS menu_access_permissions (
        menu_key TEXT PRIMARY KEY,
        max_level INTEGER NOT NULL CHECK(max_level BETWEEN -1 AND 99),
        updated_by TEXT,
        updated_at DATETIME DEFAULT CURRENT_TIMESTAMP
    )''')

    c.execute('''CREATE TABLE IF NOT EXISTS custom_themes (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL,
        effect TEXT DEFAULT 'blobs',
        category TEXT DEFAULT 'custom',
        vars_json TEXT NOT NULL,
        enabled INTEGER DEFAULT 1,
        created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
        updated_at DATETIME DEFAULT CURRENT_TIMESTAMP
    )''')

    c.execute('''CREATE TABLE IF NOT EXISTS theme_catalog_preferences (
        owner_emp_no TEXT NOT NULL,
        theme_key TEXT NOT NULL,
        is_favorite INTEGER NOT NULL DEFAULT 0,
        is_hidden INTEGER NOT NULL DEFAULT 0,
        updated_at DATETIME DEFAULT CURRENT_TIMESTAMP,
        PRIMARY KEY (owner_emp_no, theme_key)
    )''')

    c.execute('''CREATE TABLE IF NOT EXISTS approvals (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        doc_type TEXT, title TEXT, drafter TEXT,
        approver_1 TEXT, approver_2 TEXT, status TEXT DEFAULT '대기',
        receivers TEXT DEFAULT '', cc_receivers TEXT DEFAULT '',
        doc_data TEXT, filename TEXT, filepath TEXT, filesize TEXT DEFAULT '',
        created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
        updated_at DATETIME DEFAULT CURRENT_TIMESTAMP
    )''')

    c.execute('''CREATE TABLE IF NOT EXISTS expense_reports (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        approval_id INTEGER UNIQUE,
        title TEXT,
        drafter TEXT,
        approver_1 TEXT,
        approver_2 TEXT,
        doc_status TEXT DEFAULT '대기',
        payment_status TEXT DEFAULT '결재중',
        total_amount INTEGER DEFAULT 0,
        item_count INTEGER DEFAULT 0,
        report_year TEXT,
        report_month TEXT,
        submitted_at DATETIME,
        approved_at DATETIME,
        paid_at DATETIME,
        paid_by TEXT,
        source_filename TEXT,
        source_filepath TEXT,
        payment_account TEXT,
        memo TEXT,
        created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
        updated_at DATETIME DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (approval_id) REFERENCES approvals(id)
    )''')

    c.execute('''CREATE TABLE IF NOT EXISTS expense_items (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        report_id INTEGER,
        approval_id INTEGER,
        row_no INTEGER,
        expense_date TEXT,
        category TEXT,
        vendor TEXT,
        description TEXT,
        payment_method TEXT,
        amount INTEGER DEFAULT 0,
        note TEXT,
        raw_json TEXT,
        created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (report_id) REFERENCES expense_reports(id),
        FOREIGN KEY (approval_id) REFERENCES approvals(id)
    )''')

    c.execute('''CREATE TABLE IF NOT EXISTS ai_mail_workgroups (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        owner_emp_no TEXT NOT NULL,
        name TEXT NOT NULL,
        features_json TEXT NOT NULL DEFAULT '{}',
        created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
        updated_at DATETIME DEFAULT CURRENT_TIMESTAMP,
        UNIQUE(owner_emp_no, name)
    )''')

    c.execute('''CREATE TABLE IF NOT EXISTS ai_mail_recipients (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        group_id INTEGER NOT NULL,
        email TEXT NOT NULL COLLATE NOCASE,
        recipient_name TEXT NOT NULL,
        memo TEXT DEFAULT '',
        created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
        updated_at DATETIME DEFAULT CURRENT_TIMESTAMP,
        UNIQUE(group_id, email),
        FOREIGN KEY (group_id) REFERENCES ai_mail_workgroups(id) ON DELETE CASCADE
    )''')

    c.execute('''CREATE TABLE IF NOT EXISTS ai_mail_senders (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        owner_emp_no TEXT NOT NULL,
        label TEXT NOT NULL,
        email TEXT NOT NULL COLLATE NOCASE,
        encrypted_app_password TEXT NOT NULL,
        is_active INTEGER NOT NULL DEFAULT 1,
        last_tested_at DATETIME,
        last_test_status TEXT,
        last_test_error TEXT,
        created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
        updated_at DATETIME DEFAULT CURRENT_TIMESTAMP,
        UNIQUE(owner_emp_no, email)
    )''')

    c.execute('''CREATE TABLE IF NOT EXISTS ai_mail_templates (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        owner_emp_no TEXT NOT NULL,
        name TEXT NOT NULL,
        subject TEXT NOT NULL,
        body_html TEXT NOT NULL,
        body_text TEXT DEFAULT '',
        is_active INTEGER NOT NULL DEFAULT 1,
        created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
        updated_at DATETIME DEFAULT CURRENT_TIMESTAMP,
        UNIQUE(owner_emp_no, name)
    )''')

    c.execute('''CREATE TABLE IF NOT EXISTS ai_mail_template_assets (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        template_id INTEGER NOT NULL,
        original_name TEXT NOT NULL,
        stored_name TEXT NOT NULL,
        filepath TEXT NOT NULL,
        mime_type TEXT NOT NULL,
        content_id TEXT NOT NULL UNIQUE,
        size_bytes INTEGER NOT NULL DEFAULT 0,
        created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (template_id) REFERENCES ai_mail_templates(id) ON DELETE CASCADE
    )''')

    c.execute('''CREATE TABLE IF NOT EXISTS ai_mail_campaigns (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        owner_emp_no TEXT NOT NULL,
        group_id INTEGER NOT NULL,
        sender_id INTEGER NOT NULL,
        template_id INTEGER,
        group_name TEXT NOT NULL DEFAULT '',
        sender_label TEXT NOT NULL DEFAULT '',
        sender_email TEXT NOT NULL DEFAULT '',
        name TEXT NOT NULL,
        subject TEXT NOT NULL,
        body_html TEXT NOT NULL,
        body_text TEXT DEFAULT '',
        attachment_mode TEXT NOT NULL DEFAULT 'none',
        status TEXT NOT NULL DEFAULT 'staged',
        total_count INTEGER NOT NULL DEFAULT 0,
        processed_count INTEGER NOT NULL DEFAULT 0,
        sent_count INTEGER NOT NULL DEFAULT 0,
        failed_count INTEGER NOT NULL DEFAULT 0,
        cancelled_count INTEGER NOT NULL DEFAULT 0,
        cancel_requested INTEGER NOT NULL DEFAULT 0,
        cancel_requested_at DATETIME,
        cancel_requested_by TEXT,
        cancel_reason TEXT,
        allow_missing_attachment INTEGER NOT NULL DEFAULT 0,
        send_interval REAL NOT NULL DEFAULT 1.0,
        preflight_ok INTEGER NOT NULL DEFAULT 0,
        preflight_json TEXT DEFAULT '{}',
        error_code TEXT,
        error_message TEXT,
        created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
        queued_at DATETIME,
        started_at DATETIME,
        finished_at DATETIME,
        updated_at DATETIME DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (group_id) REFERENCES ai_mail_workgroups(id),
        FOREIGN KEY (sender_id) REFERENCES ai_mail_senders(id),
        FOREIGN KEY (template_id) REFERENCES ai_mail_templates(id)
    )''')

    c.execute('''CREATE TABLE IF NOT EXISTS ai_mail_campaign_recipients (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        campaign_id INTEGER NOT NULL,
        source_recipient_id INTEGER,
        email TEXT NOT NULL COLLATE NOCASE,
        recipient_name TEXT NOT NULL,
        memo TEXT DEFAULT '',
        status TEXT NOT NULL DEFAULT 'pending',
        attachment_count INTEGER NOT NULL DEFAULT 0,
        attachment_bytes INTEGER NOT NULL DEFAULT 0,
        attempt_count INTEGER NOT NULL DEFAULT 0,
        message_id TEXT,
        error_code TEXT,
        error_message TEXT,
        smtp_response TEXT,
        started_at DATETIME,
        sent_at DATETIME,
        updated_at DATETIME DEFAULT CURRENT_TIMESTAMP,
        UNIQUE(campaign_id, email),
        FOREIGN KEY (campaign_id) REFERENCES ai_mail_campaigns(id) ON DELETE CASCADE,
        FOREIGN KEY (source_recipient_id) REFERENCES ai_mail_recipients(id)
    )''')

    c.execute('''CREATE TABLE IF NOT EXISTS ai_mail_campaign_attachments (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        campaign_id INTEGER NOT NULL,
        campaign_recipient_id INTEGER,
        kind TEXT NOT NULL,
        match_method TEXT NOT NULL DEFAULT 'auto',
        match_status TEXT NOT NULL DEFAULT 'pending',
        original_name TEXT NOT NULL,
        stored_name TEXT NOT NULL,
        filepath TEXT NOT NULL,
        mime_type TEXT,
        size_bytes INTEGER NOT NULL DEFAULT 0,
        sha256 TEXT,
        diagnostic TEXT,
        created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (campaign_id) REFERENCES ai_mail_campaigns(id) ON DELETE CASCADE,
        FOREIGN KEY (campaign_recipient_id) REFERENCES ai_mail_campaign_recipients(id) ON DELETE SET NULL
    )''')

    c.execute('''CREATE TABLE IF NOT EXISTS ai_mail_campaign_events (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        campaign_id INTEGER NOT NULL,
        event_type TEXT NOT NULL,
        level TEXT NOT NULL DEFAULT 'info',
        message TEXT NOT NULL,
        details_json TEXT DEFAULT '{}',
        created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (campaign_id) REFERENCES ai_mail_campaigns(id) ON DELETE CASCADE
    )''')

    # 명세서 발송 작업공간: 화면의 임시 배열이 아니라 사용자별로 영구 저장한다.
    c.execute('''CREATE TABLE IF NOT EXISTS payroll_workgroups (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        owner_emp_no TEXT NOT NULL,
        name TEXT NOT NULL,
        form_type TEXT NOT NULL DEFAULT 'form_basic',
        subject TEXT NOT NULL,
        body_html TEXT NOT NULL DEFAULT '',
        banner1_data TEXT,
        banner2_data TEXT,
        banner1_asset_id INTEGER,
        banner2_asset_id INTEGER,
        logo_asset_id INTEGER,
        memo TEXT DEFAULT '',
        template_id INTEGER,
        created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
        updated_at DATETIME DEFAULT CURRENT_TIMESTAMP,
        UNIQUE(owner_emp_no, name)
    )''')

    c.execute('''CREATE TABLE IF NOT EXISTS payroll_image_assets (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        owner_emp_no TEXT NOT NULL,
        asset_kind TEXT NOT NULL,
        name TEXT NOT NULL,
        source_type TEXT NOT NULL,
        source_value TEXT NOT NULL,
        created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
        updated_at DATETIME DEFAULT CURRENT_TIMESTAMP,
        UNIQUE(owner_emp_no, asset_kind, name)
    )''')

    c.execute('''CREATE TABLE IF NOT EXISTS payroll_mail_templates (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        owner_emp_no TEXT NOT NULL,
        template_key TEXT,
        name TEXT NOT NULL,
        subject TEXT NOT NULL,
        description TEXT DEFAULT '',
        source_filename TEXT,
        match_keywords TEXT,
        body_html TEXT NOT NULL,
        is_system INTEGER NOT NULL DEFAULT 0,
        is_active INTEGER NOT NULL DEFAULT 1,
        created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
        updated_at DATETIME DEFAULT CURRENT_TIMESTAMP,
        UNIQUE(owner_emp_no, name)
    )''')

    c.execute('''CREATE TABLE IF NOT EXISTS payroll_campaigns (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        owner_emp_no TEXT NOT NULL,
        group_id INTEGER,
        group_name TEXT NOT NULL DEFAULT '',
        sender_id INTEGER,
        sender_email TEXT NOT NULL DEFAULT '',
        subject TEXT NOT NULL DEFAULT '',
        source_filename TEXT NOT NULL DEFAULT '',
        status TEXT NOT NULL DEFAULT 'queued',
        total_count INTEGER NOT NULL DEFAULT 0,
        processed_count INTEGER NOT NULL DEFAULT 0,
        sent_count INTEGER NOT NULL DEFAULT 0,
        failed_count INTEGER NOT NULL DEFAULT 0,
        errors_json TEXT NOT NULL DEFAULT '[]',
        created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
        started_at DATETIME,
        finished_at DATETIME,
        updated_at DATETIME DEFAULT CURRENT_TIMESTAMP
    )''')

    c.execute('''CREATE TABLE IF NOT EXISTS payroll_campaign_recipients (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        campaign_id INTEGER NOT NULL,
        owner_emp_no TEXT NOT NULL,
        sheet_name TEXT NOT NULL DEFAULT '',
        excel_row INTEGER,
        recipient_type TEXT NOT NULL DEFAULT '',
        school_name TEXT NOT NULL DEFAULT '',
        recipient_name TEXT NOT NULL DEFAULT '',
        email TEXT NOT NULL DEFAULT '',
        status TEXT NOT NULL DEFAULT 'queued',
        error_message TEXT NOT NULL DEFAULT '',
        started_at DATETIME,
        finished_at DATETIME,
        elapsed_seconds REAL NOT NULL DEFAULT 0,
        created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (campaign_id) REFERENCES payroll_campaigns(id) ON DELETE CASCADE
    )''')

    c.execute('CREATE INDEX IF NOT EXISTS idx_ai_mail_groups_owner ON ai_mail_workgroups(owner_emp_no, updated_at)')
    c.execute('CREATE INDEX IF NOT EXISTS idx_ai_mail_recipients_group ON ai_mail_recipients(group_id, recipient_name)')
    c.execute('CREATE INDEX IF NOT EXISTS idx_ai_mail_senders_owner ON ai_mail_senders(owner_emp_no, is_active)')
    c.execute('CREATE INDEX IF NOT EXISTS idx_ai_mail_templates_owner ON ai_mail_templates(owner_emp_no, updated_at)')
    c.execute('CREATE INDEX IF NOT EXISTS idx_ai_mail_template_assets_template ON ai_mail_template_assets(template_id, id)')
    c.execute('CREATE INDEX IF NOT EXISTS idx_ai_mail_campaigns_owner ON ai_mail_campaigns(owner_emp_no, created_at)')
    c.execute('CREATE INDEX IF NOT EXISTS idx_ai_mail_campaigns_status ON ai_mail_campaigns(owner_emp_no, status, updated_at)')
    c.execute('CREATE INDEX IF NOT EXISTS idx_ai_mail_campaigns_refs ON ai_mail_campaigns(group_id, sender_id, template_id)')
    c.execute('CREATE INDEX IF NOT EXISTS idx_ai_mail_campaign_recipients_campaign ON ai_mail_campaign_recipients(campaign_id, status)')
    c.execute('CREATE INDEX IF NOT EXISTS idx_ai_mail_campaign_attachments_campaign ON ai_mail_campaign_attachments(campaign_id, kind)')
    c.execute('CREATE INDEX IF NOT EXISTS idx_ai_mail_campaign_attachments_recipient ON ai_mail_campaign_attachments(campaign_recipient_id, match_status)')
    c.execute('CREATE INDEX IF NOT EXISTS idx_ai_mail_campaign_events_campaign ON ai_mail_campaign_events(campaign_id, created_at)')
    c.execute('CREATE INDEX IF NOT EXISTS idx_payroll_groups_owner ON payroll_workgroups(owner_emp_no, updated_at)')
    c.execute('CREATE INDEX IF NOT EXISTS idx_payroll_assets_owner_kind ON payroll_image_assets(owner_emp_no, asset_kind, updated_at)')
    c.execute('CREATE INDEX IF NOT EXISTS idx_payroll_templates_owner ON payroll_mail_templates(owner_emp_no, updated_at)')
    c.execute('CREATE INDEX IF NOT EXISTS idx_payroll_campaigns_owner ON payroll_campaigns(owner_emp_no, created_at)')
    c.execute('CREATE INDEX IF NOT EXISTS idx_payroll_campaign_recipients_campaign ON payroll_campaign_recipients(campaign_id, id)')
    c.execute('CREATE INDEX IF NOT EXISTS idx_payroll_campaign_recipients_owner ON payroll_campaign_recipients(owner_emp_no, campaign_id)')
    c.execute('CREATE INDEX IF NOT EXISTS idx_theme_preferences_owner ON theme_catalog_preferences(owner_emp_no, is_hidden, is_favorite)')

    c.execute('''CREATE TABLE IF NOT EXISTS gallery (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        title TEXT NOT NULL,
        filename TEXT NOT NULL,
        thumb_name TEXT NOT NULL,
        file_type TEXT NOT NULL,
        created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
        tab_id INTEGER DEFAULT 1
    )''')

    c.execute('''CREATE TABLE IF NOT EXISTS gallery_tabs (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL
    )''')
    
    c.execute('''CREATE TABLE IF NOT EXISTS gall2_tabs (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )''')

    c.execute('''CREATE TABLE IF NOT EXISTS gall2 (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        title TEXT,
        filename TEXT NOT NULL,
        thumb_name TEXT NOT NULL,
        file_type TEXT,
        tab_id INTEGER NOT NULL DEFAULT 1,
        post_id INTEGER,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (tab_id) REFERENCES gall2_tabs (id)
    )''')

    c.execute('''CREATE TABLE IF NOT EXISTS gall2_posts (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        title TEXT NOT NULL,
        content TEXT,
        author TEXT,
        tab_id INTEGER NOT NULL DEFAULT 1,
        upload_token TEXT UNIQUE,
        school_id INTEGER,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (tab_id) REFERENCES gall2_tabs (id)
    )''')
    gall2_post_columns = {
        row['name'] if hasattr(row, 'keys') else row[1]
        for row in c.execute('PRAGMA table_info(gall2_posts)').fetchall()
    }
    if 'school_id' not in gall2_post_columns:
        c.execute('ALTER TABLE gall2_posts ADD COLUMN school_id INTEGER')
    # 학교갤러리는 센터별이 아니라 모든 센터장이 공유하는 단일 범위(0)다.
    c.execute('UPDATE gall2_posts SET school_id=0 WHERE school_id IS NOT NULL AND school_id<>0')
    c.execute('CREATE INDEX IF NOT EXISTS idx_gall2_posts_school_id ON gall2_posts(school_id, created_at)')
    
    c.execute('''CREATE TABLE IF NOT EXISTS schools (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        access_key TEXT UNIQUE,
        year TEXT NOT NULL,                
        school_name TEXT NOT NULL,         
        contract_subject TEXT,
        office_phone TEXT,                 
        office_location TEXT,              
        school_address TEXT,
        school_phone TEXT,
        school_email TEXT,
        neulbom_assistant TEXT,            
        neulbom_manager TEXT,              
        center_director_id TEXT,           
        created_at DATETIME DEFAULT CURRENT_TIMESTAMP
    )''')

    c.execute('''CREATE TABLE IF NOT EXISTS school_posts (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        school_id INTEGER NOT NULL,        
        category TEXT NOT NULL,            
        title TEXT NOT NULL,
        content TEXT,
        author TEXT,
        filename TEXT,                     
        filepath TEXT,                     
        created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (school_id) REFERENCES schools (id) ON DELETE CASCADE
    )''')

    # 🚀 [수정] 데이터베이스의 하위 호환성을 완벽히 보장하여 기존 db 파일을 마이그레이션할 때 
    # `is_read` 컬럼이 누락되어 카운트가 비정상 차감되거나 작동하지 않는 현상을 완전히 차단하기 위해 alter 구문에 is_read를 강제 주입했습니다.
    alter_queries = [
        "ALTER TABLE messages ADD COLUMN filename TEXT",
        "ALTER TABLE messages ADD COLUMN filepath TEXT",
        "ALTER TABLE messages ADD COLUMN room_id TEXT", 
        "ALTER TABLE messages ADD COLUMN is_read INTEGER DEFAULT 0",
        "ALTER TABLE messages ADD COLUMN message_uid TEXT",
        "ALTER TABLE messages ADD COLUMN reply_to_uid TEXT",
        "ALTER TABLE messages ADD COLUMN edited_at DATETIME",
        "ALTER TABLE messages ADD COLUMN deleted_for_all INTEGER DEFAULT 0",
        "ALTER TABLE messages ADD COLUMN deleted_at DATETIME",
        "ALTER TABLE daily_attendance ADD COLUMN reason TEXT",
        "ALTER TABLE daily_attendance ADD COLUMN position TEXT",
        "ALTER TABLE gallery ADD COLUMN tab_id INTEGER DEFAULT 1",
        "ALTER TABLE users ADD COLUMN profile_icon TEXT DEFAULT '👤'",
        "ALTER TABLE users ADD COLUMN address TEXT",
        "ALTER TABLE users ADD COLUMN bank_account TEXT",
        "ALTER TABLE users ADD COLUMN department TEXT",
        "ALTER TABLE users ADD COLUMN profile_path TEXT",
        "ALTER TABLE users ADD COLUMN custom_department TEXT DEFAULT ''",
        "ALTER TABLE users ADD COLUMN custom_team TEXT DEFAULT ''",
        "ALTER TABLE users ADD COLUMN applied_at DATETIME",
        "ALTER TABLE users ADD COLUMN approved_at DATETIME",
        "ALTER TABLE approvals ADD COLUMN receivers TEXT DEFAULT ''",
        "ALTER TABLE approvals ADD COLUMN cc_receivers TEXT DEFAULT ''",
        "ALTER TABLE approvals ADD COLUMN filesize TEXT DEFAULT ''",
        "ALTER TABLE attendance ADD COLUMN approval_id INTEGER",
        "ALTER TABLE expense_reports ADD COLUMN memo TEXT",
        "ALTER TABLE expense_reports ADD COLUMN payment_account TEXT",
        "ALTER TABLE ai_mail_campaigns ADD COLUMN group_name TEXT NOT NULL DEFAULT ''",
        "ALTER TABLE ai_mail_campaigns ADD COLUMN sender_label TEXT NOT NULL DEFAULT ''",
        "ALTER TABLE ai_mail_campaigns ADD COLUMN sender_email TEXT NOT NULL DEFAULT ''",
        "ALTER TABLE ai_mail_campaigns ADD COLUMN cancel_requested_at DATETIME",
        "ALTER TABLE ai_mail_campaigns ADD COLUMN cancel_requested_by TEXT",
        "ALTER TABLE ai_mail_campaigns ADD COLUMN cancel_reason TEXT",
        "ALTER TABLE ai_mail_campaigns ADD COLUMN allow_missing_attachment INTEGER NOT NULL DEFAULT 0",
        "ALTER TABLE payroll_mail_templates ADD COLUMN template_key TEXT",
        "ALTER TABLE payroll_mail_templates ADD COLUMN description TEXT DEFAULT ''",
        "ALTER TABLE payroll_mail_templates ADD COLUMN source_filename TEXT",
        "ALTER TABLE payroll_mail_templates ADD COLUMN is_system INTEGER NOT NULL DEFAULT 0",
        "ALTER TABLE payroll_mail_templates ADD COLUMN match_keywords TEXT",
        "ALTER TABLE payroll_campaign_recipients ADD COLUMN school_name TEXT NOT NULL DEFAULT ''",
        "ALTER TABLE payroll_workgroups ADD COLUMN banner1_asset_id INTEGER",
        "ALTER TABLE payroll_workgroups ADD COLUMN banner2_asset_id INTEGER",
        "ALTER TABLE payroll_workgroups ADD COLUMN logo_asset_id INTEGER",
        "ALTER TABLE custom_themes ADD COLUMN category TEXT DEFAULT 'custom'",
        "ALTER TABLE custom_themes ADD COLUMN updated_at DATETIME DEFAULT CURRENT_TIMESTAMP",
        "ALTER TABLE schools ADD COLUMN contract_subject TEXT",
        "ALTER TABLE schools ADD COLUMN school_address TEXT",
        "ALTER TABLE schools ADD COLUMN school_phone TEXT",
        "ALTER TABLE schools ADD COLUMN school_email TEXT",
        "ALTER TABLE schools ADD COLUMN access_key TEXT",
        "ALTER TABLE gall2 ADD COLUMN post_id INTEGER",
        "ALTER TABLE usage_logs ADD COLUMN session_id TEXT",
        "ALTER TABLE usage_user_totals ADD COLUMN login_count INTEGER NOT NULL DEFAULT 0",
        "ALTER TABLE usage_user_totals ADD COLUMN logout_count INTEGER NOT NULL DEFAULT 0",
        "ALTER TABLE usage_user_totals ADD COLUMN last_login DATETIME",
        "ALTER TABLE usage_user_totals ADD COLUMN last_logout DATETIME",
        "ALTER TABLE usage_user_totals ADD COLUMN updated_at DATETIME DEFAULT CURRENT_TIMESTAMP"
    ]
    
    for q in alter_queries:
        try:
            c.execute(q)
        except sqlite3.OperationalError:
            pass 

    # 기존 회원은 보유 중인 입사일을 승인일로 이관하고, 기록이 없던 신청일은
    # 마이그레이션 시각으로 채워 이후부터 날짜가 누락되지 않게 한다.
    c.execute('''
        UPDATE users
        SET applied_at = CASE
            WHEN TRIM(COALESCE(join_date, '')) <> '' THEN join_date || ' 00:00:00'
            ELSE DATETIME('now', 'localtime')
        END
        WHERE applied_at IS NULL OR TRIM(applied_at) = ''
    ''')
    c.execute('''
        UPDATE users
        SET approved_at = join_date || ' 00:00:00'
        WHERE status = '승인'
          AND TRIM(COALESCE(join_date, '')) <> ''
          AND (approved_at IS NULL OR TRIM(approved_at) = '')
    ''')

    c.execute('''
        CREATE UNIQUE INDEX IF NOT EXISTS idx_attendance_approval_id
        ON attendance(approval_id)
        WHERE approval_id IS NOT NULL
    ''')

    # 학교 상세 주소에 순번(id)을 노출하지 않도록 기존/신규 학교에
    # 예측 불가능한 공개 접근 키를 부여한다.
    schools_without_key = c.execute(
        "SELECT id FROM schools WHERE access_key IS NULL OR TRIM(access_key) = ''"
    ).fetchall()
    for school_row in schools_without_key:
        c.execute(
            "UPDATE schools SET access_key = ? WHERE id = ?",
            (secrets.token_urlsafe(24), school_row[0])
        )
    c.execute('''
        CREATE UNIQUE INDEX IF NOT EXISTS idx_schools_access_key
        ON schools(access_key)
        WHERE access_key IS NOT NULL
    ''')

    c.execute('CREATE INDEX IF NOT EXISTS idx_usage_logs_created_at ON usage_logs(created_at)')
    c.execute('CREATE INDEX IF NOT EXISTS idx_usage_logs_emp_created ON usage_logs(emp_no, created_at)')
    c.execute('CREATE INDEX IF NOT EXISTS idx_usage_logs_menu_created ON usage_logs(menu_name, created_at)')
    c.execute('CREATE INDEX IF NOT EXISTS idx_login_activity_created_at ON login_activity(created_at)')
    c.execute('CREATE INDEX IF NOT EXISTS idx_login_activity_emp_created ON login_activity(emp_no, created_at)')
    c.execute('CREATE INDEX IF NOT EXISTS idx_usage_user_menu_last ON usage_user_menu_totals(emp_no, last_used)')
    c.execute('''
        CREATE UNIQUE INDEX IF NOT EXISTS idx_users_emp_no_unique
        ON users(emp_no)
        WHERE emp_no IS NOT NULL AND TRIM(emp_no) != ''
    ''')

    # 기존 원본 로그에서 화면 접속으로 판단되는 기록만 1회 이관한다.
    # API 폴링, 파일/썸네일, 상태 조회는 과거 이용통계를 부풀렸으므로 누계에서 제외한다.
    usage_backfill = c.execute(
        "SELECT value FROM admin_settings WHERE key='usage_stats_v2_backfilled'"
    ).fetchone()
    if not usage_backfill:
        legacy_page_filter = '''
            method='GET'
            AND COALESCE(path, '') <> ''
            AND path NOT LIKE '/api/%'
            AND path NOT LIKE '%/api/%'
            AND path NOT LIKE '/widget/%'
            AND path NOT LIKE '/uploads/%'
            AND path NOT LIKE '%/thumb/%'
            AND path NOT LIKE '%/attachment/%'
            AND path NOT LIKE '%/download/%'
            AND path NOT LIKE '%/file/%'
            AND path NOT LIKE '%/weblink-file/%'
            AND path NOT LIKE '/get_%'
            AND path NOT LIKE '/check_%'
            AND path <> '/user/my_info'
            AND LOWER(path) NOT LIKE '%.ico'
            AND LOWER(path) NOT LIKE '%.jpg'
            AND LOWER(path) NOT LIKE '%.jpeg'
            AND LOWER(path) NOT LIKE '%.png'
            AND LOWER(path) NOT LIKE '%.gif'
            AND LOWER(path) NOT LIKE '%.svg'
            AND LOWER(path) NOT LIKE '%.webp'
            AND LOWER(path) NOT LIKE '%.css'
            AND LOWER(path) NOT LIKE '%.js'
            AND LOWER(path) NOT LIKE '%.json'
            AND LOWER(path) NOT LIKE '%.pdf'
            AND LOWER(path) NOT LIKE '%.xlsx'
            AND EXISTS (
                SELECT 1 FROM users u
                WHERE CAST(u.emp_no AS TEXT)=CAST(usage_logs.emp_no AS TEXT)
            )
            AND COALESCE(endpoint, '') NOT LIKE 'api_%'
            AND COALESCE(endpoint, '') NOT LIKE 'get_%'
            AND COALESCE(endpoint, '') NOT LIKE 'serve_%'
            AND COALESCE(endpoint, '') NOT LIKE 'download_%'
            AND COALESCE(endpoint, '') NOT LIKE 'check_%'
            AND COALESCE(endpoint, '') NOT LIKE '%.api_%'
            AND COALESCE(endpoint, '') NOT LIKE '%.get_%'
            AND COALESCE(endpoint, '') NOT LIKE '%.serve_%'
            AND COALESCE(endpoint, '') NOT LIKE '%.download_%'
            AND COALESCE(endpoint, '') NOT LIKE '%.widget_%'
            AND COALESCE(endpoint, '') NOT LIKE '%.bootstrap'
            AND COALESCE(endpoint, '') NOT LIKE '%.%status%'
        '''
        legacy_events = f'''
            SELECT COALESCE(emp_no, user_name, 'unknown') AS emp_no,
                   MAX(COALESCE(user_name, emp_no, '알 수 없음')) AS user_name,
                   COALESCE(menu_name, '기타') AS menu_name,
                   MIN(created_at) AS event_at
            FROM usage_logs
            WHERE {legacy_page_filter}
            GROUP BY COALESCE(emp_no, user_name, 'unknown'),
                     COALESCE(menu_name, '기타'),
                     path,
                     STRFTIME('%Y-%m-%d %H:%M:%S', created_at)
        '''
        legacy_users = c.execute(f'''
            SELECT emp_no,
                   MAX(user_name) AS user_name,
                   COUNT(*) AS access_count,
                   MIN(event_at) AS first_used,
                   MAX(event_at) AS last_used
            FROM ({legacy_events})
            GROUP BY emp_no
        ''').fetchall()
        for row in legacy_users:
            c.execute('''
                INSERT INTO usage_user_totals (
                    emp_no, user_name, access_count, first_used, last_used, updated_at
                ) VALUES (?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
                ON CONFLICT(emp_no) DO UPDATE SET
                    user_name=excluded.user_name,
                    access_count=excluded.access_count,
                    first_used=excluded.first_used,
                    last_used=excluded.last_used,
                    updated_at=CURRENT_TIMESTAMP
            ''', tuple(row))

        legacy_menus = c.execute(f'''
            SELECT emp_no,
                   MAX(user_name) AS user_name,
                   menu_name,
                   COUNT(*) AS access_count,
                   MIN(event_at) AS first_used,
                   MAX(event_at) AS last_used
            FROM ({legacy_events})
            GROUP BY emp_no, menu_name
        ''').fetchall()
        for row in legacy_menus:
            c.execute('''
                INSERT INTO usage_user_menu_totals (
                    emp_no, user_name, menu_name, access_count, first_used, last_used
                ) VALUES (?, ?, ?, ?, ?, ?)
                ON CONFLICT(emp_no, menu_name) DO UPDATE SET
                    user_name=excluded.user_name,
                    access_count=excluded.access_count,
                    first_used=excluded.first_used,
                    last_used=excluded.last_used
            ''', tuple(row))

        legacy_logins = c.execute('''
            SELECT COALESCE(emp_no, user_name, 'unknown') AS emp_no,
                   MAX(COALESCE(user_name, emp_no, '알 수 없음')) AS user_name,
                   SUM(CASE WHEN action='login' THEN 1 ELSE 0 END) AS login_count,
                   SUM(CASE WHEN action='logout' THEN 1 ELSE 0 END) AS logout_count,
                   MAX(CASE WHEN action='login' THEN created_at END) AS last_login,
                   MAX(CASE WHEN action='logout' THEN created_at END) AS last_logout
            FROM login_activity
            GROUP BY COALESCE(emp_no, user_name, 'unknown')
        ''').fetchall()
        for row in legacy_logins:
            c.execute('''
                INSERT INTO usage_user_totals (
                    emp_no, user_name, login_count, logout_count,
                    last_login, last_logout, updated_at
                ) VALUES (?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
                ON CONFLICT(emp_no) DO UPDATE SET
                    user_name=excluded.user_name,
                    login_count=excluded.login_count,
                    logout_count=excluded.logout_count,
                    last_login=excluded.last_login,
                    last_logout=excluded.last_logout,
                    updated_at=CURRENT_TIMESTAMP
            ''', tuple(row))

        c.execute('''
            INSERT INTO admin_settings (key, value, updated_at)
            VALUES ('usage_stats_v2_backfilled', '1', CURRENT_TIMESTAMP)
        ''')

    migrate_legacy_certificates(conn)
    from .contract_repository import ensure_contract_schema_and_migrate
    contract_count = ensure_contract_schema_and_migrate(conn)
    from .verified_contract_repository import ensure_verified_contract_schema
    ensure_verified_contract_schema(conn)

    tabs_count = c.execute("SELECT count(*) FROM gallery_tabs").fetchone()[0]
    if tabs_count == 0:
        c.execute("INSERT INTO gallery_tabs (id, name) VALUES (1, '기본 갤러리')")

    gall2_tabs_count = c.execute("SELECT count(*) FROM gall2_tabs").fetchone()[0]
    if gall2_tabs_count == 0:
        c.execute("INSERT INTO gall2_tabs (id, name) VALUES (1, '기본 갤러리 2')")

    conn.commit()
    conn.close()
    if contract_count:
        from .contract_repository import archive_legacy_contract_database
        archive_legacy_contract_database()
    print("DATABASE INITIALIZED SUCCESSFULLY")

if __name__ == "__main__":
    init_db()
