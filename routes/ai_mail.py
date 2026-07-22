import base64
import csv
import functools
import hashlib
import hmac
import html
import io
import json
import mimetypes
import os
import re
import secrets
import socket
import smtplib
import sqlite3
import threading
import time
import unicodedata
import uuid
from datetime import datetime
from email import policy
from email.message import EmailMessage
from email.utils import make_msgid, parseaddr

from bs4 import BeautifulSoup, Comment
from cryptography.fernet import Fernet, InvalidToken
from flask import Blueprint, current_app, jsonify, render_template, request, send_file, session
from openpyxl import Workbook, load_workbook
from PIL import Image, UnidentifiedImageError

from .database import AI_MAIL_UPLOADS, get_db


ai_mail_bp = Blueprint('ai_mail', __name__)

MAX_RECIPIENTS = 300
MAX_TEMPLATES = 100
MAX_MESSAGE_BYTES = 18 * 1024 * 1024
MAX_IMPORT_BYTES = 5 * 1024 * 1024
MAX_TEMPLATE_IMAGE_BYTES = 3 * 1024 * 1024
MAX_TEMPLATE_TOTAL_IMAGE_BYTES = 20 * 1024 * 1024
MAX_ATTACHMENT_FILES = 350
EMAIL_RE = re.compile(r'^[^\s@]+@[^\s@]+\.[^\s@]{2,}$')
VARIABLE_NAMES = ('수신자명', '메일주소', '메모', '작업그룹명')
ATTACHMENT_MODES = {'none', 'common', 'smart', 'smart_and_common'}
BLOCKED_ATTACHMENT_EXTENSIONS = {
    '.ade', '.adp', '.app', '.bat', '.cmd', '.com', '.cpl', '.dll', '.exe',
    '.hta', '.inf', '.ins', '.isp', '.jar', '.js', '.jse', '.lib', '.lnk',
    '.mde', '.msc', '.msi', '.msp', '.mst', '.pif', '.ps1', '.scr', '.sct',
    '.shb', '.sys', '.vb', '.vbe', '.vbs', '.vxd', '.wsc', '.wsf', '.wsh'
}
IMAGE_MIME_TYPES = {'image/png', 'image/jpeg', 'image/gif', 'image/webp'}
MUTATING_METHODS = {'POST', 'PUT', 'PATCH', 'DELETE'}

_worker_lock = threading.Lock()
_worker_threads = {}
_worker_cancel_events = {}
_worker_servers = {}


def _success(message='', **payload):
    result = {'status': 'success', 'message': message}
    result.update(payload)
    return jsonify(result)


def _error(message, http_status=400, **payload):
    result = {'status': 'error', 'message': message}
    result.update(payload)
    return jsonify(result), http_status


def _owner_emp_no():
    return str(session.get('emp_no') or '').strip()


def _login_required(view):
    @functools.wraps(view)
    def wrapped(*args, **kwargs):
        if not _owner_emp_no():
            if request.path.startswith('/api/') or '/api/' in request.path:
                return _error('로그인이 필요합니다.', 401, code='AUTH_REQUIRED')
            return '로그인이 필요합니다.', 401
        return view(*args, **kwargs)
    return wrapped


def _csrf_token():
    token = session.get('ai_mail_csrf_token')
    if not token:
        token = secrets.token_urlsafe(32)
        session['ai_mail_csrf_token'] = token
    return token


def _csrf_required(view):
    @functools.wraps(view)
    def wrapped(*args, **kwargs):
        expected = session.get('ai_mail_csrf_token') or ''
        supplied = request.headers.get('X-CSRF-Token', '')
        if not expected or not supplied or not hmac.compare_digest(str(expected), str(supplied)):
            return _error('CSRF 보안 토큰이 없거나 일치하지 않습니다.', 403, code='CSRF_INVALID')
        return view(*args, **kwargs)
    return wrapped


def _mutating(view):
    return _login_required(_csrf_required(view))


def _clean_text(value, limit=None):
    text = str(value or '').strip()
    return text[:limit] if limit else text


def _json_data():
    return request.get_json(silent=True) or {}


def _request_value(name, default=''):
    if request.is_json:
        return _json_data().get(name, default)
    return request.form.get(name, default)


def _parse_json_value(value, default):
    if value is None or value == '':
        return default
    if isinstance(value, (dict, list, int, float, bool)):
        return value
    try:
        return json.loads(value)
    except (TypeError, ValueError):
        return default


def _is_valid_email(value):
    email = _clean_text(value, 254).lower()
    parsed_name, parsed_email = parseaddr(email)
    return bool(not parsed_name and parsed_email == email and EMAIL_RE.fullmatch(email))


def _fernet():
    secret = current_app.secret_key or os.environ.get('SECRET_KEY')
    if not secret:
        raise RuntimeError('SECRET_KEY가 설정되지 않아 메일 자격증명을 보관할 수 없습니다.')
    digest = hashlib.sha256(str(secret).encode('utf-8')).digest()
    return Fernet(base64.urlsafe_b64encode(digest))


def _encrypt_password(password):
    compact = re.sub(r'\s+', '', _clean_text(password))
    if not compact:
        raise ValueError('구글 앱 비밀번호를 입력해주세요.')
    return _fernet().encrypt(compact.encode('utf-8')).decode('ascii')


def _decrypt_password(token):
    try:
        return _fernet().decrypt(str(token).encode('ascii')).decode('utf-8')
    except (InvalidToken, ValueError, TypeError) as exc:
        raise RuntimeError('발송자 앱 비밀번호를 복호화할 수 없습니다.') from exc


def _features(value):
    parsed = _parse_json_value(value, {})
    if not isinstance(parsed, (dict, list)):
        raise ValueError('선택기능 형식이 올바르지 않습니다.')
    return parsed


def _loads_object(value):
    parsed = _parse_json_value(value, {})
    return parsed if isinstance(parsed, dict) else {}


def _owned_row(conn, table, row_id, owner_column='owner_emp_no'):
    return conn.execute(
        f'SELECT * FROM {table} WHERE id=? AND {owner_column}=?',
        (row_id, _owner_emp_no())
    ).fetchone()


def _owned_group(conn, group_id):
    return _owned_row(conn, 'ai_mail_workgroups', group_id)


def _owned_campaign(conn, campaign_id):
    return _owned_row(conn, 'ai_mail_campaigns', campaign_id)


def _db():
    """AI 메일 연결만 FK 검사와 SQLite 잠금 대기를 활성화한다."""
    conn = get_db()
    conn.execute('PRAGMA foreign_keys=ON')
    conn.execute('PRAGMA busy_timeout=5000')
    return conn


def _group_dict(row, recipient_count=None):
    item = dict(row)
    item['features'] = _parse_json_value(item.pop('features_json', '{}'), {})
    if recipient_count is not None:
        item['recipient_count'] = recipient_count
    return item


def _sender_dict(row):
    item = dict(row)
    item.pop('encrypted_app_password', None)
    item['has_password'] = True
    item['is_active'] = bool(item.get('is_active'))
    return item


def _template_dict(row, assets=None):
    item = dict(row)
    item['is_active'] = bool(item.get('is_active'))
    if assets is not None:
        item['assets'] = [_template_asset_dict(asset) for asset in assets]
        preview = item.get('body_html') or ''
        for asset in assets:
            preview = preview.replace(
                f"cid:{asset['content_id']}",
                f"/ai-mail/api/template-assets/{asset['id']}"
            )
        item['preview_html'] = preview
    return item


def _templates_with_assets(conn, rows):
    result = []
    for row in rows:
        assets = conn.execute(
            'SELECT * FROM ai_mail_template_assets WHERE template_id=? ORDER BY id',
            (row['id'],)
        ).fetchall()
        result.append(_template_dict(row, assets))
    return result


def _campaign_dict(row):
    item = dict(row)
    item['cancel_requested'] = bool(item.get('cancel_requested'))
    item['preflight_ok'] = bool(item.get('preflight_ok'))
    item['allow_missing_attachment'] = bool(item.get('allow_missing_attachment'))
    item['preflight'] = _parse_json_value(item.pop('preflight_json', '{}'), {})
    return item


@ai_mail_bp.route('/', strict_slashes=False)
@_login_required
def index():
    return render_template('ai_mail.html')


@ai_mail_bp.route('/api/bootstrap')
@_login_required
def bootstrap():
    owner = _owner_emp_no()
    conn = _db()
    try:
        groups = conn.execute('''
            SELECT g.*, COUNT(r.id) AS recipient_count
            FROM ai_mail_workgroups g
            LEFT JOIN ai_mail_recipients r ON r.group_id=g.id
            WHERE g.owner_emp_no=?
            GROUP BY g.id
            ORDER BY g.updated_at DESC, g.id DESC
        ''', (owner,)).fetchall()
        senders = conn.execute('''
            SELECT * FROM ai_mail_senders WHERE owner_emp_no=?
            ORDER BY is_active DESC, updated_at DESC, id DESC
        ''', (owner,)).fetchall()
        templates = conn.execute('''
            SELECT * FROM ai_mail_templates WHERE owner_emp_no=? AND is_active=1
            ORDER BY updated_at DESC, id DESC
        ''', (owner,)).fetchall()
        campaigns = conn.execute('''
            SELECT * FROM ai_mail_campaigns WHERE owner_emp_no=? AND status!='staged'
            ORDER BY created_at DESC, id DESC LIMIT 30
        ''', (owner,)).fetchall()
        return _success(
            csrf_token=_csrf_token(),
            groups=[_group_dict(row, row['recipient_count']) for row in groups],
            senders=[_sender_dict(row) for row in senders],
            templates=_templates_with_assets(conn, templates),
            recent_campaigns=[_campaign_dict(row) for row in campaigns],
            limits={
                'recipients_per_group': MAX_RECIPIENTS,
                'templates_per_owner': MAX_TEMPLATES,
                'message_bytes': MAX_MESSAGE_BYTES,
                'message_mb': 18
            },
            variables=[f'{{{{{name}}}}}' for name in VARIABLE_NAMES]
        )
    finally:
        conn.close()


@ai_mail_bp.route('/api/groups', methods=['GET'])
@_login_required
def list_groups():
    conn = _db()
    try:
        rows = conn.execute('''
            SELECT g.*, COUNT(r.id) AS recipient_count
            FROM ai_mail_workgroups g
            LEFT JOIN ai_mail_recipients r ON r.group_id=g.id
            WHERE g.owner_emp_no=?
            GROUP BY g.id
            ORDER BY g.updated_at DESC, g.id DESC
        ''', (_owner_emp_no(),)).fetchall()
        return _success(groups=[_group_dict(row, row['recipient_count']) for row in rows])
    finally:
        conn.close()


@ai_mail_bp.route('/api/groups', methods=['POST'])
@_mutating
def create_group():
    name = _clean_text(_request_value('name'), 120)
    if not name:
        return _error('작업그룹 이름을 입력해주세요.', code='GROUP_NAME_REQUIRED')
    try:
        features = _features(_request_value('features', _request_value('features_json', {})))
    except ValueError as exc:
        return _error(str(exc), code='FEATURES_INVALID')

    conn = _db()
    try:
        cursor = conn.execute('''
            INSERT INTO ai_mail_workgroups (owner_emp_no, name, features_json)
            VALUES (?, ?, ?)
        ''', (_owner_emp_no(), name, json.dumps(features, ensure_ascii=False)))
        conn.commit()
        row = _owned_group(conn, cursor.lastrowid)
        return _success('작업그룹이 등록되었습니다.', group=_group_dict(row, 0))
    except sqlite3.IntegrityError:
        conn.rollback()
        return _error('같은 이름의 작업그룹이 이미 있습니다.', 409, code='GROUP_DUPLICATE')
    finally:
        conn.close()


@ai_mail_bp.route('/api/groups/<int:group_id>', methods=['GET'])
@_login_required
def get_group(group_id):
    conn = _db()
    try:
        group = _owned_group(conn, group_id)
        if not group:
            return _error('작업그룹을 찾을 수 없습니다.', 404, code='GROUP_NOT_FOUND')
        recipients = conn.execute('''
            SELECT * FROM ai_mail_recipients WHERE group_id=?
            ORDER BY recipient_name COLLATE NOCASE, id
        ''', (group_id,)).fetchall()
        return _success(
            group=_group_dict(group, len(recipients)),
            recipients=[dict(row) for row in recipients]
        )
    finally:
        conn.close()


@ai_mail_bp.route('/api/groups/<int:group_id>', methods=['PUT', 'PATCH'])
@_mutating
def update_group(group_id):
    data = _json_data() if request.is_json else request.form
    conn = _db()
    try:
        group = _owned_group(conn, group_id)
        if not group:
            return _error('작업그룹을 찾을 수 없습니다.', 404, code='GROUP_NOT_FOUND')
        name = _clean_text(data.get('name', group['name']), 120)
        if not name:
            return _error('작업그룹 이름을 입력해주세요.', code='GROUP_NAME_REQUIRED')
        raw_features = data.get('features', data.get('features_json', group['features_json']))
        try:
            features = _features(raw_features)
        except ValueError as exc:
            return _error(str(exc), code='FEATURES_INVALID')
        conn.execute('''
            UPDATE ai_mail_workgroups
            SET name=?, features_json=?, updated_at=CURRENT_TIMESTAMP
            WHERE id=? AND owner_emp_no=?
        ''', (name, json.dumps(features, ensure_ascii=False), group_id, _owner_emp_no()))
        conn.commit()
        count = conn.execute('SELECT COUNT(*) FROM ai_mail_recipients WHERE group_id=?', (group_id,)).fetchone()[0]
        return _success('작업그룹이 수정되었습니다.', group=_group_dict(_owned_group(conn, group_id), count))
    except sqlite3.IntegrityError:
        conn.rollback()
        return _error('같은 이름의 작업그룹이 이미 있습니다.', 409, code='GROUP_DUPLICATE')
    finally:
        conn.close()


@ai_mail_bp.route('/api/groups/<int:group_id>', methods=['DELETE'])
@_mutating
def delete_group(group_id):
    conn = _db()
    try:
        group = _owned_group(conn, group_id)
        if not group:
            return _error('작업그룹을 찾을 수 없습니다.', 404, code='GROUP_NOT_FOUND')
        campaign_count = conn.execute(
            'SELECT COUNT(*) FROM ai_mail_campaigns WHERE group_id=? AND owner_emp_no=?',
            (group_id, _owner_emp_no())
        ).fetchone()[0]
        if campaign_count:
            return _error(
                '발송 이력이 있는 작업그룹은 삭제할 수 없습니다. 이름을 변경해 보관해주세요.',
                409,
                code='GROUP_HAS_HISTORY'
            )
        conn.execute('DELETE FROM ai_mail_recipients WHERE group_id=?', (group_id,))
        conn.execute('DELETE FROM ai_mail_workgroups WHERE id=? AND owner_emp_no=?', (group_id, _owner_emp_no()))
        conn.commit()
        return _success('작업그룹이 삭제되었습니다.')
    finally:
        conn.close()


@ai_mail_bp.route('/api/groups/<int:group_id>/copy', methods=['POST'])
@_mutating
def copy_group(group_id):
    new_name = _clean_text(_request_value('name'), 120)
    if not new_name:
        return _error('새 작업그룹 이름을 입력해주세요.', code='GROUP_NAME_REQUIRED')
    conn = _db()
    try:
        source = _owned_group(conn, group_id)
        if not source:
            return _error('복사할 작업그룹을 찾을 수 없습니다.', 404, code='GROUP_NOT_FOUND')
        conn.execute('BEGIN IMMEDIATE')
        cursor = conn.execute('''
            INSERT INTO ai_mail_workgroups (owner_emp_no, name, features_json)
            VALUES (?, ?, ?)
        ''', (_owner_emp_no(), new_name, source['features_json']))
        new_group_id = cursor.lastrowid
        conn.execute('''
            INSERT INTO ai_mail_recipients (group_id, email, recipient_name, memo)
            SELECT ?, email, recipient_name, memo
            FROM ai_mail_recipients WHERE group_id=?
        ''', (new_group_id, group_id))
        count = conn.execute('SELECT COUNT(*) FROM ai_mail_recipients WHERE group_id=?', (new_group_id,)).fetchone()[0]
        conn.commit()
        return _success(
            '수신자를 복사해 새 작업그룹을 만들었습니다.',
            group=_group_dict(_owned_group(conn, new_group_id), count)
        )
    except sqlite3.IntegrityError:
        conn.rollback()
        return _error('같은 이름의 작업그룹이 이미 있습니다.', 409, code='GROUP_DUPLICATE')
    finally:
        conn.close()


@ai_mail_bp.route('/api/groups/<int:group_id>/recipients', methods=['GET'])
@_login_required
def list_recipients(group_id):
    conn = _db()
    try:
        if not _owned_group(conn, group_id):
            return _error('작업그룹을 찾을 수 없습니다.', 404, code='GROUP_NOT_FOUND')
        rows = conn.execute('''
            SELECT * FROM ai_mail_recipients WHERE group_id=?
            ORDER BY recipient_name COLLATE NOCASE, id
        ''', (group_id,)).fetchall()
        return _success(recipients=[dict(row) for row in rows], count=len(rows), limit=MAX_RECIPIENTS)
    finally:
        conn.close()


@ai_mail_bp.route('/api/groups/<int:group_id>/recipients', methods=['POST'])
@_mutating
def create_recipient(group_id):
    email = _clean_text(_request_value('email'), 254).lower()
    recipient_name = _clean_text(_request_value('recipient_name', _request_value('name')), 160)
    memo = _clean_text(_request_value('memo'), 1000)
    if not _is_valid_email(email):
        return _error('수신자 메일주소 형식이 올바르지 않습니다.', code='EMAIL_INVALID')
    if not recipient_name:
        return _error('수신자명을 입력해주세요.', code='RECIPIENT_NAME_REQUIRED')
    conn = _db()
    try:
        if not _owned_group(conn, group_id):
            return _error('작업그룹을 찾을 수 없습니다.', 404, code='GROUP_NOT_FOUND')
        conn.execute('BEGIN IMMEDIATE')
        count = conn.execute('SELECT COUNT(*) FROM ai_mail_recipients WHERE group_id=?', (group_id,)).fetchone()[0]
        if count >= MAX_RECIPIENTS:
            conn.rollback()
            return _error('작업그룹에는 수신자를 최대 300명까지 등록할 수 있습니다.', 409, code='RECIPIENT_LIMIT')
        cursor = conn.execute('''
            INSERT INTO ai_mail_recipients (group_id, email, recipient_name, memo)
            VALUES (?, ?, ?, ?)
        ''', (group_id, email, recipient_name, memo))
        conn.execute('UPDATE ai_mail_workgroups SET updated_at=CURRENT_TIMESTAMP WHERE id=?', (group_id,))
        conn.commit()
        row = conn.execute('SELECT * FROM ai_mail_recipients WHERE id=?', (cursor.lastrowid,)).fetchone()
        return _success('수신자가 등록되었습니다.', recipient=dict(row), count=count + 1)
    except sqlite3.IntegrityError:
        conn.rollback()
        return _error('같은 메일주소가 이 작업그룹에 이미 등록되어 있습니다.', 409, code='RECIPIENT_DUPLICATE')
    finally:
        conn.close()


@ai_mail_bp.route('/api/groups/<int:group_id>/recipients/<int:recipient_id>', methods=['PUT', 'PATCH'])
@_mutating
def update_recipient(group_id, recipient_id):
    data = _json_data() if request.is_json else request.form
    conn = _db()
    try:
        if not _owned_group(conn, group_id):
            return _error('작업그룹을 찾을 수 없습니다.', 404, code='GROUP_NOT_FOUND')
        row = conn.execute('SELECT * FROM ai_mail_recipients WHERE id=? AND group_id=?', (recipient_id, group_id)).fetchone()
        if not row:
            return _error('수신자를 찾을 수 없습니다.', 404, code='RECIPIENT_NOT_FOUND')
        email = _clean_text(data.get('email', row['email']), 254).lower()
        name = _clean_text(data.get('recipient_name', data.get('name', row['recipient_name'])), 160)
        memo = _clean_text(data.get('memo', row['memo']), 1000)
        if not _is_valid_email(email):
            return _error('수신자 메일주소 형식이 올바르지 않습니다.', code='EMAIL_INVALID')
        if not name:
            return _error('수신자명을 입력해주세요.', code='RECIPIENT_NAME_REQUIRED')
        conn.execute('''
            UPDATE ai_mail_recipients
            SET email=?, recipient_name=?, memo=?, updated_at=CURRENT_TIMESTAMP
            WHERE id=? AND group_id=?
        ''', (email, name, memo, recipient_id, group_id))
        conn.execute('UPDATE ai_mail_workgroups SET updated_at=CURRENT_TIMESTAMP WHERE id=?', (group_id,))
        conn.commit()
        updated = conn.execute('SELECT * FROM ai_mail_recipients WHERE id=?', (recipient_id,)).fetchone()
        return _success('수신자 정보가 수정되었습니다.', recipient=dict(updated))
    except sqlite3.IntegrityError:
        conn.rollback()
        return _error('같은 메일주소가 이 작업그룹에 이미 등록되어 있습니다.', 409, code='RECIPIENT_DUPLICATE')
    finally:
        conn.close()


@ai_mail_bp.route('/api/groups/<int:group_id>/recipients/<int:recipient_id>', methods=['DELETE'])
@_mutating
def delete_recipient(group_id, recipient_id):
    conn = _db()
    try:
        if not _owned_group(conn, group_id):
            return _error('작업그룹을 찾을 수 없습니다.', 404, code='GROUP_NOT_FOUND')
        cursor = conn.execute('DELETE FROM ai_mail_recipients WHERE id=? AND group_id=?', (recipient_id, group_id))
        if not cursor.rowcount:
            return _error('수신자를 찾을 수 없습니다.', 404, code='RECIPIENT_NOT_FOUND')
        conn.execute('UPDATE ai_mail_workgroups SET updated_at=CURRENT_TIMESTAMP WHERE id=?', (group_id,))
        conn.commit()
        count = conn.execute('SELECT COUNT(*) FROM ai_mail_recipients WHERE group_id=?', (group_id,)).fetchone()[0]
        return _success('수신자가 삭제되었습니다.', count=count)
    finally:
        conn.close()


def _uploaded_size(file_storage):
    stream = file_storage.stream
    position = stream.tell()
    stream.seek(0, os.SEEK_END)
    size = stream.tell()
    stream.seek(position)
    return size


def _normalized_header(value):
    return re.sub(r'[\s_\-]+', '', unicodedata.normalize('NFKC', _clean_text(value))).lower()


def _recipient_import_rows(file_storage):
    filename = _clean_text(file_storage.filename)
    ext = os.path.splitext(filename)[1].lower()
    if ext not in {'.xlsx', '.xlsm', '.csv'}:
        raise ValueError('xlsx, xlsm, csv 파일만 등록할 수 있습니다.')
    if _uploaded_size(file_storage) > MAX_IMPORT_BYTES:
        raise ValueError('수신자 일괄등록 파일은 5MB 이하만 가능합니다.')

    file_storage.stream.seek(0)
    if ext == '.csv':
        raw = file_storage.stream.read()
        decoded = None
        for encoding in ('utf-8-sig', 'cp949', 'euc-kr'):
            try:
                decoded = raw.decode(encoding)
                break
            except UnicodeDecodeError:
                continue
        if decoded is None:
            raise ValueError('CSV 문자 인코딩을 읽을 수 없습니다.')
        rows = list(csv.reader(io.StringIO(decoded)))
    else:
        workbook = load_workbook(file_storage.stream, data_only=True, read_only=True)
        sheet = workbook.active
        rows = [list(row) for row in sheet.iter_rows(values_only=True)]

    rows = [row for row in rows if any(_clean_text(cell) for cell in row)]
    if not rows:
        raise ValueError('읽을 수 있는 수신자 내역이 없습니다.')

    aliases = {
        'email': {'메일주소', '이메일', 'email', 'e-mail'},
        'recipient_name': {'수신자명', '수신자', '이름', '성명'},
        'memo': {'메모', '비고', '참고'}
    }
    normalized_aliases = {key: {_normalized_header(v) for v in values} for key, values in aliases.items()}
    best_index, best_score = 0, -1
    for index, row in enumerate(rows[:15]):
        cells = {_normalized_header(cell) for cell in row}
        score = sum(bool(cells & values) for values in normalized_aliases.values())
        if score > best_score:
            best_index, best_score = index, score
    headers = [_normalized_header(cell) for cell in rows[best_index]]
    column_map = {}
    for field, values in normalized_aliases.items():
        for index, header in enumerate(headers):
            if header in values:
                column_map[field] = index
                break
    if 'email' not in column_map or 'recipient_name' not in column_map:
        raise ValueError("제목행에 '메일주소', '수신자명' 열이 필요합니다.")

    parsed, errors, seen = [], [], set()
    for excel_row, row in enumerate(rows[best_index + 1:], start=best_index + 2):
        def cell(field):
            index = column_map.get(field)
            return _clean_text(row[index]) if index is not None and index < len(row) else ''

        email = cell('email').lower()
        name = cell('recipient_name')
        memo = cell('memo')[:1000]
        if not email and not name and not memo:
            continue
        row_errors = []
        if not _is_valid_email(email):
            row_errors.append('메일주소 형식 오류')
        if not name:
            row_errors.append('수신자명 누락')
        if email in seen:
            row_errors.append('파일 내 중복 메일주소')
        if row_errors:
            errors.append({'row': excel_row, 'email': email, 'name': name, 'message': ', '.join(row_errors)})
            continue
        seen.add(email)
        parsed.append({'email': email, 'recipient_name': name[:160], 'memo': memo})
    return parsed, errors


@ai_mail_bp.route('/api/recipient-template.xlsx')
@_login_required
def download_recipient_template():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = '수신자'
    sheet.append(['메일주소', '수신자명', '메모'])
    sheet.append(['school@example.com', '청곡초등학교', '7월 내역서'])
    sheet.freeze_panes = 'A2'
    sheet.column_dimensions['A'].width = 32
    sheet.column_dimensions['B'].width = 28
    sheet.column_dimensions['C'].width = 45
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return send_file(
        buffer,
        as_attachment=True,
        download_name='AI메일_수신자_일괄등록_양식.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@ai_mail_bp.route('/api/groups/<int:group_id>/recipients/import', methods=['POST'])
@_mutating
def import_recipients(group_id):
    file_storage = request.files.get('file') or request.files.get('excel')
    if not file_storage or not file_storage.filename:
        return _error('일괄등록 엑셀 또는 CSV 파일을 첨부해주세요.', code='IMPORT_FILE_REQUIRED')
    try:
        parsed, row_errors = _recipient_import_rows(file_storage)
    except (ValueError, OSError) as exc:
        return _error(str(exc), code='IMPORT_PARSE_ERROR')
    except Exception as exc:
        return _error(f'일괄등록 파일을 읽지 못했습니다: {_clean_text(exc, 300)}', code='IMPORT_PARSE_ERROR')

    if row_errors:
        return _error(
            f'수정이 필요한 행이 {len(row_errors)}건 있어 등록하지 않았습니다.',
            code='IMPORT_ROW_ERRORS',
            errors=row_errors[:100],
            error_count=len(row_errors)
        )
    if not parsed:
        return _error('등록할 수신자가 없습니다.', code='IMPORT_EMPTY')

    conn = _db()
    try:
        if not _owned_group(conn, group_id):
            return _error('작업그룹을 찾을 수문을 수 없습니다.', 404, code='GROUP_NOT_FOUND')
        conn.execute('BEGIN IMMEDIATE')
        existing_rows = conn.execute('SELECT email FROM ai_mail_recipients WHERE group_id=?', (group_id,)).fetchall()
        existing = {row['email'].lower() for row in existing_rows}
        new_rows = [row for row in parsed if row['email'] not in existing]
        skipped = [row['email'] for row in parsed if row['email'] in existing]
        if len(existing) + len(new_rows) > MAX_RECIPIENTS:
            conn.rollback()
            return _error(
                f'등록 후 수신자가 {len(existing) + len(new_rows)}명이 되어 300명 한도를 초과합니다.',
                409,
                code='RECIPIENT_LIMIT',
                current_count=len(existing),
                new_count=len(new_rows),
                limit=MAX_RECIPIENTS
            )
        conn.executemany('''
            INSERT INTO ai_mail_recipients (group_id, email, recipient_name, memo)
            VALUES (?, ?, ?, ?)
        ''', [(group_id, row['email'], row['recipient_name'], row['memo']) for row in new_rows])
        conn.execute('UPDATE ai_mail_workgroups SET updated_at=CURRENT_TIMESTAMP WHERE id=?', (group_id,))
        conn.commit()
        return _success(
            f'수신자 {len(new_rows)}명을 일괄등록했습니다.',
            imported_count=len(new_rows),
            skipped_count=len(skipped),
            skipped_emails=skipped,
            total_count=len(existing) + len(new_rows)
        )
    except sqlite3.IntegrityError as exc:
        conn.rollback()
        return _error(f'일괄등록 중 중복 데이터가 발견되었습니다: {_clean_text(exc, 200)}', 409, code='RECIPIENT_DUPLICATE')
    finally:
        conn.close()


def _smtp_error_info(exc):
    message = _clean_text(exc, 1000) or exc.__class__.__name__
    smtp_code = getattr(exc, 'smtp_code', None)
    smtp_error = getattr(exc, 'smtp_error', None)
    if isinstance(smtp_error, bytes):
        smtp_error = smtp_error.decode('utf-8', errors='replace')
    if smtp_error:
        message = _clean_text(smtp_error, 1000)

    if isinstance(exc, smtplib.SMTPAuthenticationError):
        return 'AUTH_FAILED', '구글 계정 또는 앱 비밀번호 인증에 실패했습니다.', smtp_code, message, False
    if isinstance(exc, smtplib.SMTPRecipientsRefused):
        return 'RECIPIENT_REFUSED', '수신자 메일 주소가 거부되었습니다.', smtp_code, message, False
    if isinstance(exc, smtplib.SMTPSenderRefused):
        return 'SENDER_REFUSED', '발송자 메일 주소가 거부되었습니다.', smtp_code, message, False
    if isinstance(exc, smtplib.SMTPDataError):
        transient = bool(smtp_code and 400 <= int(smtp_code) < 500)
        code = 'SMTP_TEMPORARY' if transient else 'MESSAGE_REJECTED'
        return code, '메일 본문 또는 첨부파일이 메일 서버에서 거부되었습니다.', smtp_code, message, transient
    if isinstance(exc, smtplib.SMTPResponseException):
        transient = bool(smtp_code and 400 <= int(smtp_code) < 500)
        code = 'SMTP_TEMPORARY' if transient else 'SMTP_REJECTED'
        return code, '메일 서버가 요청을 거부했습니다.', smtp_code, message, transient
    if isinstance(exc, (smtplib.SMTPServerDisconnected, smtplib.SMTPConnectError, socket.timeout, TimeoutError, OSError)):
        return 'SMTP_CONNECTION', '메일 서버 연결에 실패했습니다.', smtp_code, message, True
    return 'SMTP_UNKNOWN', '메일 발송 중 알 수 없는 오류가 발생했습니다.', smtp_code, message, False


def _smtp_login(sender_email, app_password):
    server = smtplib.SMTP('smtp.gmail.com', 587, timeout=20)
    try:
        server.ehlo()
        server.starttls()
        server.ehlo()
        server.login(sender_email, app_password)
        return server
    except Exception:
        try:
            server.quit()
        except Exception:
            pass
        raise


@ai_mail_bp.route('/api/senders', methods=['GET'])
@_login_required
def list_senders():
    conn = _db()
    try:
        rows = conn.execute('''
            SELECT * FROM ai_mail_senders WHERE owner_emp_no=?
            ORDER BY is_active DESC, updated_at DESC, id DESC
        ''', (_owner_emp_no(),)).fetchall()
        return _success(senders=[_sender_dict(row) for row in rows])
    finally:
        conn.close()


@ai_mail_bp.route('/api/senders', methods=['POST'])
@_mutating
def create_sender():
    email = _clean_text(_request_value('email'), 254).lower()
    label = _clean_text(_request_value('label'), 120) or email
    password = _request_value('app_password', _request_value('password'))
    if not _is_valid_email(email):
        return _error('발송자 메일주소 형식이 올바르지 않습니다.', code='EMAIL_INVALID')
    try:
        encrypted = _encrypt_password(password)
    except (ValueError, RuntimeError) as exc:
        return _error(str(exc), code='SENDER_SECRET_INVALID')
    conn = _db()
    try:
        cursor = conn.execute('''
            INSERT INTO ai_mail_senders (
                owner_emp_no, label, email, encrypted_app_password, is_active
            ) VALUES (?, ?, ?, ?, 1)
        ''', (_owner_emp_no(), label, email, encrypted))
        conn.commit()
        return _success('발송자 계정이 등록되었습니다.', sender=_sender_dict(_owned_row(conn, 'ai_mail_senders', cursor.lastrowid)))
    except sqlite3.IntegrityError:
        conn.rollback()
        return _error('같은 발송자 메일주소가 이미 등록되어 있습니다.', 409, code='SENDER_DUPLICATE')
    finally:
        conn.close()


@ai_mail_bp.route('/api/senders/<int:sender_id>', methods=['PUT', 'PATCH'])
@_mutating
def update_sender(sender_id):
    data = _json_data() if request.is_json else request.form
    conn = _db()
    try:
        sender = _owned_row(conn, 'ai_mail_senders', sender_id)
        if not sender:
            return _error('발송자 계정을 찾을 수 없습니다.', 404, code='SENDER_NOT_FOUND')
        email = _clean_text(data.get('email', sender['email']), 254).lower()
        label = _clean_text(data.get('label', sender['label']), 120) or email
        is_active = 1 if str(data.get('is_active', sender['is_active'])).lower() not in {'0', 'false', 'off', 'no'} else 0
        if not _is_valid_email(email):
            return _error('발송자 메일주소 형식이 올바르지 않습니다.', code='EMAIL_INVALID')
        encrypted = sender['encrypted_app_password']
        password = data.get('app_password', data.get('password'))
        if password not in (None, ''):
            try:
                encrypted = _encrypt_password(password)
            except (ValueError, RuntimeError) as exc:
                return _error(str(exc), code='SENDER_SECRET_INVALID')
        conn.execute('''
            UPDATE ai_mail_senders
            SET label=?, email=?, encrypted_app_password=?, is_active=?,
                last_test_status=NULL, last_test_error=NULL, updated_at=CURRENT_TIMESTAMP
            WHERE id=? AND owner_emp_no=?
        ''', (label, email, encrypted, is_active, sender_id, _owner_emp_no()))
        conn.commit()
        return _success('발송자 계정이 수정되었습니다.', sender=_sender_dict(_owned_row(conn, 'ai_mail_senders', sender_id)))
    except sqlite3.IntegrityError:
        conn.rollback()
        return _error('같은 발송자 메일주소가 이미 등록되어 있습니다.', 409, code='SENDER_DUPLICATE')
    finally:
        conn.close()


@ai_mail_bp.route('/api/senders/<int:sender_id>', methods=['DELETE'])
@_mutating
def deactivate_sender(sender_id):
    conn = _db()
    try:
        sender = _owned_row(conn, 'ai_mail_senders', sender_id)
        if not sender:
            return _error('발송자 계정을 찾을 수 없습니다.', 404, code='SENDER_NOT_FOUND')
        conn.execute('''
            UPDATE ai_mail_senders SET is_active=0, updated_at=CURRENT_TIMESTAMP
            WHERE id=? AND owner_emp_no=?
        ''', (sender_id, _owner_emp_no()))
        conn.commit()
        return _success('발송자 계정을 비활성화했습니다.', sender=_sender_dict(_owned_row(conn, 'ai_mail_senders', sender_id)))
    finally:
        conn.close()


@ai_mail_bp.route('/api/senders/<int:sender_id>/test', methods=['POST'])
@_mutating
def test_sender(sender_id):
    conn = _db()
    try:
        sender = _owned_row(conn, 'ai_mail_senders', sender_id)
        if not sender:
            return _error('발송자 계정을 찾을 수 없습니다.', 404, code='SENDER_NOT_FOUND')
        try:
            password = _decrypt_password(sender['encrypted_app_password'])
            server = _smtp_login(sender['email'], password)
            try:
                server.quit()
            except Exception:
                pass
            conn.execute('''
                UPDATE ai_mail_senders
                SET last_tested_at=CURRENT_TIMESTAMP, last_test_status='success',
                    last_test_error=NULL, updated_at=CURRENT_TIMESTAMP
                WHERE id=?
            ''', (sender_id,))
            conn.commit()
            return _success('구글 SMTP 인증에 성공했습니다.', sender=_sender_dict(_owned_row(conn, 'ai_mail_senders', sender_id)))
        except Exception as exc:
            code, friendly, smtp_code, detail, _ = _smtp_error_info(exc)
            full_error = f'{friendly} ({detail})' if detail else friendly
            conn.execute('''
                UPDATE ai_mail_senders
                SET last_tested_at=CURRENT_TIMESTAMP, last_test_status='error',
                    last_test_error=?, updated_at=CURRENT_TIMESTAMP
                WHERE id=?
            ''', (_clean_text(full_error, 1000), sender_id))
            conn.commit()
            return _error(
                friendly,
                400,
                code=code,
                smtp_code=smtp_code,
                detail=detail,
                sender=_sender_dict(_owned_row(conn, 'ai_mail_senders', sender_id))
            )
    finally:
        conn.close()


ALLOWED_HTML_TAGS = {
    'a', 'b', 'blockquote', 'br', 'div', 'em', 'h1', 'h2', 'h3', 'h4', 'hr',
    'i', 'img', 'li', 'ol', 'p', 's', 'span', 'strong', 'table', 'tbody', 'td',
    'th', 'thead', 'tr', 'u', 'ul'
}
DROP_HTML_TAGS = {
    'base', 'button', 'embed', 'form', 'iframe', 'input', 'link', 'meta',
    'object', 'script', 'style', 'textarea'
}
ALLOWED_STYLE_PROPERTIES = {
    'background', 'background-color', 'border', 'border-bottom', 'border-collapse',
    'border-color', 'border-left', 'border-radius', 'border-right', 'border-style',
    'border-top', 'border-width', 'color', 'display', 'font-family', 'font-size',
    'font-style', 'font-weight', 'height', 'letter-spacing', 'line-height', 'margin',
    'margin-bottom', 'margin-left', 'margin-right', 'margin-top', 'max-width',
    'min-width', 'padding', 'padding-bottom', 'padding-left', 'padding-right',
    'padding-top', 'text-align', 'text-decoration', 'vertical-align', 'white-space',
    'width'
}


def _safe_style(value):
    safe = []
    for declaration in _clean_text(value).split(';'):
        if ':' not in declaration:
            continue
        prop, raw_value = declaration.split(':', 1)
        prop = prop.strip().lower()
        raw_value = raw_value.strip()
        lowered = raw_value.lower().replace(' ', '')
        if prop not in ALLOWED_STYLE_PROPERTIES:
            continue
        if any(token in lowered for token in ('expression(', 'javascript:', 'vbscript:', '@import', '-moz-binding')):
            continue
        if 'url(' in lowered:
            continue
        safe.append(f'{prop}:{raw_value}')
    return ';'.join(safe)


def _safe_link(value, image=False):
    raw_value = str(value or '').strip()
    lowered = raw_value.lower()
    allowed = ('https://', 'cid:') if image else ('https://', 'http://', 'mailto:')
    
    if image and re.match(r'^data:image/(?:png|jpeg|gif|webp);base64,', lowered):
        return raw_value

    value = _clean_text(raw_value, 10000)
    return value if lowered.startswith(allowed) else ''


def _sanitize_html(value):
    raw = str(value or '')
    if len(raw.encode('utf-8')) > 28 * 1024 * 1024:
        raise ValueError('이미지를 포함한 템플릿 HTML이 28MB를 초과합니다.')
    soup = BeautifulSoup(raw, 'html.parser')
    for comment in soup.find_all(string=lambda text: isinstance(text, Comment)):
        comment.extract()
    for tag in list(soup.find_all(True)):
        name = tag.name.lower()
        if name in DROP_HTML_TAGS:
            tag.decompose()
            continue
        if name not in ALLOWED_HTML_TAGS:
            tag.unwrap()
            continue
        cleaned = {}
        style = _safe_style(tag.attrs.get('style', ''))
        if style:
            cleaned['style'] = style
        if name == 'a':
            href = _safe_link(tag.attrs.get('href', ''))
            if href:
                cleaned['href'] = href
                cleaned['target'] = '_blank'
                cleaned['rel'] = 'noopener noreferrer'
        elif name == 'img':
            src = _safe_link(tag.attrs.get('src', ''), image=True)
            if src:
                cleaned['src'] = src
            alt = _clean_text(tag.attrs.get('alt', ''), 300)
            if alt:
                cleaned['alt'] = alt
            for dimension in ('width', 'height'):
                raw_dimension = _clean_text(tag.attrs.get(dimension, ''), 10)
                if raw_dimension.isdigit():
                    cleaned[dimension] = raw_dimension
        elif name in {'td', 'th'}:
            for attr in ('colspan', 'rowspan'):
                raw_span = _clean_text(tag.attrs.get(attr, ''), 3)
                if raw_span.isdigit() and 1 <= int(raw_span) <= 20:
                    cleaned[attr] = raw_span
        tag.attrs = cleaned
    return str(soup)


def _plain_from_html(body_html):
    soup = BeautifulSoup(body_html or '', 'html.parser')
    return soup.get_text('\n', strip=True)[:200000]


def _storage_dir(kind, record_id):
    directory = os.path.abspath(os.path.join(AI_MAIL_UPLOADS, kind, str(record_id)))
    root = os.path.abspath(AI_MAIL_UPLOADS)
    if os.path.commonpath([root, directory]) != root:
        raise ValueError('저장 경로가 올바르지 않습니다.')
    os.makedirs(directory, exist_ok=True)
    return directory


def _safe_saved_path(path):
    root = os.path.abspath(AI_MAIL_UPLOADS)
    target = os.path.abspath(path or '')
    if not path or os.path.commonpath([root, target]) != root:
        raise ValueError('파일 경로가 올바르지 않습니다.')
    return target


def _delete_saved_file(path):
    try:
        target = _safe_saved_path(path)
        if os.path.isfile(target):
            os.remove(target)
    except (OSError, ValueError):
        pass


def _image_format_and_mime(stream):
    position = stream.tell()
    stream.seek(0)
    try:
        image = Image.open(stream)
        image.verify()
        image_format = (image.format or '').upper()
    except (UnidentifiedImageError, OSError, ValueError) as exc:
        raise ValueError('올바른 이미지 파일이 아닙니다.') from exc
    finally:
        stream.seek(position)
    formats = {
        'PNG': ('.png', 'image/png'),
        'JPEG': ('.jpg', 'image/jpeg'),
        'GIF': ('.gif', 'image/gif'),
        'WEBP': ('.webp', 'image/webp')
    }
    if image_format not in formats:
        raise ValueError('PNG, JPG, GIF, WEBP 이미지만 사용할 수 있습니다.')
    return formats[image_format]


def _template_asset_dict(row):
    item = dict(row)
    item.pop('filepath', None)
    item.pop('stored_name', None)
    item['cid_url'] = f"cid:{item['content_id']}"
    item['preview_url'] = f"/ai-mail/api/template-assets/{item['id']}"
    return item


def _save_template_image_bytes(conn, template_id, data, mime_type, original_name):
    if len(data) > MAX_TEMPLATE_IMAGE_BYTES:
        raise ValueError('템플릿 이미지는 1개당 3MB 이하만 가능합니다.')
    total = conn.execute(
        'SELECT COALESCE(SUM(size_bytes), 0) FROM ai_mail_template_assets WHERE template_id=?',
        (template_id,)
    ).fetchone()[0]
    if int(total or 0) + len(data) > MAX_TEMPLATE_TOTAL_IMAGE_BYTES:
        raise ValueError('템플릿 이미지 합계는 20MB 이하만 가능합니다.')
    subtype = mime_type.split('/', 1)[1]
    extension = '.jpg' if subtype == 'jpeg' else f'.{subtype}'
    stored_name = f'{uuid.uuid4().hex}{extension}'
    target = os.path.join(_storage_dir('templates', template_id), stored_name)
    with open(target, 'wb') as output:
        output.write(data)
    content_id = f'ai-mail-template-{template_id}-{uuid.uuid4().hex}'
    cursor = conn.execute('''
        INSERT INTO ai_mail_template_assets (
            template_id, original_name, stored_name, filepath, mime_type,
            content_id, size_bytes
        ) VALUES (?, ?, ?, ?, ?, ?, ?)
    ''', (
        template_id,
        _clean_text(os.path.basename(original_name), 255) or f'image{extension}',
        stored_name,
        target,
        mime_type,
        content_id,
        len(data)
    ))
    return conn.execute('SELECT * FROM ai_mail_template_assets WHERE id=?', (cursor.lastrowid,)).fetchone()


def _extract_embedded_images(conn, template_id, body_html):
    soup = BeautifulSoup(body_html or '', 'html.parser')
    saved_paths = []
    data_pattern = re.compile(r'^data:(image/(?:png|jpeg|gif|webp));base64,(.+)$', re.I | re.S)
    try:
        for index, image_tag in enumerate(soup.find_all('img'), start=1):
            src = _clean_text(image_tag.get('src', ''))
            match = data_pattern.match(src)
            if not match:
                continue
            try:
                data = base64.b64decode(re.sub(r'\s+', '', match.group(2)), validate=True)
            except (ValueError, TypeError) as exc:
                raise ValueError(f'{index}번째 템플릿 이미지 데이터가 손상되었습니다.') from exc
            asset = _save_template_image_bytes(
                conn, template_id, data, match.group(1).lower(), f'embedded_{index}'
            )
            saved_paths.append(asset['filepath'])
            image_tag['src'] = f"cid:{asset['content_id']}"
        return str(soup), saved_paths
    except Exception:
        for path in saved_paths:
            _delete_saved_file(path)
        raise


def _validate_template_cids(conn, template_id, body_html):
    allowed = {
        row['content_id']
        for row in conn.execute(
            'SELECT content_id FROM ai_mail_template_assets WHERE template_id=?',
            (template_id,)
        ).fetchall()
    }
    soup = BeautifulSoup(body_html or '', 'html.parser')
    for tag in soup.find_all('img'):
        src = _clean_text(tag.get('src', ''))
        if src.lower().startswith('cid:') and src[4:] not in allowed:
            tag.decompose()
    return str(soup)


def _restore_template_asset_urls(conn, template_id, body_html):
    restored = str(body_html or '')
    for asset in conn.execute(
        'SELECT id, content_id FROM ai_mail_template_assets WHERE template_id=?',
        (template_id,)
    ).fetchall():
        restored = restored.replace(
            f"/ai-mail/api/template-assets/{asset['id']}",
            f"cid:{asset['content_id']}"
        )
    return restored


def _enforce_final_html_size(body_html):
    if len((body_html or '').encode('utf-8')) > 2 * 1024 * 1024:
        raise ValueError('이미지 분리 후 메일 본문 HTML은 2MB 이하여야 합니다.')


@ai_mail_bp.route('/api/templates', methods=['GET'])
@_login_required
def list_templates():
    conn = _db()
    try:
        rows = conn.execute('''
            SELECT * FROM ai_mail_templates WHERE owner_emp_no=? AND is_active=1
            ORDER BY is_active DESC, updated_at DESC, id DESC
        ''', (_owner_emp_no(),)).fetchall()
        return _success(templates=_templates_with_assets(conn, rows), limit=MAX_TEMPLATES)
    finally:
        conn.close()


@ai_mail_bp.route('/api/templates', methods=['POST'])
@_mutating
def create_template():
    name = _clean_text(_request_value('name'), 120)
    subject = _clean_text(_request_value('subject'), 300)
    raw_html = _request_value('body_html')
    body_text = _clean_text(_request_value('body_text'), 200000)
    if not name:
        return _error('템플릿 이름을 입력해주세요.', code='TEMPLATE_NAME_REQUIRED')
    if not subject:
        return _error('메일 제목을 입력해주세요.', code='SUBJECT_REQUIRED')
    try:
        sanitized = _sanitize_html(raw_html)
    except ValueError as exc:
        return _error(str(exc), code='HTML_INVALID')
    if not BeautifulSoup(sanitized, 'html.parser').get_text(strip=True) and '<img' not in sanitized.lower():
        return _error('메일 내용을 입력해주세요.', code='BODY_REQUIRED')

    conn = _db()
    saved_paths = []
    try:
        conn.execute('BEGIN IMMEDIATE')
        count = conn.execute('''
            SELECT COUNT(*) FROM ai_mail_templates WHERE owner_emp_no=? AND is_active=1
        ''', (_owner_emp_no(),)).fetchone()[0]
        if count >= MAX_TEMPLATES:
            conn.rollback()
            return _error('템플릿은 최대 100개까지 등록할 수 있습니다.', 409, code='TEMPLATE_LIMIT')
        cursor = conn.execute('''
            INSERT INTO ai_mail_templates (
                owner_emp_no, name, subject, body_html, body_text, is_active
            ) VALUES (?, ?, ?, '', '', 1)
        ''', (_owner_emp_no(), name, subject))
        template_id = cursor.lastrowid
        sanitized, saved_paths = _extract_embedded_images(conn, template_id, sanitized)
        sanitized = _validate_template_cids(conn, template_id, sanitized)
        _enforce_final_html_size(sanitized)
        final_text = body_text or _plain_from_html(sanitized)
        conn.execute('''
            UPDATE ai_mail_templates SET body_html=?, body_text=? WHERE id=?
        ''', (sanitized, final_text, template_id))
        conn.commit()
        row = _owned_row(conn, 'ai_mail_templates', template_id)
        assets = conn.execute('SELECT * FROM ai_mail_template_assets WHERE template_id=?', (template_id,)).fetchall()
        return _success('메일 템플릿이 저장되었습니다.', template=_template_dict(row, assets))
    except sqlite3.IntegrityError:
        conn.rollback()
        for path in saved_paths:
            _delete_saved_file(path)
        return _error('같은 이름의 템플릿이 이미 있습니다.', 409, code='TEMPLATE_DUPLICATE')
    except (ValueError, OSError) as exc:
        conn.rollback()
        for path in saved_paths:
            _delete_saved_file(path)
        return _error(str(exc), code='TEMPLATE_IMAGE_ERROR')
    finally:
        conn.close()


@ai_mail_bp.route('/api/templates/<int:template_id>', methods=['GET'])
@_login_required
def get_template(template_id):
    conn = _db()
    try:
        row = _owned_row(conn, 'ai_mail_templates', template_id)
        if not row:
            return _error('템플릿을 찾을 수 없습니다.', 404, code='TEMPLATE_NOT_FOUND')
        assets = conn.execute('SELECT * FROM ai_mail_template_assets WHERE template_id=? ORDER BY id', (template_id,)).fetchall()
        return _success(template=_template_dict(row, assets))
    finally:
        conn.close()


@ai_mail_bp.route('/api/templates/<int:template_id>', methods=['PUT', 'PATCH'])
@_mutating
def update_template(template_id):
    data = _json_data() if request.is_json else request.form
    conn = _db()
    saved_paths = []
    try:
        current = _owned_row(conn, 'ai_mail_templates', template_id)
        if not current:
            return _error('템플릿을 찾을 수 없습니다.', 404, code='TEMPLATE_NOT_FOUND')
        name = _clean_text(data.get('name', current['name']), 120)
        subject = _clean_text(data.get('subject', current['subject']), 300)
        if not name or not subject:
            return _error('템플릿 이름과 메일 제목은 필수입니다.', code='TEMPLATE_FIELDS_REQUIRED')
        raw_html = _restore_template_asset_urls(
            conn, template_id, data.get('body_html', current['body_html'])
        )
        try:
            sanitized = _sanitize_html(raw_html)
        except ValueError as exc:
            return _error(str(exc), code='HTML_INVALID')
        desired_active = 1 if str(data.get('is_active', current['is_active'])).lower() not in {'0', 'false', 'off', 'no'} else 0
        conn.execute('BEGIN IMMEDIATE')
        if desired_active and not current['is_active']:
            active_count = conn.execute('''
                SELECT COUNT(*) FROM ai_mail_templates WHERE owner_emp_no=? AND is_active=1
            ''', (_owner_emp_no(),)).fetchone()[0]
            if active_count >= MAX_TEMPLATES:
                conn.rollback()
                return _error('템플릿은 최대 100개까지 활성화할 수 있습니다.', 409, code='TEMPLATE_LIMIT')
        sanitized, saved_paths = _extract_embedded_images(conn, template_id, sanitized)
        sanitized = _validate_template_cids(conn, template_id, sanitized)
        _enforce_final_html_size(sanitized)
        body_text = _clean_text(data.get('body_text'), 200000) or _plain_from_html(sanitized)
        conn.execute('''
            UPDATE ai_mail_templates
            SET name=?, subject=?, body_html=?, body_text=?, is_active=?,
                updated_at=CURRENT_TIMESTAMP
            WHERE id=? AND owner_emp_no=?
        ''', (name, subject, sanitized, body_text, desired_active, template_id, _owner_emp_no()))
        conn.commit()
        updated = _owned_row(conn, 'ai_mail_templates', template_id)
        assets = conn.execute('SELECT * FROM ai_mail_template_assets WHERE template_id=? ORDER BY id', (template_id,)).fetchall()
        return _success('메일 템플릿이 수정되었습니다.', template=_template_dict(updated, assets))
    except sqlite3.IntegrityError:
        conn.rollback()
        for path in saved_paths:
            _delete_saved_file(path)
        return _error('같은 이름의 템플릿이 이미 있습니다.', 409, code='TEMPLATE_DUPLICATE')
    except (ValueError, OSError) as exc:
        conn.rollback()
        for path in saved_paths:
            _delete_saved_file(path)
        return _error(str(exc), code='TEMPLATE_IMAGE_ERROR')
    finally:
        conn.close()


@ai_mail_bp.route('/api/templates/<int:template_id>/images', methods=['POST'])
@_mutating
def upload_template_image(template_id):
    file_storage = request.files.get('file') or request.files.get('image')
    if not file_storage or not file_storage.filename:
        return _error('이미지 파일을 첨부해주세요.', code='IMAGE_REQUIRED')
    if _uploaded_size(file_storage) > MAX_TEMPLATE_IMAGE_BYTES:
        return _error('템플릿 이미지는 1개당 3MB 이하만 가능합니다.', code='IMAGE_TOO_LARGE')
    try:
        extension, mime_type = _image_format_and_mime(file_storage.stream)
    except ValueError as exc:
        return _error(str(exc), code='IMAGE_INVALID')
    file_storage.stream.seek(0)
    data = file_storage.stream.read()
    conn = _db()
    path = None
    try:
        if not _owned_row(conn, 'ai_mail_templates', template_id):
            return _error('템플릿을 찾을 수 없습니다.', 404, code='TEMPLATE_NOT_FOUND')
        asset = _save_template_image_bytes(conn, template_id, data, mime_type, file_storage.filename)
        path = asset['filepath']
        conn.commit()
        return _success('템플릿 이미지가 등록되었습니다.', asset=_template_asset_dict(asset))
    except (ValueError, OSError) as exc:
        conn.rollback()
        if path:
            _delete_saved_file(path)
        return _error(str(exc), code='TEMPLATE_IMAGE_ERROR')
    finally:
        conn.close()


@ai_mail_bp.route('/api/template-assets/<int:asset_id>')
@_login_required
def download_template_asset(asset_id):
    conn = _db()
    try:
        asset = conn.execute('''
            SELECT a.* FROM ai_mail_template_assets a
            JOIN ai_mail_templates t ON t.id=a.template_id
            WHERE a.id=? AND t.owner_emp_no=?
        ''', (asset_id, _owner_emp_no())).fetchone()
        if not asset:
            return _error('템플릿 이미지를 찾을 수 없습니다.', 404, code='ASSET_NOT_FOUND')
        path = _safe_saved_path(asset['filepath'])
        if not os.path.isfile(path):
            return _error('템플릿 이미지 파일이 손실되었습니다.', 404, code='ASSET_FILE_MISSING')
        return send_file(path, mimetype=asset['mime_type'], download_name=asset['original_name'])
    finally:
        conn.close()


@ai_mail_bp.route('/api/templates/<int:template_id>/images/<int:asset_id>', methods=['DELETE'])
@_mutating
def delete_template_image(template_id, asset_id):
    conn = _db()
    try:
        if not _owned_row(conn, 'ai_mail_templates', template_id):
            return _error('템플릿을 찾을 수 없습니다.', 404, code='TEMPLATE_NOT_FOUND')
        asset = conn.execute(
            'SELECT * FROM ai_mail_template_assets WHERE id=? AND template_id=?',
            (asset_id, template_id)
        ).fetchone()
        if not asset:
            return _error('템플릿 이미지를 찾을 수 없습니다.', 404, code='ASSET_NOT_FOUND')
        template = _owned_row(conn, 'ai_mail_templates', template_id)
        if f"cid:{asset['content_id']}" in (template['body_html'] or ''):
            return _error('메일 본문에서 사용 중인 이미지입니다. 본문에서 먼저 제거해주세요.', 409, code='ASSET_IN_USE')
        conn.execute('DELETE FROM ai_mail_template_assets WHERE id=? AND template_id=?', (asset_id, template_id))
        conn.commit()
        _delete_saved_file(asset['filepath'])
        return _success('템플릿 이미지가 삭제되었습니다.')
    finally:
        conn.close()


@ai_mail_bp.route('/api/templates/<int:template_id>', methods=['DELETE'])
@_mutating
def delete_template(template_id):
    conn = _db()
    try:
        template = _owned_row(conn, 'ai_mail_templates', template_id)
        if not template:
            return _error('템플릿을 찾을 수 없습니다.', 404, code='TEMPLATE_NOT_FOUND')
        campaign_count = conn.execute('''
            SELECT COUNT(*) FROM ai_mail_campaigns WHERE template_id=? AND owner_emp_no=?
        ''', (template_id, _owner_emp_no())).fetchone()[0]
        if campaign_count:
            conn.execute('''
                UPDATE ai_mail_templates SET is_active=0, updated_at=CURRENT_TIMESTAMP
                WHERE id=? AND owner_emp_no=?
            ''', (template_id, _owner_emp_no()))
            conn.commit()
            return _success('발송 이력을 보존하기 위해 템플릿을 비활성화했습니다.')
        assets = conn.execute('SELECT filepath FROM ai_mail_template_assets WHERE template_id=?', (template_id,)).fetchall()
        conn.execute('DELETE FROM ai_mail_template_assets WHERE template_id=?', (template_id,))
        conn.execute('DELETE FROM ai_mail_templates WHERE id=? AND owner_emp_no=?', (template_id, _owner_emp_no()))
        conn.commit()
        for asset in assets:
            _delete_saved_file(asset['filepath'])
        return _success('템플릿이 삭제되었습니다.')
    finally:
        conn.close()


def _as_int(value, default=None):
    try:
        return int(value)
    except (TypeError, ValueError):
        return default


def _as_bool(value, default=False):
    if value is None or value == '':
        return bool(default)
    if isinstance(value, bool):
        return value
    return str(value).strip().lower() in {'1', 'true', 'yes', 'on', 'y'}


def _group_feature_enabled(group, feature):
    configured = _parse_json_value(group['features_json'], {})
    aliases = {
        'smart_attachment': {'smart_attachment', 'smart', 'individual_attachment', '개별첨부'},
        'common_attachment': {'common_attachment', 'common', 'bulk_attachment', '공통첨부'},
        'templates': {'templates', 'template', 'mail_template', '템플릿'}
    }
    accepted = aliases.get(feature, {feature})
    if isinstance(configured, list):
        return bool({_clean_text(item).lower() for item in configured} & accepted)
    if isinstance(configured, dict):
        return any(_as_bool(configured.get(key)) for key in accepted if key in configured)
    return False


def _parse_recipient_ids(value):
    parsed = _parse_json_value(value, value)
    if parsed in (None, '', []):
        return []
    if isinstance(parsed, str):
        parsed = [item for item in re.split(r'[,\s]+', parsed) if item]
    if not isinstance(parsed, list):
        raise ValueError('수신자 선택 형식이 올바르지 않습니다.')
    result = []
    for value in parsed:
        number = _as_int(value)
        if number is None or number <= 0:
            raise ValueError('수신자 선택값에 잘못된 ID가 있습니다.')
        if number not in result:
            result.append(number)
    return result


def _attachment_mode_from_request():
    mode = _clean_text(_request_value('attachment_mode')).lower()
    if not mode:
        modes = _parse_json_value(
            _request_value('attachment_modes', _request_value('attachment_modes[]')),
            []
        )
        if isinstance(modes, str):
            modes = [modes]
        normalized = {_clean_text(item).lower() for item in (modes or [])}
        has_common = 'common' in normalized
        has_smart = bool(normalized & {'smart', 'individual', 'personal'})
        if has_common and has_smart:
            mode = 'smart_and_common'
        elif has_common:
            mode = 'common'
        elif has_smart:
            mode = 'smart'
        else:
            mode = 'none'
    aliases = {'individual': 'smart', 'personal': 'smart', 'both': 'smart_and_common'}
    return aliases.get(mode, mode)


def _campaign_event(conn, campaign_id, event_type, message, level='info', details=None):
    conn.execute('''
        INSERT INTO ai_mail_campaign_events (
            campaign_id, event_type, level, message, details_json
        ) VALUES (?, ?, ?, ?, ?)
    ''', (
        campaign_id,
        _clean_text(event_type, 80),
        _clean_text(level, 20) or 'info',
        _clean_text(message, 1000),
        json.dumps(details or {}, ensure_ascii=False)
    ))


def _attachment_dict(row):
    item = dict(row)
    item.pop('filepath', None)
    item.pop('stored_name', None)
    item['download_url'] = f"/ai-mail/api/campaigns/{item['campaign_id']}/attachments/{item['id']}/download"
    return item


def _safe_original_filename(value):
    name = _clean_text(os.path.basename(str(value or '').replace('\\', '/')), 255)
    if not name or name in {'.', '..'}:
        raise ValueError('첨부파일명이 올바르지 않습니다.')
    if '\x00' in name:
        raise ValueError('첨부파일명이 올바르지 않습니다.')
    return name


def _save_campaign_attachment(conn, campaign_id, file_storage, kind):
    if kind not in {'common', 'smart'}:
        raise ValueError('첨부파일 방식은 common 또는 smart여야 합니다.')
    original_name = _safe_original_filename(file_storage.filename)
    extension = os.path.splitext(original_name)[1].lower()
    if extension in BLOCKED_ATTACHMENT_EXTENSIONS:
        raise ValueError(f'보안상 첨부할 수 없는 파일 형식입니다: {original_name}')
    size = _uploaded_size(file_storage)
    if size <= 0:
        raise ValueError(f'빈 파일은 첨부할 수 없습니다: {original_name}')
    if size > MAX_MESSAGE_BYTES:
        raise ValueError(f'18MB를 초과한 파일은 첨부할 수 없습니다: {original_name}')
    count = conn.execute(
        'SELECT COUNT(*) FROM ai_mail_campaign_attachments WHERE campaign_id=?',
        (campaign_id,)
    ).fetchone()[0]
    if count >= MAX_ATTACHMENT_FILES:
        raise ValueError(f'한 발송작업에는 파일을 최대 {MAX_ATTACHMENT_FILES}개까지 등록할 수 있습니다.')
    stored_name = f'{uuid.uuid4().hex}{extension}'
    target = os.path.join(_storage_dir('campaigns', campaign_id), stored_name)
    file_storage.stream.seek(0)
    file_storage.save(target)
    digest = hashlib.sha256()
    with open(target, 'rb') as saved:
        for chunk in iter(lambda: saved.read(1024 * 1024), b''):
            digest.update(chunk)
    mime_type = mimetypes.guess_type(original_name)[0] or 'application/octet-stream'
    try:
        cursor = conn.execute('''
            INSERT INTO ai_mail_campaign_attachments (
                campaign_id, campaign_recipient_id, kind, match_method,
                match_status, original_name, stored_name, filepath, mime_type,
                size_bytes, sha256
            ) VALUES (?, NULL, ?, 'auto', ?, ?, ?, ?, ?, ?, ?)
        ''', (
            campaign_id,
            kind,
            'matched' if kind == 'common' else 'pending',
            original_name,
            stored_name,
            target,
            mime_type,
            size,
            digest.hexdigest()
        ))
        return conn.execute(
            'SELECT * FROM ai_mail_campaign_attachments WHERE id=?',
            (cursor.lastrowid,)
        ).fetchone()
    except Exception:
        _delete_saved_file(target)
        raise


def _attachment_files_from_request():
    common_files = [item for item in request.files.getlist('common_files') if item and item.filename]
    smart_files = [item for item in request.files.getlist('smart_files') if item and item.filename]
    single_files = [item for item in request.files.getlist('file') if item and item.filename]
    generic_files = [item for item in request.files.getlist('files') if item and item.filename]
    single_kind = _clean_text(_request_value('mode', _request_value('kind'))).lower()
    if single_kind not in {'common', 'smart'}:
        campaign_mode = _attachment_mode_from_request()
        single_kind = 'common' if campaign_mode == 'common' else 'smart'
    if single_kind == 'common':
        common_files.extend(single_files + generic_files)
    else:
        smart_files.extend(single_files + generic_files)
    return common_files, smart_files


@ai_mail_bp.route('/api/campaigns', methods=['POST'])
@_mutating
def create_campaign():
    group_id = _as_int(_request_value('group_id'))
    sender_id = _as_int(_request_value('sender_id'))
    template_id = _as_int(_request_value('template_id'))
    mode = _attachment_mode_from_request()
    allow_missing_attachment = _as_bool(
        _request_value('allow_missing_attachment', _request_value('allow_missing_attachments', False))
    )
    if mode not in ATTACHMENT_MODES:
        return _error('첨부파일 발송 방식이 올바르지 않습니다.', code='ATTACHMENT_MODE_INVALID')
    try:
        recipient_ids = _parse_recipient_ids(_request_value('recipient_ids', []))
    except ValueError as exc:
        return _error(str(exc), code='RECIPIENT_SELECTION_INVALID')
    raw_interval = _request_value('send_interval', _request_value('interval_seconds', 1.0))
    try:
        send_interval = max(0.5, min(float(raw_interval), 30.0))
    except (TypeError, ValueError):
        return _error('발송 간격은 0.5~30초 사이의 숫자여야 합니다.', code='INTERVAL_INVALID')

    conn = _db()
    saved_paths = []
    try:
        group = _owned_group(conn, group_id) if group_id else None
        if not group:
            return _error('작업그룹을 찾을 수 없습니다.', 404, code='GROUP_NOT_FOUND')
        if mode in {'smart', 'smart_and_common'} and not _group_feature_enabled(group, 'smart_attachment'):
            return _error('이 작업그룹에 개별파일 자동분류 기능이 선택되어 있지 않습니다.', 403, code='FEATURE_NOT_ENABLED')
        if mode in {'common', 'smart_and_common'} and not _group_feature_enabled(group, 'common_attachment'):
            return _error('이 작업그룹에 공통파일 일괄첨부 기능이 선택되어 있지 않습니다.', 403, code='FEATURE_NOT_ENABLED')
        sender = _owned_row(conn, 'ai_mail_senders', sender_id) if sender_id else None
        if not sender or not sender['is_active']:
            return _error('사용 가능한 발송자 계정을 선택해주세요.', 404, code='SENDER_NOT_FOUND')
        template = None
        if template_id:
            if not _group_feature_enabled(group, 'templates'):
                return _error('이 작업그룹에 메일 템플릿 기능이 선택되어 있지 않습니다.', 403, code='FEATURE_NOT_ENABLED')
            template = _owned_row(conn, 'ai_mail_templates', template_id)
            if not template or not template['is_active']:
                return _error('사용 가능한 메일 템플릿을 찾을 수 없습니다.', 404, code='TEMPLATE_NOT_FOUND')

        if recipient_ids:
            placeholders = ','.join('?' for _ in recipient_ids)
            recipients = conn.execute(f'''
                SELECT * FROM ai_mail_recipients
                WHERE group_id=? AND id IN ({placeholders})
                ORDER BY id
            ''', [group_id, *recipient_ids]).fetchall()
            if len(recipients) != len(recipient_ids):
                return _error('선택한 수신자 중 이 작업그룹에 속하지 않는 항목이 있습니다.', 403, code='RECIPIENT_OWNERSHIP_INVALID')
        else:
            recipients = conn.execute('''
                SELECT * FROM ai_mail_recipients WHERE group_id=? ORDER BY id
            ''', (group_id,)).fetchall()
        if not recipients:
            return _error('발송할 수신자를 선택해주세요.', code='RECIPIENTS_REQUIRED')
        if len(recipients) > MAX_RECIPIENTS:
            return _error('한 발송작업의 수신자는 최대 300명입니다.', code='RECIPIENT_LIMIT')

        subject = _clean_text(_request_value('subject', template['subject'] if template else ''), 300)
        raw_html = _request_value('body_html', template['body_html'] if template else '')
        if template:
            raw_html = _restore_template_asset_urls(conn, template_id, raw_html)
        try:
            body_html = _sanitize_html(raw_html)
        except ValueError as exc:
            return _error(str(exc), code='HTML_INVALID')
        if re.search(r'<img[^>]+src=["\']data:image/', body_html, re.I):
            return _error('본문 이미지는 먼저 메일 템플릿으로 저장한 뒤 선택해주세요.', code='INLINE_IMAGE_TEMPLATE_REQUIRED')
        if template:
            body_html = _validate_template_cids(conn, template_id, body_html)
        _enforce_final_html_size(body_html)
        body_text = _clean_text(_request_value('body_text', template['body_text'] if template else ''), 200000) or _plain_from_html(body_html)
        if not subject:
            return _error('메일 제목을 입력해주세요.', code='SUBJECT_REQUIRED')
        if not body_text and '<img' not in body_html.lower():
            return _error('메일 내용을 입력해주세요.', code='BODY_REQUIRED')
        name = _clean_text(_request_value('name'), 160) or f'{group["name"]} - {datetime.now().strftime("%Y-%m-%d %H:%M")}'

        conn.execute('BEGIN IMMEDIATE')
        cursor = conn.execute('''
            INSERT INTO ai_mail_campaigns (
                owner_emp_no, group_id, sender_id, template_id, group_name,
                sender_label, sender_email, name, subject, body_html, body_text,
                attachment_mode, status, total_count, allow_missing_attachment,
                send_interval
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'staged', ?, ?, ?)
        ''', (
            _owner_emp_no(), group_id, sender_id, template_id, group['name'],
            sender['label'], sender['email'], name, subject, body_html, body_text,
            mode, len(recipients), 1 if allow_missing_attachment else 0, send_interval
        ))
        campaign_id = cursor.lastrowid
        conn.executemany('''
            INSERT INTO ai_mail_campaign_recipients (
                campaign_id, source_recipient_id, email, recipient_name, memo
            ) VALUES (?, ?, ?, ?, ?)
        ''', [
            (campaign_id, row['id'], row['email'], row['recipient_name'], row['memo'])
            for row in recipients
        ])
        common_files, smart_files = _attachment_files_from_request()
        if len(common_files) + len(smart_files) > MAX_ATTACHMENT_FILES:
            raise ValueError(f'파일은 최대 {MAX_ATTACHMENT_FILES}개까지 등록할 수 있습니다.')
        for file_storage in common_files:
            attachment = _save_campaign_attachment(conn, campaign_id, file_storage, 'common')
            saved_paths.append(attachment['filepath'])
        for file_storage in smart_files:
            attachment = _save_campaign_attachment(conn, campaign_id, file_storage, 'smart')
            saved_paths.append(attachment['filepath'])
        _campaign_event(conn, campaign_id, 'staged', '메일 발송작업을 임시 저장했습니다.', details={'recipient_count': len(recipients)})
        conn.commit()
        campaign = _owned_campaign(conn, campaign_id)
        attachments = conn.execute('SELECT * FROM ai_mail_campaign_attachments WHERE campaign_id=? ORDER BY id', (campaign_id,)).fetchall()
        return _success(
            '메일 발송작업을 임시 저장했습니다. 발송 전 점검을 실행해주세요.',
            campaign=_campaign_dict(campaign),
            attachments=[_attachment_dict(row) for row in attachments]
        )
    except (ValueError, OSError, sqlite3.IntegrityError) as exc:
        conn.rollback()
        for path in saved_paths:
            _delete_saved_file(path)
        return _error(f'발송작업을 저장하지 못했습니다: {_clean_text(exc, 500)}', code='CAMPAIGN_CREATE_FAILED')
    finally:
        conn.close()


@ai_mail_bp.route('/api/campaigns/<int:campaign_id>/attachments', methods=['POST'])
@_mutating
def add_campaign_attachments(campaign_id):
    common_files, smart_files = _attachment_files_from_request()
    if not common_files and not smart_files:
        return _error('첨부할 파일을 선택해주세요.', code='ATTACHMENT_REQUIRED')
    conn = _db()
    saved_paths = []
    try:
        campaign = _owned_campaign(conn, campaign_id)
        if not campaign:
            return _error('발송작업을 찾을 수문을 수 없습니다.', 404, code='CAMPAIGN_NOT_FOUND')
        if campaign['status'] != 'staged':
            return _error('발송 전 임시 저장 상태에서만 첨부파일을 추가할 수 있습니다.', 409, code='CAMPAIGN_LOCKED')
        conn.execute('BEGIN IMMEDIATE')
        added = []
        for file_storage in common_files:
            row = _save_campaign_attachment(conn, campaign_id, file_storage, 'common')
            saved_paths.append(row['filepath'])
            added.append(_attachment_dict(row))
        for file_storage in smart_files:
            row = _save_campaign_attachment(conn, campaign_id, file_storage, 'smart')
            saved_paths.append(row['filepath'])
            added.append(_attachment_dict(row))
        conn.execute('''
            UPDATE ai_mail_campaigns
            SET preflight_ok=0, preflight_json='{}', updated_at=CURRENT_TIMESTAMP
            WHERE id=? AND owner_emp_no=?
        ''', (campaign_id, _owner_emp_no()))
        _campaign_event(conn, campaign_id, 'attachments_added', f'첨부파일 {len(added)}개를 추가했습니다.')
        conn.commit()
        return _success(f'첨부파일 {len(added)}개를 추가했습니다.', attachments=added)
    except (ValueError, OSError, sqlite3.IntegrityError) as exc:
        conn.rollback()
        for path in saved_paths:
            _delete_saved_file(path)
        return _error(str(exc), code='ATTACHMENT_SAVE_FAILED')
    finally:
        conn.close()


def _canonical_school_name(value):
    text = re.sub(
        r'[^0-9a-zA-Zㄱ-ㆎ가-힣]+',
        '',
        unicodedata.normalize('NFKC', _clean_text(value)).lower()
    )
    replacements = (
        ('초등학교', '초'), ('초교', '초'),
        ('중학교', '중'), ('중교', '중'),
        ('고등학교', '고'), ('고교', '고')
    )
    for suffix, replacement in replacements:
        if text.endswith(suffix):
            return text[:-len(suffix)] + replacement
    return text


def _filename_match_key(filename):
    stem = os.path.splitext(_safe_original_filename(filename))[0]
    normalized = unicodedata.normalize('NFKC', stem).strip()
    # 요구된 구분자(밑줄, 하이픈, 별표/전각 별표, 공백) 앞을 수신자명으로 본다.
    prefix = re.split(r'[\s_\-*\uff0a]+', normalized, maxsplit=1)[0]
    return _canonical_school_name(prefix)


def _preflight_campaign(conn, campaign):
    campaign_id = campaign['id']
    recipients = conn.execute('''
        SELECT * FROM ai_mail_campaign_recipients
        WHERE campaign_id=? ORDER BY id
    ''', (campaign_id,)).fetchall()
    attachments = conn.execute('''
        SELECT * FROM ai_mail_campaign_attachments
        WHERE campaign_id=? ORDER BY id
    ''', (campaign_id,)).fetchall()
    mode = campaign['attachment_mode']
    errors = []
    warnings = []
    unmatched_files = []
    ambiguous_files = []
    missing_recipients = []
    invalid_recipients = []
    oversized_recipients = []
    missing_files = []

    recipient_keys = {}
    for recipient in recipients:
        key = _canonical_school_name(recipient['recipient_name'])
        if key:
            recipient_keys.setdefault(key, []).append(recipient)
        if not _is_valid_email(recipient['email']):
            invalid_recipients.append({
                'id': recipient['id'], 'name': recipient['recipient_name'],
                'email': recipient['email'], 'code': 'EMAIL_INVALID'
            })

    for attachment in attachments:
        if attachment['kind'] == 'common':
            conn.execute('''
                UPDATE ai_mail_campaign_attachments
                SET campaign_recipient_id=NULL, match_status='matched',
                    match_method='auto', diagnostic=NULL
                WHERE id=? AND campaign_id=?
            ''', (attachment['id'], campaign_id))
            continue
        if attachment['match_method'] == 'manual' and attachment['campaign_recipient_id']:
            valid_manual = conn.execute('''
                SELECT id FROM ai_mail_campaign_recipients
                WHERE id=? AND campaign_id=?
            ''', (attachment['campaign_recipient_id'], campaign_id)).fetchone()
            if valid_manual:
                conn.execute('''
                    UPDATE ai_mail_campaign_attachments
                    SET match_status='matched', diagnostic=NULL
                    WHERE id=? AND campaign_id=?
                ''', (attachment['id'], campaign_id))
                continue

        key = _filename_match_key(attachment['original_name'])
        candidates = recipient_keys.get(key, []) if key else []
        if len(candidates) == 1:
            conn.execute('''
                UPDATE ai_mail_campaign_attachments
                SET campaign_recipient_id=?, match_method='auto',
                    match_status='matched', diagnostic=NULL
                WHERE id=? AND campaign_id=?
            ''', (candidates[0]['id'], attachment['id'], campaign_id))
        elif len(candidates) > 1:
            names = [f"{row['recipient_name']} <{row['email']}>" for row in candidates]
            diagnostic = '같은 수신자명이 여러 명입니다.'
            conn.execute('''
                UPDATE ai_mail_campaign_attachments
                SET campaign_recipient_id=NULL, match_method='auto',
                    match_status='ambiguous', diagnostic=?
                WHERE id=? AND campaign_id=?
            ''', (diagnostic, attachment['id'], campaign_id))
            ambiguous_files.append({
                'id': attachment['id'], 'filename': attachment['original_name'],
                'match_key': key, 'candidates': names, 'message': diagnostic
            })
        else:
            diagnostic = '파일명 앞부분과 일치하는 수신자명이 없습니다.'
            conn.execute('''
                UPDATE ai_mail_campaign_attachments
                SET campaign_recipient_id=NULL, match_method='auto',
                    match_status='unmatched', diagnostic=?
                WHERE id=? AND campaign_id=?
            ''', (diagnostic, attachment['id'], campaign_id))
            unmatched_files.append({
                'id': attachment['id'], 'filename': attachment['original_name'],
                'match_key': key, 'message': diagnostic
            })

    attachments = conn.execute('''
        SELECT * FROM ai_mail_campaign_attachments
        WHERE campaign_id=? ORDER BY id
    ''', (campaign_id,)).fetchall()
    common = [row for row in attachments if row['kind'] == 'common']
    common_bytes = 0
    for attachment in attachments:
        try:
            path = _safe_saved_path(attachment['filepath'])
            exists = os.path.isfile(path)
        except ValueError:
            exists = False
        if not exists:
            missing_files.append({
                'id': attachment['id'], 'filename': attachment['original_name'],
                'code': 'ATTACHMENT_FILE_MISSING'
            })
        elif attachment['kind'] == 'common':
            common_bytes += int(attachment['size_bytes'] or 0)

    inline_assets = []
    inline_bytes = 0
    cid_references = set(re.findall(r'cid:([a-zA-Z0-9_.@-]+)', campaign['body_html'] or ''))
    if cid_references:
        if not campaign['template_id']:
            errors.append({'code': 'INLINE_IMAGE_TEMPLATE_MISSING', 'message': '템플릿 연결 없이 본문 이미지가 포함되었습니다.'})
        else:
            asset_rows = conn.execute('''
                SELECT * FROM ai_mail_template_assets WHERE template_id=?
            ''', (campaign['template_id'],)).fetchall()
            asset_map = {row['content_id']: row for row in asset_rows}
            for cid in cid_references:
                asset = asset_map.get(cid)
                if not asset:
                    errors.append({'code': 'INLINE_IMAGE_MISSING', 'message': f'본문 이미지를 찾을 수 없습니다: {cid}'})
                    continue
                try:
                    asset_path = _safe_saved_path(asset['filepath'])
                except ValueError:
                    asset_path = ''
                if not asset_path or not os.path.isfile(asset_path):
                    errors.append({'code': 'INLINE_IMAGE_FILE_MISSING', 'message': f'본문 이미지 파일이 손실되었습니다: {asset["original_name"]}'})
                    continue
                inline_assets.append(asset)
                inline_bytes += int(asset['size_bytes'] or 0)

    smart_required = mode in {'smart', 'smart_and_common'}
    for recipient in recipients:
        smart_rows = [
            row for row in attachments
            if row['kind'] == 'smart'
            and row['match_status'] == 'matched'
            and row['campaign_recipient_id'] == recipient['id']
        ]
        if smart_required and not smart_rows:
            missing_recipients.append({
                'id': recipient['id'], 'name': recipient['recipient_name'],
                'email': recipient['email'], 'code': 'SMART_ATTACHMENT_MISSING'
            })
        relevant_common = common if mode in {'common', 'smart_and_common'} else []
        relevant_smart = smart_rows if smart_required else []
        raw_bytes = sum(int(row['size_bytes'] or 0) for row in relevant_common + relevant_smart)
        encoded_estimate = (
            int((raw_bytes + inline_bytes) * 4 / 3)
            + len((campaign['body_html'] or '').encode('utf-8'))
            + (len(relevant_common) + len(relevant_smart) + len(inline_assets) + 1) * 8192
        )
        conn.execute('''
            UPDATE ai_mail_campaign_recipients
            SET attachment_count=?, attachment_bytes=?, updated_at=CURRENT_TIMESTAMP
            WHERE id=? AND campaign_id=?
        ''', (len(relevant_common) + len(relevant_smart), raw_bytes, recipient['id'], campaign_id))
        if encoded_estimate > MAX_MESSAGE_BYTES:
            oversized_recipients.append({
                'id': recipient['id'], 'name': recipient['recipient_name'],
                'email': recipient['email'], 'estimated_bytes': encoded_estimate,
                'limit_bytes': MAX_MESSAGE_BYTES, 'code': 'MESSAGE_TOO_LARGE'
            })

    if mode in {'common', 'smart_and_common'} and not common:
        errors.append({'code': 'COMMON_ATTACHMENT_REQUIRED', 'message': '공통 첨부파일이 없습니다.'})
    if invalid_recipients:
        errors.append({'code': 'INVALID_RECIPIENTS', 'message': f'메일주소 오류 {len(invalid_recipients)}건을 수정해주세요.'})
    if unmatched_files:
        errors.append({'code': 'UNMATCHED_FILES', 'message': f'수신자와 일치하지 않는 파일이 {len(unmatched_files)}개 있습니다.'})
    if ambiguous_files:
        errors.append({'code': 'AMBIGUOUS_FILES', 'message': f'수신자를 하나로 결정할 수 없는 파일이 {len(ambiguous_files)}개 있습니다.'})
    if missing_recipients:
        message = f'개별 첨부파일이 없는 수신자가 {len(missing_recipients)}명입니다.'
        item = {'code': 'MISSING_SMART_ATTACHMENTS', 'message': message}
        if campaign['allow_missing_attachment']:
            warnings.append(item)
        else:
            errors.append(item)
    if missing_files:
        errors.append({'code': 'ATTACHMENT_FILES_MISSING', 'message': f'저장소에서 손실된 첨부파일이 {len(missing_files)}개 있습니다.'})
    if oversized_recipients:
        errors.append({'code': 'MESSAGE_TOO_LARGE', 'message': f'18MB 제한을 초과하는 메일이 {len(oversized_recipients)}건 있습니다.'})
    sender = conn.execute('''
        SELECT * FROM ai_mail_senders WHERE id=? AND owner_emp_no=?
    ''', (campaign['sender_id'], campaign['owner_emp_no'])).fetchone()
    if not sender or not sender['is_active']:
        errors.append({'code': 'SENDER_INACTIVE', 'message': '발송자 계정이 비활성화되었거나 삭제되었습니다.'})
    elif sender['email'].lower() != (campaign['sender_email'] or '').lower():
        errors.append({'code': 'SENDER_CHANGED', 'message': '발송자 메일주소가 작업 저장 후 변경되었습니다. 새 발송작업을 만들어주세요.'})

    invalid_ids = {item['id'] for item in invalid_recipients}
    missing_ids = {item['id'] for item in missing_recipients}
    oversized_ids = {item['id'] for item in oversized_recipients}
    matches = []
    for recipient in recipients:
        assigned = []
        if mode in {'common', 'smart_and_common'}:
            assigned.extend(common)
        if smart_required:
            assigned.extend([
                row for row in attachments
                if row['kind'] == 'smart'
                and row['match_status'] == 'matched'
                and row['campaign_recipient_id'] == recipient['id']
            ])
        if recipient['id'] in invalid_ids:
            match_status, match_message = 'error', '메일주소 형식을 수정해야 합니다.'
        elif recipient['id'] in oversized_ids:
            match_status, match_message = 'error', '18MB 메일 크기 제한을 초과합니다.'
        elif recipient['id'] in missing_ids:
            match_status = 'warning' if campaign['allow_missing_attachment'] else 'error'
            match_message = '개별 첨부파일이 없습니다.'
        else:
            match_status, match_message = 'ready', '발송 준비 완료'
        matches.append({
            'id': recipient['id'],
            'source_recipient_id': recipient['source_recipient_id'],
            'recipient_name': recipient['recipient_name'],
            'name': recipient['recipient_name'],
            'email': recipient['email'],
            'status': match_status,
            'message': match_message,
            'attachments': [
                {
                    'id': row['id'], 'original_name': row['original_name'],
                    'size_bytes': row['size_bytes'], 'kind': row['kind'],
                    'download_url': f"/ai-mail/api/campaigns/{campaign_id}/attachments/{row['id']}/download"
                }
                for row in assigned
            ]
        })

    report = {
        'ok': not errors,
        'errors': errors,
        'warnings': warnings,
        'unmatched_files': unmatched_files,
        'ambiguous_files': ambiguous_files,
        'missing_recipients': missing_recipients,
        'invalid_recipients': invalid_recipients,
        'oversized_recipients': oversized_recipients,
        'missing_files': missing_files,
        'matches': matches,
        'counts': {
            'recipients': len(recipients),
            'attachments': len(attachments),
            'common_files': len(common),
            'matched_smart_files': sum(1 for row in attachments if row['kind'] == 'smart' and row['match_status'] == 'matched'),
            'unmatched_files': len(unmatched_files),
            'ambiguous_files': len(ambiguous_files),
            'missing_recipients': len(missing_recipients)
        }
    }
    conn.execute('''
        UPDATE ai_mail_campaigns
        SET preflight_ok=?, preflight_json=?, updated_at=CURRENT_TIMESTAMP
        WHERE id=?
    ''', (1 if report['ok'] else 0, json.dumps(report, ensure_ascii=False), campaign_id))
    _campaign_event(
        conn, campaign_id, 'preflight',
        '발송 전 점검을 통과했습니다.' if report['ok'] else f'발송 전 점검에서 오류 {len(errors)}건을 찾았습니다.',
        'info' if report['ok'] else 'warning',
        {'error_count': len(errors), 'warning_count': len(warnings)}
    )
    return report


@ai_mail_bp.route('/api/campaigns/<int:campaign_id>/preflight', methods=['POST'])
@_mutating
def preflight_campaign(campaign_id):
    conn = _db()
    try:
        campaign = _owned_campaign(conn, campaign_id)
        if not campaign:
            return _error('발송작업을 찾을 수 없습니다.', 404, code='CAMPAIGN_NOT_FOUND')
        if campaign['status'] != 'staged':
            return _error('임시 저장 상태에서만 발송 전 점검을 실행할 수 있습니다.', 409, code='CAMPAIGN_LOCKED')
        conn.execute('BEGIN IMMEDIATE')
        report = _preflight_campaign(conn, campaign)
        conn.commit()
        refreshed = _owned_campaign(conn, campaign_id)
        message = '발송 준비가 완료되었습니다.' if report['ok'] else '발송 전 점검 오류를 확인해주세요.'
        if report['ok']:
            return _success(message, preflight=report, campaign=_campaign_dict(refreshed))
        return _error(message, 400, code='PREFLIGHT_FAILED', preflight=report, campaign=_campaign_dict(refreshed))
    except (ValueError, OSError, sqlite3.DatabaseError) as exc:
        conn.rollback()
        return _error(f'발송 전 점검을 실행하지 못했습니다: {_clean_text(exc, 500)}', code='PREFLIGHT_ERROR')
    finally:
        conn.close()


@ai_mail_bp.route('/api/campaigns', methods=['GET'])
@_login_required
def list_campaigns():
    page = max(_as_int(request.args.get('page'), 1), 1)
    per_page = min(max(_as_int(request.args.get('per_page'), 30), 1), 100)
    status_filter = _clean_text(request.args.get('status'), 40)
    params = [_owner_emp_no()]
    where = 'owner_emp_no=?'
    if status_filter:
        where += ' AND status=?'
        params.append(status_filter)
    elif not _as_bool(request.args.get('include_staged')):
        where += " AND status!='staged'"
    conn = _db()
    try:
        total = conn.execute(f'SELECT COUNT(*) FROM ai_mail_campaigns WHERE {where}', params).fetchone()[0]
        rows = conn.execute(f'''
            SELECT id, owner_emp_no, group_id, sender_id, template_id, group_name,
                   sender_label, sender_email, name, subject, attachment_mode, status,
                   total_count, processed_count, sent_count, failed_count,
                   cancelled_count, cancel_requested, cancel_requested_at,
                   cancel_requested_by, cancel_reason, allow_missing_attachment,
                   send_interval, preflight_ok, preflight_json, error_code,
                   error_message, created_at, queued_at, started_at, finished_at,
                   updated_at
            FROM ai_mail_campaigns WHERE {where}
            ORDER BY created_at DESC, id DESC LIMIT ? OFFSET ?
        ''', [*params, per_page, (page - 1) * per_page]).fetchall()
        return _success(
            campaigns=[_campaign_dict(row) for row in rows],
            pagination={
                'page': page, 'per_page': per_page, 'total': total,
                'total_pages': max(1, (total + per_page - 1) // per_page)
            }
        )
    finally:
        conn.close()


@ai_mail_bp.route('/api/campaigns/<int:campaign_id>', methods=['GET'])
@_login_required
def get_campaign(campaign_id):
    conn = _db()
    try:
        campaign = _owned_campaign(conn, campaign_id)
        if not campaign:
            return _error('발송작업을 찾을 수 없습니다.', 404, code='CAMPAIGN_NOT_FOUND')
        recipients = conn.execute('''
            SELECT * FROM ai_mail_campaign_recipients
            WHERE campaign_id=? ORDER BY id
        ''', (campaign_id,)).fetchall()
        attachments = conn.execute('''
            SELECT * FROM ai_mail_campaign_attachments
            WHERE campaign_id=? ORDER BY kind, original_name, id
        ''', (campaign_id,)).fetchall()
        events = conn.execute('''
            SELECT id, campaign_id, event_type, level, message, details_json, created_at
            FROM ai_mail_campaign_events WHERE campaign_id=?
            ORDER BY id DESC LIMIT 100
        ''', (campaign_id,)).fetchall()
        event_items = []
        for event in events:
            item = dict(event)
            item['details'] = _parse_json_value(item.pop('details_json', '{}'), {})
            event_items.append(item)
        return _success(
            campaign=_campaign_dict(campaign),
            recipients=[dict(row) for row in recipients],
            attachments=[_attachment_dict(row) for row in attachments],
            events=event_items
        )
    finally:
        conn.close()


@ai_mail_bp.route('/api/campaigns/<int:campaign_id>/logs', methods=['GET'])
@_login_required
def campaign_logs(campaign_id):
    page = max(_as_int(request.args.get('page'), 1), 1)
    per_page = min(max(_as_int(request.args.get('per_page'), 50), 1), 100)
    status_filter = _clean_text(request.args.get('status'), 30)
    conn = _db()
    try:
        if not _owned_campaign(conn, campaign_id):
            return _error('발송작업을 찾을 수 없습니다.', 404, code='CAMPAIGN_NOT_FOUND')
        where = 'campaign_id=?'
        params = [campaign_id]
        if status_filter:
            where += ' AND status=?'
            params.append(status_filter)
        total = conn.execute(f'SELECT COUNT(*) FROM ai_mail_campaign_recipients WHERE {where}', params).fetchone()[0]
        rows = conn.execute(f'''
            SELECT * FROM ai_mail_campaign_recipients WHERE {where}
            ORDER BY id LIMIT ? OFFSET ?
        ''', [*params, per_page, (page - 1) * per_page]).fetchall()
        return _success(
            logs=[dict(row) for row in rows],
            pagination={
                'page': page, 'per_page': per_page, 'total': total,
                'total_pages': max(1, (total + per_page - 1) // per_page)
            }
        )
    finally:
        conn.close()


@ai_mail_bp.route('/api/campaigns/<int:campaign_id>/status', methods=['GET'])
@_login_required
def campaign_status(campaign_id):
    conn = _db()
    try:
        row = conn.execute('''
            SELECT id, owner_emp_no, status, total_count, processed_count,
                   sent_count, failed_count, cancelled_count, cancel_requested,
                   cancel_requested_at, cancel_requested_by, cancel_reason,
                   preflight_ok, preflight_json, error_code, error_message,
                   queued_at, started_at, finished_at, updated_at
            FROM ai_mail_campaigns WHERE id=? AND owner_emp_no=?
        ''', (campaign_id, _owner_emp_no())).fetchone()
        if not row:
            return _error('발송작업을 찾을 수 없습니다.', 404, code='CAMPAIGN_NOT_FOUND')
        campaign = _campaign_dict(row)
        total = int(row['total_count'] or 0)
        processed = int(row['processed_count'] or 0)
        progress = round(processed * 100 / total, 1) if total else 0
        return _success(campaign=campaign, progress_percent=progress)
    finally:
        conn.close()


@ai_mail_bp.route('/api/campaigns/<int:campaign_id>/attachments/<int:attachment_id>', methods=['PUT', 'PATCH'])
@_mutating
def update_campaign_attachment(campaign_id, attachment_id):
    data = _json_data() if request.is_json else request.form
    kind = _clean_text(data.get('kind', data.get('mode'))).lower()
    target_id = _as_int(data.get('campaign_recipient_id', data.get('recipient_id')))
    if kind not in {'common', 'smart'}:
        return _error('첨부파일 방식은 common 또는 smart여야 합니다.', code='ATTACHMENT_KIND_INVALID')
    conn = _db()
    try:
        campaign = _owned_campaign(conn, campaign_id)
        if not campaign:
            return _error('발송작업을 찾을 수 없습니다.', 404, code='CAMPAIGN_NOT_FOUND')
        if campaign['status'] != 'staged':
            return _error('임시 저장 상태에서만 첨부파일 분류를 변경할 수 있습니다.', 409, code='CAMPAIGN_LOCKED')
        attachment = conn.execute('''
            SELECT * FROM ai_mail_campaign_attachments WHERE id=? AND campaign_id=?
        ''', (attachment_id, campaign_id)).fetchone()
        if not attachment:
            return _error('첨부파일을 찾을 수 없습니다.', 404, code='ATTACHMENT_NOT_FOUND')
        campaign_recipient_id = None
        method, status, diagnostic = 'auto', 'matched' if kind == 'common' else 'pending', None
        if kind == 'smart' and target_id:
            recipient = conn.execute('''
                SELECT * FROM ai_mail_campaign_recipients
                WHERE campaign_id=? AND (id=? OR source_recipient_id=?)
                ORDER BY CASE WHEN id=? THEN 0 ELSE 1 END LIMIT 1
            ''', (campaign_id, target_id, target_id, target_id)).fetchone()
            if not recipient:
                return _error('선택한 수신자가 이 발송작업에 속하지 않습니다.', 403, code='RECIPIENT_OWNERSHIP_INVALID')
            campaign_recipient_id = recipient['id']
            method, status = 'manual', 'matched'
        conn.execute('''
            UPDATE ai_mail_campaign_attachments
            SET kind=?, campaign_recipient_id=?, match_method=?, match_status=?, diagnostic=?
            WHERE id=? AND campaign_id=?
        ''', (kind, campaign_recipient_id, method, status, diagnostic, attachment_id, campaign_id))
        conn.execute('''
            UPDATE ai_mail_campaigns SET preflight_ok=0, preflight_json='{}',
                updated_at=CURRENT_TIMESTAMP WHERE id=? AND owner_emp_no=?
        ''', (campaign_id, _owner_emp_no()))
        conn.commit()
        updated = conn.execute('SELECT * FROM ai_mail_campaign_attachments WHERE id=?', (attachment_id,)).fetchone()
        return _success('첨부파일 분류를 변경했습니다.', attachment=_attachment_dict(updated))
    finally:
        conn.close()


@ai_mail_bp.route('/api/campaigns/<int:campaign_id>/attachments/<int:attachment_id>', methods=['DELETE'])
@_mutating
def delete_campaign_attachment(campaign_id, attachment_id):
    conn = _db()
    try:
        campaign = _owned_campaign(conn, campaign_id)
        if not campaign:
            return _error('발송작업을 찾을 수 없습니다.', 404, code='CAMPAIGN_NOT_FOUND')
        if campaign['status'] != 'staged':
            return _error('임시 저장 상태에서만 첨부파일을 삭제할 수 있습니다.', 409, code='CAMPAIGN_LOCKED')
        attachment = conn.execute('SELECT * FROM ai_mail_campaign_attachments WHERE id=? AND campaign_id=?', (attachment_id, campaign_id)).fetchone()
        if not attachment:
            return _error('첨부파일을 찾을 수 없습니다.', 404, code='ATTACHMENT_NOT_FOUND')
        conn.execute('DELETE FROM ai_mail_campaign_attachments WHERE id=? AND campaign_id=?', (attachment_id, campaign_id))
        conn.execute('''
            UPDATE ai_mail_campaigns SET preflight_ok=0, preflight_json='{}',
                updated_at=CURRENT_TIMESTAMP WHERE id=? AND owner_emp_no=?
        ''', (campaign_id, _owner_emp_no()))
        conn.commit()
        _delete_saved_file(attachment['filepath'])
        return _success('첨부파일을 삭제했습니다.')
    finally:
        conn.close()


@ai_mail_bp.route('/api/campaigns/<int:campaign_id>/attachments/<int:attachment_id>/download')
@_login_required
def download_campaign_attachment(campaign_id, attachment_id):
    conn = _db()
    try:
        if not _owned_campaign(conn, campaign_id):
            return _error('발송작업을 찾을 수 없습니다.', 404, code='CAMPAIGN_NOT_FOUND')
        attachment = conn.execute('SELECT * FROM ai_mail_campaign_attachments WHERE id=? AND campaign_id=?', (attachment_id, campaign_id)).fetchone()
        if not attachment:
            return _error('첨부파일을 찾을 수 없습니다.', 404, code='ATTACHMENT_NOT_FOUND')
        path = _safe_saved_path(attachment['filepath'])
        if not os.path.isfile(path):
            return _error('첨부파일이 저장소에서 손실되었습니다.', 404, code='ATTACHMENT_FILE_MISSING')
        return send_file(path, as_attachment=True, download_name=attachment['original_name'], mimetype=attachment['mime_type'])
    finally:
        conn.close()


@ai_mail_bp.route('/api/campaigns/<int:campaign_id>', methods=['DELETE'])
@_mutating
def discard_staged_campaign(campaign_id):
    conn = _db()
    try:
        campaign = _owned_campaign(conn, campaign_id)
        if not campaign:
            return _error('발송작업을 찾을 수 없습니다.', 404, code='CAMPAIGN_NOT_FOUND')
        if campaign['status'] != 'staged':
            return _error('발송 이력은 삭제할 수 없습니다. 아직 발송하지 않은 임시 저장 작업만 폐기할 수 있습니다.', 409, code='CAMPAIGN_HISTORY_PROTECTED')
        files = conn.execute('''
            SELECT filepath FROM ai_mail_campaign_attachments WHERE campaign_id=?
        ''', (campaign_id,)).fetchall()
        conn.execute('BEGIN IMMEDIATE')
        conn.execute('DELETE FROM ai_mail_campaign_events WHERE campaign_id=?', (campaign_id,))
        conn.execute('DELETE FROM ai_mail_campaign_attachments WHERE campaign_id=?', (campaign_id,))
        conn.execute('DELETE FROM ai_mail_campaign_recipients WHERE campaign_id=?', (campaign_id,))
        conn.execute('DELETE FROM ai_mail_campaigns WHERE id=? AND owner_emp_no=? AND status=\'staged\'', (campaign_id, _owner_emp_no()))
        conn.commit()
        for row in files:
            _delete_saved_file(row['filepath'])
        return _success('임시 저장한 발송작업을 폐기했습니다.')
    except sqlite3.DatabaseError as exc:
        conn.rollback()
        return _error(f'발송작업을 폐기하지 못했습니다: {_clean_text(exc, 300)}', code='CAMPAIGN_DISCARD_FAILED')
    finally:
        conn.close()


def _send_campaign_worker(app, campaign_id, cancel_event=None):
    cancel_event = cancel_event or threading.Event()
    with app.app_context():
        conn = _db()
        try:
            # 1. Update status to 'running'
            conn.execute('''
                UPDATE ai_mail_campaigns 
                SET status=CASE WHEN cancel_requested=1 THEN 'cancel_requested' ELSE 'running' END,
                    started_at=CURRENT_TIMESTAMP, updated_at=CURRENT_TIMESTAMP
                WHERE id=?
            ''', (campaign_id,))
            conn.commit()

            campaign = conn.execute('SELECT * FROM ai_mail_campaigns WHERE id=?', (campaign_id,)).fetchone()
            sender = conn.execute('SELECT * FROM ai_mail_senders WHERE id=?', (campaign['sender_id'],)).fetchone()
            recipients = conn.execute('SELECT * FROM ai_mail_campaign_recipients WHERE campaign_id=? ORDER BY id', (campaign_id,)).fetchall()
            attachments = conn.execute('SELECT * FROM ai_mail_campaign_attachments WHERE campaign_id=?', (campaign_id,)).fetchall()
            
            inline_assets = []
            if campaign['template_id']:
                inline_assets = conn.execute('SELECT * FROM ai_mail_template_assets WHERE template_id=?', (campaign['template_id'],)).fetchall()

            password = _decrypt_password(sender['encrypted_app_password'])
            
            processed_count = 0
            sent_count = 0
            failed_count = 0
            cancelled_count = 0

            server = None
            def _get_server():
                nonlocal server
                if server:
                    try:
                        server.noop()
                    except Exception:
                        server = None
                if not server:
                    server = _smtp_login(sender['email'], password)
                    with _worker_lock:
                        _worker_servers[campaign_id] = server
                return server

            for recipient in recipients:
                if cancel_event.is_set():
                    break
                check_cancel = conn.execute('SELECT cancel_requested FROM ai_mail_campaigns WHERE id=?', (campaign_id,)).fetchone()
                if check_cancel and check_cancel['cancel_requested']:
                    cancel_event.set()
                    break

                try:
                    r_name = recipient['recipient_name'] or ''
                    r_email = recipient['email'] or ''
                    r_memo = recipient['memo'] or ''
                    g_name = campaign['group_name'] or ''
                    
                    subject = campaign['subject'].replace('{{수신자명}}', r_name).replace('{{메일주소}}', r_email).replace('{{메모}}', r_memo).replace('{{작업그룹명}}', g_name)
                    body_text = (campaign['body_text'] or '').replace('{{수신자명}}', r_name).replace('{{메일주소}}', r_email).replace('{{메모}}', r_memo).replace('{{작업그룹명}}', g_name)
                    body_html = (campaign['body_html'] or '').replace('{{수신자명}}', r_name).replace('{{메일주소}}', r_email).replace('{{메모}}', r_memo).replace('{{작업그룹명}}', g_name)

                    msg = EmailMessage()
                    msg['Subject'] = subject
                    msg['From'] = f"{sender['label']} <{sender['email']}>" if sender['label'] else sender['email']
                    msg['To'] = f"{r_name} <{r_email}>"

                    msg.set_content(body_text)
                    if body_html:
                        msg.add_alternative(body_html, subtype='html')
                        for asset in inline_assets:
                            path = _safe_saved_path(asset['filepath'])
                            if os.path.isfile(path):
                                with open(path, 'rb') as f:
                                    img_data = f.read()
                                subtype = asset['mime_type'].split('/')[1] if '/' in asset['mime_type'] else 'jpeg'
                                msg.get_payload()[1].add_related(img_data, maintype='image', subtype=subtype, cid=f"<{asset['content_id']}>")

                    for att in attachments:
                        is_common = att['kind'] == 'common'
                        is_smart_matched = att['kind'] == 'smart' and att['campaign_recipient_id'] == recipient['id']
                        
                        if is_common or is_smart_matched:
                            path = _safe_saved_path(att['filepath'])
                            if os.path.isfile(path):
                                with open(path, 'rb') as f:
                                    att_data = f.read()
                                maintype, subtype = att['mime_type'].split('/', 1) if '/' in att['mime_type'] else ('application', 'octet-stream')
                                msg.add_attachment(att_data, maintype=maintype, subtype=subtype, filename=att['original_name'])

                    if cancel_event.is_set():
                        raise InterruptedError('사용자 요청으로 발송이 중단되었습니다.')
                    srv = _get_server()
                    if cancel_event.is_set():
                        try:
                            srv.close()
                        except Exception:
                            pass
                        raise InterruptedError('사용자 요청으로 발송이 중단되었습니다.')
                    srv.send_message(msg)

                    conn.execute("UPDATE ai_mail_campaign_recipients SET status='sent', sent_at=CURRENT_TIMESTAMP, updated_at=CURRENT_TIMESTAMP WHERE id=?", (recipient['id'],))
                    sent_count += 1

                except Exception as exc:
                    error_msg = _clean_text(str(exc), 500)
                    if cancel_event.is_set():
                        conn.execute("UPDATE ai_mail_campaign_recipients SET status='cancelled', error_message=?, updated_at=CURRENT_TIMESTAMP WHERE id=?", ('사용자 요청으로 발송이 중단되었습니다.', recipient['id']))
                        cancelled_count += 1
                    else:
                        conn.execute("UPDATE ai_mail_campaign_recipients SET status='failed', error_message=?, updated_at=CURRENT_TIMESTAMP WHERE id=?", (error_msg, recipient['id']))
                        failed_count += 1

                finally:
                    processed_count += 1
                    conn.execute('''
                        UPDATE ai_mail_campaigns 
                        SET processed_count=?, sent_count=?, failed_count=?, cancelled_count=?, updated_at=CURRENT_TIMESTAMP
                        WHERE id=?
                    ''', (processed_count, sent_count, failed_count, cancelled_count, campaign_id))
                    conn.commit()
                    if cancel_event.wait(float(campaign['send_interval'] or 1.0)):
                        break

            if server:
                try:
                    server.quit()
                except Exception:
                    pass

            final_check = conn.execute('SELECT cancel_requested FROM ai_mail_campaigns WHERE id=?', (campaign_id,)).fetchone()
            if cancel_event.is_set() or (final_check and final_check['cancel_requested']):
                cancel_event.set()
                cancelled_cursor = conn.execute('''
                    UPDATE ai_mail_campaign_recipients
                    SET status='cancelled', error_message=?, updated_at=CURRENT_TIMESTAMP
                    WHERE campaign_id=? AND status='pending'
                ''', ('사용자 요청으로 발송이 중단되었습니다.', campaign_id))
                cancelled_count += max(0, cancelled_cursor.rowcount)
                processed_count = sent_count + failed_count + cancelled_count
                final_status = 'cancelled'
                conn.execute('''
                    UPDATE ai_mail_campaigns
                    SET processed_count=?, sent_count=?, failed_count=?, cancelled_count=?
                    WHERE id=?
                ''', (processed_count, sent_count, failed_count, cancelled_count, campaign_id))
                _campaign_event(conn, campaign_id, 'cancelled', '사용자 요청으로 발송이 중단되었습니다.')
            elif failed_count > 0:
                final_status = 'completed_with_errors'
                _campaign_event(conn, campaign_id, 'completed_with_errors', f'발송을 완료했으나 {failed_count}건 실패했습니다.')
            else:
                final_status = 'completed'
                _campaign_event(conn, campaign_id, 'completed', '모든 메일 발송을 성공적으로 완료했습니다.')

            conn.execute('''
                UPDATE ai_mail_campaigns 
                SET status=?, finished_at=CURRENT_TIMESTAMP, updated_at=CURRENT_TIMESTAMP 
                WHERE id=?
            ''', (final_status, campaign_id))
            conn.commit()

        except Exception as e:
            error_msg = _clean_text(str(e), 500)
            conn.execute('''
                UPDATE ai_mail_campaigns 
                SET status='failed', error_message=?, finished_at=CURRENT_TIMESTAMP, updated_at=CURRENT_TIMESTAMP 
                WHERE id=?
            ''', (error_msg, campaign_id))
            _campaign_event(conn, campaign_id, 'error', f'시스템 오류 발생: {error_msg}')
            conn.commit()
        finally:
            conn.close()
            with _worker_lock:
                _worker_threads.pop(campaign_id, None)
                _worker_cancel_events.pop(campaign_id, None)
                _worker_servers.pop(campaign_id, None)


@ai_mail_bp.route('/api/campaigns/<int:campaign_id>/start', methods=['POST'])
@_mutating
def start_campaign(campaign_id):
    conn = _db()
    try:
        campaign = _owned_campaign(conn, campaign_id)
        if not campaign:
            return _error('발송작업을 찾을 수 없습니다.', 404, code='CAMPAIGN_NOT_FOUND')
        if campaign['status'] != 'staged':
            return _error('임시 저장 상태의 작업만 발송을 시작할 수 있습니다.', 409, code='CAMPAIGN_LOCKED')
        if not campaign['preflight_ok']:
            return _error('발송 전 사전검사를 먼저 통과해야 합니다.', 400, code='PREFLIGHT_REQUIRED')

        conn.execute('BEGIN IMMEDIATE')
        # 상태를 대기(queued)로 변경하고 발송 준비
        conn.execute('''
            UPDATE ai_mail_campaigns
            SET status='queued', queued_at=CURRENT_TIMESTAMP, updated_at=CURRENT_TIMESTAMP
            WHERE id=? AND owner_emp_no=?
        ''', (campaign_id, _owner_emp_no()))
        
        _campaign_event(conn, campaign_id, 'started', '메일 발송 대기열에 등록되었습니다.')
        conn.commit()
        
        # 실제 발송을 처리할 백그라운드 스레드 실행
        app = current_app._get_current_object()
        with _worker_lock:
            if campaign_id not in _worker_threads:
                cancel_event = threading.Event()
                _worker_cancel_events[campaign_id] = cancel_event
                thread = threading.Thread(target=_send_campaign_worker, args=(app, campaign_id, cancel_event))
                thread.daemon = True
                _worker_threads[campaign_id] = thread
                thread.start()
        
        refreshed = _owned_campaign(conn, campaign_id)
        return _success('발송을 시작했습니다.', campaign=_campaign_dict(refreshed))
    except sqlite3.DatabaseError as exc:
        conn.rollback()
        return _error(f'발송 시작을 처리하지 못했습니다: {_clean_text(exc, 300)}', code='CAMPAIGN_START_FAILED')
    finally:
        conn.close()


@ai_mail_bp.route('/api/campaigns/<int:campaign_id>/cancel', methods=['POST'])
@_mutating
def cancel_campaign(campaign_id):
    conn = _db()
    try:
        campaign = _owned_campaign(conn, campaign_id)
        if not campaign:
            return _error('발송작업을 찾을 수 없습니다.', 404, code='CAMPAIGN_NOT_FOUND')

        status = str(campaign['status'] or '').strip().lower()
        if status in {'completed', 'success', 'completed_with_errors', 'failed', 'cancelled'}:
            return _success('이미 종료된 발송작업입니다.', campaign=_campaign_dict(campaign))
        if status == 'staged':
            return _error('아직 시작하지 않은 발송작업입니다.', 409, code='CAMPAIGN_NOT_STARTED')

        with _worker_lock:
            worker = _worker_threads.get(campaign_id)
            worker_active = bool(worker and worker.is_alive())
            cancel_event = _worker_cancel_events.get(campaign_id)
            active_server = _worker_servers.get(campaign_id)

        conn.execute('BEGIN IMMEDIATE')
        if worker_active:
            conn.execute('''
                UPDATE ai_mail_campaigns
                SET status='cancel_requested', cancel_requested=1,
                    cancel_requested_at=CURRENT_TIMESTAMP, cancel_requested_by=?,
                    cancel_reason=?, updated_at=CURRENT_TIMESTAMP
                WHERE id=? AND owner_emp_no=?
            ''', (_owner_emp_no(), '사용자 요청', campaign_id, _owner_emp_no()))
            _campaign_event(conn, campaign_id, 'cancel_requested', '사용자가 발송 중단을 요청했습니다.')
            message = '발송 중단을 요청했습니다.'
        else:
            cancelled_cursor = conn.execute('''
                UPDATE ai_mail_campaign_recipients
                SET status='cancelled', error_message=?, updated_at=CURRENT_TIMESTAMP
                WHERE campaign_id=? AND status='pending'
            ''', ('사용자 요청으로 발송이 중단되었습니다.', campaign_id))
            cancelled_count = int(campaign['cancelled_count'] or 0) + max(0, cancelled_cursor.rowcount)
            processed_count = int(campaign['sent_count'] or 0) + int(campaign['failed_count'] or 0) + cancelled_count
            conn.execute('''
                UPDATE ai_mail_campaigns
                SET status='cancelled', cancel_requested=1,
                    cancel_requested_at=CURRENT_TIMESTAMP, cancel_requested_by=?,
                    cancel_reason=?, processed_count=?, cancelled_count=?,
                    finished_at=CURRENT_TIMESTAMP, updated_at=CURRENT_TIMESTAMP
                WHERE id=? AND owner_emp_no=?
            ''', (
                _owner_emp_no(), '사용자 요청', processed_count, cancelled_count,
                campaign_id, _owner_emp_no(),
            ))
            _campaign_event(conn, campaign_id, 'cancelled', '실행 중인 발송 작업이 없어 즉시 중단 처리했습니다.')
            message = '발송을 중단했습니다.'
        conn.commit()

        if cancel_event:
            cancel_event.set()
        if active_server:
            try:
                active_server.close()
            except Exception:
                pass

        refreshed = _owned_campaign(conn, campaign_id)
        return _success(message, campaign=_campaign_dict(refreshed))
    except sqlite3.DatabaseError as exc:
        conn.rollback()
        return _error(f'발송 중단을 처리하지 못했습니다: {_clean_text(exc, 300)}', code='CAMPAIGN_CANCEL_FAILED')
    finally:
        conn.close()
