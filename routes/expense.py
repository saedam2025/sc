from flask import Blueprint, render_template, request, jsonify, session, send_file, abort, redirect, url_for
import csv
import hmac
import html
from io import BytesIO, StringIO
import json
import os
import re
import smtplib
import tempfile
import time
import zipfile
from datetime import date, datetime
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from PIL import Image, ImageOps, UnidentifiedImageError

from .database import get_db
from .menu_access import INSTRUCTOR_EXPENSE_ACCESS_SESSION, menu_is_allowed
from .storage import APP_ROOT, UPLOADS_ROOT
from .security import is_admin_session
from .secure_files import (
    decode_filename_token,
    delete_file,
    encode_filename_token,
    encrypted_response,
    encrypted_storage_name,
    encrypt_stream,
    encrypt_upload,
    original_filename,
    plaintext_size,
    read_decrypted,
    temporary_decrypted_path,
)

expense_bp = Blueprint('expense', __name__)
INSTRUCTOR_EXPENSE_PASSWORD = os.environ.get('EXPENSE_INSTRUCTOR_PASSWORD', '0070')

EXCEL_EXTENSIONS = {'.xlsx', '.xlsm', '.csv'}
RECEIPT_IMAGE_EXTENSIONS = {'.jpg', '.png', '.gif'}
RECEIPT_ALLOWED_EXTENSIONS = RECEIPT_IMAGE_EXTENSIONS
RECEIPT_IMAGE_FORMATS = {'JPEG', 'PNG', 'GIF'}
MAX_EXCEL_FILES = 1
MAX_RECEIPT_FILES = 20
MAX_EXCEL_FILE_SIZE = 10 * 1024 * 1024
MAX_RECEIPT_TOTAL_SIZE = 15 * 1024 * 1024
MAX_BULK_EMAIL_REPORTS = 50
MAX_BULK_EMAIL_ZIP_SIZE = 17 * 1024 * 1024
RECEIPT_IMAGE_MAX_SIZE = (1920, 1080)
RECEIPT_IMAGE_QUALITY = 85
UPLOAD_FOLDER = str(UPLOADS_ROOT)
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
MAIL_SETTINGS_PATH = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'mail_settings.json'))
EXPENSE_TEMPLATE_PATH = str(
    APP_ROOT
    / '-== 참고자료'
    / '지출결의서_기본양식.xlsx'
)
CENTER_EXPENSE_TEMPLATE_PATH = str(
    APP_ROOT
    / '-== 참고자료'
    / '지출결의서_센터장용_샘플양식.xlsx'
)

HEADER_ALIASES = {
    'expense_date': ['일자', '날짜', '사용일', '사용일자', '지출일', '지출일자', '거래일자', '집행일자'],
    'category': ['구분', '분류', '항목', '계정', '계정과목', '비목', '지출항목'],
    'vendor': ['거래처', '사용처', '사용출처', '가맹점', '지급처', '상호', '업체명', '구입처'],
    'description': ['내용', '내역', '적요', '품목', '세부내용', '사용내역', '지출내용'],
    'payment_method': ['결제수단', '지급방법', '결제방법', '카드', '현금', '계좌'],
    'amount': ['금액', '지출금액', '청구금액', '합계', '총액', '결제금액', '공급가액'],
    'note': ['비고', '메모', '참고', '증빙', '영수증']
}
TOTAL_ROW_LABELS = {'합계', '총계', '총합계', '소계', '계', '합'}
INFORMATION_ROW_KEYWORDS = (
    '안내', '안내문구', '비고', '참고', '참고사항', '주의', '주의사항',
    '유의', '유의사항', '알림', '작성방법', '작성요령', '입력방법', '영수증'
)
INFORMATION_ROW_PREFIXES = ('*', '※', '☞', '▶', '◆', '◇', '●', '○')
PLACEHOLDER_ROW_CHARACTERS = frozenset('-_–—·•')


def ensure_expense_schema():
    conn = get_db()
    try:
        conn.execute('''CREATE TABLE IF NOT EXISTS expense_reports (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            approval_id INTEGER UNIQUE,
            title TEXT,
            drafter TEXT,
            approver_1 TEXT,
            approver_2 TEXT,
            doc_status TEXT DEFAULT '대기',
            payment_status TEXT DEFAULT '결재중',
            total_amount INTEGER DEFAULT 0,
            item_count INTEGER DEFAULT 0,
            expense_org_type TEXT,
            expense_school_name TEXT,
            expense_manager TEXT,
            expense_kind TEXT,
            expense_date TEXT,
            report_year TEXT,
            report_month TEXT,
            report_day TEXT,
            submitted_at DATETIME,
            approved_at DATETIME,
            paid_at DATETIME,
            paid_by TEXT,
            source_filename TEXT,
            source_filepath TEXT,
            receipt_filename TEXT,
            receipt_filepath TEXT,
            submitter_email TEXT,
            payment_account TEXT,
            memo TEXT,
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            updated_at DATETIME DEFAULT CURRENT_TIMESTAMP
        )''')
        conn.execute('''CREATE TABLE IF NOT EXISTS expense_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            report_id INTEGER,
            approval_id INTEGER,
            row_no INTEGER,
            expense_date TEXT,
            category TEXT,
            vendor TEXT,
            description TEXT,
            payment_method TEXT,
            amount INTEGER DEFAULT 0,
            note TEXT,
            raw_json TEXT,
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP
        )''')
        for column_sql in [
            "ALTER TABLE expense_reports ADD COLUMN memo TEXT",
            "ALTER TABLE expense_reports ADD COLUMN expense_org_type TEXT",
            "ALTER TABLE expense_reports ADD COLUMN expense_school_name TEXT",
            "ALTER TABLE expense_reports ADD COLUMN expense_manager TEXT",
            "ALTER TABLE expense_reports ADD COLUMN expense_kind TEXT",
            "ALTER TABLE expense_reports ADD COLUMN expense_date TEXT",
            "ALTER TABLE expense_reports ADD COLUMN report_day TEXT",
            "ALTER TABLE expense_reports ADD COLUMN receipt_filename TEXT",
            "ALTER TABLE expense_reports ADD COLUMN receipt_filepath TEXT",
            "ALTER TABLE expense_reports ADD COLUMN submitter_email TEXT",
            "ALTER TABLE expense_reports ADD COLUMN payment_account TEXT"
        ]:
            try:
                conn.execute(column_sql)
            except Exception:
                pass
        conn.commit()
    finally:
        conn.close()


def _clean_text(value):
    if value is None:
        return ''
    if isinstance(value, datetime):
        return value.strftime('%Y-%m-%d')
    return str(value).strip()


def _parse_amount(value):
    if value is None or value == '':
        return 0
    if isinstance(value, (int, float)):
        return int(round(value))
    text = str(value).strip()
    if not text:
        return 0
    text = text.replace(',', '').replace('원', '').replace('₩', '').replace(' ', '')
    if text.startswith('(') and text.endswith(')'):
        text = '-' + text[1:-1]
    try:
        return int(round(float(text)))
    except Exception:
        digits = ''.join(ch for ch in text if ch.isdigit() or ch in '.-')
        try:
            return int(round(float(digits))) if digits else 0
        except Exception:
            return 0


def _parse_amount_strict(value):
    if value is None or _clean_text(value) == '':
        return None, '금액이 비어 있습니다. 예: 128000 또는 128,000'
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return int(round(value)), ''

    text = str(value).strip()
    normalized = text.replace(',', '').replace('원', '').replace('₩', '').replace(' ', '')
    if normalized.startswith('(') and normalized.endswith(')'):
        normalized = '-' + normalized[1:-1]
    try:
        return int(round(float(normalized))), ''
    except Exception:
        return None, f"금액 형식이 올바르지 않습니다: {text} (예: 128000 또는 128,000)"


def _normalize_date(value):
    if not value:
        return ''
    if isinstance(value, datetime):
        return value.strftime('%Y-%m-%d')
    text = str(value).strip()
    for fmt in ('%Y-%m-%d', '%Y.%m.%d', '%Y/%m/%d', '%y-%m-%d', '%y.%m.%d', '%y/%m/%d'):
        try:
            return datetime.strptime(text[:10], fmt).strftime('%Y-%m-%d')
        except Exception:
            pass
    return text


def _normalize_date_strict(value):
    if value is None or _clean_text(value) == '':
        return '', '날짜가 비어 있습니다. 예: 2026-02-22'
    if isinstance(value, datetime):
        return value.strftime('%Y-%m-%d'), ''
    if isinstance(value, date):
        return value.strftime('%Y-%m-%d'), ''

    text = str(value).strip()
    for fmt in ('%Y-%m-%d', '%Y.%m.%d', '%Y/%m/%d', '%y-%m-%d', '%y.%m.%d', '%y/%m/%d'):
        try:
            return datetime.strptime(text, fmt).strftime('%Y-%m-%d'), ''
        except Exception:
            pass
    return '', f"날짜 형식이 올바르지 않습니다: {text} (예: 2026-02-22)"


def _find_header_row(rows):
    best_index = 0
    best_score = -1
    for idx, row in enumerate(rows[:15]):
        normalized = [_clean_text(cell).replace(' ', '') for cell in row]
        score = 0
        for aliases in HEADER_ALIASES.values():
            if any(alias.replace(' ', '') in normalized for alias in aliases):
                score += 1
        if score > best_score:
            best_index = idx
            best_score = score
    return best_index


def _build_column_map(headers):
    normalized_headers = [_clean_text(h).replace(' ', '') for h in headers]
    column_map = {}
    for field, aliases in HEADER_ALIASES.items():
        for alias in aliases:
            alias_key = alias.replace(' ', '')
            if alias_key in normalized_headers:
                column_map[field] = normalized_headers.index(alias_key)
                break
    return column_map


def _is_total_row(row, column_map):
    amount_idx = column_map.get('amount')
    for idx, cell in enumerate(row):
        if idx == amount_idx:
            continue
        text = _clean_text(cell).replace(' ', '')
        if not text:
            continue
        # 비어 있는 기본 양식에서는 합계 수식 결과가 0이므로 금액의 참/거짓과
        # 관계없이 합계 라벨만으로 집계 행을 제외해야 한다.
        if text in TOTAL_ROW_LABELS or text.endswith('합계') or text.endswith('총계') or text.startswith('합계'):
            return True
    return False


def _is_placeholder_text(value):
    text = _clean_text(value).replace(' ', '')
    return bool(text) and all(char in PLACEHOLDER_ROW_CHARACTERS for char in text)


def _is_informational_row(row, column_map):
    def value_for(field):
        idx = column_map.get(field)
        if idx is None or idx >= len(row):
            return ''
        return row[idx]

    raw_date = value_for('expense_date')
    raw_amount = value_for('amount')
    vendor = _clean_text(value_for('vendor'))
    description = _clean_text(value_for('description'))
    _, date_error = _normalize_date_strict(raw_date)
    _, amount_error = _parse_amount_strict(raw_amount)

    # 날짜가 정상이거나 금액·사용처·사용내역이 모두 갖춰진 행은 실제 지출행으로 보고
    # 나머지 필드의 오류가 있으면 기존 검증에서 정확히 안내한다.
    if not date_error:
        return False
    if (
        not amount_error
        and vendor and not _is_placeholder_text(vendor)
        and description and not _is_placeholder_text(description)
    ):
        return False

    texts = [_clean_text(cell) for cell in row if _clean_text(cell)]
    substantive_texts = [text for text in texts if not _is_placeholder_text(text)]
    if not substantive_texts:
        return True

    core_texts = [
        _clean_text(value_for(field))
        for field in ('expense_date', 'vendor', 'description', 'amount')
    ]
    core_substantive_texts = [
        text for text in core_texts
        if text and not _is_placeholder_text(text)
    ]
    if not core_substantive_texts:
        return True

    first_text = substantive_texts[0].lstrip()
    compact_first_text = re.sub(r'\s+', '', first_text)
    return (
        first_text.startswith(INFORMATION_ROW_PREFIXES)
        or any(compact_first_text.startswith(keyword) for keyword in INFORMATION_ROW_KEYWORDS)
    )


def _is_total_item(item):
    amount = int(item['amount'] or 0) if 'amount' in item.keys() else 0
    for field in ('expense_date', 'category', 'vendor', 'description', 'payment_method', 'note'):
        text = _clean_text(item[field] if field in item.keys() else '').replace(' ', '')
        if amount and text and (text in TOTAL_ROW_LABELS or text.endswith('합계') or text.endswith('총계') or text.startswith('합계')):
            return True
    return False


def _mail_credentials():
    sender_email = os.environ.get('MAIL_USERNAME', '').strip()
    sender_password = os.environ.get('MAIL_PASSWORD', '').strip()

    if (not sender_email or not sender_password) and os.path.exists(MAIL_SETTINGS_PATH):
        try:
            with open(MAIL_SETTINGS_PATH, encoding='utf-8') as f:
                settings = json.load(f)
            sender_email = sender_email or _clean_text(settings.get('MAIL_USERNAME') or settings.get('email'))
            sender_password = sender_password or _clean_text(settings.get('MAIL_PASSWORD') or settings.get('password'))
        except Exception:
            pass

    return sender_email, sender_password.replace(' ', '')


def _normalize_existing_expense_totals(conn, report_id=None):
    if report_id:
        reports = conn.execute("SELECT id FROM expense_reports WHERE id=?", (report_id,)).fetchall()
    else:
        reports = conn.execute("SELECT id FROM expense_reports").fetchall()
    for report in reports:
        items = conn.execute("SELECT * FROM expense_items WHERE report_id=?", (report['id'],)).fetchall()
        total_item_ids = [item['id'] for item in items if _is_total_item(item)]
        if total_item_ids:
            placeholders = ','.join('?' for _ in total_item_ids)
            conn.execute(f"DELETE FROM expense_items WHERE id IN ({placeholders})", total_item_ids)
            items = [item for item in items if item['id'] not in total_item_ids]
        total_amount = sum(int(item['amount'] or 0) for item in items)
        item_count = len(items)
        conn.execute(
            "UPDATE expense_reports SET total_amount=?, item_count=?, updated_at=CURRENT_TIMESTAMP WHERE id=? AND (total_amount != ? OR item_count != ?)",
            (total_amount, item_count, report['id'], total_amount, item_count)
        )


def _rows_from_xlsx(path):
    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        sheet = workbook.active
        return [list(row) for row in sheet.iter_rows(values_only=True)]
    finally:
        # read_only 워크북은 명시적으로 닫지 않으면 Windows에서 임시파일을
        # 계속 점유하여 복호화 파일 정리 단계가 PermissionError로 실패한다.
        workbook.close()


def _rows_from_csv(path):
    rows = []
    encodings = ['utf-8-sig', 'cp949', 'euc-kr']
    last_error = None
    for encoding in encodings:
        try:
            with open(path, newline='', encoding=encoding) as f:
                rows = list(csv.reader(f))
            return rows
        except Exception as exc:
            last_error = exc
    if last_error:
        raise last_error
    return rows


def parse_expense_file_with_errors(path):
    ext = os.path.splitext(path)[1].lower()
    if ext not in EXCEL_EXTENSIONS:
        return [], ["등록할 수 없는 엑셀 파일 형식입니다."]

    rows = _rows_from_csv(path) if ext == '.csv' else _rows_from_xlsx(path)
    rows = [row for row in rows if any(_clean_text(cell) for cell in row)]
    if not rows:
        return [], ["엑셀에 읽을 수 있는 지출 내역이 없습니다."]

    header_index = _find_header_row(rows)
    headers = rows[header_index]
    column_map = _build_column_map(headers)
    if 'amount' not in column_map:
        return [], ["금액 열을 찾을 수 없습니다."]
    if 'expense_date' not in column_map:
        return [], ["날짜 열을 찾을 수 없습니다."]

    items = []
    errors = []
    for row_no, row in enumerate(rows[header_index + 1:], start=1):
        if _is_total_row(row, column_map) or _is_informational_row(row, column_map):
            continue

        def value_for(field):
            idx = column_map.get(field)
            if idx is None or idx >= len(row):
                return ''
            return row[idx]

        description = _clean_text(value_for('description'))
        vendor = _clean_text(value_for('vendor'))
        raw_date = value_for('expense_date')
        raw_amount = value_for('amount')
        has_row_data = any(_clean_text(value_for(field)) for field in ('expense_date', 'vendor', 'description', 'amount', 'category', 'payment_method', 'note'))

        if not has_row_data:
            continue

        expense_date, date_error = _normalize_date_strict(raw_date)
        amount, amount_error = _parse_amount_strict(raw_amount)
        vendor_error = '사용처가 비어 있습니다.' if not vendor else ''
        description_error = '사용내역이 비어 있습니다.' if not description else ''
        row_errors = [err for err in (date_error, amount_error, vendor_error, description_error) if err]
        invalid_fields = []
        if date_error:
            invalid_fields.append('expense_date')
        if amount_error:
            invalid_fields.append('amount')
        if vendor_error:
            invalid_fields.append('vendor')
        if description_error:
            invalid_fields.append('description')

        raw = {str(_clean_text(headers[i]) or f'컬럼{i + 1}'): _clean_text(row[i]) for i in range(min(len(headers), len(row)))}
        item = {
            'row_no': row_no,
            'expense_date': expense_date or _clean_text(raw_date),
            'category': _clean_text(value_for('category')),
            'vendor': vendor,
            'description': description,
            'payment_method': _clean_text(value_for('payment_method')),
            'amount': amount if amount is not None else 0,
            'amount_text': _clean_text(raw_amount),
            'note': _clean_text(value_for('note')),
            'raw_json': json.dumps(raw, ensure_ascii=False)
        }
        if row_errors:
            error_text = f"{row_no}행: " + " / ".join(row_errors)
            errors.append(error_text)
            item['has_error'] = True
            item['invalid_fields'] = invalid_fields
            item['validation_errors'] = row_errors
        items.append(item)
    return items, errors


def parse_expense_file(path):
    items, _ = parse_expense_file_with_errors(path)
    return [item for item in items if not item.get('has_error')]


def _payment_status_for_doc(doc_status):
    if doc_status == '완료':
        return '지급대기'
    if doc_status == '반려':
        return '반려'
    return '결재중'


def _can_manage_expenses():
    """지출결의 관리 권한은 메뉴권한관리의 설정을 따른다."""
    return menu_is_allowed('expense_main')


def _can_view_all_expenses():
    """지출결의 관리 메뉴에 접근할 수 있는 회원은 전체 내역을 조회한다."""
    return _can_manage_expenses()


def _html_text(value):
    return html.escape(_clean_text(value), quote=True)


def _expense_items_email_html(items):
    if not items:
        return '<p style="color:#64748b;">등록된 지출 세부내역이 없습니다.</p>'

    rows = []
    for index, item in enumerate(items, start=1):
        item_get = item.get if hasattr(item, 'get') else lambda key, default='': item[key] if key in item.keys() else default
        rows.append(f"""
            <tr>
                <td style="border:1px solid #cbd5e1; padding:7px; text-align:center;">{index}</td>
                <td style="border:1px solid #cbd5e1; padding:7px; text-align:center;">{_html_text(item_get('expense_date', ''))}</td>
                <td style="border:1px solid #cbd5e1; padding:7px;">{_html_text(item_get('vendor', ''))}</td>
                <td style="border:1px solid #cbd5e1; padding:7px;">{_html_text(item_get('description', ''))}</td>
                <td style="border:1px solid #cbd5e1; padding:7px; text-align:right; font-weight:700;">{int(item_get('amount', 0) or 0):,}원</td>
            </tr>
        """)

    return f"""
        <h3 style="margin:18px 0 8px; color:#0f766e;">지출 세부내역</h3>
        <table cellpadding="0" cellspacing="0" style="border-collapse:collapse; border:1px solid #cbd5e1; width:100%; max-width:760px; font-size:13px;">
            <thead>
                <tr>
                    <th style="background:#f8fafc; border:1px solid #cbd5e1; padding:8px; text-align:center;">순번</th>
                    <th style="background:#f8fafc; border:1px solid #cbd5e1; padding:8px; text-align:center;">일자</th>
                    <th style="background:#f8fafc; border:1px solid #cbd5e1; padding:8px; text-align:center;">사용처</th>
                    <th style="background:#f8fafc; border:1px solid #cbd5e1; padding:8px; text-align:center;">사용내역</th>
                    <th style="background:#f8fafc; border:1px solid #cbd5e1; padding:8px; text-align:center;">금액</th>
                </tr>
            </thead>
            <tbody>{''.join(rows)}</tbody>
        </table>
    """


def _record_value(record, key, default=''):
    if hasattr(record, 'get'):
        return record.get(key, default)
    return record[key] if key in record.keys() else default


def _combined_expense_status(report):
    if _clean_text(_record_value(report, 'doc_status')) == '반려':
        return '반려'
    if _clean_text(_record_value(report, 'payment_status')) == '지급완료':
        return '지급완료'
    if _clean_text(_record_value(report, 'doc_status')) == '완료':
        return '지급대기'
    return '결재대기'


def _safe_archive_component(value, fallback):
    text = _clean_text(value)
    text = re.sub(r'[\\/:*?"<>|\x00-\x1f]+', '_', text)
    text = re.sub(r'\s+', ' ', text).strip(' .')
    return (text or fallback)[:90]


def _unique_archive_path(used_paths, candidate):
    normalized = candidate.casefold()
    if normalized not in used_paths:
        used_paths.add(normalized)
        return candidate

    root, ext = os.path.splitext(candidate)
    number = 2
    while True:
        unique_candidate = f"{root}_{number}{ext}"
        normalized = unique_candidate.casefold()
        if normalized not in used_paths:
            used_paths.add(normalized)
            return unique_candidate
        number += 1


def _expense_report_archive_html(report, items):
    org_text = (
        _record_value(report, 'expense_school_name') or '학교'
        if _clean_text(_record_value(report, 'expense_org_type')) == '학교'
        else '본사'
    )
    memo_html = _html_text(_record_value(report, 'memo') or '-').replace('\n', '<br>')
    source_names = _html_text(_record_value(report, 'source_filename') or '-')
    receipt_names = _html_text(_record_value(report, 'receipt_filename') or '-')
    total_amount = int(_record_value(report, 'total_amount', 0) or 0)
    return f"""<!doctype html>
<html lang="ko">
<head>
    <meta charset="utf-8">
    <title>{_html_text(_record_value(report, 'title') or '지출결의서')}</title>
</head>
<body style="font-family:'Malgun Gothic',sans-serif; color:#0f172a; line-height:1.55; padding:24px;">
    <h1 style="text-align:center; font-size:24px; margin:0 0 22px;">지출결의서</h1>
    <table cellpadding="0" cellspacing="0" style="border-collapse:collapse; border:1px solid #cbd5e1; width:100%; max-width:900px; font-size:13px;">
        <tr><th style="background:#f8fafc; border:1px solid #cbd5e1; padding:9px; width:140px;">문서명</th><td style="border:1px solid #cbd5e1; padding:9px;" colspan="3">{_html_text(_record_value(report, 'title') or '-')}</td></tr>
        <tr><th style="background:#f8fafc; border:1px solid #cbd5e1; padding:9px;">학교/본사</th><td style="border:1px solid #cbd5e1; padding:9px;">{_html_text(org_text)}</td><th style="background:#f8fafc; border:1px solid #cbd5e1; padding:9px; width:140px;">담당자</th><td style="border:1px solid #cbd5e1; padding:9px;">{_html_text(_record_value(report, 'expense_manager') or _record_value(report, 'drafter') or '-')}</td></tr>
        <tr><th style="background:#f8fafc; border:1px solid #cbd5e1; padding:9px;">결의일</th><td style="border:1px solid #cbd5e1; padding:9px;">{_html_text(_record_value(report, 'expense_date') or '-')}</td><th style="background:#f8fafc; border:1px solid #cbd5e1; padding:9px;">상태</th><td style="border:1px solid #cbd5e1; padding:9px;">{_html_text(_combined_expense_status(report))}</td></tr>
        <tr><th style="background:#f8fafc; border:1px solid #cbd5e1; padding:9px;">결의서 내역</th><td style="border:1px solid #cbd5e1; padding:9px;">{_html_text(_record_value(report, 'expense_kind') or '-')}</td><th style="background:#f8fafc; border:1px solid #cbd5e1; padding:9px;">총 금액</th><td style="border:1px solid #cbd5e1; padding:9px; font-weight:700;">{total_amount:,}원</td></tr>
        <tr><th style="background:#f8fafc; border:1px solid #cbd5e1; padding:9px;">지급계좌번호</th><td style="border:1px solid #cbd5e1; padding:9px;" colspan="3">{_html_text(_record_value(report, 'payment_account') or '-')}</td></tr>
        <tr><th style="background:#f8fafc; border:1px solid #cbd5e1; padding:9px;">첨부 엑셀</th><td style="border:1px solid #cbd5e1; padding:9px;" colspan="3">{source_names}</td></tr>
        <tr><th style="background:#f8fafc; border:1px solid #cbd5e1; padding:9px;">영수증 증빙</th><td style="border:1px solid #cbd5e1; padding:9px;" colspan="3">{receipt_names}</td></tr>
        <tr><th style="background:#f8fafc; border:1px solid #cbd5e1; padding:9px;">비고</th><td style="border:1px solid #cbd5e1; padding:9px;" colspan="3">{memo_html}</td></tr>
    </table>
    {_expense_items_email_html(items)}
</body>
</html>"""


def _safe_expense_attachment_path(path):
    if not path or not os.path.isfile(path):
        return ''
    try:
        upload_root = os.path.realpath(UPLOAD_FOLDER)
        real_path = os.path.realpath(path)
        if os.path.commonpath([upload_root, real_path]) != upload_root:
            return ''
        return real_path
    except (OSError, ValueError):
        return ''


def _archive_attachment_name(original_name, stored_path):
    original_base = os.path.basename(_clean_text(original_name))
    return _safe_archive_component(original_base, os.path.basename(stored_path) or '첨부파일')


def _build_expense_archive(report_entries):
    archive_buffer = BytesIO()
    missing_files = []
    file_count = 0
    used_paths = set()
    summary_buffer = StringIO()
    summary_writer = csv.writer(summary_buffer)
    summary_writer.writerow(['순번', '문서명', '결의일', '학교/본사', '담당자', '상태', '금액'])

    with zipfile.ZipFile(archive_buffer, 'w', compression=zipfile.ZIP_DEFLATED) as archive:
        for index, (report, items) in enumerate(report_entries, start=1):
            report_id = int(_record_value(report, 'id', 0) or 0)
            title = _clean_text(_record_value(report, 'title')) or f'지출결의서_{report_id}'
            folder = _safe_archive_component(f"{index:02d}_{report_id}_{title}", f"{index:02d}_지출결의서")
            html_path = _unique_archive_path(used_paths, f"{folder}/지출결의서_내용.html")
            archive.writestr(html_path, _expense_report_archive_html(report, items).encode('utf-8'))
            file_count += 1

            org_text = (
                _record_value(report, 'expense_school_name') or '학교'
                if _clean_text(_record_value(report, 'expense_org_type')) == '학교'
                else '본사'
            )
            summary_writer.writerow([
                index,
                title,
                _record_value(report, 'expense_date') or '',
                org_text,
                _record_value(report, 'expense_manager') or _record_value(report, 'drafter') or '',
                _combined_expense_status(report),
                int(_record_value(report, 'total_amount', 0) or 0)
            ])

            attachment_groups = [
                ('첨부엑셀', _record_value(report, 'source_filename'), _record_value(report, 'source_filepath')),
                ('영수증', _record_value(report, 'receipt_filename'), _record_value(report, 'receipt_filepath'))
            ]
            for group_name, filename_text, filepath_text in attachment_groups:
                for detail in _attachment_details(filename_text, filepath_text):
                    display_name = detail['name'] or os.path.basename(detail['path'])
                    safe_path = _safe_expense_attachment_path(detail['path'])
                    if not safe_path:
                        missing_files.append(f"{title} - {display_name}")
                        continue
                    archive_name = _archive_attachment_name(display_name, safe_path)
                    archive_path = _unique_archive_path(
                        used_paths,
                        f"{folder}/{group_name}/{archive_name}"
                    )
                    archive.writestr(archive_path, read_decrypted(safe_path))
                    file_count += 1

        summary_path = _unique_archive_path(used_paths, '선택_지출결의서_목록.csv')
        archive.writestr(summary_path, summary_buffer.getvalue().encode('utf-8-sig'))
        file_count += 1

    archive_name = f"지출결의서_선택자료_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"
    return archive_buffer.getvalue(), archive_name, file_count, missing_files


def _send_expense_archive_email(to_email, archive_bytes, archive_name, reports):
    sender_email, sender_password = _mail_credentials()
    if not sender_email or not sender_password:
        return False, '메일 계정이 설정되어 있지 않습니다.'

    total_amount = sum(int(_record_value(report, 'total_amount', 0) or 0) for report in reports)
    report_rows = ''.join(
        f"<li>{_html_text(_record_value(report, 'title') or '지출결의서')} "
        f"({_combined_expense_status(report)}, {int(_record_value(report, 'total_amount', 0) or 0):,}원)</li>"
        for report in reports[:20]
    )
    extra_text = f"<li>외 {len(reports) - 20}건</li>" if len(reports) > 20 else ''
    body = f"""
    <div style="font-family:'Malgun Gothic',sans-serif; color:#0f172a; line-height:1.6;">
        <h2 style="margin:0 0 14px; color:#0f766e;">지출결의서 선택 자료</h2>
        <p>요청하신 지출결의서 <b>{len(reports):,}건</b>을 압축파일로 첨부합니다.</p>
        <ul>{report_rows}{extra_text}</ul>
        <p><b>전체 금액:</b> {total_amount:,}원</p>
        <p style="color:#64748b; font-size:12px;">압축파일에는 지출결의서 내용, 첨부 엑셀 및 영수증 증빙파일이 포함되어 있습니다.</p>
        <p style="margin-top:18px; color:#64748b; font-size:12px;">본 메일은 새담 인트라넷에서 발송되었습니다.</p>
    </div>
    """
    msg = MIMEMultipart()
    msg['From'] = f"새담 인트라넷 <{sender_email}>"
    msg['To'] = to_email
    msg['Subject'] = f"[새담 인트라넷] 지출결의서 선택 자료 {len(reports)}건"
    msg.attach(MIMEText(body, 'html'))
    archive_part = MIMEApplication(archive_bytes, _subtype='zip')
    archive_part.add_header('Content-Disposition', 'attachment', filename=('utf-8', '', archive_name))
    msg.attach(archive_part)

    try:
        with smtplib.SMTP('smtp.gmail.com', 587) as server:
            server.starttls()
            server.login(sender_email, sender_password)
            server.sendmail(sender_email, to_email, msg.as_string())
        return True, ''
    except Exception as exc:
        return False, f"메일 서버 전송 오류: {exc}"


def _send_expense_status_email(report, status_label, note='', items=None):
    to_email = _clean_text(report['submitter_email'] if 'submitter_email' in report.keys() else '')
    if not to_email or '@' not in to_email:
        return False, '신청자 이메일 없음'

    sender_email, sender_password = _mail_credentials()
    if not sender_email or not sender_password:
        return False, '메일 계정 설정 없음'

    period_info = _expense_period_from_items(items) if items else {}
    period_text = period_info.get('period_text') if period_info.get('has_dates') else ''
    if not period_text:
        period_text = f"{_clean_text(report['report_year'])}-{_clean_text(report['report_month'])}"

    subject = f"[새담 인트라넷] 지출결의서 {status_label} 안내"
    body = f"""
    <div style="font-family:'Malgun Gothic',sans-serif; color:#0f172a; line-height:1.6;">
        <h2 style="margin:0 0 14px; color:#0f766e;">지출결의서 {status_label}</h2>
        <p>아래 지출결의서 처리 결과를 안내드립니다.</p>
        <table cellpadding="0" cellspacing="0" style="border-collapse:collapse; border:1px solid #cbd5e1; min-width:520px; font-size:13px; line-height:1.45;">
            <tr><th style="background:#f8fafc; border:1px solid #cbd5e1; text-align:left; width:120px; padding:9px 12px;">문서명</th><td style="border:1px solid #cbd5e1; padding:9px 12px;">{_html_text(report['title'])}</td></tr>
            <tr><th style="background:#f8fafc; border:1px solid #cbd5e1; text-align:left; width:120px; padding:9px 12px;">담당자</th><td style="border:1px solid #cbd5e1; padding:9px 12px;">{_html_text(report['expense_manager'])}</td></tr>
            <tr><th style="background:#f8fafc; border:1px solid #cbd5e1; text-align:left; width:120px; padding:9px 12px;">지출기간</th><td style="border:1px solid #cbd5e1; padding:9px 12px;">{_html_text(period_text)}</td></tr>
            <tr><th style="background:#f8fafc; border:1px solid #cbd5e1; text-align:left; width:120px; padding:9px 12px;">금액</th><td style="border:1px solid #cbd5e1; padding:9px 12px;">{int(report['total_amount'] or 0):,}원</td></tr>
            <tr><th style="background:#f8fafc; border:1px solid #cbd5e1; text-align:left; width:120px; padding:9px 12px;">처리결과</th><td style="border:1px solid #cbd5e1; padding:9px 12px;">{status_label}</td></tr>
        </table>
        {f'<p style="margin-top:14px;"><b>비고:</b> {note}</p>' if note else ''}
        {_expense_items_email_html(items or [])}
        <p style="margin-top:18px; color:#64748b; font-size:12px;">본 메일은 새담 인트라넷에서 자동 발송되었습니다.</p>
    </div>
    """
    msg = MIMEMultipart()
    msg['From'] = f"새담 인트라넷 <{sender_email}>"
    msg['To'] = to_email
    msg['Subject'] = subject
    msg.attach(MIMEText(body, 'html'))

    try:
        with smtplib.SMTP('smtp.gmail.com', 587) as server:
            server.starttls()
            server.login(sender_email, sender_password)
            server.sendmail(sender_email, to_email, msg.as_string())
        return True, ''
    except Exception as exc:
        return False, str(exc)


def _send_expense_submission_email(report, items):
    to_email = _clean_text(report.get('submitter_email') if hasattr(report, 'get') else report['submitter_email'])
    if not to_email or '@' not in to_email:
        return False, '신청자 이메일 없음'

    sender_email, sender_password = _mail_credentials()
    if not sender_email or not sender_password:
        return False, '메일 계정 설정 없음'

    subject = "[새담 인트라넷] 지출결의서 접수 확인"
    total_amount = int(report.get('total_amount', 0) or 0)
    period_text = _clean_text(report.get('period_text')) or _expense_period_from_items(items).get('period_text')
    body = f"""
    <div style="font-family:'Malgun Gothic',sans-serif; color:#0f172a; line-height:1.6;">
        <h2 style="margin:0 0 14px; color:#0f766e;">지출결의서가 접수되었습니다.</h2>
        <p>아래는 전송하신 지출결의서 접수 내용입니다.</p>
        <table cellpadding="0" cellspacing="0" style="border-collapse:collapse; border:1px solid #cbd5e1; min-width:520px; font-size:13px; line-height:1.45;">
            <tr><th style="background:#f8fafc; border:1px solid #cbd5e1; text-align:left; width:120px; padding:9px 12px;">문서명</th><td style="border:1px solid #cbd5e1; padding:9px 12px;">{_html_text(report.get('title'))}</td></tr>
            <tr><th style="background:#f8fafc; border:1px solid #cbd5e1; text-align:left; width:120px; padding:9px 12px;">담당자</th><td style="border:1px solid #cbd5e1; padding:9px 12px;">{_html_text(report.get('expense_manager'))}</td></tr>
            <tr><th style="background:#f8fafc; border:1px solid #cbd5e1; text-align:left; width:120px; padding:9px 12px;">지출기간</th><td style="border:1px solid #cbd5e1; padding:9px 12px;">{_html_text(period_text)}</td></tr>
            <tr><th style="background:#f8fafc; border:1px solid #cbd5e1; text-align:left; width:120px; padding:9px 12px;">항목수</th><td style="border:1px solid #cbd5e1; padding:9px 12px;">{len(items):,}건</td></tr>
            <tr><th style="background:#f8fafc; border:1px solid #cbd5e1; text-align:left; width:120px; padding:9px 12px;">합계</th><td style="border:1px solid #cbd5e1; padding:9px 12px; font-weight:700;">{total_amount:,}원</td></tr>
        </table>
        {_expense_items_email_html(items)}
        <p style="margin-top:18px; color:#64748b; font-size:12px;">본 메일은 새담 인트라넷에서 자동 발송되었습니다.</p>
    </div>
    """
    msg = MIMEMultipart()
    msg['From'] = f"새담 인트라넷 <{sender_email}>"
    msg['To'] = to_email
    msg['Subject'] = subject
    msg.attach(MIMEText(body, 'html'))

    try:
        with smtplib.SMTP('smtp.gmail.com', 587) as server:
            server.starttls()
            server.login(sender_email, sender_password)
            server.sendmail(sender_email, to_email, msg.as_string())
        return True, ''
    except Exception as exc:
        return False, str(exc)


def _delete_file_paths(path_text):
    deleted = 0
    for path in [p.strip() for p in (path_text or '').split(',') if p.strip()]:
        if os.path.exists(path) and delete_file(path):
            deleted += 1
    return deleted


def _format_file_size(size):
    if size is None:
        return '-'
    if size >= 1024 * 1024:
        return f"{size / 1024 / 1024:.2f}MB"
    if size >= 1024:
        return f"{size / 1024:.1f}KB"
    return f"{size}B"


def _attachment_details(filename_text, filepath_text):
    names = [name.strip() for name in (filename_text or '').split(',') if name.strip()]
    paths = [path.strip() for path in (filepath_text or '').split(',') if path.strip()]
    details = []
    for index in range(max(len(names), len(paths))):
        path = paths[index] if index < len(paths) else ''
        name = decode_filename_token(names[index]) if index < len(names) else os.path.basename(path)
        size = plaintext_size(path) if path and os.path.exists(path) else None
        details.append({
            'name': name,
            'path': path,
            'size': size,
            'size_text': _format_file_size(size)
        })
    return details


def _safe_upload_name(filename):
    return encrypted_storage_name(original_filename(filename))


def _file_size(file):
    stream = file.stream
    current = stream.tell()
    stream.seek(0, os.SEEK_END)
    size = stream.tell()
    stream.seek(current)
    return size


def _validate_uploaded_files(files, max_count, allowed_extensions, label, max_file_size=None, max_total_size=None):
    if len(files) > max_count:
        return f"{label}은 최대 {max_count}개까지 첨부할 수 있습니다."
    total_size = 0
    for file in files:
        ext = os.path.splitext(file.filename or '')[1].lower()
        if allowed_extensions and ext not in allowed_extensions:
            return f"{label}에 등록할 수 없는 파일 형식입니다: {file.filename}"
        file_size = _file_size(file)
        total_size += file_size
        if max_file_size and file_size > max_file_size:
            return f"{label} 파일은 1개당 10MB 이하만 첨부할 수 있습니다: {file.filename}"
    if max_total_size and total_size > max_total_size:
        return f"{label}의 총 용량은 15MB 이하만 첨부할 수 있습니다."
    return ''


def _validate_receipt_images(files):
    """확장자만 바꾼 파일도 차단하고 이후 저장을 위해 스트림 위치를 복원한다."""
    for file in files:
        try:
            file.stream.seek(0)
            with Image.open(file.stream) as image:
                image_format = str(image.format or '').upper()
                image.verify()
            if image_format not in RECEIPT_IMAGE_FORMATS:
                return f"영수증 증빙파일은 JPG, PNG, GIF 이미지만 첨부할 수 있습니다: {file.filename}"
        except (UnidentifiedImageError, OSError, ValueError):
            return f"영수증 증빙파일이 올바른 이미지가 아닙니다: {file.filename}"
        finally:
            try:
                file.stream.seek(0)
            except Exception:
                pass
    return ''


def _parse_uploaded_expense_file(file):
    ext = os.path.splitext(file.filename or '')[1].lower()
    fd, path = tempfile.mkstemp(prefix='saedam-expense-parse-', suffix=ext)
    os.close(fd)
    try:
        file.stream.seek(0)
        file.save(path)
        return parse_expense_file(path)
    finally:
        try:
            os.remove(path)
        except Exception:
            pass


def _parse_uploaded_expense_file_with_errors(file):
    ext = os.path.splitext(file.filename or '')[1].lower()
    fd, path = tempfile.mkstemp(prefix='saedam-expense-parse-', suffix=ext)
    os.close(fd)
    try:
        file.stream.seek(0)
        file.save(path)
        return parse_expense_file_with_errors(path)
    finally:
        try:
            os.remove(path)
        except Exception:
            pass


def _expense_validation_message(errors):
    visible_errors = errors[:10]
    suffix = f"\n외 {len(errors) - 10}건의 오류가 더 있습니다." if len(errors) > 10 else ''
    return "엑셀 내역을 확인해주세요.\n" + "\n".join(visible_errors) + suffix


def _save_regular_uploaded_files(files):
    saved = []
    for file in files:
        if not file or not file.filename:
            continue
        original_name = original_filename(file.filename)
        safe_name = _safe_upload_name(original_name)
        path = os.path.join(UPLOAD_FOLDER, safe_name)
        try:
            encrypt_upload(file, path)
            saved.append((encode_filename_token(original_name), path))
        except Exception:
            _delete_file_paths(','.join(item[1] for item in saved))
            raise
    return saved


def _save_receipt_files(files):
    saved = []
    for file in files:
        if not file or not file.filename:
            continue
        original_name = original_filename(file.filename)
        ext = os.path.splitext(original_name)[1].lower()
        is_image = ext in RECEIPT_IMAGE_EXTENSIONS
        save_name = _safe_upload_name(original_name)

        if not is_image:
            _delete_file_paths(','.join(item[1] for item in saved))
            raise ValueError(f'허용되지 않은 영수증 파일 형식입니다: {original_name}')
        if is_image:
            save_root = os.path.splitext(save_name)[0]
            path = os.path.join(UPLOAD_FOLDER, f"{save_root}.jpg")
            try:
                file.stream.seek(0)
                img = Image.open(file.stream)
                img = ImageOps.exif_transpose(img)
                img.thumbnail(RECEIPT_IMAGE_MAX_SIZE, Image.Resampling.LANCZOS)
                if img.mode != 'RGB':
                    img = img.convert('RGB')
                buffer = BytesIO()
                img.save(buffer, format='JPEG', optimize=True, quality=RECEIPT_IMAGE_QUALITY)
                buffer.seek(0)
                encrypt_stream(buffer, path)
                saved.append((encode_filename_token(original_name), path))
                continue
            except (UnidentifiedImageError, OSError, ValueError):
                _delete_file_paths(','.join(item[1] for item in saved))
                raise ValueError(f'올바른 영수증 이미지가 아닙니다: {original_name}')
    return saved


def _split_month(expense_month):
    text = (expense_month or '').strip()
    try:
        dt = datetime.strptime(text, '%Y-%m')
    except Exception:
        dt = datetime.now()
    return str(dt.year), f"{dt.month:02d}", f"{dt.year}-{dt.month:02d}-01"


def _expense_period_from_items(items):
    dates = []
    for item in items or []:
        value = item.get('expense_date') if hasattr(item, 'get') else (item['expense_date'] if 'expense_date' in item.keys() else '')
        text = _clean_text(value)
        if not text:
            continue
        try:
            dates.append(datetime.strptime(text[:10], '%Y-%m-%d'))
        except Exception:
            pass

    if dates:
        start = min(dates)
        end = max(dates)
    else:
        now = datetime.now()
        start = end = datetime(now.year, now.month, 1)

    if start.year == end.year and start.month == end.month:
        period_text = f"{start.year}년 {start.month}월"
    elif start.year == end.year:
        period_text = f"{start.year}년 {start.month}월~{end.month}월"
    else:
        period_text = f"{start.year}년 {start.month}월~{end.year}년 {end.month}월"

    return {
        "has_dates": bool(dates),
        "start": start,
        "end": end,
        "period_text": period_text,
        "report_year": str(start.year),
        "report_month": f"{start.month:02d}",
        "report_day": f"{start.day:02d}",
        "expense_date": start.strftime('%Y-%m-%d')
    }


def _active_school_names(conn):
    try:
        rows = conn.execute('''
            SELECT DISTINCT school_name
            FROM schools
            WHERE COALESCE(is_active, 1)=1
              AND school_name IS NOT NULL
              AND TRIM(school_name) != ''
            ORDER BY school_name ASC
        ''').fetchall()
    except Exception:
        rows = conn.execute('''
            SELECT DISTINCT school_name
            FROM schools
            WHERE school_name IS NOT NULL
              AND TRIM(school_name) != ''
            ORDER BY school_name ASC
        ''').fetchall()
    return [row['school_name'] for row in rows]


def _build_expense_template_workbook():
    """배포본에 정적 양식이 없어도 같은 기본 엑셀 양식을 생성한다."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = '지출결의서'
    sheet.sheet_view.showGridLines = False

    teal = '0F766E'
    teal_soft = 'F0FDFA'
    slate = '334155'
    border_color = 'CBD5E1'
    white = 'FFFFFF'
    thin = Side(style='thin', color=border_color)
    cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    sheet.merge_cells('A1:G1')
    sheet['A1'] = '지출결의서'
    sheet['A1'].font = Font(name='맑은 고딕', size=20, bold=True, color=teal)
    sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')

    sheet.merge_cells('A2:G2')
    sheet['A2'] = '아래 입력란에 날짜·사용내역·사용출처·지출금액을 입력하고, 관련 영수증을 같은 순서로 첨부해 주세요.'
    sheet['A2'].font = Font(name='맑은 고딕', size=10, color=slate)
    sheet['A2'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    headers = ['날짜', '구분', '사용내역', '사용출처', '결제수단', '지출금액', '비고']
    for column, value in enumerate(headers, start=1):
        cell = sheet.cell(row=4, column=column, value=value)
        cell.font = Font(name='맑은 고딕', size=10, bold=True, color=white)
        cell.fill = PatternFill('solid', fgColor=teal)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = cell_border

    for row in range(5, 25):
        for column in range(1, 8):
            cell = sheet.cell(row=row, column=column)
            cell.font = Font(name='맑은 고딕', size=10, color=slate)
            cell.fill = PatternFill('solid', fgColor=white)
            cell.border = cell_border
            cell.alignment = Alignment(
                horizontal='right' if column == 6 else ('center' if column in {1, 2, 5} else 'left'),
                vertical='center',
                wrap_text=column in {3, 4, 7},
            )
        sheet.cell(row=row, column=1).number_format = 'yyyy-mm-dd'
        sheet.cell(row=row, column=6).number_format = '#,##0'

    category_validation = DataValidation(
        type='list',
        formula1='"교통비,식비,소모품비,회의비,통신비,기타"',
        allow_blank=True,
    )
    payment_validation = DataValidation(
        type='list',
        formula1='"법인카드,개인카드,현금,계좌이체,기타"',
        allow_blank=True,
    )
    sheet.add_data_validation(category_validation)
    sheet.add_data_validation(payment_validation)
    category_validation.add('B5:B24')
    payment_validation.add('E5:E24')

    sheet.merge_cells('A25:E25')
    sheet['A25'] = '합계'
    sheet['F25'] = '=SUM(F5:F24)'
    for column in range(1, 8):
        cell = sheet.cell(row=25, column=column)
        cell.font = Font(name='맑은 고딕', size=10, bold=True, color=teal)
        cell.fill = PatternFill('solid', fgColor=teal_soft)
        cell.border = cell_border
        cell.alignment = Alignment(horizontal='right', vertical='center')
    sheet['A25'].alignment = Alignment(horizontal='center', vertical='center')
    sheet['F25'].number_format = '#,##0'

    sheet.merge_cells('A27:G27')
    sheet['A27'] = '※ 작성 안내: 날짜는 2026-07-29 형식, 지출금액은 숫자로 입력해 주세요. 행이 부족하면 입력 행의 서식을 복사해 추가할 수 있습니다.'
    sheet['A27'].font = Font(name='맑은 고딕', size=9, color=slate)
    sheet['A27'].fill = PatternFill('solid', fgColor=teal_soft)
    sheet['A27'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    column_widths = {'A': 14, 'B': 14, 'C': 28, 'D': 24, 'E': 14, 'F': 15, 'G': 24}
    for column, width in column_widths.items():
        sheet.column_dimensions[column].width = width
    sheet.row_dimensions[1].height = 34
    sheet.row_dimensions[2].height = 32
    sheet.row_dimensions[4].height = 25
    for row in range(5, 25):
        sheet.row_dimensions[row].height = 22
    sheet.row_dimensions[25].height = 24
    sheet.row_dimensions[27].height = 32

    sheet.print_area = 'A1:G27'
    sheet.page_setup.orientation = 'landscape'
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.freeze_panes = 'A5'

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def _build_center_expense_template_workbook():
    """센터장용 5열 샘플 양식을 생성한다."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = '지출결의서'
    sheet.sheet_view.showGridLines = False

    navy, blue, blue_soft = '1F4E78', '5B9BD5', 'D9EAF7'
    green_soft, yellow_soft = 'E2F0D9', 'FFF2CC'
    slate, white, border_color = '334155', 'FFFFFF', 'D9D9D9'
    thin = Side(style='thin', color=border_color)
    cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    sheet.merge_cells('A1:E1')
    sheet['A1'] = '지출결의서'
    sheet['A1'].font = Font(name='맑은 고딕', size=20, bold=True, color=white)
    sheet['A1'].fill = PatternFill('solid', fgColor=navy)
    sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')

    sheet.merge_cells('A2:E2')
    sheet['A2'] = '아래 입력란에 날짜·사용내역·사용출처·지출금액을 입력하고, 관련 영수증을 같은 순서로 첨부해 주세요.'
    sheet['A2'].font = Font(name='맑은 고딕', size=10, color=navy)
    sheet['A2'].fill = PatternFill('solid', fgColor=blue_soft)
    sheet['A2'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    headers = ['날짜', '사용내역', '사용출처', '지출금액', '비고']
    for column, value in enumerate(headers, start=1):
        cell = sheet.cell(row=4, column=column, value=value)
        cell.font = Font(name='맑은 고딕', size=10, bold=True, color=white)
        cell.fill = PatternFill('solid', fgColor=blue)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = cell_border

    for row in range(5, 25):
        for column in range(1, 6):
            cell = sheet.cell(row=row, column=column)
            cell.font = Font(name='맑은 고딕', size=10, color=slate)
            cell.border = cell_border
            cell.alignment = Alignment(
                horizontal='right' if column == 4 else ('center' if column == 1 else 'left'),
                vertical='center',
                wrap_text=column in {2, 3, 5},
            )
        sheet.cell(row=row, column=1).number_format = 'yyyy-mm-dd'
        sheet.cell(row=row, column=4).number_format = '#,##0'

    sample_values = [date(2026, 7, 15), '강사 출석부용 황파일 구입', '문구점', 5700, '']
    for column, value in enumerate(sample_values, start=1):
        sheet.cell(row=5, column=column, value=value)

    sheet.merge_cells('A25:C25')
    sheet['A25'] = '합계'
    sheet['D25'] = '=SUM(D5:D24)'
    for column in range(1, 6):
        cell = sheet.cell(row=25, column=column)
        cell.font = Font(name='맑은 고딕', size=10, bold=True, color='385723')
        cell.fill = PatternFill('solid', fgColor=green_soft)
        cell.border = cell_border
        cell.alignment = Alignment(horizontal='right', vertical='center')
    sheet['A25'].alignment = Alignment(horizontal='center', vertical='center')
    sheet['D25'].number_format = '#,##0'

    sheet.merge_cells('A27:E27')
    sheet['A27'] = '※ 작성 안내: 날짜는 2026-07-29 형식, 지출금액은 숫자로 입력해 주세요. 행이 부족하면 입력 행의 서식을 복사해 추가할 수 있습니다.'
    sheet['A27'].font = Font(name='맑은 고딕', size=9, color='9C6500')
    sheet['A27'].fill = PatternFill('solid', fgColor=yellow_soft)
    sheet['A27'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    for column, width in {'A': 14, 'B': 30, 'C': 24, 'D': 15, 'E': 24}.items():
        sheet.column_dimensions[column].width = width
    for row, height in {1: 34, 2: 32, 4: 25, 25: 24, 27: 32}.items():
        sheet.row_dimensions[row].height = height
    for row in range(5, 25):
        sheet.row_dimensions[row].height = 22
    sheet.print_area = 'A1:E27'
    sheet.page_setup.orientation = 'landscape'
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.freeze_panes = 'A5'

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


@expense_bp.route('/template')
def expense_template():
    channel = str(request.args.get('channel') or 'headquarters').strip().lower()
    if channel == 'center':
        if os.path.isfile(CENTER_EXPENSE_TEMPLATE_PATH):
            return send_file(
                CENTER_EXPENSE_TEMPLATE_PATH,
                as_attachment=True,
                download_name='지출결의서_센터장용_샘플양식.xlsx',
            )
        return send_file(
            _build_center_expense_template_workbook(),
            as_attachment=True,
            download_name='지출결의서_센터장용_샘플양식.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

    if not os.path.isfile(EXPENSE_TEMPLATE_PATH):
        return send_file(
            _build_expense_template_workbook(),
            as_attachment=True,
            download_name='지출결의서_기본양식.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
    return send_file(
        EXPENSE_TEMPLATE_PATH,
        as_attachment=True,
        download_name='지출결의서_기본양식.xlsx'
    )


@expense_bp.route('/api/preview', methods=['POST'])
def preview_expense_upload():
    submit_channel = str(request.form.get('expense_submit_channel') or 'headquarters').strip()
    if submit_channel == 'center' and not is_admin_session():
        conn = get_db()
        assigned_school = _assigned_center_school(conn)
        conn.close()
        if not assigned_school:
            return jsonify({"status": "error", "message": "담당 학교의 센터장만 지출결의 엑셀을 올릴 수 있습니다."}), 403
    if submit_channel == 'instructor' and not session.get(INSTRUCTOR_EXPENSE_ACCESS_SESSION):
        return jsonify({"status": "error", "message": "강사용 전송페이지 비밀번호 인증이 필요합니다."}), 403

    excel_files = request.files.getlist('expense_excel')
    excel_files = [f for f in excel_files if f and f.filename]
    if not excel_files:
        return jsonify({"status": "error", "message": "엑셀파일이 첨부되지 않았습니다."}), 400

    excel_error = _validate_uploaded_files(
        excel_files,
        MAX_EXCEL_FILES,
        EXCEL_EXTENSIONS,
        '지출결의 엑셀',
        max_file_size=MAX_EXCEL_FILE_SIZE
    )
    if excel_error:
        return jsonify({"status": "error", "message": excel_error}), 400

    try:
        items, parse_errors = _parse_uploaded_expense_file_with_errors(excel_files[0])
    except Exception:
        return jsonify({
            "status": "error",
            "message": "엑셀 파일을 열 수 없습니다. 파일이 손상되지 않았는지 확인하고 Excel에서 .xlsx 형식으로 다시 저장해주세요."
        }), 400

    if not items:
        return jsonify({
            "status": "error",
            "message": "엑셀에 입력된 지출내역이 없습니다. 기본 양식의 5행부터 날짜·사용내역·사용출처·지출금액을 입력해주세요."
        }), 400

    valid_items = [item for item in items if not item.get('has_error')]
    period_info = _expense_period_from_items(valid_items)
    if not period_info["has_dates"] and not parse_errors:
        return jsonify({
            "status": "error",
            "message": "엑셀에서 지출일자를 읽지 못했습니다. 날짜/사용일/지출일 열을 확인해주세요."
        }), 400
    return jsonify({
        "status": "warning" if parse_errors else "success",
        "has_errors": bool(parse_errors),
        "message": _expense_validation_message(parse_errors) if parse_errors else "",
        "errors": parse_errors,
        "items": items,
        "item_count": len(valid_items),
        "total_amount": sum(item['amount'] for item in valid_items),
        "period_text": period_info["period_text"] if period_info["has_dates"] else "",
        "title": f"{period_info['period_text']} 지출결의서" if period_info["has_dates"] else ""
    })


def _assigned_center_school(conn, emp_no=None):
    emp_no = str(emp_no or session.get('emp_no') or '').strip()
    if not emp_no:
        return None
    return conn.execute(
        """
        SELECT id, school_name
        FROM schools
        WHERE ? IN (center_director_id, center_director_id_2)
          AND COALESCE(is_active, 1) = 1
        ORDER BY year DESC, id DESC
        LIMIT 1
        """,
        (emp_no,),
    ).fetchone()


def _expense_submit_page_context(prefer_assigned_school=False):
    conn = get_db()
    school_list = _active_school_names(conn)
    school_name = ''
    manager = ''
    email = ''
    payment_account = ''
    if prefer_assigned_school and session.get('emp_no'):
        current_user = conn.execute(
            "SELECT name, email, bank_account FROM users WHERE emp_no = ?",
            (session.get('emp_no'),),
        ).fetchone()
        assigned_school = _assigned_center_school(conn)
        if current_user:
            manager = current_user['name'] or ''
            email = current_user['email'] or ''
            payment_account = current_user['bank_account'] or ''
        if assigned_school:
            school_name = assigned_school['school_name'] or ''
    conn.close()
    return {
        'school_list': school_list,
        'current_month': datetime.now().strftime('%Y-%m'),
        'expense_submit_school_name': school_name,
        'expense_submit_manager': manager,
        'expense_submit_email': email,
        'expense_submit_payment_account': payment_account,
    }


@expense_bp.route('/submit/center')
def submit_expense_center():
    ensure_expense_schema()
    return render_template(
        'expense_submit_center.html',
        **_expense_submit_page_context(prefer_assigned_school=True),
    )


@expense_bp.route('/submit/instructor', methods=['GET', 'POST'])
def submit_expense_instructor():
    ensure_expense_schema()
    if request.method == 'POST':
        password = str(request.form.get('password') or '')
        if hmac.compare_digest(password, INSTRUCTOR_EXPENSE_PASSWORD):
            session[INSTRUCTOR_EXPENSE_ACCESS_SESSION] = True
            return redirect(url_for('expense.submit_expense_instructor'))
        return render_template(
            'expense_submit_instructor_login.html',
            password_error='비밀번호가 올바르지 않습니다.',
        ), 401

    if not session.get(INSTRUCTOR_EXPENSE_ACCESS_SESSION):
        return render_template('expense_submit_instructor_login.html')

    return render_template(
        'expense_submit_instructor.html',
        **_expense_submit_page_context(),
    )


@expense_bp.route('/submit', methods=['GET', 'POST'])
def submit_expense():
    ensure_expense_schema()
    if request.method == 'GET':
        return render_template(
            'expense_submit.html',
            **_expense_submit_page_context(),
        )

    org_type = request.form.get('expense_org_type', '').strip()
    school_name = request.form.get('expense_school_name', '').strip()
    manager = request.form.get('expense_manager', '').strip()
    submitter_email = request.form.get('submitter_email', '').strip()
    payment_account = request.form.get('payment_account', '').strip()
    expense_kind = request.form.get('expense_kind', '').strip()
    memo = request.form.get('memo', '').strip()
    submit_channel = str(request.form.get('expense_submit_channel') or 'headquarters').strip()

    if submit_channel not in {'headquarters', 'center', 'instructor'}:
        return jsonify({"status": "error", "message": "올바르지 않은 지출결의 전송 경로입니다."}), 400
    if submit_channel == 'instructor':
        if not session.get(INSTRUCTOR_EXPENSE_ACCESS_SESSION):
            return jsonify({"status": "error", "message": "강사용 전송페이지 비밀번호 인증이 필요합니다."}), 403
    elif submit_channel == 'center' and not is_admin_session():
        conn = get_db()
        assigned_school = _assigned_center_school(conn)
        conn.close()
        if not assigned_school:
            return jsonify({"status": "error", "message": "담당 학교의 센터장만 지출결의서를 전송할 수 있습니다."}), 403
        if org_type != '학교' or school_name != str(assigned_school['school_name'] or ''):
            return jsonify({"status": "error", "message": "센터장으로 지정된 담당 학교의 지출결의서만 전송할 수 있습니다."}), 403
    elif submit_channel == 'headquarters' and not session.get('emp_no'):
        return jsonify({"status": "error", "message": "인트라넷 로그인이 필요합니다."}), 401

    if org_type not in {'본사', '학교'}:
        return jsonify({"status": "error", "message": "본사 또는 학교를 선택해주세요."}), 400
    if org_type == '학교' and not school_name:
        return jsonify({"status": "error", "message": "학교명을 입력해주세요."}), 400
    if org_type == '학교':
        conn = get_db()
        active_schools = _active_school_names(conn)
        conn.close()
        if active_schools and school_name not in active_schools:
            return jsonify({"status": "error", "message": "활성 상태의 학교업무공간에 등록된 학교명만 선택할 수 있습니다."}), 400
    if not manager:
        return jsonify({"status": "error", "message": "담당자 성명을 입력해주세요."}), 400
    if not re.match(r'^[^\s@]+@[^\s@]+\.[^\s@]{2,}$', submitter_email):
        return jsonify({"status": "error", "message": "결과를 받을 이메일을 정확히 입력해주세요."}), 400
    if not expense_kind:
        return jsonify({"status": "error", "message": "결의서 내역을 선택해주세요."}), 400
    if not payment_account:
        return jsonify({"status": "error", "message": "지급계좌번호를 입력해주세요."}), 400

    excel_files = request.files.getlist('expense_excel')
    excel_files = [f for f in excel_files if f and f.filename]
    if not excel_files:
        return jsonify({"status": "error", "message": "엑셀파일이 첨부되지 않았습니다."}), 400
    excel_error = _validate_uploaded_files(
        excel_files,
        MAX_EXCEL_FILES,
        EXCEL_EXTENSIONS,
        '지출결의 엑셀',
        max_file_size=MAX_EXCEL_FILE_SIZE
    )
    if excel_error:
        return jsonify({"status": "error", "message": excel_error}), 400

    receipt_files = request.files.getlist('receipt_files')
    receipt_files = [f for f in receipt_files if f and f.filename]
    if not receipt_files:
        return jsonify({"status": "error", "message": "영수증 증빙 이미지가 첨부되지 않았습니다."}), 400
    receipt_error = _validate_uploaded_files(
        receipt_files,
        MAX_RECEIPT_FILES,
        RECEIPT_ALLOWED_EXTENSIONS,
        '영수증 증빙파일',
        max_total_size=MAX_RECEIPT_TOTAL_SIZE
    )
    if receipt_error:
        if '파일 형식' in receipt_error:
            receipt_error = "영수증 증빙파일은 JPG, PNG, GIF 이미지만 첨부할 수 있습니다."
        return jsonify({"status": "error", "message": receipt_error}), 400
    receipt_image_error = _validate_receipt_images(receipt_files)
    if receipt_image_error:
        return jsonify({"status": "error", "message": receipt_image_error}), 400

    saved_excels = []
    saved_receipts = []
    try:
        saved_excels = _save_regular_uploaded_files(excel_files)
        saved_receipts = _save_receipt_files(receipt_files)
    except (UnidentifiedImageError, OSError, ValueError):
        _delete_file_paths(','.join(path for _, path in saved_excels))
        _delete_file_paths(','.join(path for _, path in saved_receipts))
        return jsonify({"status": "error", "message": "영수증 증빙파일을 이미지로 처리할 수 없습니다. JPG, PNG, GIF 파일인지 확인해주세요."}), 400

    items = []
    parse_errors = []
    unreadable_files = []
    for stored_name, path in saved_excels:
        display_name = decode_filename_token(stored_name)
        try:
            with temporary_decrypted_path(path, display_name) as temporary_path:
                parsed, errors = parse_expense_file_with_errors(temporary_path)
        except Exception:
            parsed, errors = [], []
            unreadable_files.append(display_name)
        items.extend(parsed)
        parse_errors.extend(errors)

    if unreadable_files:
        _delete_file_paths(','.join(path for _, path in saved_excels))
        _delete_file_paths(','.join(path for _, path in saved_receipts))
        return jsonify({
            "status": "error",
            "message": "엑셀 파일을 열 수 없습니다. 파일이 손상되지 않았는지 확인하고 Excel에서 .xlsx 형식으로 다시 저장해주세요."
        }), 400

    if parse_errors:
        _delete_file_paths(','.join(path for _, path in saved_excels))
        _delete_file_paths(','.join(path for _, path in saved_receipts))
        return jsonify({
            "status": "error",
            "message": _expense_validation_message(parse_errors),
            "errors": parse_errors
        }), 400

    if not items:
        _delete_file_paths(','.join(path for _, path in saved_excels))
        _delete_file_paths(','.join(path for _, path in saved_receipts))
        return jsonify({"status": "error", "message": "엑셀에 입력된 지출내역이 없습니다. 기본 양식의 5행부터 날짜·사용내역·사용출처·지출금액을 입력해주세요."}), 400

    if expense_kind:
        for item in items:
            if not item.get('category'):
                item['category'] = expense_kind

    period_info = _expense_period_from_items(items)
    if not period_info["has_dates"]:
        _delete_file_paths(','.join(path for _, path in saved_excels))
        _delete_file_paths(','.join(path for _, path in saved_receipts))
        return jsonify({"status": "error", "message": "엑셀에서 지출일자를 읽지 못했습니다. 날짜/사용일/지출일 열을 확인해주세요."}), 400
    report_year = period_info["report_year"]
    report_month = period_info["report_month"]
    report_day = period_info["report_day"]
    expense_date = period_info["expense_date"]
    title = f"{period_info['period_text']} 지출결의서"
    total_amount = sum(item['amount'] for item in items)
    excel_names = ','.join(name for name, _ in saved_excels)
    excel_paths = ','.join(path for _, path in saved_excels)
    receipt_names = ','.join(name for name, _ in saved_receipts)
    receipt_paths = ','.join(path for _, path in saved_receipts)

    conn = get_db()
    cursor = conn.execute('''
        INSERT INTO expense_reports (
            approval_id, title, drafter, approver_1, approver_2, doc_status, payment_status,
            total_amount, item_count, expense_org_type, expense_school_name, expense_manager,
            expense_kind, expense_date, report_year, report_month, report_day, submitted_at,
            source_filename, source_filepath, receipt_filename, receipt_filepath, submitter_email, payment_account, memo, updated_at
        )
        VALUES (NULL, ?, ?, '', '', '대기', '결재대기', ?, ?, ?, ?, ?, ?, ?, ?, ?, ?,
                CURRENT_TIMESTAMP, ?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
    ''', (
        title, manager, total_amount, len(items), org_type, school_name, manager,
        expense_kind, expense_date, report_year, report_month,
        report_day, excel_names, excel_paths, receipt_names, receipt_paths, submitter_email, payment_account, memo
    ))
    report_id = cursor.lastrowid
    for item in items:
        conn.execute('''
            INSERT INTO expense_items (
                report_id, approval_id, row_no, expense_date, category, vendor,
                description, payment_method, amount, note, raw_json
            )
            VALUES (?, NULL, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            report_id, item['row_no'], item['expense_date'], item['category'], item['vendor'],
            item['description'], item['payment_method'], item['amount'], item['note'], item['raw_json']
        ))
    conn.commit()
    conn.close()

    report_for_mail = {
        "title": title,
        "expense_manager": manager,
        "report_year": report_year,
        "report_month": report_month,
        "period_text": period_info["period_text"],
        "total_amount": total_amount,
        "submitter_email": submitter_email
    }
    mail_sent, mail_error = _send_expense_submission_email(report_for_mail, items)
    message = "지출결의서가 접수되었습니다."
    message += " 접수 확인 메일 발송 완료." if mail_sent else f" 접수 확인 메일 발송 실패: {mail_error}"

    return jsonify({
        "status": "success",
        "message": message,
        "report_id": report_id,
        "title": title,
        "period_text": period_info["period_text"],
        "expense_manager": manager,
        "submitter_email": submitter_email,
        "mail_sent": mail_sent,
        "mail_error": '' if mail_sent else mail_error,
        "total_amount": total_amount,
        "item_count": len(items)
    })


def sync_expense_from_approval(approval_id, conn=None):
    owns_connection = conn is None
    if conn is None:
        conn = get_db()

    try:
        doc = conn.execute("SELECT * FROM approvals WHERE id=?", (approval_id,)).fetchone()
        if not doc or doc['doc_type'] != '지출결의서':
            if owns_connection:
                conn.close()
            return

        filenames = [name.strip() for name in (doc['filename'] or '').split(',') if name.strip()]
        filepaths = [path.strip() for path in (doc['filepath'] or '').split(',') if path.strip()]

        items = []
        source_names = []
        source_paths = []
        for idx, path in enumerate(filepaths):
            source_name = decode_filename_token(filenames[idx]) if idx < len(filenames) else os.path.basename(path)
            ext = os.path.splitext(source_name)[1].lower()
            if ext not in EXCEL_EXTENSIONS:
                continue
            try:
                with temporary_decrypted_path(path, source_name) as temporary_path:
                    parsed = parse_expense_file(temporary_path)
            except Exception:
                parsed = []
            if parsed:
                source_names.append(filenames[idx] if idx < len(filenames) else encode_filename_token(source_name))
                source_paths.append(path)
                items.extend(parsed)

        total_amount = sum(item['amount'] for item in items)
        item_count = len(items)
        submitted_at = doc['created_at']
        try:
            doc_data = json.loads(doc['doc_data']) if doc['doc_data'] else {}
        except Exception:
            doc_data = {}
        expense_date = _normalize_date(doc_data.get('expense_date')) or (items[0]['expense_date'] if items and items[0].get('expense_date') else '')
        basis_date = expense_date or submitted_at
        submitted_dt = None
        try:
            submitted_dt = datetime.strptime(str(basis_date)[:10], '%Y-%m-%d')
        except Exception:
            try:
                submitted_dt = datetime.strptime(str(submitted_at)[:19], '%Y-%m-%d %H:%M:%S')
            except Exception:
                submitted_dt = datetime.now()

        report_year = str(submitted_dt.year)
        report_month = f"{submitted_dt.month:02d}"
        report_day = f"{submitted_dt.day:02d}"
        approved_at = doc['updated_at'] if doc['status'] == '완료' else None
        expense_org_type = _clean_text(doc_data.get('expense_org_type')) or '본사'
        expense_school_name = _clean_text(doc_data.get('expense_school_name'))
        expense_manager = _clean_text(doc_data.get('expense_manager')) or doc['drafter']
        expense_kind = _clean_text(doc_data.get('expense_kind'))
        if expense_kind:
            for item in items:
                if not item.get('category'):
                    item['category'] = expense_kind

        existing = conn.execute("SELECT * FROM expense_reports WHERE approval_id=?", (approval_id,)).fetchone()
        existing_payment = existing['payment_status'] if existing else ''
        payment_status = existing_payment if existing_payment == '지급완료' else _payment_status_for_doc(doc['status'])

        conn.execute('''
            INSERT INTO expense_reports (
                approval_id, title, drafter, approver_1, approver_2, doc_status, payment_status,
                total_amount, item_count, expense_org_type, expense_school_name, expense_manager,
                expense_kind, expense_date, report_year, report_month, report_day, submitted_at, approved_at,
                source_filename, source_filepath, updated_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
            ON CONFLICT(approval_id) DO UPDATE SET
                title=excluded.title,
                drafter=excluded.drafter,
                approver_1=excluded.approver_1,
                approver_2=excluded.approver_2,
                doc_status=excluded.doc_status,
                payment_status=excluded.payment_status,
                total_amount=excluded.total_amount,
                item_count=excluded.item_count,
                expense_org_type=excluded.expense_org_type,
                expense_school_name=excluded.expense_school_name,
                expense_manager=excluded.expense_manager,
                expense_kind=excluded.expense_kind,
                expense_date=excluded.expense_date,
                report_year=excluded.report_year,
                report_month=excluded.report_month,
                report_day=excluded.report_day,
                submitted_at=excluded.submitted_at,
                approved_at=excluded.approved_at,
                source_filename=excluded.source_filename,
                source_filepath=excluded.source_filepath,
                updated_at=CURRENT_TIMESTAMP
        ''', (
            approval_id, doc['title'], doc['drafter'], doc['approver_1'], doc['approver_2'], doc['status'],
            payment_status, total_amount, item_count, expense_org_type, expense_school_name, expense_manager,
            expense_kind, expense_date, report_year, report_month, report_day, submitted_at, approved_at,
            ','.join(source_names), ','.join(source_paths)
        ))

        report = conn.execute("SELECT id FROM expense_reports WHERE approval_id=?", (approval_id,)).fetchone()
        report_id = report['id']
        conn.execute("DELETE FROM expense_items WHERE report_id=?", (report_id,))
        for item in items:
            conn.execute('''
                INSERT INTO expense_items (
                    report_id, approval_id, row_no, expense_date, category, vendor,
                    description, payment_method, amount, note, raw_json
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                report_id, approval_id, item['row_no'], item['expense_date'], item['category'], item['vendor'],
                item['description'], item['payment_method'], item['amount'], item['note'], item['raw_json']
            ))

        if owns_connection:
            conn.commit()
            conn.close()
    except Exception:
        if owns_connection:
            conn.close()
        raise


def _query_reports(conn, filters, can_see_all, current_user):
    where = ["1=1"]
    params = []
    if not can_see_all:
        where.append("drafter=?")
        params.append(current_user)
    if filters.get('drafter'):
        where.append("drafter=?")
        params.append(filters['drafter'])
    if filters.get('org_type'):
        where.append("expense_org_type=?")
        params.append(filters['org_type'])
    if filters.get('school_name'):
        where.append("expense_school_name=?")
        params.append(filters['school_name'])
    if filters.get('manager'):
        where.append("expense_manager=?")
        params.append(filters['manager'])
    if filters.get('expense_kind'):
        where.append("expense_kind=?")
        params.append(filters['expense_kind'])
    if filters.get('year'):
        where.append("report_year=?")
        params.append(filters['year'])
    if filters.get('month'):
        where.append("report_month=?")
        params.append(filters['month'].zfill(2))
    if filters.get('day'):
        where.append("report_day=?")
        params.append(filters['day'].zfill(2))
    if filters.get('payment_status'):
        where.append("payment_status=?")
        params.append(filters['payment_status'])
    if filters.get('q'):
        where.append("""(
            title LIKE ? OR drafter LIKE ? OR expense_manager LIKE ? OR expense_org_type LIKE ?
            OR expense_school_name LIKE ? OR expense_kind LIKE ? OR expense_date LIKE ?
            OR report_year LIKE ? OR report_month LIKE ? OR report_day LIKE ?
            OR source_filename LIKE ? OR receipt_filename LIKE ?
        )""")
        like = f"%{filters['q']}%"
        params.extend([like, like, like, like, like, like, like, like, like, like, like, like])

    return conn.execute(f'''
        SELECT *
        FROM expense_reports
        WHERE {' AND '.join(where)}
        ORDER BY submitted_at DESC, id DESC
    ''', params).fetchall()


@expense_bp.route('/')
def index():
    ensure_expense_schema()
    current_user = session.get('user_name', '')
    can_see_all = _can_view_all_expenses()
    can_manage_expenses = _can_manage_expenses()

    filters = {
        'q': request.args.get('q', '').strip(),
        'drafter': request.args.get('drafter', '').strip(),
        'org_type': request.args.get('org_type', '').strip(),
        'school_name': request.args.get('school_name', '').strip(),
        'manager': request.args.get('manager', '').strip(),
        'expense_kind': request.args.get('expense_kind', '').strip(),
        'year': request.args.get('year', '').strip(),
        'month': request.args.get('month', '').strip(),
        'day': request.args.get('day', '').strip(),
        'payment_status': request.args.get('payment_status', '').strip()
    }

    conn = get_db()
    _normalize_existing_expense_totals(conn)
    conn.commit()
    rows = _query_reports(conn, filters, can_see_all, current_user)
    reports = [dict(row) for row in rows]

    drafters = conn.execute("SELECT DISTINCT drafter FROM expense_reports WHERE drafter IS NOT NULL AND drafter != '' ORDER BY drafter").fetchall()
    managers = conn.execute("SELECT DISTINCT expense_manager FROM expense_reports WHERE expense_manager IS NOT NULL AND expense_manager != '' ORDER BY expense_manager").fetchall()
    schools = conn.execute("SELECT DISTINCT expense_school_name FROM expense_reports WHERE expense_school_name IS NOT NULL AND expense_school_name != '' ORDER BY expense_school_name").fetchall()
    years = conn.execute("SELECT DISTINCT report_year FROM expense_reports WHERE report_year IS NOT NULL AND report_year != '' ORDER BY report_year DESC").fetchall()
    conn.close()

    approved_count = sum(1 for r in reports if r.get('doc_status') == '완료')
    paid_count = sum(1 for r in reports if r.get('payment_status') == '지급완료')
    
    # [추가됨] 결재대기 건수 계산 (완료되거나 반려되지 않은 문서들의 수)
    approval_waiting_count = sum(1 for r in reports if r.get('doc_status') not in ('완료', '반려'))

    approval_completion_rate = round((approved_count / len(reports) * 100), 1) if reports else 0
    payment_completion_rate = round((paid_count / len(reports) * 100), 1) if reports else 0

    summary = {
        'total_count': len(reports),
        'total_amount': sum(int(r.get('total_amount') or 0) for r in reports),
        'approval_waiting_count': approval_waiting_count,  # <--- 이 부분이 추가되었습니다.
        'waiting_count': sum(1 for r in reports if r.get('payment_status') == '지급대기'),
        'waiting_amount': sum(int(r.get('total_amount') or 0) for r in reports if r.get('payment_status') == '지급대기'),
        'approved_count': approved_count,
        'paid_count': paid_count,
        'paid_amount': sum(int(r.get('total_amount') or 0) for r in reports if r.get('payment_status') == '지급완료'),
        'approval_completion_rate': approval_completion_rate,
        'payment_completion_rate': payment_completion_rate
    }

    return render_template(
        'expense.html',
        reports=reports,
        filters=filters,
        summary=summary,
        drafters=[r['drafter'] for r in drafters],
        managers=[r['expense_manager'] for r in managers],
        schools=[r['expense_school_name'] for r in schools],
        years=[r['report_year'] for r in years],
        can_see_all=can_see_all,
        can_manage_expenses=can_manage_expenses
    )


@expense_bp.route('/api/report/<int:report_id>')
def report_detail(report_id):
    ensure_expense_schema()
    conn = get_db()
    _normalize_existing_expense_totals(conn, report_id)
    conn.commit()
    report = conn.execute("SELECT * FROM expense_reports WHERE id=?", (report_id,)).fetchone()
    if not report:
        conn.close()
        return jsonify({"status": "error", "message": "지출결의 내역을 찾을 수 없습니다."}), 404
    if not _can_view_all_expenses() and report['drafter'] != session.get('user_name'):
        conn.close()
        return jsonify({"status": "error", "message": "조회 권한이 없습니다."}), 403
    items = conn.execute("SELECT * FROM expense_items WHERE report_id=? ORDER BY row_no ASC, id ASC", (report_id,)).fetchall()
    source_files = _attachment_details(report['source_filename'], report['source_filepath'])
    receipt_files = _attachment_details(report['receipt_filename'], report['receipt_filepath'])
    conn.close()
    for kind, files in (('source', source_files), ('receipt', receipt_files)):
        for index, item in enumerate(files):
            item['url'] = f"/expense/api/report/{report_id}/attachment/{kind}/{index}"
            item.pop('path', None)
    return jsonify({
        "status": "success",
        "report": dict(report),
        "items": [dict(item) for item in items],
        "source_files": source_files,
        "receipt_files": receipt_files
    })


@expense_bp.route('/api/report/<int:report_id>/attachment/<kind>/<int:file_index>')
def report_attachment(report_id, kind, file_index):
    if kind not in {'source', 'receipt'}:
        abort(404)
    conn = get_db()
    report = conn.execute("SELECT * FROM expense_reports WHERE id=?", (report_id,)).fetchone()
    conn.close()
    if not report:
        abort(404)
    if not _can_view_all_expenses() and report['drafter'] != session.get('user_name'):
        abort(403)
    prefix = 'source' if kind == 'source' else 'receipt'
    files = _attachment_details(report[f'{prefix}_filename'], report[f'{prefix}_filepath'])
    if file_index < 0 or file_index >= len(files):
        abort(404)
    item = files[file_index]
    return encrypted_response(item['path'], item['name'], as_attachment=True)


@expense_bp.route('/api/reports/email', methods=['POST'])
def email_selected_reports():
    ensure_expense_schema()
    data = request.get_json(silent=True) or {}
    to_email = _clean_text(data.get('email'))
    raw_ids = data.get('ids') or []

    if not re.fullmatch(r'^[^\s@]+@[^\s@]+\.[^\s@]{2,}$', to_email):
        return jsonify({"status": "error", "message": "받는 사람 이메일 주소를 정확히 입력해주세요."}), 400
    if not isinstance(raw_ids, list):
        return jsonify({"status": "error", "message": "전송할 지출결의서를 다시 선택해주세요."}), 400

    report_ids = []
    for value in raw_ids:
        try:
            report_id = int(value)
        except (TypeError, ValueError):
            continue
        if report_id > 0 and report_id not in report_ids:
            report_ids.append(report_id)

    if not report_ids:
        return jsonify({"status": "error", "message": "전송할 지출결의서를 선택해주세요."}), 400
    if len(report_ids) > MAX_BULK_EMAIL_REPORTS:
        return jsonify({
            "status": "error",
            "message": f"한 번에 최대 {MAX_BULK_EMAIL_REPORTS}건까지 전송할 수 있습니다."
        }), 400

    current_user = session.get('user_name', '')
    can_view_all = _can_view_all_expenses()
    report_entries = []
    conn = get_db()
    try:
        for report_id in report_ids:
            report = conn.execute(
                "SELECT * FROM expense_reports WHERE id=?",
                (report_id,)
            ).fetchone()
            if not report:
                return jsonify({
                    "status": "error",
                    "message": f"선택한 지출결의서 #{report_id}을(를) 찾을 수 없습니다."
                }), 404
            if not can_view_all and report['drafter'] != current_user:
                return jsonify({
                    "status": "error",
                    "message": "본인이 작성한 지출결의서만 이메일로 전송할 수 있습니다."
                }), 403
            items = conn.execute(
                "SELECT * FROM expense_items WHERE report_id=? ORDER BY row_no ASC, id ASC",
                (report_id,)
            ).fetchall()
            report_entries.append((report, items))
    finally:
        conn.close()

    try:
        archive_bytes, archive_name, file_count, missing_files = _build_expense_archive(report_entries)
    except Exception:
        return jsonify({
            "status": "error",
            "message": "지출결의서 압축파일을 만드는 중 오류가 발생했습니다."
        }), 500

    if missing_files:
        visible_missing = ', '.join(missing_files[:5])
        suffix = f" 외 {len(missing_files) - 5}개" if len(missing_files) > 5 else ''
        return jsonify({
            "status": "error",
            "message": f"첨부파일을 찾을 수 없어 발송하지 않았습니다: {visible_missing}{suffix}"
        }), 400

    if len(archive_bytes) > MAX_BULK_EMAIL_ZIP_SIZE:
        return jsonify({
            "status": "error",
            "message": (
                f"압축파일이 {_format_file_size(len(archive_bytes))}로 메일 첨부 제한을 초과했습니다. "
                "선택 항목 수를 줄여 나누어 전송해주세요."
            )
        }), 400

    reports = [entry[0] for entry in report_entries]
    sent, mail_error = _send_expense_archive_email(to_email, archive_bytes, archive_name, reports)
    if not sent:
        return jsonify({
            "status": "error",
            "message": mail_error or "메일을 전송하지 못했습니다."
        }), 500

    return jsonify({
        "status": "success",
        "message": f"{to_email} 주소로 지출결의서 {len(reports)}건과 파일 {file_count}개를 전송했습니다.",
        "report_count": len(reports),
        "file_count": file_count,
        "archive_size": len(archive_bytes)
    })


@expense_bp.route('/api/report/<int:report_id>/paid', methods=['POST'])
def mark_paid(report_id):
    ensure_expense_schema()
    if not _can_manage_expenses():
        return jsonify({"status": "error", "message": "지급 처리 권한이 없습니다."}), 403
    current_user = session.get('user_name', '')
    conn = get_db()
    report = conn.execute("SELECT * FROM expense_reports WHERE id=?", (report_id,)).fetchone()
    if not report:
        conn.close()
        return jsonify({"status": "error", "message": "지출결의 내역을 찾을 수 없습니다."}), 404
    if report['doc_status'] != '완료':
        conn.close()
        return jsonify({"status": "error", "message": "결재 완료 문서만 지급완료 처리할 수 있습니다."}), 400
    if report['payment_status'] == '지급완료':
        conn.close()
        return jsonify({"status": "success", "message": "이미 지급완료 상태입니다."})

    cursor = conn.execute('''
        UPDATE expense_reports
        SET payment_status='지급완료', paid_at=CURRENT_TIMESTAMP, paid_by=?, updated_at=CURRENT_TIMESTAMP
        WHERE id=? AND payment_status != '지급완료'
    ''', (current_user, report_id))
    if cursor.rowcount == 0:
        conn.close()
        return jsonify({"status": "success", "message": "이미 지급완료 상태입니다."})
    updated_report = conn.execute("SELECT * FROM expense_reports WHERE id=?", (report_id,)).fetchone()
    items = conn.execute("SELECT * FROM expense_items WHERE report_id=? ORDER BY row_no ASC, id ASC", (report_id,)).fetchall()
    conn.commit()
    conn.close()
    mail_sent, mail_error = _send_expense_status_email(updated_report, '지급완료', items=items) if updated_report else (False, '보고서 없음')
    message = "지급완료로 처리했습니다."
    if updated_report and updated_report['submitter_email']:
        message += " 메일 발송 완료." if mail_sent else f" 메일 발송 실패: {mail_error}"
    return jsonify({"status": "success", "message": message})


@expense_bp.route('/api/report/<int:report_id>/approve', methods=['POST'])
def mark_approved(report_id):
    ensure_expense_schema()
    if not _can_manage_expenses():
        return jsonify({"status": "error", "message": "결재 처리 권한이 없습니다."}), 403
    current_user = session.get('user_name', '')
    conn = get_db()
    report = conn.execute("SELECT * FROM expense_reports WHERE id=?", (report_id,)).fetchone()
    if not report:
        conn.close()
        return jsonify({"status": "error", "message": "지출결의 내역을 찾을 수 없습니다."}), 404
    if report['doc_status'] == '완료':
        conn.close()
        return jsonify({"status": "success", "message": "이미 결재완료 상태입니다."})

    cursor = conn.execute('''
        UPDATE expense_reports
        SET doc_status='완료', payment_status='지급대기', approver_1=?, approved_at=CURRENT_TIMESTAMP, updated_at=CURRENT_TIMESTAMP
        WHERE id=? AND doc_status != '완료'
    ''', (current_user, report_id))
    if cursor.rowcount == 0:
        conn.close()
        return jsonify({"status": "success", "message": "이미 결재완료 상태입니다."})
    conn.commit()
    conn.close()
    return jsonify({"status": "success", "message": "결재완료로 처리했습니다."})


@expense_bp.route('/api/report/<int:report_id>/reject', methods=['POST'])
def mark_rejected(report_id):
    ensure_expense_schema()
    if not _can_manage_expenses():
        return jsonify({"status": "error", "message": "반려 처리 권한이 없습니다."}), 403
    data = request.get_json(silent=True) or {}
    reject_reason = _clean_text(data.get('reason'))
    current_user = session.get('user_name', '')
    conn = get_db()
    report = conn.execute("SELECT * FROM expense_reports WHERE id=?", (report_id,)).fetchone()
    if not report:
        conn.close()
        return jsonify({"status": "error", "message": "지출결의 내역을 찾을 수 없습니다."}), 404

    memo = _clean_text(report['memo'] if 'memo' in report.keys() else '')
    if reject_reason:
        memo = (memo + "\n" if memo else "") + f"[반려사유] {reject_reason}"

    conn.execute('''
        UPDATE expense_reports
        SET doc_status='반려', payment_status='반려', approver_1=?, memo=?, updated_at=CURRENT_TIMESTAMP
        WHERE id=?
    ''', (current_user, memo, report_id))
    updated_report = conn.execute("SELECT * FROM expense_reports WHERE id=?", (report_id,)).fetchone()
    items = conn.execute("SELECT * FROM expense_items WHERE report_id=? ORDER BY row_no ASC, id ASC", (report_id,)).fetchall()
    mail_sent, mail_error = _send_expense_status_email(updated_report, '반려', reject_reason, items=items) if updated_report else (False, '보고서 없음')
    conn.commit()
    conn.close()
    message = "반려로 처리했습니다."
    if updated_report and updated_report['submitter_email']:
        message += " 메일 발송 완료." if mail_sent else f" 메일 발송 실패: {mail_error}"
    return jsonify({"status": "success", "message": message})


@expense_bp.route('/api/reports/delete', methods=['POST'])
def delete_reports():
    ensure_expense_schema()
    if not _can_manage_expenses():
        return jsonify({"status": "error", "message": "삭제 권한이 없습니다."}), 403
    data = request.get_json(silent=True) or {}
    ids = data.get('ids') or []
    ids = [int(v) for v in ids if str(v).isdigit()]
    if not ids:
        return jsonify({"status": "error", "message": "삭제할 지출결의서를 선택해주세요."}), 400

    conn = get_db()
    file_paths_to_delete = []
    deleted_reports = 0
    try:
        for report_id in ids:
            report = conn.execute("SELECT * FROM expense_reports WHERE id=?", (report_id,)).fetchone()
            if not report:
                continue
            for column in ('source_filepath', 'receipt_filepath'):
                if column in report.keys():
                    file_paths_to_delete.extend(_split_csv(report[column]))
            conn.execute("DELETE FROM expense_items WHERE report_id=?", (report_id,))
            conn.execute("DELETE FROM expense_reports WHERE id=?", (report_id,))
            deleted_reports += 1
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()
    deleted_files = _delete_file_paths(','.join(file_paths_to_delete))
    return jsonify({
        "status": "success",
        "message": f"{deleted_reports}건을 삭제했습니다. 첨부파일 {deleted_files}개도 함께 삭제했습니다."
    })


@expense_bp.route('/api/report/<int:report_id>/waiting', methods=['POST'])
def mark_waiting(report_id):
    ensure_expense_schema()
    if not _can_manage_expenses():
        return jsonify({"status": "error", "message": "지급 처리 권한이 없습니다."}), 403
    conn = get_db()
    report = conn.execute("SELECT * FROM expense_reports WHERE id=?", (report_id,)).fetchone()
    if not report:
        conn.close()
        return jsonify({"status": "error", "message": "지출결의 내역을 찾을 수 없습니다."}), 404
    if report['doc_status'] != '완료':
        conn.close()
        return jsonify({"status": "error", "message": "결재 완료 문서만 지급대기로 되돌릴 수 있습니다."}), 400

    conn.execute('''
        UPDATE expense_reports
        SET payment_status='지급대기', paid_at=NULL, paid_by=NULL, updated_at=CURRENT_TIMESTAMP
        WHERE id=?
    ''', (report_id,))
    conn.commit()
    conn.close()
    return jsonify({"status": "success", "message": "지급대기로 되돌렸습니다."})
