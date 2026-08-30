import io
import os
import sqlite3
import sys
import tempfile
import unittest
from datetime import date, timedelta
from email.mime.text import MIMEText
from email.utils import parseaddr
from pathlib import Path
from unittest import mock

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from flask import Flask, url_for
from openpyxl import load_workbook

from routes import document
from routes.database import ensure_certificate_schema


class CertificateManagementFlowTest(unittest.TestCase):
    def setUp(self):
        descriptor, self.db_path = tempfile.mkstemp(suffix=".db")
        os.close(descriptor)

        def test_db():
            connection = sqlite3.connect(self.db_path)
            connection.row_factory = sqlite3.Row
            return connection

        self.original_get_db = document.get_db
        self.original_seal_folder = document.CERT_SEAL_FOLDER
        self.original_logo_folder = document.CERT_LOGO_FOLDER
        self.original_pdf_folder = document.PDF_FOLDER
        self.asset_directory = tempfile.TemporaryDirectory()
        document.CERT_SEAL_FOLDER = os.path.join(self.asset_directory.name, "seals")
        document.CERT_LOGO_FOLDER = os.path.join(self.asset_directory.name, "logos")
        document.PDF_FOLDER = os.path.join(self.asset_directory.name, "pdfs")
        os.makedirs(document.CERT_SEAL_FOLDER, exist_ok=True)
        os.makedirs(document.CERT_LOGO_FOLDER, exist_ok=True)
        os.makedirs(document.PDF_FOLDER, exist_ok=True)
        document.get_db = test_db
        connection = test_db()
        connection.execute("""
            CREATE TABLE ai_mail_senders (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                owner_emp_no TEXT NOT NULL,
                label TEXT NOT NULL,
                email TEXT NOT NULL,
                encrypted_app_password TEXT NOT NULL,
                is_active INTEGER NOT NULL DEFAULT 1,
                last_tested_at DATETIME,
                last_test_status TEXT,
                last_test_error TEXT,
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                updated_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                provider TEXT NOT NULL DEFAULT 'gmail',
                smtp_host TEXT,
                smtp_port INTEGER,
                smtp_security TEXT,
                smtp_username TEXT
            )
        """)
        connection.execute("""
            CREATE TABLE users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                emp_no TEXT NOT NULL,
                name TEXT NOT NULL
            )
        """)
        connection.execute("""
            INSERT INTO users (emp_no, name) VALUES ('admin', '관리자테스터')
        """)
        ensure_certificate_schema(connection)
        self.sender_id = connection.execute("""
            INSERT INTO ai_mail_senders (
                owner_emp_no, label, email, encrypted_app_password,
                provider, smtp_host, smtp_port, smtp_security, smtp_username
            ) VALUES ('admin', '테스트 발송', 'sender@example.com', 'encrypted',
                      'gmail', 'smtp.gmail.com', 465, 'ssl', 'sender@example.com')
        """).lastrowid
        connection.commit()
        connection.close()

        self.app = Flask(
            __name__,
            template_folder=str(ROOT / "templates"),
            static_folder=str(ROOT / "static"),
        )
        self.app.secret_key = "certificate-management-test-secret"
        self.app.register_blueprint(document.document_bp, url_prefix="/document")
        self.client = self.app.test_client()
        with self.client.session_transaction() as session:
            session["emp_no"] = "admin"
            session["user_name"] = "admin"
            session["user_level"] = 1
            session["ai_mail_csrf_token"] = "csrf-test"

    def tearDown(self):
        document.get_db = self.original_get_db
        document.CERT_SEAL_FOLDER = self.original_seal_folder
        document.CERT_LOGO_FOLDER = self.original_logo_folder
        document.PDF_FOLDER = self.original_pdf_folder
        self.asset_directory.cleanup()
        try:
            os.remove(self.db_path)
        except FileNotFoundError:
            pass

    @property
    def headers(self):
        return {"X-CSRF-Token": "csrf-test"}

    def test_company_workgroup_link_and_request_snapshot(self):
        response = self.client.post(
            "/document/api/companies",
            data={
                "company_name": "테스트 주식회사",
                "representative_name": "홍대표",
                "business_number": "123-45-67890",
                "address": "서울시 테스트구",
                "phone": "02-123-4567",
                "seal": (io.BytesIO(b"seal-image"), "회사 도장 파일.png"),
                "logo": (io.BytesIO(b"logo-image"), "회사 로고 파일.png"),
            },
            headers=self.headers,
        )
        self.assertEqual(response.status_code, 200)
        company_id = response.get_json()["company_id"]

        response = self.client.post(
            "/document/api/workgroups",
            json={
                "name": "테스트 발급그룹",
                "company_id": company_id,
                "sender_id": self.sender_id,
                "allow_instructor": True,
                "allow_employee": True,
            },
            headers=self.headers,
        )
        self.assertEqual(response.status_code, 200)

        settings = self.client.get("/document/api/settings").get_json()
        company = settings["companies"][0]
        self.assertEqual(company["seal_filename"], "회사 도장 파일.png")
        self.assertEqual(company["logo_filename"], "회사 로고 파일.png")
        self.assertTrue(company["seal_path"].endswith(".png.sdf"))
        self.assertTrue(company["logo_path"].endswith(".png.sdf"))
        self.assertNotIn("회사 도장", os.path.basename(company["seal_path"]))
        old_logo_url = company["logo_url"]
        old_seal_url = company["seal_url"]

        response = self.client.post(
            f"/document/api/companies/{company_id}",
            data={
                "company_name": "테스트 주식회사",
                "representative_name": "홍대표",
                "business_number": "123-45-67890",
                "address": "서울시 테스트구",
                "phone": "02-123-4567",
                "seal": (io.BytesIO(b"new-seal-image"), "회사 도장 파일.png"),
                "logo": (io.BytesIO(b"new-logo-image"), "회사 로고 파일.png"),
            },
            headers=self.headers,
        )
        self.assertEqual(response.status_code, 200)
        settings = self.client.get("/document/api/settings").get_json()
        company = settings["companies"][0]
        self.assertNotEqual(company["logo_url"], old_logo_url)
        self.assertNotEqual(company["seal_url"], old_seal_url)
        self.assertIn("?v=", company["logo_url"])
        self.assertIn("?v=", company["seal_url"])

        group = settings["workgroups"][0]
        with self.client.session_transaction() as session:
            session.clear()

        auth_response = self.client.get(group["instructor_path"])
        self.assertEqual(auth_response.status_code, 200)
        auth_page = auth_response.get_data(as_text=True)
        self.assertIn("테스트 주식회사", auth_page)
        self.assertIn("02-123-4567", auth_page)
        self.assertIn(company["logo_url"], auth_page)

        with self.client.session_transaction() as session:
            session["certificate_apply_verified"] = document.CERTIFICATE_FORM_AUTH_TOKEN

        response = self.client.get(group["instructor_path"])
        self.assertEqual(response.status_code, 200)
        page = response.get_data(as_text=True)
        self.assertIn("테스트 주식회사", page)
        self.assertIn("02-123-4567", page)
        self.assertIn(company["logo_url"], page)

        logo_response = self.client.get(company["logo_url"])
        self.assertEqual(logo_response.status_code, 200)
        self.assertEqual(logo_response.data, b"new-logo-image")
        self.assertIn("no-store", logo_response.headers.get("Cache-Control", ""))
        logo_response.close()

        response = self.client.post(
            group["instructor_path"],
            data={
                "신청구분": "강사",
                "증명서종류": "강사활동증명서",
                "성명": "신청테스터",
                "이메일주소": "applicant@example.com",
            },
        )
        self.assertEqual(response.status_code, 200)
        with self.client.session_transaction() as session:
            session["certificate_apply_verified"] = document.CERTIFICATE_FORM_AUTH_TOKEN
        repeat_response = self.client.post(
            group["instructor_path"],
            data={
                "신청구분": "강사",
                "증명서종류": "강사활동증명서",
                "성명": "신청테스터",
                "이메일주소": "applicant@example.com",
            },
        )
        self.assertEqual(repeat_response.status_code, 200)
        connection = document.get_db()
        repeat_count = connection.execute("""
            SELECT COUNT(*) FROM certificate_requests
            WHERE applicant_name='신청테스터'
        """).fetchone()[0]
        row = connection.execute("""
            SELECT workgroup_id, company_id, workgroup_name, company_name
            FROM certificate_requests ORDER BY id DESC LIMIT 1
        """).fetchone()
        connection.close()
        self.assertEqual(repeat_count, 2)
        self.assertEqual(row["workgroup_id"], group["id"])
        self.assertEqual(row["company_id"], company_id)
        self.assertEqual(row["company_name"], "테스트 주식회사")
        self.assertEqual(row["workgroup_name"], "테스트 발급그룹")

    def test_excellent_instructor_roster_lookup_and_application(self):
        company_response = self.client.post(
            "/document/api/companies",
            data={
                "company_name": "우수강사 테스트 회사",
                "representative_name": "김대표",
                "phone": "070-456-1234",
            },
            headers=self.headers,
        )
        self.assertEqual(company_response.status_code, 200)
        company_id = company_response.get_json()["company_id"]

        group_response = self.client.post(
            "/document/api/workgroups",
            json={
                "name": "우수강사 전용그룹",
                "company_id": company_id,
                "sender_id": self.sender_id,
                "allow_instructor": False,
                "allow_employee": False,
                "allow_excellent_instructor": True,
            },
            headers=self.headers,
        )
        self.assertEqual(group_response.status_code, 200)
        settings = self.client.get("/document/api/settings").get_json()
        group = settings["workgroups"][0]
        self.assertTrue(group["allow_excellent_instructor"])
        self.assertIn("/document/apply-excellent/", group["excellent_instructor_path"])

        class FakeCell:
            def __init__(self, value):
                self.value = value

        class FakeWorksheet:
            headers = ["성명", "주민번호", "학교명", "직책", "강의과목"]

            def __init__(self, rows):
                self.rows = rows

            def iter_rows(self, min_row=1, max_row=None, values_only=False):
                if min_row == 1 and max_row == 1:
                    yield tuple(FakeCell(value) for value in self.headers)
                    return
                for row in self.rows:
                    yield row

        class FakeWorkbook:
            def __init__(self, rows):
                self.active = FakeWorksheet(rows)

            def close(self):
                pass

        today = date.today()
        valid_from = (today - timedelta(days=30)).isoformat()
        valid_until = (today + timedelta(days=30)).isoformat()
        first_rows = [
            ("홍길동", "900101-1234567", "가람초등학교", "방과후 강사", "바둑"),
            ("홍길동", "9001011234567", "나래초등학교", "늘봄학교 선택형 강사", "로봇코딩"),
        ]
        with mock.patch.object(
            document, "load_workbook", return_value=FakeWorkbook(first_rows)
        ):
            preview_response = self.client.post(
                "/document/api/excellent-instructors/preview",
                data={"file": (io.BytesIO(b"xlsx-preview"), "우수강사명단.xlsx")},
                headers=self.headers,
            )
        self.assertEqual(preview_response.status_code, 200)
        preview_data = preview_response.get_json()
        self.assertEqual(preview_data["count"], 2)
        self.assertEqual(preview_data["items"][0]["number"], 1)
        self.assertEqual(preview_data["items"][0]["school_name"], "가람초등학교")
        self.assertEqual(
            preview_data["items"][0]["resident_number_masked"], "900101-1******"
        )
        with mock.patch.object(
            document, "load_workbook", return_value=FakeWorkbook(first_rows)
        ):
            upload_response = self.client.post(
                "/document/api/excellent-instructors/upload",
                data={
                    "file": (io.BytesIO(b"xlsx"), "우수강사명단.xlsx"),
                    "group_name": "2026년 1차 우수강사",
                    "company_id": str(company_id),
                    "valid_from": valid_from,
                    "valid_until": valid_until,
                },
                headers=self.headers,
            )
        self.assertEqual(upload_response.status_code, 200)
        self.assertEqual(upload_response.get_json()["count"], 2)
        first_group_id = upload_response.get_json()["group_id"]

        next_valid_from = (today + timedelta(days=60)).isoformat()
        next_valid_until = (today + timedelta(days=90)).isoformat()
        second_rows = [
            (
                f"김강사{index:02d}", "910202-2234567", f"테스트{index:02d}초등학교",
                "방과후 강사", f"과목{index:02d}",
            )
            for index in range(1, 56)
        ]
        with mock.patch.object(
            document, "load_workbook", return_value=FakeWorkbook(second_rows)
        ):
            second_upload = self.client.post(
                "/document/api/excellent-instructors/upload",
                data={
                    "file": (io.BytesIO(b"xlsx-2"), "다음학기명단.xlsx"),
                    "group_name": "2026년 2차 우수강사",
                    "company_id": str(company_id),
                    "valid_from": next_valid_from,
                    "valid_until": next_valid_until,
                },
                headers=self.headers,
            )
        self.assertEqual(second_upload.status_code, 200)
        self.assertEqual(second_upload.get_json()["count"], 55)
        second_group_id = second_upload.get_json()["group_id"]

        settings = self.client.get("/document/api/settings").get_json()
        self.assertEqual(settings["excellent_roster_summary"]["group_count"], 2)
        self.assertEqual(settings["excellent_roster_summary"]["valid_group_count"], 1)
        self.assertEqual(settings["excellent_roster_summary"]["member_count"], 57)

        groups_response = self.client.get("/document/api/excellent-instructor-groups")
        self.assertEqual(groups_response.status_code, 200)
        groups = {item["id"]: item for item in groups_response.get_json()["groups"]}
        self.assertEqual(groups[first_group_id]["member_count"], 2)
        self.assertEqual(groups[first_group_id]["company_id"], company_id)
        self.assertEqual(groups[first_group_id]["company_name"], "우수강사 테스트 회사")
        self.assertTrue(groups[first_group_id]["is_valid"])
        self.assertEqual(groups[first_group_id]["created_by_name"], "관리자테스터")
        self.assertRegex(groups[first_group_id]["created_date"], r"^\d{4}-\d{2}-\d{2}$")
        self.assertEqual(groups[second_group_id]["member_count"], 55)
        self.assertTrue(groups[second_group_id]["is_upcoming"])

        roster_response = self.client.get(
            f"/document/api/excellent-instructor-groups/{first_group_id}/members"
        )
        self.assertEqual(roster_response.status_code, 200)
        roster_data = roster_response.get_json()
        self.assertEqual(roster_data["total"], 2)
        self.assertEqual(roster_data["items"][0]["resident_number_masked"], "900101-1******")
        first_member_id = roster_data["items"][0]["id"]

        paged_response = self.client.get(
            f"/document/api/excellent-instructor-groups/{second_group_id}/members"
        )
        self.assertEqual(paged_response.status_code, 200)
        paged_data = paged_response.get_json()
        self.assertEqual(paged_data["total"], 55)
        self.assertEqual(len(paged_data["items"]), 50)
        self.assertEqual(paged_data["total_pages"], 2)
        second_member_id = paged_data["items"][0]["id"]

        edit_member = self.client.patch(
            f"/document/api/excellent-instructors/{first_member_id}",
            json={
                "applicant_name": "홍길동",
                "resident_number": "900101-1234567",
                "school_name": "가람초등학교",
                "position": "우수 방과후 강사",
                "subject": "창의바둑",
            },
            headers=self.headers,
        )
        self.assertEqual(edit_member.status_code, 200)
        delete_member = self.client.delete(
            f"/document/api/excellent-instructors/{second_member_id}",
            headers=self.headers,
        )
        self.assertEqual(delete_member.status_code, 200)
        second_members = self.client.get(
            f"/document/api/excellent-instructor-groups/{second_group_id}/members"
        ).get_json()
        self.assertEqual(second_members["total"], 54)

        delete_group = self.client.delete(
            f"/document/api/excellent-instructor-groups/{second_group_id}",
            headers=self.headers,
        )
        self.assertEqual(delete_group.status_code, 200)
        remaining_groups = self.client.get(
            "/document/api/excellent-instructor-groups"
        ).get_json()["groups"]
        self.assertEqual([item["id"] for item in remaining_groups], [first_group_id])
        excellent_path = settings["workgroups"][0]["excellent_instructor_path"]

        with self.client.session_transaction() as session:
            session.clear()
        login_page = self.client.get(excellent_path)
        self.assertEqual(login_page.status_code, 200)
        self.assertIn("우수강사인증서 발급 신청", login_page.get_data(as_text=True))
        unauthorized_lookup = self.client.post(
            f"{excellent_path}/lookup",
            json={"name": "홍길동", "resident_number": "900101-1234567"},
        )
        self.assertEqual(unauthorized_lookup.status_code, 401)
        password_response = self.client.post(excellent_path, data={"password": "0070"})
        self.assertEqual(password_response.status_code, 302)

        form_page = self.client.get(excellent_path)
        self.assertEqual(form_page.status_code, 200)
        form_html = form_page.get_data(as_text=True)
        self.assertIn("신청 대상 조회", form_html)
        self.assertIn("우수강사인증서 발급 신청", form_html)

        with self.client.session_transaction() as session:
            session["emp_no"] = "admin"
            session["user_name"] = "admin"
            session["user_level"] = 1
            session["ai_mail_csrf_token"] = "csrf-test"
        expired_response = self.client.patch(
            f"/document/api/excellent-instructor-groups/{first_group_id}",
            json={
                "name": "2026년 1차 우수강사",
                "valid_from": (today - timedelta(days=60)).isoformat(),
                "valid_until": (today - timedelta(days=31)).isoformat(),
            },
            headers=self.headers,
        )
        self.assertEqual(expired_response.status_code, 200)
        with self.client.session_transaction() as session:
            session.clear()
            session["certificate_apply_excellent_verified"] = document.CERTIFICATE_FORM_AUTH_TOKEN
        expired_lookup = self.client.post(
            f"{excellent_path}/lookup",
            json={"name": "홍길동", "resident_number": "900101-1234567"},
        )
        self.assertEqual(expired_lookup.status_code, 200)
        expired_data = expired_lookup.get_json()
        self.assertEqual(expired_data["summary"]["expired"], 1)
        self.assertEqual(expired_data["groups"][0]["application_status"], "expired")
        self.assertFalse(expired_data["groups"][0]["selectable"])
        with self.client.session_transaction() as session:
            session["emp_no"] = "admin"
            session["user_name"] = "admin"
            session["user_level"] = 1
            session["ai_mail_csrf_token"] = "csrf-test"
        restore_response = self.client.patch(
            f"/document/api/excellent-instructor-groups/{first_group_id}",
            json={
                "name": "2026년 1차 우수강사 수정",
                "valid_from": valid_from,
                "valid_until": valid_until,
            },
            headers=self.headers,
        )
        self.assertEqual(restore_response.status_code, 200)

        third_rows = [
            ("홍길동", "900101-1234567", "다솜초등학교", "방과후 강사", "창의수학"),
        ]
        with mock.patch.object(
            document, "load_workbook", return_value=FakeWorkbook(third_rows)
        ):
            third_upload = self.client.post(
                "/document/api/excellent-instructors/upload",
                data={
                    "file": (io.BytesIO(b"xlsx-3"), "별도그룹명단.xlsx"),
                    "group_name": "2026년 별도 우수강사",
                    "company_id": str(company_id),
                    "valid_from": valid_from,
                    "valid_until": valid_until,
                },
                headers=self.headers,
            )
        self.assertEqual(third_upload.status_code, 200)
        third_group_id = third_upload.get_json()["group_id"]
        with self.client.session_transaction() as session:
            session.clear()
            session["certificate_apply_excellent_verified"] = document.CERTIFICATE_FORM_AUTH_TOKEN

        lookup_response = self.client.post(
            f"{excellent_path}/lookup",
            json={"name": "홍길동", "resident_number": "900101-1234567"},
        )
        self.assertEqual(lookup_response.status_code, 200)
        lookup_data = lookup_response.get_json()
        groups_by_id = {group["id"]: group for group in lookup_data["groups"]}
        self.assertEqual(groups_by_id[first_group_id]["application_status"], "available")
        self.assertEqual(groups_by_id[third_group_id]["application_status"], "available")
        items = groups_by_id[first_group_id]["items"]
        self.assertEqual(len(items), 2)
        self.assertEqual(
            {item["school_name"] for item in items},
            {"가람초등학교", "나래초등학교"},
        )

        with mock.patch.object(document, "send_admin_alert"):
            multiple_apply_response = self.client.post(
                excellent_path,
                data={
                    "대상항목": [str(item["id"]) for item in items],
                    "성명": "홍길동",
                    "주민번호": "900101-1234567",
                    "용도": "기관제출용",
                    "이메일주소": "excellent@example.com",
                    "자택주소": "서울시 테스트구",
                    "근무시작일": "2025-03-01",
                    "종료일선택": "현재까지",
                },
            )
        self.assertEqual(multiple_apply_response.status_code, 400)
        self.assertIn(
            "한번에 한 건씩 발급 가능합니다",
            multiple_apply_response.get_data(as_text=True),
        )

        with mock.patch.object(document, "send_admin_alert"):
            apply_response = self.client.post(
                excellent_path,
                data={
                    "대상항목": [str(items[0]["id"])],
                    "성명": "홍길동",
                    "주민번호": "900101-1234567",
                    "용도": "기관제출용",
                    "이메일주소": "excellent@example.com",
                    "자택주소": "서울시 테스트구",
                    "근무시작일": "2025-03-01",
                    "종료일선택": "현재까지",
                },
            )
        self.assertEqual(apply_response.status_code, 200)
        connection = document.get_db()
        saved = connection.execute("""
            SELECT id, applicant_type, certificate_type, applicant_name,
                   resident_number, workplace, subject_or_duty, position,
                   work_end_date, company_name
            FROM certificate_requests ORDER BY id DESC LIMIT 1
        """).fetchone()
        first_request_id = saved["id"]
        request_groups = connection.execute("""
            SELECT group_id FROM excellent_instructor_request_groups
            WHERE request_id=? ORDER BY group_id
        """, (first_request_id,)).fetchall()
        connection.close()
        self.assertEqual([row["group_id"] for row in request_groups], [first_group_id])
        self.assertEqual(saved["applicant_type"], "강사")
        self.assertEqual(saved["certificate_type"], "우수강사인증서")
        self.assertEqual(saved["resident_number"], "900101-1234567")
        self.assertEqual(saved["workplace"], "가람초등학교")
        self.assertEqual(saved["subject_or_duty"], "창의바둑")
        self.assertEqual(saved["work_end_date"], "현재까지")
        self.assertEqual(saved["company_name"], "우수강사 테스트 회사")

        with self.client.session_transaction() as session:
            session["certificate_apply_excellent_verified"] = document.CERTIFICATE_FORM_AUTH_TOKEN
        pending_lookup = self.client.post(
            f"{excellent_path}/lookup",
            json={"name": "홍길동", "resident_number": "900101-1234567"},
        )
        self.assertEqual(pending_lookup.status_code, 200)
        pending_groups = {
            group["id"]: group for group in pending_lookup.get_json()["groups"]
        }
        self.assertEqual(pending_groups[first_group_id]["application_status"], "pending")
        self.assertFalse(pending_groups[first_group_id]["selectable"])
        self.assertEqual(pending_groups[third_group_id]["application_status"], "available")

        with mock.patch.object(document, "send_admin_alert"):
            duplicate_response = self.client.post(
                excellent_path,
                data={
                    "대상항목": [str(items[0]["id"])],
                    "성명": "홍길동",
                    "주민번호": "900101-1234567",
                    "용도": "기관제출용",
                    "이메일주소": "excellent@example.com",
                    "자택주소": "서울시 테스트구",
                    "근무시작일": "2025-03-01",
                    "종료일선택": "현재까지",
                },
            )
        self.assertEqual(duplicate_response.status_code, 409)
        self.assertIn("이미 신청중", duplicate_response.get_data(as_text=True))

        with self.client.session_transaction() as session:
            session["emp_no"] = "admin"
            session["user_name"] = "admin"
            session["user_level"] = 1
            session["ai_mail_csrf_token"] = "csrf-test"
        pending_roster = self.client.get(
            f"/document/api/excellent-instructor-groups/{first_group_id}/members"
        ).get_json()
        self.assertEqual(
            {item["application_status"] for item in pending_roster["items"]},
            {"신청중"},
        )
        connection = document.get_db()
        connection.execute("""
            UPDATE certificate_requests
            SET status='발급완료', issued_date='2026-08-28',
                issue_number='제26-9999호'
            WHERE id=?
        """, (first_request_id,))
        connection.commit()
        connection.close()

        issued_roster = self.client.get(
            f"/document/api/excellent-instructor-groups/{first_group_id}/members"
        ).get_json()
        self.assertEqual(
            {item["application_status"] for item in issued_roster["items"]},
            {"발급완료"},
        )

        with self.client.session_transaction() as session:
            session.clear()
            session["certificate_apply_excellent_verified"] = document.CERTIFICATE_FORM_AUTH_TOKEN
        issued_lookup = self.client.post(
            f"{excellent_path}/lookup",
            json={"name": "홍길동", "resident_number": "900101-1234567"},
        )
        self.assertEqual(issued_lookup.status_code, 200)
        issued_groups = {
            group["id"]: group for group in issued_lookup.get_json()["groups"]
        }
        self.assertEqual(issued_groups[first_group_id]["application_status"], "issued")
        self.assertEqual(
            issued_groups[first_group_id]["history"]["issue_number"], "제26-9999호"
        )
        self.assertEqual(issued_groups[third_group_id]["application_status"], "available")

        third_items = issued_groups[third_group_id]["items"]
        with mock.patch.object(document, "send_admin_alert"):
            different_group_response = self.client.post(
                excellent_path,
                data={
                    "대상항목": [str(item["id"]) for item in third_items],
                    "성명": "홍길동",
                    "주민번호": "900101-1234567",
                    "용도": "기관제출용",
                    "이메일주소": "excellent@example.com",
                    "자택주소": "서울시 테스트구",
                    "근무시작일": "2025-03-01",
                    "종료일선택": "현재까지",
                },
            )
        self.assertEqual(different_group_response.status_code, 200)

        instructor_template = (ROOT / "templates" / "certificate" / "form.html").read_text(
            encoding="utf-8"
        )
        self.assertNotIn('<option value="우수강사인증서">', instructor_template)
        sample_path = ROOT / "static" / "templates" / "우수강사명단등록.xlsx"
        sample_workbook = load_workbook(sample_path, read_only=True, data_only=True)
        try:
            self.assertEqual(sample_workbook.sheetnames, ["등록양식", "작성안내"])
            self.assertEqual(
                [cell.value for cell in sample_workbook["등록양식"][1]],
                ["학교명", "성명", "주민번호", "직책", "강의과목"],
            )
        finally:
            sample_workbook.close()
        for template_name in (
            "excellent_form.html", "form_login.html", "settings.html", "admin.html",
        ):
            source = (ROOT / "templates" / "certificate" / template_name).read_text(
                encoding="utf-8"
            )
            self.app.jinja_env.parse(source)
        settings_source = (
            ROOT / "templates" / "certificate" / "settings.html"
        ).read_text(encoding="utf-8")
        self.assertIn('id="closeRosterMembers"', settings_source)
        self.assertIn('id="rosterFileName"', settings_source)
        self.assertIn('id="rosterDropZone"', settings_source)
        self.assertIn("previewRosterSelection()", settings_source)
        self.assertNotIn('id="rosterPreviewBtn"', settings_source)
        self.assertIn('data-tab="roster"', settings_source)
        self.assertIn('class="cs-panel" id="panel-roster"', settings_source)
        self.assertNotIn('id="rosterModal"', settings_source)
        self.assertIn("document.body.appendChild(modal)", settings_source)
        self.assertIn("tbody tr:nth-child(odd) td", settings_source)

    def test_excellent_rosters_and_history_are_isolated_by_company(self):
        company_ids = {}
        for company_name in ("사단법인 새담청소년교육문화원", "(주)에듀탑스쿨"):
            response = self.client.post(
                "/document/api/companies",
                data={
                    "company_name": company_name,
                    "representative_name": "테스트대표",
                },
                headers=self.headers,
            )
            self.assertEqual(response.status_code, 200)
            company_ids[company_name] = response.get_json()["company_id"]

        for company_name, company_id in company_ids.items():
            response = self.client.post(
                "/document/api/workgroups",
                json={
                    "name": f"{company_name} 우수강사",
                    "company_id": company_id,
                    "sender_id": self.sender_id,
                    "allow_instructor": False,
                    "allow_employee": False,
                    "allow_excellent_instructor": True,
                },
                headers=self.headers,
            )
            self.assertEqual(response.status_code, 200)

        settings = self.client.get("/document/api/settings").get_json()
        paths = {
            item["company_id"]: item["excellent_instructor_path"]
            for item in settings["workgroups"]
        }
        today = date.today().isoformat()
        connection = document.get_db()
        group_ids = {}
        member_ids = {}
        for company_name, school_name in (
            ("사단법인 새담청소년교육문화원", "새담초등학교"),
            ("(주)에듀탑스쿨", "에듀탑초등학교"),
        ):
            company_id = company_ids[company_name]
            group_id = connection.execute("""
                INSERT INTO excellent_instructor_roster_groups (
                    name, company_id, valid_from, valid_until, created_by
                ) VALUES (?, ?, ?, ?, 'admin')
            """, (f"{company_name} 명단", company_id, today, today)).lastrowid
            member_id = connection.execute("""
                INSERT INTO excellent_instructor_eligibility (
                    group_id, applicant_name, resident_number,
                    resident_number_normalized, school_name, position,
                    subject, uploaded_by
                ) VALUES (?, '동일강사', '900101-1234567', '9001011234567',
                          ?, '방과후 강사', '바둑', 'admin')
            """, (group_id, school_name)).lastrowid
            group_ids[company_id] = group_id
            member_ids[company_id] = member_id
        connection.commit()
        connection.close()

        lookups = {}
        for company_name, company_id in company_ids.items():
            response = self.client.post(
                f"{paths[company_id]}/lookup",
                json={"name": "동일강사", "resident_number": "900101-1234567"},
            )
            self.assertEqual(response.status_code, 200)
            lookups[company_id] = response.get_json()
            self.assertEqual(
                [group["id"] for group in lookups[company_id]["groups"]],
                [group_ids[company_id]],
            )

        saedam_id = company_ids["사단법인 새담청소년교육문화원"]
        edutop_id = company_ids["(주)에듀탑스쿨"]
        self.assertEqual(
            lookups[saedam_id]["groups"][0]["items"][0]["school_name"],
            "새담초등학교",
        )
        self.assertEqual(
            lookups[edutop_id]["groups"][0]["items"][0]["school_name"],
            "에듀탑초등학교",
        )

        cross_company_apply = self.client.post(
            paths[edutop_id],
            data={
                "대상항목": str(member_ids[saedam_id]),
                "성명": "동일강사",
                "주민번호": "900101-1234567",
                "용도": "기관제출용",
                "이메일주소": "same@example.com",
                "자택주소": "서울시 테스트구",
                "근무시작일": "2025-03-01",
                "종료일선택": "현재까지",
            },
        )
        self.assertEqual(cross_company_apply.status_code, 400)
        self.assertIn(
            "선택한 신청 대상 정보를 확인할 수 없습니다",
            cross_company_apply.get_data(as_text=True),
        )

    def test_issued_pdf_url_does_not_expose_personal_filename(self):
        filename = "제26-0006호_이보연.pdf"
        pdf_path = os.path.join(document.PDF_FOLDER, filename)
        with open(pdf_path, "wb") as pdf_file:
            pdf_file.write(b"test-pdf")

        connection = document.get_db()
        cursor = connection.execute(
            """
            INSERT INTO certificate_requests (
                applied_date, applicant_type, certificate_type, applicant_name,
                status, issued_date, issue_number, filename
            ) VALUES (?, ?, ?, ?, '발급완료', ?, ?, ?)
            """,
            (
                "2026-08-27", "강사", "경력증명서", "이보연",
                "2026-08-27", "제26-0006호", filename,
            ),
        )
        request_id = cursor.lastrowid
        connection.commit()
        connection.close()

        with self.app.test_request_context():
            pdf_url = url_for("document.serve_pdf", idx=request_id)
        self.assertEqual(pdf_url, f"/document/pdf/{request_id}")
        self.assertNotIn(filename, pdf_url)

        template = (ROOT / "templates" / "certificate" / "admin.html").read_text(
            encoding="utf-8"
        )
        self.assertIn("url_for('document.serve_pdf', idx=item.index)", template)
        self.assertNotIn("url_for('document.serve_pdf', filename=item.파일명)", template)

        response = self.client.get(f"/document/pdf/{request_id}")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data, b"test-pdf")
        self.assertEqual(
            response.headers.get("Content-Disposition"),
            "inline; filename=\"certificate.pdf\"; filename*=UTF-8''certificate.pdf",
        )
        self.assertEqual(self.client.get("/document/pdf/999999").status_code, 404)

    def test_zeptomail_uses_encoded_from_header_and_explicit_envelope_sender(self):
        class FakeSmtp:
            def __init__(self):
                self.message = None
                self.from_addr = None
                self.closed = False

            def send_message(self, message, from_addr=None, **_kwargs):
                self.message = message
                self.from_addr = from_addr

            def quit(self):
                self.closed = True

            def close(self):
                self.closed = True

        smtp = FakeSmtp()
        sender = {
            "provider": "zeptomail",
            "label": "한글 증명서 발송 <certificate@saedam.org>",
            "email": "certificate@saedam.org",
        }
        message = MIMEText("테스트", "plain", "utf-8")
        message["From"] = document._sender_from_header(sender)
        message["To"] = "recipient@example.com"
        message["Subject"] = "증명서 테스트"

        with mock.patch.object(document, "_smtp_login_for_sender", return_value=smtp), \
             mock.patch.object(document, "_verify_smtp_sender"):
            document._send_registered_message(sender, message)

        self.assertEqual(smtp.from_addr, "certificate@saedam.org")
        self.assertEqual(parseaddr(str(smtp.message["From"]))[1], "certificate@saedam.org")
        self.assertTrue(smtp.closed)

    def test_legacy_excellent_roster_is_migrated_to_a_group(self):
        descriptor, legacy_db_path = tempfile.mkstemp(suffix=".db")
        os.close(descriptor)
        connection = sqlite3.connect(legacy_db_path)
        connection.row_factory = sqlite3.Row
        try:
            connection.execute("""
                CREATE TABLE excellent_instructor_eligibility (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    applicant_name TEXT NOT NULL,
                    resident_number TEXT NOT NULL,
                    resident_number_normalized TEXT NOT NULL,
                    school_name TEXT NOT NULL,
                    position TEXT NOT NULL,
                    subject TEXT NOT NULL,
                    uploaded_by TEXT NOT NULL DEFAULT '',
                    created_at DATETIME DEFAULT CURRENT_TIMESTAMP
                )
            """)
            connection.execute("""
                CREATE TABLE excellent_instructor_roster_settings (
                    id INTEGER PRIMARY KEY CHECK (id = 1),
                    valid_from TEXT NOT NULL DEFAULT '',
                    valid_until TEXT NOT NULL DEFAULT '',
                    updated_by TEXT NOT NULL DEFAULT '',
                    updated_at DATETIME DEFAULT CURRENT_TIMESTAMP
                )
            """)
            connection.execute("""
                INSERT INTO excellent_instructor_roster_settings (
                    id, valid_from, valid_until
                ) VALUES (1, '2026-01-01', '2026-12-31')
            """)
            connection.execute("""
                INSERT INTO excellent_instructor_eligibility (
                    applicant_name, resident_number, resident_number_normalized,
                    school_name, position, subject, uploaded_by
                ) VALUES (
                    '기존강사', '900101-1234567', '9001011234567',
                    '기존초등학교', '방과후 강사', '바둑', 'admin'
                )
            """)
            ensure_certificate_schema(connection)
            legacy_request_id = connection.execute("""
                INSERT INTO certificate_requests (
                    applied_date, applicant_type, certificate_type,
                    applicant_name, resident_number, workplace,
                    subject_or_duty, status
                ) VALUES (
                    '2026-06-01', '강사', '우수강사인증서',
                    '기존강사', '900101-1234567', '기존초등학교',
                    '바둑', '발급완료'
                )
            """).lastrowid
            connection.execute("""
                DELETE FROM certificate_schema_meta
                WHERE key='excellent_request_group_backfill_v1'
            """)
            ensure_certificate_schema(connection)
            migrated = connection.execute("""
                SELECT e.group_id, e.is_active, e.updated_at,
                       g.name, g.valid_from, g.valid_until
                FROM excellent_instructor_eligibility e
                JOIN excellent_instructor_roster_groups g ON g.id=e.group_id
            """).fetchone()
            group_count = connection.execute(
                "SELECT COUNT(*) FROM excellent_instructor_roster_groups"
            ).fetchone()[0]
            self.assertEqual(group_count, 1)
            self.assertEqual(migrated["name"], "기존 등록 명단")
            self.assertEqual(migrated["valid_from"], "2026-01-01")
            self.assertEqual(migrated["valid_until"], "2026-12-31")
            self.assertEqual(migrated["is_active"], 1)
            self.assertTrue(migrated["updated_at"])
            mapped = connection.execute("""
                SELECT request_id, group_id, applicant_name,
                       resident_number_normalized
                FROM excellent_instructor_request_groups
                WHERE request_id=?
            """, (legacy_request_id,)).fetchone()
            self.assertIsNotNone(mapped)
            self.assertEqual(mapped["group_id"], migrated["group_id"])
            self.assertEqual(mapped["applicant_name"], "기존강사")
            self.assertEqual(mapped["resident_number_normalized"], "9001011234567")
        finally:
            connection.close()
            try:
                os.remove(legacy_db_path)
            except FileNotFoundError:
                pass

if __name__ == "__main__":
    unittest.main()
