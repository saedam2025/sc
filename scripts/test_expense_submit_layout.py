import os
import sys
import tempfile
from io import BytesIO


PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, PROJECT_ROOT)

from app import app
from routes.database import get_db
from routes.menu_access import (
    MENU_CATALOG,
    SCHOOL_CENTER_SHARED_ACTION_MENUS,
    SCHOOL_WORKSPACE_CATEGORY_MENU_KEYS,
    resolve_request_menu,
)
from openpyxl import load_workbook
from werkzeug.datastructures import FileStorage
import routes.expense as expense_routes
from routes.secure_files import decode_filename_token, temporary_decrypted_path


def login_admin(client):
    with client.session_transaction() as session:
        session['emp_no'] = 'admin'
        session['user_name'] = 'admin'
        session['user_level'] = 1


with app.app_context():
    conn = get_db()
    school = conn.execute(
        """
        SELECT access_key, school_name
        FROM schools
        WHERE COALESCE(is_active, 1) = 1
        ORDER BY id ASC
        LIMIT 1
        """
    ).fetchone()
    conn.close()

assert school is not None, '테스트할 활성 학교가 없습니다.'
assert expense_routes.RECEIPT_ALLOWED_EXTENSIONS == {'.jpg', '.png', '.gif'}

pdf_receipt = FileStorage(stream=BytesIO(b'%PDF-1.4'), filename='receipt.pdf')
pdf_error = expense_routes._validate_uploaded_files(
    [pdf_receipt],
    expense_routes.MAX_RECEIPT_FILES,
    expense_routes.RECEIPT_ALLOWED_EXTENSIONS,
    '영수증 증빙파일',
)
assert '등록할 수 없는 파일 형식' in pdf_error

fake_image = FileStorage(stream=BytesIO(b'not-an-image'), filename='receipt.jpg')
assert '올바른 이미지가 아닙니다' in expense_routes._validate_receipt_images([fake_image])

client = app.test_client()
login_admin(client)

management_response = client.get('/expense/')
assert management_response.status_code == 200, management_response.get_data(as_text=True)[:1000]
management_html = management_response.get_data(as_text=True)
assert 'href="/expense/submit" class="expense-link-btn"' in management_html
assert 'href="/expense/submit" class="expense-link-btn" target="_blank"' not in management_html
assert 'href="/expense/submit/instructor" class="expense-link-btn" target="_blank" rel="noopener"' in management_html
assert 'id="copyInstructorExpenseLink"' in management_html

submit_response = client.get('/expense/submit')
assert submit_response.status_code == 200, submit_response.get_data(as_text=True)[:1000]
submit_html = submit_response.get_data(as_text=True)
assert '<header class="navbar">' in submit_html
assert submit_html.count('id="expenseSubmitForm"') == 1
assert 'class="submit-page embedded"' not in submit_html
assert '본부용 지출결의서 전송' in submit_html
assert 'name="expense_submit_channel" value="headquarters"' in submit_html
assert '지출결의관리페이지' in submit_html
assert 'href="/expense/"' in submit_html
assert '<option value="" selected disabled>선택하세요</option>' in submit_html
assert 'id="expenseManager" name="expense_manager" required placeholder="성명" value=""' in submit_html
assert 'id="submitterEmail" name="submitter_email" type="email" required' in submit_html
assert 'id="paymentAccount" name="payment_account" required' in submit_html
assert 'placeholder="은행명 / 계좌번호 / 예금주" value=""' in submit_html
assert 'id="receiptFiles" name="receipt_files" type="file"' in submit_html
assert 'accept=".jpg,.png,.gif"' in submit_html
assert 'JPG, PNG, GIF 이미지' in submit_html
assert '.pdf,.jpg' not in submit_html
assert 'multiple hidden>' in submit_html
assert '.upload-grid { display: grid; grid-template-columns: repeat(2, minmax(0, 1fr)); gap: 12px; margin: 16px 0; padding: 0 25px; }' in submit_html
assert 'name="expense_excel" type="file" accept=".xlsx,.xlsm,.csv" hidden required' not in submit_html

center_submit_response = client.get('/expense/submit/center')
assert center_submit_response.status_code == 200, center_submit_response.get_data(as_text=True)[:1000]
center_submit_html = center_submit_response.get_data(as_text=True)
assert '<header class="navbar">' not in center_submit_html
assert center_submit_html.count('id="expenseSubmitForm"') == 1
assert '센터장용 지출결의서 전송' in center_submit_html
assert 'name="expense_submit_channel" value="center"' in center_submit_html
assert "formData.append('expense_submit_channel', submitChannel);" in center_submit_html
assert "missingFiles.push('엑셀파일이 첨부되지 않았습니다.');" in center_submit_html
assert "missingFiles.push('영수증 증빙 이미지가 첨부되지 않았습니다.');" in center_submit_html
assert 'color: var(--submit-primary)' in center_submit_html
assert 'color-mix(in srgb, var(--submit-primary) 42%' in center_submit_html
assert '<option value="학교" selected>학교</option>' in center_submit_html
assert '<option value="" disabled>선택하세요</option>' in center_submit_html
assert '지출결의관리페이지' not in center_submit_html
assert '지출결의서 전송 사용법' in center_submit_html
assert "입력된 샘플형식대로 지출내역을 작성합니다." in center_submit_html
assert '추후 처리결과도 이메일로 안내됩니다.' in center_submit_html
assert 'class="submit-guide-visual" aria-hidden="true"' in center_submit_html
assert 'fa-book-open-reader' in center_submit_html
assert 'fa-lightbulb' in center_submit_html

instructor_submit_response = client.get('/expense/submit/instructor')
assert instructor_submit_response.status_code == 200, instructor_submit_response.get_data(as_text=True)[:1000]
instructor_login_html = instructor_submit_response.get_data(as_text=True)
assert '<header class="navbar">' not in instructor_login_html
assert '강사용 전송페이지 비밀번호를 입력해주세요.' in instructor_login_html
assert 'id="expenseSubmitForm"' not in instructor_login_html

bad_password_response = client.post('/expense/submit/instructor', data={'password': 'wrong'})
assert bad_password_response.status_code == 401
assert '비밀번호가 올바르지 않습니다.' in bad_password_response.get_data(as_text=True)

password_response = client.post(
    '/expense/submit/instructor',
    data={'password': '0070'},
    follow_redirects=True,
)
assert password_response.status_code == 200, password_response.get_data(as_text=True)[:1000]
instructor_submit_html = password_response.get_data(as_text=True)
assert '<header class="navbar">' not in instructor_submit_html
assert instructor_submit_html.count('id="expenseSubmitForm"') == 1
assert '강사용 지출결의서 전송' in instructor_submit_html
assert 'name="expense_submit_channel" value="instructor"' in instructor_submit_html
assert '<option value="" selected disabled>선택하세요</option>' in instructor_submit_html
assert '지출결의관리페이지' not in instructor_submit_html
assert '지출결의서 전송 사용법' in instructor_submit_html

# 인트라넷 로그인이 전혀 없는 강사도 전용 비밀번호만으로 폼과
# 엑셀 미리보기 API에 접근할 수 있어야 한다.
public_client = app.test_client()
public_login_response = public_client.get('/expense/submit/instructor')
assert public_login_response.status_code == 200
assert '강사용 전송페이지 비밀번호를 입력해주세요.' in public_login_response.get_data(as_text=True)
public_form_response = public_client.post(
    '/expense/submit/instructor',
    data={'password': '0070'},
    follow_redirects=True,
)
assert public_form_response.status_code == 200
assert 'id="expenseSubmitForm"' in public_form_response.get_data(as_text=True)
public_preview_response = public_client.post('/expense/api/preview', data={})
assert public_preview_response.status_code == 400
assert public_preview_response.is_json

school_response = client.get(f"/school/{school['access_key']}?category=community")
assert school_response.status_code == 200, school_response.get_data(as_text=True)[:1000]
school_html = school_response.get_data(as_text=True)
assert 'href="/expense/submit/center" target="_blank" rel="noopener"' in school_html
assert 'onclick="return openSchoolCategoryWindow(event, this.href);"' in school_html
assert 'class="submit-page embedded"' not in school_html
assert 'id="expenseSubmitForm"' not in school_html
assert '본부공지사항' in school_html

legacy_expense_response = client.get(f"/school/{school['access_key']}?category=expense")
assert legacy_expense_response.status_code == 302
assert f"/school/{school['access_key']}?category=community" in legacy_expense_response.headers['Location']

expected_center_menu_keys = set(SCHOOL_WORKSPACE_CATEGORY_MENU_KEYS.values())
assert len(expected_center_menu_keys) == 2
assert expected_center_menu_keys.issubset(MENU_CATALOG)
expected_shared_action_keys = set(SCHOOL_CENTER_SHARED_ACTION_MENUS.values())
assert len(expected_shared_action_keys) == 5
assert expected_shared_action_keys.issubset(MENU_CATALOG)
assert resolve_request_menu('/expense/submit/center') == SCHOOL_WORKSPACE_CATEGORY_MENU_KEYS['expense']
assert resolve_request_menu('/expense/submit/instructor') is None

permissions_response = client.get('/admin/menu-permissions')
assert permissions_response.status_code == 200, permissions_response.get_data(as_text=True)[:1000]
permissions_html = permissions_response.get_data(as_text=True)
for menu_key in expected_center_menu_keys | expected_shared_action_keys:
    assert f'name="{menu_key}"' in permissions_html
assert '[센터장] 수강안내문~만족도조사 (8개 메뉴 일괄)' in permissions_html
for action_label in ('접근', '읽기', '쓰기', '삭제', '댓글'):
    assert f'[센터장] 본부공지사항·자료실 - {action_label}' in permissions_html

template_path = os.path.join(PROJECT_ROOT, '-== 참고자료', '지출결의서_기본양식.xlsx')
blank_items, blank_errors = expense_routes.parse_expense_file_with_errors(template_path)
assert blank_items == []
assert blank_errors == []

blank_preview_response = client.post(
    '/expense/api/preview',
    data={'expense_excel': (open(template_path, 'rb'), '지출결의서_기본양식.xlsx')},
    content_type='multipart/form-data',
)
assert blank_preview_response.status_code == 400
assert '엑셀에 입력된 지출내역이 없습니다.' in blank_preview_response.get_json()['message']

sample_path = os.path.join(PROJECT_ROOT, '-== 메뉴얼', '예당초_지출결의서(2026-06).xlsx')
sample_items, sample_errors = expense_routes.parse_expense_file_with_errors(sample_path)
assert len(sample_items) > 0
assert sample_errors == []

# 저장 시 암호화된 엑셀을 다시 복호화하여 읽는 실제 전송 경로도 검증한다.
original_upload_folder = expense_routes.UPLOAD_FOLDER
try:
    with tempfile.TemporaryDirectory(prefix='saedam-expense-roundtrip-') as upload_folder:
        expense_routes.UPLOAD_FOLDER = upload_folder
        with open(sample_path, 'rb') as source:
            saved = expense_routes._save_regular_uploaded_files([
                FileStorage(stream=source, filename='예당초_지출결의서(2026-06).xlsx')
            ])
        display_name = decode_filename_token(saved[0][0])
        with temporary_decrypted_path(saved[0][1], display_name) as decrypted_path:
            roundtrip_items, roundtrip_errors = expense_routes.parse_expense_file_with_errors(decrypted_path)
        assert len(roundtrip_items) == len(sample_items)
        assert roundtrip_errors == []
finally:
    expense_routes.UPLOAD_FOLDER = original_upload_folder

missing_account_response = client.post('/expense/submit', data={
    'expense_org_type': '본사',
    'expense_manager': '테스트',
    'submitter_email': 'test@example.com',
    'expense_kind': '교구재비',
    'payment_account': '',
})
assert missing_account_response.status_code == 400
assert missing_account_response.get_json()['message'] == '지급계좌번호를 입력해주세요.'

missing_excel_response = client.post('/expense/submit', data={
    'expense_org_type': '본사',
    'expense_manager': '테스트',
    'submitter_email': 'test@example.com',
    'expense_kind': '교구재비',
    'payment_account': '테스트은행 1234 테스트',
})
assert missing_excel_response.status_code == 400
assert missing_excel_response.get_json()['message'] == '엑셀파일이 첨부되지 않았습니다.'

with open(sample_path, 'rb') as source:
    missing_receipt_response = client.post(
        '/expense/submit',
        data={
            'expense_org_type': '본사',
            'expense_manager': '테스트',
            'submitter_email': 'test@example.com',
            'expense_kind': '교구재비',
            'payment_account': '테스트은행 1234 테스트',
            'expense_excel': (source, '예당초_지출결의서(2026-06).xlsx'),
        },
        content_type='multipart/form-data',
    )
assert missing_receipt_response.status_code == 400
assert missing_receipt_response.get_json()['message'] == '영수증 증빙 이미지가 첨부되지 않았습니다.'

with open(sample_path, 'rb') as source:
    invalid_receipt_response = client.post(
        '/expense/submit',
        data={
            'expense_org_type': '본사',
            'expense_manager': '테스트',
            'submitter_email': 'test@example.com',
            'expense_kind': '교구재비',
            'payment_account': '테스트은행 1234 테스트',
            'expense_excel': (source, '예당초_지출결의서(2026-06).xlsx'),
            'receipt_files': (BytesIO(b'%PDF-1.4'), 'receipt.pdf'),
        },
        content_type='multipart/form-data',
    )
assert invalid_receipt_response.status_code == 400
assert invalid_receipt_response.get_json()['message'] == '영수증 증빙파일은 JPG, PNG, GIF 이미지만 첨부할 수 있습니다.'

with open(sample_path, 'rb') as source:
    fake_image_response = client.post(
        '/expense/submit',
        data={
            'expense_org_type': '본사',
            'expense_manager': '테스트',
            'submitter_email': 'test@example.com',
            'expense_kind': '교구재비',
            'payment_account': '테스트은행 1234 테스트',
            'expense_excel': (source, '예당초_지출결의서(2026-06).xlsx'),
            'receipt_files': (BytesIO(b'not-an-image'), 'receipt.jpg'),
        },
        content_type='multipart/form-data',
    )
assert fake_image_response.status_code == 400
assert '올바른 이미지가 아닙니다' in fake_image_response.get_json()['message']

original_template_path = expense_routes.EXPENSE_TEMPLATE_PATH
expense_routes.EXPENSE_TEMPLATE_PATH = os.path.join(PROJECT_ROOT, 'missing-expense-template.xlsx')
try:
    fallback_response = client.get('/expense/template')
finally:
    expense_routes.EXPENSE_TEMPLATE_PATH = original_template_path

assert fallback_response.status_code == 200
assert fallback_response.mimetype == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
fallback_workbook = load_workbook(BytesIO(fallback_response.data), data_only=False)
fallback_sheet = fallback_workbook.active
assert fallback_sheet.title == '지출결의서'
assert [fallback_sheet.cell(4, column).value for column in range(1, 8)] == [
    '날짜', '구분', '사용내역', '사용출처', '결제수단', '지출금액', '비고'
]
assert fallback_sheet['F25'].value == '=SUM(F5:F24)'

print('Expense submit layout test: PASS')
